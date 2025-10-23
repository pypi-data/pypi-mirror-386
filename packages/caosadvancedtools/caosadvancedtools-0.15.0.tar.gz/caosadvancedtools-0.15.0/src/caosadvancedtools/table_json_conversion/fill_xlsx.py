# encoding: utf-8
#
# This file is a part of the LinkAhead Project.
#
# Copyright (C) 2024 Indiscale GmbH <info@indiscale.com>
# Copyright (C) 2024 Henrik tom Wörden <h.tomwoerden@indiscale.com>
# Copyright (C) 2024 Daniel Hornung <d.hornung@indiscale.com>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.
"""Class and function to fill an XLSX template from actual data."""

from __future__ import annotations

import datetime
import warnings
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Optional, TextIO, Union
from warnings import warn

from jsonschema.exceptions import ValidationError
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from ._validation_utils import validate_jsonschema
from .xlsx_utils import (
    array_schema_from_model_schema,
    get_foreign_key_columns,
    get_row_type_column_index,
    is_exploded_sheet,
    next_row_index,
    p2s,
    read_or_dict,
    ColumnType,
    RowType
)


class TemplateFiller:
    """Class to fill XLSX templates.  Has an index for all relevant columns."""

    def __init__(self, workbook: Workbook, graceful: bool = False):
        self._workbook = workbook
        self._graceful = graceful
        self._create_index()

    @property
    def workbook(self):
        """Return the workbook of this TemplateFiller."""
        return self._workbook

    def fill_data(self, data: dict):
        """Fill the data into the workbook."""
        self._handle_data(data=data)

    class Context:
        """Context for an entry: simple properties of all ancestors, organized in a dict.

        This is similar to a dictionary with all scalar element properties at the tree nodes up to
        the root.  Siblings in lists and dicts are ignored.  Additionally the context knows where
        its current position is.

        Lookup of elements can easily be achieved by giving the path (as ``list[str]`` or
        stringified path).

        """

        def __init__(self, current_path: Optional[list[str]] = None,
                     props: Optional[dict[str, Any]] = None):
            self._current_path = current_path if current_path is not None else []
            self._props = props if props is not None else {}  # this is flat

        def copy(self) -> TemplateFiller.Context:
            """Deep copy."""
            result = TemplateFiller.Context(current_path=self._current_path.copy(),
                                            props=self._props.copy())
            return result

        def next_level(self, next_level: str) -> TemplateFiller.Context:
            """Return a copy of this Context, with the path appended by ``next_level``."""
            result = self.copy()
            result._current_path.append(next_level)  # pylint: disable=protected-access
            return result

        def __getitem__(self, path: Union[list[str], str], owner=None) -> Any:
            if isinstance(path, list):
                path = p2s(path)
            return self._props[path]

        def __setitem__(self, propname: str, value):
            fullpath = p2s(self._current_path + [propname])
            self._props[fullpath] = value

        def fill_from_data(self, data: dict[str, Any]):
            # TODO recursive for dicts and list?
            """Fill current level with all scalar elements of ``data``."""
            for name, value in data.items():
                if not isinstance(value, (dict, list)):
                    self[name] = value
                elif isinstance(value, dict):
                    if not value or isinstance(list(value.items())[0], list):
                        continue
                    old_path = self._current_path
                    new_path = self._current_path.copy() + [name]
                    self._current_path = new_path
                    self.fill_from_data(data=value)
                    self._current_path = old_path

    def _create_index(self):
        """Create a sheet index for the workbook.

        Index the sheets by all path arrays leading to them.  Also create a simple column index by
        column type and path.

        This method creates and populates the dict ``self._sheet_index``.
        """
        self._sheet_index = {}
        for sheetname in self._workbook.sheetnames:
            sheet = self._workbook[sheetname]
            type_column = [x.value for x in list(sheet.columns)[
                get_row_type_column_index(sheet)]]
            # 0-indexed, as everything outside of sheet.cell(...):
            coltype_idx = type_column.index(RowType.COL_TYPE.name)
            path_indices = [i for i, typ in enumerate(type_column) if typ == RowType.PATH.name]

            # Get the paths, use without the leaf component for sheet indexing, with type prefix and
            # leaf for column indexing.
            for col_idx, col in enumerate(sheet.columns):
                if col[coltype_idx].value == RowType.COL_TYPE.name:
                    continue
                path = []
                for path_idx in path_indices:
                    if col[path_idx].value is not None:
                        path.append(col[path_idx].value)
                # col_key = p2s([col[coltype_idx].value] + path)
                # col_index[col_key] = SimpleNamespace(column=col, col_index=col_idx)
                if col[coltype_idx].value not in [ColumnType.SCALAR.name, ColumnType.LIST.name,
                                                  ColumnType.MULTIPLE_CHOICE.name]:
                    continue

                path_str = p2s(path)
                assert path_str not in self._sheet_index
                self._sheet_index[path_str] = SimpleNamespace(
                    sheetname=sheetname, sheet=sheet, col_index=col_idx,
                    col_type=col[coltype_idx].value)

    def _handle_data(self, data: dict, current_path: Optional[list[str]] = None,
                     context: Optional[TemplateFiller.Context] = None,
                     only_collect_insertables: bool = False,
                     utc: bool = False,
                     ) -> Optional[dict[str, Any]]:
        """Handle the data and write it into ``workbook``.

Parameters
----------
data: dict
  The data at the current path position.  Elements may be dicts, lists or simple scalar values.

current_path: list[str], optional
  If this is None or empty, we are at the top level.  This means that all children shall be entered
  into their respective sheets and not into a sheet at this level.  ``current_path`` and ``context``
  must either both be given, or none of them.

context: TemplateFiller.Context, optional
  Directory of scalar element properties at the tree nodes up to the root.  Siblings in lists
  and dicts are ignored.  ``context`` and ``current_path`` must either both be given, or none of
  them.

only_collect_insertables: bool, optional
  If True, do not insert anything on this level, but return a dict with entries to be inserted.

utc: bool, optional
  If True, store times as UTC. Else store as local time on a best-effort base.

Returns
-------
out: union[dict, None]
  If ``only_collect_insertables`` is True, return a dict (path string -> value)
        """
        # FIXME The `utc` parameter is neither used, tested nor propagated recursively.
        assert (current_path is None) is (context is None), (
            "`current_path` and `context` must either both be given, or none of them.")
        if current_path is None:
            current_path = []
        if context is None:
            context = TemplateFiller.Context()
        context.fill_from_data(data)

        insertables: dict[str, Any] = {}
        for name, content in data.items():
            # TODO is this the best way to do it????
            if name == "file":
                continue
            path = current_path + [name]
            next_context = context.next_level(name)
            # preprocessing
            if isinstance(content, list):
                if not content:  # empty list
                    continue
                # List elements must be all of the same type.
                entry_types = set(type(entry) for entry in content)
                if not len(entry_types) == 1:
                    msg = (f"More than one entry type found!  Types: {entry_types}\n"
                           f"Current path: {current_path}\n"
                           f"Current name: {name}\n"
                           f"Insertables (row) so far: {insertables}\n"
                           f"Content:\n--- Beginning of content ---\n{content}\n"
                           "--- End of content ---\n"
                           )
                    raise ValueError(msg)

                if isinstance(content[0], dict):  # all elements are dicts
                    # Heuristic to detect enum entries (only id and name):
                    if all(set(entry.keys()) == {"id", "name"} for entry in content):
                        # Convert to list of names, do not recurse
                        content = [entry["name"] for entry in content]
                    else:
                        # An array of objects: must go into exploded sheet
                        for entry in content:
                            self._handle_data(data=entry, current_path=path, context=next_context)
                        continue
            # Heuristic to detect enum entries (dict with only id and name):
            elif isinstance(content, dict) and set(content.keys()) == {"id", "name"}:
                content = [content["name"]]
            # "Normal" dicts
            elif isinstance(content, dict):  # we recurse and simply use the result
                if not current_path:  # Special handling for top level
                    self._handle_data(content, current_path=path, context=next_context)
                    continue
                insert = self._handle_data(content, current_path=path, context=next_context.copy(),
                                           only_collect_insertables=True)
                assert isinstance(insert, dict)
                assert not any(key in insertables for key in insert)
                insertables.update(insert)
                continue
            else:  # scalars
                content = [content]  # make list for unified treatment below

            # collecting the data
            assert isinstance(content, list)
            to_insert = self._try_multiple_choice(path, values=content)
            if not to_insert:
                if len(content) > 1:
                    content = [ILLEGAL_CHARACTERS_RE.sub("", str(x)) for x in content]
                    value = ";".join(content)  # TODO we need escaping of values
                else:
                    value = content[0]
                    if isinstance(value, str):
                        value = ILLEGAL_CHARACTERS_RE.sub("", value)
                    if isinstance(value, datetime.datetime):
                        if value.tzinfo is not None:
                            if utc:
                                value = value.astimezone(datetime.timezone.utc).replace(tzinfo=None)
                            else:
                                # Remove timezone, store in local timezone.
                                value = value.astimezone().replace(tzinfo=None)

                path_str = p2s(path)
                assert path_str not in insertables
                to_insert = {path_str: value}
            insertables.update(to_insert)
        if only_collect_insertables:
            return insertables
        if not current_path:  # Top level returns, because there are only sheets for the children.
            return None

        # actual data insertion
        insert_row = None
        sheet = None
        for path_str, value in insertables.items():
            if self._graceful and path_str not in self._sheet_index:
                if not (value is None or path_str.endswith(".id") or path_str.endswith(".name")):
                    warn(f"Ignoring path with missing sheet index: {path_str}")
                continue
            sheet_meta = self._sheet_index[path_str]
            if sheet is None:
                sheet = sheet_meta.sheet
            assert sheet is sheet_meta.sheet, "All entries must be in the same sheet."
            col_index = sheet_meta.col_index
            if insert_row is None:
                insert_row = next_row_index(sheet)

            sheet.cell(row=insert_row+1, column=col_index+1, value=value)

        # Insert foreign keys
        if insert_row is not None and sheet is not None and is_exploded_sheet(sheet):
            try:
                foreigns = get_foreign_key_columns(sheet)
            except ValueError:
                print(f"Sheet: {sheet}")
                raise
            for index, path in ((f.index, f.path) for f in foreigns.values()):
                value = context[path]
                sheet.cell(row=insert_row+1, column=index+1, value=value)

        return None

    def _try_multiple_choice(self, path: list[str], values: list[str]) -> Optional[dict[str, str]]:
        """Try to create sheet content for a multiple choice property.

Parameters
----------
path: list[str]
  The Path to this property.
values: list[str]
  A list of possible choices, should be unique.

Returns
-------
to_insert: Optional[dict[str, str]]
  A path-value dict.  None if this doesn't seem to be a multiple choice data set.
        """
        try:
            assert len(set(values)) == len(values)
            to_insert = {}
            found_sheet = None
            for value in values:
                assert isinstance(value, str)
                path_str = p2s(path + [value])
                assert path_str in self._sheet_index
                sheet_meta = self._sheet_index[path_str]
                # All matches shall be on the same sheet
                assert found_sheet is None or found_sheet == sheet_meta.sheetname
                found_sheet = sheet_meta.sheetname
                # Correct type
                assert sheet_meta.col_type == ColumnType.MULTIPLE_CHOICE.name
                to_insert[path_str] = "x"
        except AssertionError:
            return None
        return to_insert


def fill_template(data: Union[dict, str, TextIO], template: str, result: Union[str, Path],
                  validation_schema: Optional[Union[dict, str, TextIO]] = None) -> None:
    """Insert json data into an xlsx file, according to a template.

This function fills the json data into the template stored at ``template`` and stores the result as
``result``.

Parameters
----------
data: Union[dict, str, TextIO]
  The data, given as Python dict, path to a file or a file-like object.
template: str
  Path to the XLSX template.
result: str
  Path for the result XLSX.
validation_schema: dict, optional
  If given, validate the date against this schema first.  This raises an exception if the validation
  fails.  If no validation schema is given, try to ignore more errors in the data when filling the
  XLSX template.

Possible future development
---------------------------

- Add an ``allow_missing_foreign`` option, which would replace ``props[key]`` by ``props.get(key)``.
"""
    data = read_or_dict(data)
    assert isinstance(data, dict)

    # Validation
    if validation_schema is not None:
        validation_schema = read_or_dict(validation_schema)
        assert isinstance(validation_schema, dict)

        # convert to array_schema if given schema is a model_schema
        if 'properties' in validation_schema and validation_schema['properties'].values():
            if list(validation_schema['properties'].values())[0]["type"] != "array":
                validation_schema = array_schema_from_model_schema(read_or_dict(validation_schema))
        try:
            validate_jsonschema(data, validation_schema)
        except ValidationError as verr:
            print(verr.message)
            raise verr
    else:
        warnings.warn("No validation schema given, continue at your own risk.")

    # Filling the data
    result_wb = load_workbook(template)
    template_filler = TemplateFiller(result_wb, graceful=(validation_schema is None))
    template_filler.fill_data(data=data)

    if not isinstance(result, Path):
        result = Path(result)
    result.parent.mkdir(parents=True, exist_ok=True)
    result_wb.save(result)
