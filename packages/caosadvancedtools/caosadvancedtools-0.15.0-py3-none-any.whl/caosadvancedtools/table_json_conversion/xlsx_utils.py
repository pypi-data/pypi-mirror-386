# encoding: utf-8
#
# This file is a part of the LinkAhead Project.
#
# Copyright (C) 2024 Indiscale GmbH <info@indiscale.com>
# Copyright (C) 2024 Daniel Hornung <d.hornung@indiscale.com>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

"""General utilities to work with XLSX files with (hidden) column and row annotations and typing.

The most prominent functions are:

- ``p2s``: Path to string: ``["some", "path"] -> "some.path"``
- ``read_or_dict``: Load JSON object from path, file or dict.

This module also defines these enums:

- ColumnType
- RowType
"""

import json

from collections import OrderedDict
from copy import deepcopy
from enum import Enum
from types import SimpleNamespace
from typing import Any, Iterable, TextIO, Union

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


TRUTHY = {
    "true",
    "wahr",
    "x",
    "√",
    "yes",
    "ja",
    "y",
    "j",
}  # For multiple choice columns
FALSY = {"false", "falsch", "-", "no", "nein", "n"}  # For multiple choice columns


class ColumnType(Enum):
    """column types enum"""

    SCALAR = 1
    LIST = 2
    FOREIGN = 3
    MULTIPLE_CHOICE = 4
    IGNORE = 5


class RowType(Enum):
    """row types enum"""

    COL_TYPE = 1
    PATH = 2
    IGNORE = 3


def array_schema_from_model_schema(model_schema: dict) -> dict:
    """Convert a *data model* schema to a *data array* schema.

    Practically, this means that the top level properties are converted into lists.  In a simplified
    notation, this can be expressed as:

    ``array_schema = { elem: [elem typed data...] for elem in model_schema }``

    Parameters
    ----------
    model_schema: dict
      The schema description of the data model.  Must be a json schema *object*, with a number of
      *object* typed properties.

    Returns
    -------
    array_schema: dict
      A corresponding json schema, where the properties are arrays with the types of the input's
      top-level properties.
    """
    assert model_schema["type"] == "object"
    result = deepcopy(model_schema)
    for name, prop in result["properties"].items():
        assert prop["type"] == "object"
        new_prop = {"type": "array", "items": prop}
        result["properties"][name] = new_prop
    return result


def get_defining_paths(workbook: Workbook) -> dict[str, list[list[str]]]:
    """For all sheets in ``workbook``, list the paths which they define.

    A sheet is said to define a path, if it has data columns for properties inside that path.  For
    example, consider the following worksheet:

    +------------+-------------------+---------------+--------------+--------------------+
    | COL_TYPE   | SCALAR            | SCALAR        | LIST         | SCALAR             |
    +------------+-------------------+---------------+--------------+--------------------+
    | PATH       | Training          | Training      | Training     | Training           |
    +------------+-------------------+---------------+--------------+--------------------+
    | PATH       | url               | date          | subjects     | supervisor         |
    +------------+-------------------+---------------+--------------+--------------------+
    | PATH       |                   |               |              | email              |
    +------------+-------------------+---------------+--------------+--------------------+
    |            | example.com/mp    | 2024-02-27    | Math;Physics | steve@example.com  |
    +------------+-------------------+---------------+--------------+--------------------+
    |            | example.com/m     | 2024-02-27    | Math         | stella@example.com |
    +------------+-------------------+---------------+--------------+--------------------+

    This worksheet defines properties for the paths `["Training"]` and `["Training", "supervisor"]`, and
    thus these two path lists would be returned for the key with this sheet's sheetname.

    Parameters
    ----------
    workbook: Workbook
      The workbook to analyze.

    Returns
    -------
    out: dict[str, list[list[str]]
      A dict with worksheet names as keys and lists of paths (represented as string lists) as values.
    """
    result: dict[str, list[list[str]]] = {}
    for sheet in workbook.worksheets:
        paths = []
        added = set()
        for col in get_data_columns(sheet).values():
            rep = p2s(col.path[:-1])
            if rep not in added:
                paths.append(col.path[:-1])
                added.add(rep)
        result[sheet.title] = paths
    return result


def get_data_columns(sheet: Worksheet) -> dict[str, SimpleNamespace]:
    """Return the data paths of the worksheet.

    Returns
    -------
    out: dict[str, SimpleNamespace]
      The keys are the stringified paths.  The values are SimpleNamespace objects with ``index``,
      ``path`` and ``column`` attributes.
    """
    column_types = _get_column_types(sheet)
    path_rows = get_path_rows(sheet)
    result = OrderedDict()
    for for_idx, name in column_types.items():
        if name not in (
            ColumnType.SCALAR.name,
            ColumnType.LIST.name,
            ColumnType.MULTIPLE_CHOICE.name,
        ):
            continue
        path = []
        for row in path_rows:
            component = sheet.cell(row=row + 1, column=for_idx + 1).value
            if component is None:
                break
            assert isinstance(component, str), f"Expected string: {component}"
            path.append(component)
        result[p2s(path)] = SimpleNamespace(
            index=for_idx, path=path, column=list(sheet.columns)[for_idx]
        )
    return result


def get_foreign_key_columns(sheet: Worksheet) -> dict[str, SimpleNamespace]:
    """Return the foreign keys of the worksheet.

    Returns
    -------
    out: dict[str, SimpleNamespace]
      The keys are the stringified paths.  The values are SimpleNamespace objects with ``index``,
      ``path`` and ``column`` attributes.
    """
    column_types = _get_column_types(sheet)
    path_rows = get_path_rows(sheet)
    result = OrderedDict()
    for for_idx, name in column_types.items():
        if name != ColumnType.FOREIGN.name:
            continue
        path = []
        for row in path_rows:
            component = sheet.cell(row=row + 1, column=for_idx + 1).value
            if component is None:
                break
            assert isinstance(component, str), f"Expected string: {component}"
            path.append(component)
        result[p2s(path)] = SimpleNamespace(
            index=for_idx, path=path, column=list(sheet.columns)[for_idx]
        )
    return result


def get_path_position(sheet: Worksheet) -> tuple[list[str], str]:
    """Return a path which represents the parent element, and the sheet's "proper name".

    For top-level sheets / entries (those without foreign columns), the path is an empty list.

    A sheet's "proper name" is detected from the data column paths: it is the first component after the
    parent components.

    Returns
    -------
    parent: list[str]
      Path to the parent element.  Note that there may be list elements on the path which are **not**
      represented in this return value.

    proper_name: str
      The "proper name" of this sheet.  This defines an array where all the data lives, relative to the
      parent path.
    """
    # Parent element: longest common path shared among any foreign column and all the data columns
    parent: list[str] = []

    # longest common path in data colums
    data_paths = [el.path for el in get_data_columns(sheet).values()]
    for ii in range(min(len(path) for path in data_paths)):
        components_at_index = {path[ii] for path in data_paths}
        if len(components_at_index) > 1:
            break
    longest_data_path = data_paths[0][:ii]

    # longest common overall path
    foreign_paths = [el.path for el in get_foreign_key_columns(sheet).values()]
    ii = 0  # If no foreign_paths, proper name is the first element
    for foreign_path in foreign_paths:
        for ii in range(min([len(foreign_path), len(longest_data_path)])):
            components_at_index = {foreign_path[ii], longest_data_path[ii]}
            if len(components_at_index) > 1:
                break
        if ii > len(parent):
            parent = foreign_path[:ii]

    return parent, data_paths[0][ii]


def get_path_rows(sheet: Worksheet):
    """Return the 0-based indices of the rows which represent paths."""
    rows = []
    rt_col = get_row_type_column_index(sheet)
    for cell in list(sheet.columns)[rt_col]:
        if cell.value == RowType.PATH.name:
            rows.append(cell.row - 1)
    return rows


def get_row_type_column_index(sheet: Worksheet):
    """Return the column index (0-indexed) of the column which defines the row types."""
    for col in sheet.columns:
        for cell in col:
            if cell.value == RowType.COL_TYPE.name:
                return cell.column - 1
    raise ValueError(
        "The column which defines row types (COL_TYPE, PATH, ...) is missing"
    )


def get_column_type_row_index(sheet: Worksheet):
    """Return the row index (0-indexed) of the row which defines the column types."""
    for row in sheet.rows:
        for cell in row:
            if cell.value == RowType.COL_TYPE.name:
                return cell.row - 1
    raise ValueError(
        "The column which defines row types (COL_TYPE, SCALAR, ...) is missing"
    )


def get_subschema(path: list[str], schema: dict) -> dict:
    """Return the sub schema at ``path``."""
    if path:
        if schema["type"] == "object":
            next_schema = schema["properties"][path[0]]
            return get_subschema(path=path[1:], schema=next_schema)
        if schema["type"] == "array":
            items = schema["items"]
            if "enum" in items:
                return schema
            next_schema = items["properties"][path[0]]
            return get_subschema(path=path[1:], schema=next_schema)
    return schema


def get_worksheet_for_path(
    path: list[str], defining_path_index: dict[str, list[list[str]]]
) -> str:
    """Find the sheet name which corresponds to the given path."""
    for sheetname, paths in defining_path_index.items():
        if path in paths:
            return sheetname
    raise KeyError(f"Could not find defining worksheet for path: {path}")


def is_exploded_sheet(sheet: Worksheet) -> bool:
    """Return True if this is a an "exploded" sheet.

    An exploded sheet is a sheet whose data entries are LIST valued properties of entries in another
    sheet.  A sheet is detected as exploded if and only if it has FOREIGN columns.
    """
    column_types = _get_column_types(sheet)
    return ColumnType.FOREIGN.name in column_types.values()


def is_recursively_none(obj: Union[list, dict, None] = None):
    """Test if ``obj`` is None or recursively consists only of None-like objects.

    Parameters
    ----------
    obj: Union[list, dict, None]
      The object to be tested, a json like dict or array, or None.
    """
    if obj is None:
        return True
    if isinstance(obj, (list, dict)):
        if isinstance(obj, list):
            mylist: Iterable = obj
        else:
            mylist = obj.values()
        for element in mylist:
            if not is_recursively_none(element):
                return False
        return True
    return False


def next_row_index(sheet: Worksheet) -> int:
    """Return the index for the next data row.

    This is defined as the first row without any content.
    """
    return sheet.max_row


def p2s(path: list[str]) -> str:
    """Path to string: dot-separated.

    Special treatment for jumping levels
    ------------------------------------

    Starting form the last element, if any element starts with one or more dots ".", one preceding
    element will be *removed* for each dot.  For example, ``["A", "B1.B2", ".C", "..D.E"]`` will result
    in ``"A.D.E"`` (the elements ``B1.B2`` and ``.C`` are removed by the two leading dots in ``..D.E``).
    """
    pruned_path_reversed = []
    to_delete = []
    for idx, elem in reversed(list(enumerate(path))):
        if idx in to_delete:
            continue
        while elem.startswith("."):
            elem = elem[1:]
            idx -= 1
            to_delete.append(idx)
        pruned_path_reversed.append(elem)
    return ".".join(reversed(pruned_path_reversed))


def parse_multiple_choice(value: Any) -> bool:
    """Interpret ``value`` as a multiple choice input.

    *Truthy* values are:
    - The boolean ``True``.
    - The number "1".
    - The (case-insensitive) strings ``true``, ``wahr``, ``x``, ``√``, ``yes``, ``ja``, ``y``, ``j``.

    *Falsy* values are:
    - The boolean ``False``.
    - ``None``, empty strings, lists, dicts.
    - The number "0".
    - The (case-insensitive) strings ``false``, ``falsch``, ``-``, ``no``, ``nein``, ``n``.
    - Everything else.

    Returns
    -------
    out: bool
      The interpretation result of ``value``.
    """
    # Non-string cases first:
    # pylint: disable-next=too-many-boolean-expressions
    if (
        value is None
        or value is False
        or value == 0
        or value == []
        or value == {}
        or value == ""
    ):
        return False
    if value is True or value == 1:
        return True

    # String cases follow:
    if not isinstance(value, str):
        return False
    value = value.lower()

    if value in TRUTHY:
        return True

    # Strictly speaking, this test is not necessary, but I think it's good practice.
    if value in FALSY:
        return False
    return False


def prune_none_objects(obj: Union[list, dict]) -> Union[list, dict]:
    """Recursively replace elements of ``obj``, that are empty or recursively None, by None.

    Parameters
    ----------
    obj : Union[list, dict]
      The object to be tested, a json like dict or array.

    Returns
    -------
    out : Union[list, dict]
      A deep copy, with the None objects and empty sequences replaced by None.

    Detailed explanation
    --------------------
    - Elements of lists that are recursively None, shall be removed from the list.
    - Elements that are empty lists / dicts are replaced by None.
    """

    # FIXME Potential performance problem, because is_recursively_none is called multiple times on
    # the same objects.
    result: Union[list, dict]
    if isinstance(obj, list):
        result = []
        for value in obj:
            if isinstance(value, (dict, list)):  # json object? recurse!
                if is_recursively_none(value):
                    continue
                pruned = prune_none_objects(value)
                if not (pruned == [] or pruned == {}):
                    result.append(pruned)
            else:  # else: just assign
                result.append(value)
    elif isinstance(obj, dict):
        result = {}
        for key, value in obj.items():
            if isinstance(value, (dict, list)):
                if is_recursively_none(value):
                    result[key] = None
                    continue
                pruned = prune_none_objects(value)
                if not (pruned == [] or pruned == {}):
                    result[key] = pruned
            else:  # else: just assign
                result[key] = value
    else:
        raise ValueError(f"Object must be either list or dict, but is: {type(obj)}")
    return result


def read_or_dict(data: Union[dict, str, TextIO]) -> dict:
    """If data is a json file name or input stream, read data from there.
    If it is a dict already, just return it."""
    if isinstance(data, dict):
        return data

    if isinstance(data, str):
        with open(data, encoding="utf-8") as infile:
            data = json.load(infile)
    elif hasattr(data, "read"):
        data = json.load(data)
    else:
        raise ValueError(
            f"I don't know how to handle the datatype of `data`: {type(data)}"
        )
    assert isinstance(data, dict)
    return data


def _get_column_types(sheet: Worksheet) -> OrderedDict:
    """Return an OrderedDict: column index -> column type for the sheet."""
    result = OrderedDict()
    type_row_index = get_row_type_column_index(sheet)
    for idx, col in enumerate(sheet.columns):
        type_cell = col[type_row_index]
        result[idx] = (
            type_cell.value if type_cell.value is not None else (ColumnType.IGNORE.name)
        )
        assert (
            hasattr(ColumnType, result[idx]) or result[idx] == RowType.COL_TYPE.name
        ), f"Unexpected column type value ({idx}{type_row_index}): {type_cell.value}"
    return result
