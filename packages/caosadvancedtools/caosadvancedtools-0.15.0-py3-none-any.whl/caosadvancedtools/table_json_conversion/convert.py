# encoding: utf-8
#
# This file is a part of the LinkAhead Project.
#
# Copyright (C) 2024 Indiscale GmbH <info@indiscale.com>
# Copyright (C) 2024 Daniel Hornung <d.hornung@indiscale.com>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

"""Convert XLSX files to JSON dictionaries."""

from __future__ import annotations

import datetime
import itertools
import sys
import textwrap
from functools import lru_cache, reduce
from operator import getitem
from types import SimpleNamespace
from typing import Any, BinaryIO, Callable, TextIO, Union, Optional
from warnings import warn

import jsonschema
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import linkahead as db
from ._validation_utils import validate_jsonschema
from caosadvancedtools.table_json_conversion import xlsx_utils
from caosadvancedtools.table_json_conversion.fill_xlsx import read_or_dict
from linkahead.cached import cached_get_entity_by, DEFAULT_SIZE
from linkahead.common.datatype import get_list_datatype, is_list_datatype


def _strict_bool(value: Any) -> bool:
    """Convert value to bool, but only if it really is a valid XLSX bool."""
    if isinstance(value, bool):
        return value
    raise TypeError(f"Not a good boolean: {repr(value)}")


def _column_id_to_chars(num):
    """Converts a column id (zero based) to the corresponding string
    representation, e.g. 0 -> 'A', 97 -> 'CT'"""
    if num < 0:
        return ""
    return _column_id_to_chars(int(num / 26) - 1) + chr(int(num % 26) + 65)


def _format_exception_table(exceptions: list[tuple], worksheet_title: str,
                            column_names: Optional[Union[dict, list]] = None,
                            max_line_length: int = 120) -> str:
    """
    Given a list of tuples containing a row and column number as well as an
    exception in that order, and the title of the current worksheet, returns
    a formatted table of the exceptions.

    Optionally takes a dict of column names, if given a header will be
    generated for each column and exceptions will be clustered by column.

    Default line length is 120 and can be overwritten by max_line_length.

    Params
    ------
    exceptions:         list of tuples containing row, column, and exception
                        Data to be formatted
    worksheet_title:    str
                        Name of the current worksheet
    column_names:       dict or list, optional
                        column_names[column_num] should return the name of
                        column column_names.
                        If given, exceptions will be clustered by column.
    max_line_length:    int, default=120
                        Soft cap for the line length of the resulting table

    Return
    ------
    string_rep:         str
                        Table containing the given exceptions
    """
    max_line_length -= 40             # Estimate of Field + Type space use

    headers = {"loc": "Location", "type": "Error Type", "mess": ["Message"]}
    lengths = {key: len(headers[key]) for key in headers}
    new_data = []

    current_column = None
    exceptions.sort(key=lambda tup: tup[1])
    for row_i, col_i, excep in exceptions:
        if column_names is not None:
            # Add a line with information about the current column
            if current_column != col_i:
                current_column = col_i
                new_data.append({
                    "loc": f"\nErrors in column '{column_names[col_i]}':",
                    "type": "", "mess": [""]
                })
        # Setup for current Exception
        curr_err_data: Any = {}
        new_data.append(curr_err_data)
        # Get field
        if isinstance(row_i, int):
            curr_err_data["loc"] = f"Cell {_column_id_to_chars(col_i)}{row_i + 1}"
        else:
            curr_err_data["loc"] = f"Column {_column_id_to_chars(col_i)}"
        lengths["loc"] = max(lengths["loc"], len(curr_err_data["loc"]))
        # Add error code
        curr_err_data["type"] = type(excep).__name__
        lengths["type"] = max(lengths["type"], len(curr_err_data["type"]))
        # Format message - split into lines
        lines = str(excep).split('\n')
        new_lines = []
        for line in lines:
            new_lines += textwrap.wrap(line, max_line_length, break_long_words=False)
        for line in new_lines:
            lengths["mess"] = max(lengths["mess"], len(line))
        if new_lines == []:
            new_lines = [""]
        curr_err_data["mess"] = new_lines

    # Generate underline for each header
    dividers: dict[str, Union[str, list]] = {key: '–' * l for key, l in lengths.items()}
    dividers["mess"] = [dividers["mess"]]
    # Fill with spaces for alignment
    string_rep = f"There were errors during the validation of worksheet '{worksheet_title}':\n\n"
    for curr_err_data in [headers, dividers] + new_data:
        string_rep += ' {loc: <{fill}}  '.format(loc=curr_err_data["loc"],
                                                 fill=lengths["loc"])
        string_rep += ' {typ: <{fill}}  '.format(typ=curr_err_data["type"],
                                                 fill=lengths["type"])
        # Fill for the messages is set to 0, if we want another column or align
        # right we need to use lengths["mess"]
        string_rep += ' {mes: <{fill}}\n'.format(mes=curr_err_data["mess"][0], fill=0)
        for line in curr_err_data["mess"][1:]:
            # Front padding for lines without location and error type
            string_rep += ' ' * (lengths["loc"] + lengths["type"] + 6)
            string_rep += ' {mes: <{fill}}\n'.format(mes=line, fill=0)
    return string_rep


class ForeignError(KeyError):
    def __init__(self, *args, definitions: list, message: str = ""):
        super().__init__(message, *args)
        self.definitions = definitions


class XLSXConverter:
    """Class for conversion from XLSX to JSON.

    For a detailed description of the required formatting of the XLSX files, see ``specs.md`` in the
    documentation.
    """
    PARSER: dict[str, Callable] = {
        "string": str,
        "number": float,
        "integer": int,
        "boolean": _strict_bool,
    }

    def __init__(self, xlsx: Union[str, BinaryIO], schema: Union[dict, str, TextIO],
                 strict: bool = False):
        """
        Parameters
        ----------
        xlsx: Union[str, BinaryIO]
          Path to the XLSX file or opened file object.

        schema: Union[dict, str, TextIO]
          Schema for validation of XLSX content.

        strict: bool, default=False
          If True, fail faster.
        """
        self._workbook = load_workbook(xlsx)
        self._schema = read_or_dict(schema)
        self._defining_path_index = xlsx_utils.get_defining_paths(self._workbook)
        self._check_path_validity()
        self._check_columns(fail_fast=strict)
        self._handled_sheets: set[str] = set()
        self._result: dict = {}
        self._errors: dict = {}

    def _check_path_validity(self):
        """
        Method to check the workbook paths for completeness and correctness,
        and raises a jsonschema.ValidationError containing information on all
        faulty paths if any are found.

        If this method does not raise an error, this does not mean the workbook
        is formatted correctly, only that the contained paths are complete and
        can be found in the schema.
        """
        # Setup
        error_message = ["There were errors during path validation:"]
        only_warnings = True
        for sheetname in self._workbook.sheetnames:
            sheet = self._workbook[sheetname]
            error_message.append(f"\nIn sheet {sheetname}:")

            # Collect path information and filter out information column
            row_i_col_type = xlsx_utils.get_column_type_row_index(sheet)
            path_rows = xlsx_utils.get_path_rows(sheet)
            paths = []
            for col_i, col in enumerate(sheet.iter_cols()):
                col_type = col[row_i_col_type].value
                path = [col[row_i].value for row_i in path_rows
                        if col[row_i].value not in [None, '']]
                if col_type == 'COL_TYPE':
                    continue
                paths.append((col_type, path, col_i, col))

            # Check paths
            for col_type, path, col_i, col in paths:
                # No column type set
                if col_type in [None, '']:
                    if len(path) == 0:              # Likely a comment column
                        # Check whether the column has any visible content
                        content_in_column = False
                        for cell in col:
                            visible_content = ''.join(str(cell.value)).split()
                            if cell.value is not None and visible_content != '':
                                content_in_column = True
                        # If yes - might be an error but is not forbidden, so warn
                        if content_in_column:
                            m = (f"Warning:\tIn column {_column_id_to_chars(col_i)} "
                                 f"there is no column metadata set. This column "
                                 f"will be ignored during parsing.")
                            error_message.append(m)
                        continue
                    else:                    # Path is set but no column type
                        only_warnings = False
                        m = (f"ERROR:\t\tIn column {_column_id_to_chars(col_i)} "
                             f"the column type is missing.")
                        error_message.append(m)
                        # No continue - even if column type is missing, we can check path
                if len(path) == 0:           # Column type is set but no path
                    only_warnings = False
                    m = (f"ERROR:\t\tIn column {_column_id_to_chars(col_i)} "
                         f"the path is missing.")
                    error_message.append(m)
                    continue
                # Check path is in schema
                try:
                    subschema = xlsx_utils.get_subschema(path, self._schema)
                    schema_type = subschema.get('type', None)
                    if schema_type is None and 'enum' in subschema:
                        schema_type = 'enum'
                    if schema_type is None and 'anyOf' in subschema:
                        schema_type = 'anyOf'
                    if schema_type == 'array':      # Check item type instead
                        schema_type = subschema.get('items', {}).get('type', None)
                    if schema_type in ['object', 'array', None]:
                        m = (f"Warning:\tIn column {_column_id_to_chars(col_i)} "
                             f"the path may be incomplete.")
                        error_message.append(m)
                except KeyError as e:
                    only_warnings = False
                    m = (f"ERROR:\t\tIn column {_column_id_to_chars(col_i)} "
                         f"parsing of the path '{'.'.join(path)}' fails "
                         f"on the path component {str(e)}.\n\t\t\t"
                         f"This likely means the path is incomplete or not "
                         f"present in the schema.")
                    error_message.append(m)

            # Cleanup if no errors were found
            if error_message[-1] == f"\nIn sheet {sheetname}:":
                error_message.pop(-1)

        # Determine whether error / warning / nothing should be raised
        if error_message == ["There were errors during path validation:"]:
            return
        error_message = '\n'.join(error_message)
        if only_warnings:
            warn(error_message)
        else:
            raise jsonschema.ValidationError(error_message)

    def to_dict(self, validate: bool = False, collect_errors: bool = True) -> dict:
        """Convert the xlsx contents to a dict.

        Parameters
        ----------
        validate: bool, default=False
          If True, validate the result against the schema.

        collect_errors: bool, default=True
          If True, do not fail at the first error, but try to collect as many errors as possible.
          After an exception is raised, the errors can be collected with ``get_errors()`` and
          printed with ``get_error_str()``.

        Returns
        -------
        out: dict
          A dict representing the JSON with the extracted data.
        """
        self._handled_sheets = set()
        self._result = {}
        self._errors = {}
        if not collect_errors:
            for sheetname in self._workbook.sheetnames:
                if sheetname not in self._handled_sheets:
                    self._handle_sheet(self._workbook[sheetname], fail_later=collect_errors)
        else:
            # Collect errors from converting
            exceptions = []
            for sheetname in self._workbook.sheetnames:
                if sheetname not in self._handled_sheets:
                    try:
                        self._handle_sheet(self._workbook[sheetname], fail_later=collect_errors)
                    except jsonschema.ValidationError as e:
                        exceptions.append(e)
                        # do not collect errors from sheet again
                        self._handled_sheets.add(sheetname)
            if len(exceptions) == 1:
                raise exceptions[0]
            elif len(exceptions) > 1:
                mess = "There were errors during the validation of several worksheets:\n\n"
                mess += '\n\n'.join([str(e).replace(
                    "There were errors during the validation of worksheet",
                    "In worksheet")
                                     for e in exceptions])
                raise jsonschema.ValidationError(mess)
        if validate:
            validate_jsonschema(self._result, self._schema)
        if self._errors:
            raise RuntimeError("There were error while handling the XLSX file.")
        self._result = xlsx_utils.prune_none_objects(self._result)
        return self._result

    def get_errors(self) -> dict:
        """Return a dict with collected errors."""
        return self._errors

    def get_error_str(self) -> str:
        """Return a beautiful string with the collected errors."""
        result = ""
        for loc, value in self._errors.items():
            result += f"Sheet: {loc[0]}\tRow: {loc[1] + 1}\n"
            for item in value:
                result += f"\t\t{item[:-1]}:\t{item[-1]}\n"
        return result

    def _check_columns(self, fail_fast: bool = False):
        """Check if the columns correspond to the schema."""
        def missing(path):
            message = f"Missing column: {xlsx_utils.p2s(path)}"
            if fail_fast:
                raise ValueError(message)
            else:
                warn(message)
        for sheetname in self._workbook.sheetnames:
            sheet = self._workbook[sheetname]
            parents: dict = {}
            col_paths = []
            for col in xlsx_utils.get_data_columns(sheet).values():
                parents[xlsx_utils.p2s(col.path[:-1])] = col.path[:-1]
                col_paths.append(col.path)
            for path in parents.values():
                try:
                    subschema = xlsx_utils.get_subschema(path, self._schema)
                except KeyError as kerr:
                    kerr.args = (sheetname, *kerr.args)
                    raise
                # Unfortunately, there are a lot of special cases to handle here.
                if subschema.get("type") == "array":
                    subschema = subschema["items"]
                if "enum" in subschema:  # Was handled in parent level already
                    continue
                for child, content in subschema["properties"].items():
                    child_path = path + [child]
                    if content == {'type': 'string', 'format': 'data-url'}:
                        continue  # skip files
                    if content.get("type") == "array" and (
                            content.get("items").get("type") == "object"):
                        if child_path not in itertools.chain(*self._defining_path_index.values()):
                            missing(child_path)
                    elif content.get("type") == "array" and "enum" in content.get("items", []) and (
                            content.get("uniqueItems") is True):
                        # multiple choice
                        for choice in content["items"]["enum"]:
                            if child_path + [choice] not in col_paths:
                                missing(child_path + [choice])
                    elif content.get("type") == "object":
                        pass
                    else:
                        if child_path not in col_paths:
                            missing(child_path)

    def _handle_sheet(self, sheet: Worksheet, fail_later: bool = False) -> None:
        """Add the contents of the sheet to the result (stored in ``self._result``).

        Each row in the sheet corresponds to one entry in an array in the result.  Which array
        exactly is defined by the sheet's "proper name" and the content of the foreign columns.

        Look at ``xlsx_utils.get_path_position`` for the specification of the "proper name".


        Parameters
        ----------
        fail_later: bool, optional
          If True, do not fail with unresolvable foreign definitions, but collect all errors.
        """
        row_type_column = xlsx_utils.get_row_type_column_index(sheet)
        foreign_columns = xlsx_utils.get_foreign_key_columns(sheet)
        foreign_column_paths = {col.index: col.path for col in foreign_columns.values()}
        data_columns = xlsx_utils.get_data_columns(sheet)
        data_column_paths = {col.index: col.path for col in data_columns.values()}
        # Parent path, insert in correct order.
        try:
            parent, proper_name = xlsx_utils.get_path_position(sheet)
        except UnboundLocalError as e:
            raise jsonschema.ValidationError(f"Malformed metadata: Cannot parse "
                                             f"paths in worksheet '{sheet.title}'.") from e
        if parent:
            parent_sheetname = xlsx_utils.get_worksheet_for_path(parent, self._defining_path_index)
            if parent_sheetname not in self._handled_sheets:
                self._handle_sheet(self._workbook[parent_sheetname], fail_later=fail_later)

        # # We save single entries in lists, indexed by their foreign key contents.  Each entry
        # # consists of:
        # # - foreign: Dict with path -> value for the foreign columns
        # # - data: The actual data of this entry, a dict.
        # entries: dict[str, list[SimpleNamespace]] = {}

        exceptions: list[tuple] = []
        col_names = {}
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            # Skip non-data rows
            if row[row_type_column] is not None:
                continue
            foreign_repr = ""
            foreign = []  # A list of lists, each of which is: [path1, path2, ..., leaf, value]
            data: dict = {}     # Local data dict
            # Collect data (in dict relative to current level) and foreign data information
            for col_idx, value in enumerate(row):
                if col_idx in foreign_column_paths:
                    foreign_repr += str(value)
                    if foreign_column_paths[col_idx][-1] == "id":
                        value = str(value)
                    foreign.append(foreign_column_paths[col_idx] + [value])
                    continue

                try:
                    if col_idx in data_column_paths:
                        path = data_column_paths[col_idx]
                        col_names[col_idx] = '.'.join(path)
                        if self._is_multiple_choice(path):
                            # `path` will be modified, so we need a copy.
                            path = path.copy()
                            real_value = path.pop()  # Last component is enum value, insert above
                            # set up list
                            try:
                                _set_in_nested(mydict=data, path=path, value=[], prefix=parent,
                                               skip=1)
                            except ValueError as err:
                                if not str(err).startswith("There is already some value at"):
                                    raise
                            if not xlsx_utils.parse_multiple_choice(value):
                                continue
                            _set_in_nested(mydict=data, path=path, value=real_value, prefix=parent,
                                           skip=1, append_to_list=True)
                        else:
                            if path[-1] == "id" and isinstance(value, int):
                                value = str(value)
                            value = self._validate_and_convert(value, path)
                            _set_in_nested(mydict=data, path=path, value=value, prefix=parent,
                                           skip=1)
                        continue
                except (ValueError, KeyError, jsonschema.ValidationError) as e:
                    # Append error for entire column only once
                    if isinstance(e, KeyError) and 'column' in str(e):
                        if len([err for ri, ci, err in exceptions
                                if ci == col_idx and isinstance(err, KeyError)]) == 0:
                            exceptions.append((None, col_idx, e))
                    else:
                        exceptions.append((row_idx, col_idx, e))

            try:
                # Find current position in tree
                parent_dict = self._get_parent_dict(parent_path=parent, foreign=foreign)

                # Append data to current position's list
                if proper_name not in parent_dict:
                    parent_dict[proper_name] = []
                parent_dict[proper_name].append(data)
            except ForeignError as kerr:
                if not fail_later:
                    raise
                self._errors[(sheet.title, row_idx)] = kerr.definitions

        if exceptions:
            exception_table = _format_exception_table(exceptions, sheet.title,
                                                      col_names)
            raise jsonschema.ValidationError(exception_table)

        self._handled_sheets.add(sheet.title)

    def _is_multiple_choice(self, path: list[str]) -> bool:
        """Test if the path belongs to a multiple choice section."""
        if not path:
            return False
        subschema = self._get_subschema(path[:-1])
        if (subschema["type"] == "array"
                and subschema.get("uniqueItems") is True
                and "enum" in subschema["items"]):
            return True
        return False

    def _get_parent_dict(self, parent_path: list[str], foreign: list[list]) -> dict:
        """Return the dict into which values can be inserted.

        This method returns, from the current result-in-making, the entry at ``parent_path`` which
        matches the values given in the ``foreign`` specification.
        """
        foreign_groups = _group_foreign_paths(foreign, common=parent_path)

        current_object = self._result
        for group in foreign_groups:
            # Find list for which foreign definitions are relevant.
            current_object = reduce(getitem, group.subpath, current_object)
            assert isinstance(current_object, list)
            # Test all candidates.
            for cand in current_object:
                if all(reduce(getitem, definition[:-1], cand) == definition[-1]
                       for definition in group.definitions):
                    current_object = cand
                    break
            else:
                message = f"Cannot find an element at {parent_path} for these foreign defs:\n"
                for name, value in group.definitions:
                    message += f"    {name}: {value}\n"
                print(message, file=sys.stderr)
                error = ForeignError(definitions=group.definitions, message=message)
                raise error
        assert isinstance(current_object, dict)
        return current_object

    def _validate_and_convert(self, value: Any, path: list[str]):
        """Apply some basic validation and conversion steps.

        This includes:
        - Validation against the type given in the schema
        - List typed values are split at semicolons and validated individually
        """
        if value is None:
            return value
        subschema = self._get_subschema(path)
        # Array handling only if schema says it's an array.
        if subschema.get("type") == "array":
            items = subschema["items"]
            # Two cases: Either there are values separated by ";" or there is just a single value.
            if isinstance(value, str) and ";" in value:
                values = value.split(";")
            else:
                # We have to interpret value as a list now.
                values = [value]
            # Special case datetime comes first.
            if ("anyOf" in items):
                parsed_values = []
                for dt_val in values:
                    if isinstance(dt_val, datetime.datetime) and (
                            {'type': 'string', 'format': 'date-time'} in items["anyOf"]):
                        parsed_values.append(dt_val)
                    elif isinstance(dt_val, datetime.date) and (
                            {'type': 'string', 'format': 'date'} in items["anyOf"]):
                        parsed_values.append(dt_val)
                return parsed_values

            array_type = items.get("type")
            if array_type is None:
                raise KeyError("Array without 'type' in schema's items:\n"
                               + str(items))
            # Convert type "[foo, null]" -> "foo"
            if isinstance(array_type, list):
                if len(array_type) == 2 and array_type[-1] == "null":
                    array_type = array_type[0]
                else:
                    raise ValueError(f"Cannot handle list-valued items type. Current path: {path}")
            parsed_values = [self.PARSER[array_type](v) for v in values]
            return parsed_values
        # special case: datetime or date
        if ("anyOf" in subschema):
            if isinstance(value, datetime.datetime) and (
                    {'type': 'string', 'format': 'date-time'} in subschema["anyOf"]):
                return value
            if isinstance(value, datetime.date) and (
                    {'type': 'string', 'format': 'date'} in subschema["anyOf"]):
                return value
        # booleans might be retrieved as an integer or formula
        if subschema.get('type') == 'boolean':
            if value == 0 or isinstance(value, str) and '=false()' == value.lower():
                value = False
            if value == 1 or isinstance(value, str) and '=true()' == value.lower():
                value = True
        validate_jsonschema(value, subschema)

        # Finally: convert to target type
        return self.PARSER[subschema.get("type", "string")](value)

    def _get_subschema(self, path: list[str], schema: Optional[dict] = None) -> dict:
        """Return the sub schema at ``path``."""
        if schema is None:
            schema = self._schema
            assert schema is not None
        assert isinstance(schema, dict)

        return xlsx_utils.get_subschema(path, schema)


def _group_foreign_paths(foreign: list[list], common: list[str]) -> list[SimpleNamespace]:
    """Group the foreign keys by their base paths.

    Parameters
    ----------
    foreign: list[list]
      A list of foreign definitions, consisting of path components, property and possibly value.

    common: list[list[str]]
      A common path which defines the final target of the foreign definitions.  This helps to
      understand where the ``foreign`` paths shall be split.

    Returns
    -------
    out: list[dict[str, list[list]]]

      A list of foreign path segments, grouped by their common segments.  Each element is a
      namespace with detailed information of all those elements which form the group.  The namespace
      has the following attributes:

      - ``path``: The full path to this path segment.  This is always the previous segment's
        ``path`` plus this segment's ``subpath``.
      - ``stringpath``: The stringified ``path``, might be useful for comparison or sorting.
      - ``subpath``: The path, relative from the previous segment.
      - ``definitions``: A list of the foreign definitions for this segment, but stripped of the
        ``path`` components.
    """
    # Build a simple dict first, without subpath.
    results = {}
    for f_path in foreign:
        path = []
        for component in f_path:
            path.append(component)
            if path != common[:len(path)]:
                break
        path.pop()
        definition = f_path[len(path):]
        stringpath = xlsx_utils.p2s(path)
        if stringpath not in results:
            results[stringpath] = SimpleNamespace(stringpath=stringpath, path=path,
                                                  definitions=[definition])
        else:
            results[stringpath].definitions.append(definition)

    # Then sort by stringpath and calculate subpath.
    stringpaths = sorted(results.keys())

    resultlist = []
    last_level = 0
    for stringpath in stringpaths:
        elem = results[stringpath]
        elem.subpath = elem.path[last_level:]
        last_level = len(elem.path)
        resultlist.append(elem)

    if last_level != len(common):
        raise ValueError("Foreign keys must cover the complete `common` depth.")
    return resultlist


# pylint: disable-next=dangerous-default-value,too-many-arguments
def _set_in_nested(mydict: dict, path: list, value: Any, prefix: list = [], skip: int = 0,
                   overwrite: bool = False, append_to_list: bool = False) -> dict:
    """Set a value in a nested dict.

    Parameters
    ----------
    mydict: dict
      The dict into which the ``value`` shall be inserted.
    path: list
      A list of keys, denoting the location of the value.
    value
      The value which shall be set inside the dict.
    prefix: list
      A list of keys which shall be removed from ``path``.  A KeyError is raised if ``path`` does
      not start with the elements of ``prefix``.
    skip: int = 0
      Remove this many additional levels from the path, *after* removing the prefix.
    overwrite: bool = False
      If True, allow overwriting existing content. Otherwise, attempting to overwrite existing
      values leads to an exception.
    append_to_list: bool = False
      If True, assume that the element at ``path`` is a list and append the value to it.  If the
      list does not exist, create it.  If there is a non-list at ``path`` already, overwrite it with
      a new list, if ``overwrite`` is True, otherwise raise a ValueError.

    Returns
    -------
    mydict: dict
      The same dictionary that was given as a parameter, but modified.
    """
    for idx, el in enumerate(prefix):
        if path[idx] != el:
            raise KeyError(f"Path does not start with prefix: {prefix} not in {path}")
    path = path[len(prefix):]
    if skip:
        assert len(path) > skip, f"Path must be long enoug to remove skip={skip} elements."
        path = path[skip:]

    tmp_dict = mydict
    while len(path) > 1:
        key = path.pop(0)
        if key not in tmp_dict:
            tmp_dict[key] = {}
        if not isinstance(tmp_dict[key], dict):
            if overwrite:
                tmp_dict[key] = {}
            else:
                raise ValueError(f"There is already some value at {path}")
        tmp_dict = tmp_dict[key]
    key = path.pop()
    if append_to_list:
        if key not in tmp_dict:
            tmp_dict[key] = []
        if not isinstance(tmp_dict[key], list):
            if overwrite:
                tmp_dict[key] = []
            else:
                raise ValueError(f"There is already some non-list value at [{key}]")
        tmp_dict[key].append(value)
    else:
        if key in tmp_dict and not overwrite:
            raise ValueError(f"There is already some value at [{key}]")
        if key not in tmp_dict:
            tmp_dict[key] = {}
        tmp_dict[key] = value
    return mydict


def to_dict(xlsx: Union[str, BinaryIO], schema: Union[dict, str, TextIO],
            validate: bool = False, strict: bool = False, collect_errors: bool = True) -> dict:
    """Convert the xlsx contents to a dict, it must follow a schema.

    Parameters
    ----------
    xlsx: Union[str, BinaryIO]
      Path to the XLSX file or opened file object.

    schema: Union[dict, str, TextIO]
      Schema for validation of XLSX content.

    validate: bool, default=False
      If True, validate the result against the schema.

    strict: bool, default=False
      If True, fail faster.

    collect_errors: bool, default=True
      See `XLSXConverter.to_dict()` for details.


    Returns
    -------
    out: dict
      A dict representing the JSON with the extracted data.
    """
    converter = XLSXConverter(xlsx, schema, strict=strict)
    return converter.to_dict(validate=validate, collect_errors=collect_errors)


def records_from_json(data: dict, schema: Union[dict, str, TextIO]) -> db.Container:
    """Create Records from json data.

    Parameters
    ----------

    data: dict
      The data dictionary, as returned by ``to_dict``.

    schema : dict
      The json schema for this data.
    """

    result = db.Container()
    schema = read_or_dict(schema)
    # Top-level keys: RecordTypes for the entities.  For each RecordType, create a number of
    # Records.
    for rtname, reclist in data.items():
        for recdata in reclist:
            result.extend(deep_record_from_json(recdata, schema=schema["properties"][rtname],
                                                parent=rtname))
    result.to_xml()  # Generate negative IDs where necessary
    return result


def deep_record_from_json(recdata: dict, schema: dict, parent: str,
                          container: Optional[db.Container] = None) -> db.Container:
    """Create a single Record (and linked references recursivley) from json data.

    Parameters
    ----------

    recdata: dict
      The data dictionary for the single Record.

    schema : dict
      The json schema for this data.

    parent : str
      The name of the RecordType for this Record.

    container: db.Container, optional
      If given, use this to collect the generates entities.  Not generally necessary when calling
      this function from the outside.

    Returns
    -------

    out: db.Container
      The Record and all its references (recursively).  The first element is guaranteed to be the
      top-level element described by ``data``.
    """
    if container:
        result = container
    else:
        result = db.Container()
    if not isinstance(recdata, dict):
        raise ValueError("Expected dict in data.")
    rec = db.Record().add_parent(parent)
    result.append(rec)

    # Reserved keys
    reserved_keys = {"name", "id", "description"}
    if name := recdata.get("name"):
        rec.name = name
    if recid := recdata.get("id"):
        rec.id = int(recid)
    if description := recdata.get("description"):
        rec.description = description

    # Property keys
    for key, prop in recdata.items():
        if key in reserved_keys:
            continue
        subschema = schema["properties"][key]
        if isinstance(prop, list):
            all_refs = []
            subschema = subschema["items"]
            for element in prop:
                if isinstance(element, dict):  # if element is a dict, it is a reference
                    if "file" in element:
                        raise NotImplementedError("Files from JSON are not implemented yet.")
                        # continue
                    referenced_cont = deep_record_from_json(element, subschema,
                                                            parent=_get_ref_prop_type(key))
                    all_refs.append(referenced_cont[0])
                    result.extend(referenced_cont)
                else:
                    if "enum" in subschema and element:
                        enum_id = _get_enum_id(propname=key, enumvalue=element)
                        all_refs.append(enum_id)
                    else:
                        all_refs.append(element)
            rec.add_property(name=key, value=all_refs)
        elif isinstance(prop, dict):  # Single reference (or file)
            if "file" in prop:
                raise NotImplementedError("Files from JSON are not implemented yet.")
                # continue
            referenced_cont = deep_record_from_json(prop, subschema,
                                                    parent=_get_ref_prop_type(key))
            rec.add_property(name=key, value=referenced_cont[0])
            result.extend(referenced_cont)
        else:
            if key == "file":
                raise RuntimeError("This should not be reached if 'file' is part of the dict.")
            value = prop
            if "enum" in subschema and value:
                value = _get_enum_id(propname=key, enumvalue=prop)
            rec.add_property(name=key, value=value)
    return result


@lru_cache(maxsize=DEFAULT_SIZE)
def _get_enum_id(propname: str, enumvalue: str) -> str:
    """Retrieve the ID of the enum (cached helper function).

    Parameters
    ----------
    propname : str
      The name of the property.

    enumvalue : str
      The value of the enum

    Returns
    -------
    out : str
      The ID, as retrieved from the server.
    """
    # Get recordtype
    rtype_or_prop = cached_get_entity_by(name=propname)
    if rtype_or_prop.role == "Property":
        rtype = rtype_or_prop.datatype
        if is_list_datatype(rtype):
            rtype = get_list_datatype(rtype)
    elif rtype_or_prop.role == "RecordType":
        rtype = rtype_or_prop.name
    else:
        raise ValueError(f"Not a Property or RecordType: {propname}")

    # Get ID of enum
    enum_rec = cached_get_entity_by(query=f"FIND {rtype} with name='{enumvalue}'")

    return enum_rec.id


def _get_ref_prop_type(refname: str) -> str:
    """Return the best guess for a reference property's recordtype.

    This function attempts to retrieve it from the server, as a fallback return the input name.


    Parameters
    ----------
    refname : str
      The name of the reference property.

    Returns
    -------
    out : str
      The best gues for the recordtype's name.
    """
    try:
        refprop = cached_get_entity_by(query=f"FIND PROPERTY with name='{refname}'")
    except (ValueError, db.EmptyUniqueQueryError):  # FIXME Which errors do we want to catch?
        refprop = None

    # Fallback
    if refprop is None:
        return refname

    # Lists
    if is_list_datatype(refprop.datatype):
        return get_list_datatype(refprop.datatype)

    # Simple reference
    return refprop.datatype
