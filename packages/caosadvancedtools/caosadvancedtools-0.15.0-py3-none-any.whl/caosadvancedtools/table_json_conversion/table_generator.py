#!/usr/bin/env python3
# encoding: utf-8
#
# This file is a part of the LinkAhead Project.
#
# Copyright (C) 2024 Indiscale GmbH <info@indiscale.com>
# Copyright (C) 2024 Henrik tom Wörden <h.tomwoerden@indiscale.com>
# Copyright (C) 2024 Daniel Hornung <d.hornung@indiscale.com>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

"""
This module allows to generate template tables from JSON schemas.
"""

from __future__ import annotations

import re
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook.child import INVALID_TITLE_REGEX

from .xlsx_utils import p2s, ColumnType, RowType


class EmptySchemaException(RuntimeError):
    """Exception for when a schema is empty, so that no meaningful table can be generated."""


class TableTemplateGenerator(ABC):
    """Base class for generating tables from json schema"""

    def __init__(self):
        pass

    @abstractmethod
    def generate(
        self,
        schema: dict,
        foreign_keys: dict,
        filepath: str,
        use_ids_as_foreign: bool = False,
    ):
        """Generate a sheet definition from a given JSON schema.

        Parameters:
        -----------
        schema: dict
            Given JSON schema.

        foreign_keys: dict
            A tree-like configuration (nested dict) that defines which attributes shall be used to
            create additional columns when a list of references exists. The nested dict is
            structured like the data model, its innermost elements are leaves of the path trees
            within the JSON, they define the required keys.

            | Suppose we want to distinguish Persons that are referenced by Trainings, then
              ``foreign_keys`` must at least contain the following:
            | ``{"Training": {"Person": ["name", "email"]}}``.

            Values within the dicts can be either a list representing the keys (as in the example
            above) or a dict that allows to set additional foreign keys at higher depths.  In the
            latter case (dict instead of list) if foreign keys exist at that level (e.g. in the
            above example there might be further levels below "Person"), then the foreign keys can
            be set using the special ``__this__`` key.

            Example: ``{"Training": {"__this__": ["date"], "Person": ["name", "email"]}}``
            Here, ``date`` is the sole foreign key for Training.

            | It probably is worth extending the first example, with a case where a "Training" shall
              be distiguished by the "name" and "email" of a "Person" which it references.  The
              foreign keys for this example are specified like this:
            | ``{"Training": {"__this__": [["Person", "name"], ["Person", "email"]]}}``

        use_ids_as_foreign: bool, optional
            If True, use the id (a property named "id") as foreign key, if the key does not exist in
            the dict.  Default is False.
        """

    def _generate_sheets_from_schema(
        self,
        schema: dict,
        foreign_keys: Optional[dict] = None,
        use_ids_as_foreign: bool = False,
    ) -> dict[str, dict[str, tuple[ColumnType, Optional[str], list]]]:
        """Generate a sheet definition from a given JSON schema.

        Parameters
        ----------
        schema: dict
            given JSON schema
        foreign_keys: dict, optional
            a configuration that defines which attributes shall be used to create
            additional columns when a list of references exists. See ``foreign_keys``
            argument of TableTemplateGenerator.generate.
        use_ids_as_foreign: bool, optional
            If True, use the id (a property named "id") as foreign key, if the key does not exist in
            the dict.  Default is False.

        Returns
        -------
        sheets: dict
            A two-level dict which describes columns of template sheets.

            | The structure of this two-level dict is as follows:
            | ``sheets[sheetname][colname]= (<col_type>, <description>, [<path>, ...])``

            I.e. the outer dict contains sheet names as keys, the inner dict has column names as
            keys and tuples as values. These tuples consist of:
            - the column type
            - the description of the corresponding property
            - a list representing the path.

        """
        if not ("type" in schema or "anyOf" in schema):
            raise ValueError(
                "Inappropriate JSON schema: The following object must contain the "
                f"'type' or 'anyOf' key:\n{schema}\n"
            )
        if "properties" not in schema:
            raise ValueError(
                "Inappropriate JSON schema: The following object must contain "
                f"the 'properties' key:\n{schema}\n"
            )
        if "type" in schema:
            assert schema["type"] == "object"
        if foreign_keys is None:
            foreign_keys = {}
        # here, we treat the top level
        # sheets[sheetname][colname]= (COL_TYPE, description, [path])
        sheets: dict[str, dict[str, tuple[ColumnType, Optional[str], list]]] = {}
        for rt_name, rt_def in schema["properties"].items():
            sheets[rt_name] = self._treat_schema_element(
                schema=rt_def,
                sheets=sheets,
                path=[rt_name],
                foreign_keys=foreign_keys,
                use_ids_as_foreign=use_ids_as_foreign,
            )
        return sheets

    def _get_foreign_keys(
        self, keys: dict, path: list, use_ids_as_foreign: bool = False
    ) -> list[list[str]]:
        """Return the foreign keys that are needed at the location to which path points.

        Returns
        -------
        foreign_keys: list[list[str]]
          Contains lists of strings, each element is the path to one foreign key.
        """
        msg_missing = (
            f"A foreign key definition is missing for path:\n{path}\nKeys are:\n{keys}"
        )
        orig_path = path.copy()
        while path:
            if keys is None or path[0] not in keys:
                if (
                    use_ids_as_foreign
                ):  # Create entry ad-hoc.  TODO: don't modify passed argument?
                    keys[path[0]] = {"__this__": ["id"]}
                else:
                    raise ValueError(msg_missing)
            keys = keys[path[0]]
            path = path[1:]
        if isinstance(keys, dict) and "__this__" in keys:
            keys = keys["__this__"]
        if isinstance(keys, str):
            raise ValueError(
                "Foreign keys must be a list of strings, but a single "
                "string was given:\n"
                f"{orig_path} -> {keys}"
            )
        if not isinstance(keys, list):
            raise ValueError(msg_missing)

        # Keys must be either all lists or all strings
        types = {type(key) for key in keys}
        if len(types) > 1:
            raise ValueError(
                "The keys of this path must bei either all lists or all strings:"
                f" {orig_path}"
            )
        if types.pop() is str:
            keys = [[key] for key in keys]
        return keys

    def _treat_schema_element(
        self,
        schema: dict,
        sheets: dict,
        path: list[str],
        foreign_keys: Optional[dict] = None,
        level_in_sheet_name: int = 1,
        array_paths: Optional[list] = None,
        use_ids_as_foreign: bool = False,
    ) -> dict[str, tuple[ColumnType, Optional[str], list]]:
        """Recursively transform elements from the schema into column definitions.

        ``sheets`` is modified in place.

        Parameters
        ----------
        schema: dict
            Part of the json schema; it must be the level that contains the type definition
            (e.g. 'type' or 'oneOf' key)
        sheets: dict
            All the sheets, indexed by their name.  This is typically modified in place by this
            method.
        path: list[str]
            The relevant (sub) path for this schema part?
        array_paths: list
            A list of path along the way to the current object, where the json contains arrays.

        Returns
        -------
        columns: dict
            Describing the columns; see doc string of `_generate_sheets_from_schema`_
        """
        if not (
            "type" in schema
            or "enum" in schema
            or "oneOf" in schema
            or "anyOf" in schema
        ):
            raise ValueError(
                "Inappropriate JSON schema: The following schema part must contain "
                f"'type', 'enum', 'oneOf' or 'anyOf':\n{schema}\n"
            )

        if array_paths is None:
            # if this is not set, we are at top level and the top level element may always be an
            # array
            array_paths = [path]
        if foreign_keys is None:
            foreign_keys = {}

        ctype = ColumnType.SCALAR

        # if it is an array, value defs are in 'items'
        if schema.get("type") == "array":
            items = schema["items"]
            # list of references; special treatment
            if items.get("type") == "object" and len(path) > 1:
                # we add a new sheet with columns generated from the subtree of the schema
                sheetname = p2s(path)
                if sheetname in sheets:
                    raise ValueError(
                        "The schema would lead to two sheets with the same name, "
                        f"which is forbidden: {sheetname}"
                    )
                col_def = self._treat_schema_element(
                    schema=items,
                    sheets=sheets,
                    path=path,
                    foreign_keys=foreign_keys,
                    level_in_sheet_name=len(path),
                    array_paths=array_paths
                    + [path],  # since this level is an array extend the list
                    use_ids_as_foreign=use_ids_as_foreign,
                )
                if col_def:
                    sheets[sheetname] = col_def
                    # and add the foreign keys that are necessary up to this point
                    for array_path in array_paths:
                        foreigns = self._get_foreign_keys(
                            foreign_keys,
                            array_path,
                            use_ids_as_foreign=use_ids_as_foreign,
                        )
                        for foreign in foreigns:
                            internal_key = p2s(array_path + foreign)
                            if internal_key in sheets[sheetname]:
                                raise ValueError(
                                    "The schema would lead to two columns with the "
                                    "same name, which is forbidden:\n"
                                    f"{foreign} -> {internal_key}"
                                )
                            ref_sheet = p2s(array_path)
                            sheets[sheetname][internal_key] = (
                                ColumnType.FOREIGN,
                                f"see sheet '{ref_sheet}'",
                                array_path + foreign,
                            )
                # Columns are added to the new sheet, thus we do not return any columns for the
                # current sheet.
                return {}

            # List of enums: represent as checkbox columns
            if (
                schema.get("uniqueItems") is True
                and "enum" in items
                and len(items) == 1
            ):
                choices = items["enum"]
                assert len(path) >= 1
                prop_name = path[-1]
                result = {}
                for choice in choices:
                    name = f"{prop_name}.{choice}"
                    result[name] = (
                        ColumnType.MULTIPLE_CHOICE,
                        schema.get("description"),
                        path + [str(choice)],
                    )
                return result

            # it is a list of primitive types -> semicolon separated list
            schema = items
            ctype = ColumnType.LIST

        # This should only be the case for "new or existing reference".
        for el in schema.get("oneOf", []):
            if "type" in el:
                schema = el
                break

        if "properties" in schema:  # recurse for each property, then return
            cols = {}
            for pname in schema["properties"]:
                col_defs = self._treat_schema_element(
                    schema["properties"][pname],
                    sheets,
                    path + [pname],
                    foreign_keys,
                    level_in_sheet_name,
                    array_paths=array_paths,
                    use_ids_as_foreign=use_ids_as_foreign,
                )
                for k in col_defs:
                    if k in cols:
                        raise ValueError(
                            f"The schema would lead to two columns with the same "
                            f"name which is forbidden: {k}"
                        )
                cols.update(col_defs)
            return cols

        # The schema is a leaf.
        # definition of a single column
        default_return = {
            p2s(path[level_in_sheet_name:]): (ctype, schema.get("description"), path)
        }
        if "type" not in schema and "enum" in schema:
            return default_return
        if "type" not in schema and "anyOf" in schema:
            for d in schema["anyOf"]:
                # currently the only case where this occurs is date formats
                assert d["type"] == "string"
                assert d["format"] == "date" or d["format"] == "date-time"
            return default_return
        scalars = ["string", "number", "integer", "boolean"]
        # Also add "null" combinations, such as ["string", "null"].
        if schema["type"] in (scalars + [[scal, "null"] for scal in scalars]):
            if "format" in schema and schema["format"] == "data-url":
                return {}  # file; ignore for now
            return default_return
        raise ValueError(
            "Inappropriate JSON schema: The following part should define an"
            f" object with properties or a primitive type:\n{schema}\n"
        )


class XLSXTemplateGenerator(TableTemplateGenerator):
    """Class for generating XLSX tables from json schema definitions."""

    # def __init__(self):
    #     super().__init__()

    def generate(
        self,
        schema: dict,
        foreign_keys: dict,
        filepath: Union[str, Path],
        use_ids_as_foreign: bool = False,
    ) -> None:
        """Generate a sheet definition from a given JSON schema.

        Parameters:
        -----------
        schema: dict
            Given JSON schema
        foreign_keys: dict
            A configuration that defines which attributes shall be used to create
            additional columns when a list of references exists. See ``foreign_keys``
            argument of :meth:`TableTemplateGenerator.generate` .
        filepath: Union[str, Path]
            The XLSX file will be stored under this path.
        """
        sheets = self._generate_sheets_from_schema(
            schema, foreign_keys, use_ids_as_foreign=use_ids_as_foreign
        )
        if not sheets:
            raise EmptySchemaException("No sheets generated, schema is too empty.")
        wb = self._create_workbook_from_sheets_def(sheets)
        if not isinstance(filepath, Path):
            filepath = Path(filepath)
        parentpath = filepath.parent
        parentpath.mkdir(parents=True, exist_ok=True)
        wb.save(filepath)

    @staticmethod
    def _get_max_path_length(sheetdef: dict) -> int:
        """returns the length of the longest path contained in the sheet definition

        see TableTemplateGenerator._generate_sheets_from_schema for the structure of the sheets
        definition dict
        You need to pass the dict of a single sheet to this function.
        """
        return max(len(path) for _, _, path in sheetdef.values())

    @staticmethod
    def _get_ordered_cols(sheetdef: dict) -> list:
        """
        creates a list with tuples (colname, column type, path) where the foreign keys are first
        """
        ordered_cols = []
        # first foreign cols
        for colname, (ct, desc, path) in sheetdef.items():
            if ct == ColumnType.FOREIGN:
                ordered_cols.append((colname, ct, desc, path))
        # now the other
        for colname, (ct, desc, path) in sheetdef.items():
            if ct != ColumnType.FOREIGN:
                ordered_cols.append((colname, ct, desc, path))

        return ordered_cols

    def _create_workbook_from_sheets_def(
        self, sheets: dict[str, dict[str, tuple[ColumnType, Optional[str], list]]]
    ):
        """Create and return a nice workbook for the given sheets."""
        wb = Workbook()
        yellowfill = PatternFill(fill_type="solid", fgColor="00FFFFAA")
        # remove initial sheet
        assert wb.sheetnames == ["Sheet"]
        del wb["Sheet"]

        sheets = XLSXTemplateGenerator.normalize_sheet_titles(sheets)

        for sheetname, sheetdef in sheets.items():
            if not sheetdef:
                continue
            ws = wb.create_sheet(re.sub(INVALID_TITLE_REGEX, "_", sheetname))
            # First row will by the COL_TYPE row.
            # First column will be the indicator row with values COL_TYPE, PATH, IGNORE.
            # The COL_TYPE row will be followed by as many PATH rows as needed.

            max_path_length = self._get_max_path_length(sheetdef)
            header_index = 2 + max_path_length
            description_index = 3 + max_path_length

            # create first column
            ws.cell(1, 1, RowType.COL_TYPE.name)
            for index in range(max_path_length):
                ws.cell(2 + index, 1, RowType.PATH.name)
            ws.cell(header_index, 1, RowType.IGNORE.name)
            ws.cell(description_index, 1, RowType.IGNORE.name)

            ordered_cols = self._get_ordered_cols(sheetdef)

            # create other columns
            for index, (colname, coltype, desc, path) in enumerate(ordered_cols):
                ws.cell(1, 2 + index, coltype.name)
                for path_index, el in enumerate(path):
                    ws.cell(2 + path_index, 2 + index, el)
                ws.cell(header_index, 2 + index, colname)
                if coltype == ColumnType.FOREIGN:
                    # Visual highlighting
                    ws.cell(header_index, 2 + index).fill = yellowfill
                if desc:
                    ws.cell(description_index, 2 + index, desc)

            # hide special rows
            for index, row in enumerate(ws.rows):
                if not (row[0].value is None or row[0].value == RowType.IGNORE.name):
                    ws.row_dimensions[index + 1].hidden = True

            # hide special column
            ws.column_dimensions["A"].hidden = True

        # order sheets
        # for index, sheetname in enumerate(sorted(wb.sheetnames)):
        # wb.move_sheet(sheetname, index-wb.index(wb[sheetname]))
        # reverse sheets
        for index, sheetname in enumerate(wb.sheetnames[::-1]):
            wb.move_sheet(sheetname, index - wb.index(wb[sheetname]))

        return wb

    @staticmethod
    def normalize_sheet_titles(sheets: dict[str, Any]) -> dict[str, Any]:
        """Shorten title with more than 31 characters.  Return normalized dict.

        Specification:

        - Short titles are left unchanged.
        - Titles are unique.
        - Long titles are shortened like this:

          1. Titles are split into *parts* at the dot character (``.``)
          2. Parts are reused from the end to front, until at most 26 characters are used up.  These parts form the *tail* (joined by ``.`` characters).
          3. A *head* is added, and also joined by the ``.`` character.  The head is exactly 4 characters long and consists of these parts:

             - The first character of the first part.
             - An underscore character (``_``).
             - The lowest two-digit number starting from ``01`` (so one of 01, 02, ..., 98, 99), that leads
               to no collision with existing titles.  If all options would lead to a collision, a
               OverflowError is thrown.

        Parameters
        ----------

        sheets: dict[str, Any]
          The title -> content dict of sheets.

        Returns
        -------

        out: dict[str, Any]
          Same as the input sheets, but long titles (keys) are shortened in a sensible way.

        """
        max_len = 31
        max_counter = 99
        # Prefill with titles that need no changing
        titles: set[str] = set([key for key in sheets.keys() if len(key) <= max_len])
        result = {}
        for title, content in sheets.items():
            # Boring case: No shortening
            if len(title) <= max_len:
                result[title] = content
                continue
            # We must make it shorter now!
            parts = title.split(".")
            # Create tail
            tail = ""
            for num_parts in range(len(parts)):
                candidate = ".".join(parts[-1 - num_parts:])
                if len(candidate) <= max_len - 5:  # Short enough to prepend the head?
                    tail = candidate
                else:
                    break
            head = parts[0][0] + "_"
            # We do not keep indices of already used numbers, because we expect this function to
            # handle only small numbers anyway. (No premature optimization)
            new_title = ""
            for idx in range(1, max_counter + 2):
                candidate = f"{head}{idx:02d}.{tail}"
                if not tail:  # Prevent trailing dot.
                    candidate = f"{head}{idx:02d}"
                # Check for collision.
                if candidate in titles:
                    continue
                if idx > max_counter:
                    raise OverflowError(
                        "Too many similar sheet titles.  Refusing to create this sheet title:\n"
                        + candidate
                    )
                new_title = candidate
                titles.add(new_title)
                break
            if not new_title:
                raise RuntimeError("No new title found. This should never happen.")
            result[new_title] = content
        if not len(result) == len(sheets):
            raise RuntimeError("Differing length of result. This should never happen.")
        return result
