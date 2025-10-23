# This file is a part of the LinkAhead project.
#
# Copyright (C) 2025 IndiScale GmbH <info@indiscale.com>
# Copyright (C) 2025 Daniel Hornung <d.hornung@indiscale.com>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

"""Useful tools for unit or integration tests.
"""

from datetime import datetime
from typing import Optional, Union

from openpyxl import Workbook

from caosadvancedtools.table_json_conversion.xlsx_utils import is_recursively_none


def assert_equal_jsons(json1, json2, allow_none: bool = True, allow_empty: bool = True,
                       ignore_datetime: bool = False, ignore_id_value: bool = False,
                       allow_name_dict: bool = False,
                       path: Optional[list] = None) -> None:
    """Compare two json objects for near equality.

Raise an assertion exception if they are not equal.

Parameters
----------

allow_name_dict: bool, default=False
    If True, a string and a dict ``{"name": "string's value"}`` are considered equal.
    """
    if path is None:
        path = []
    assert isinstance(json1, dict) == isinstance(json2, dict), f"Type mismatch, path: {path}"
    if isinstance(json1, dict):
        keys = set(json1.keys()).union(json2.keys())
        for key in keys:
            this_path = path + [key]
            # Case 1: exists in both collections
            if key in json1 and key in json2:
                el1 = json1[key]
                el2 = json2[key]
                if allow_none and (el1 is None and (el2 == [] or el2 == {})
                                   or el2 is None and (el1 == [] or el1 == {})):
                    # shortcut in case of equivalent empty content
                    continue
                if allow_name_dict:  # Special exception
                    my_str = None
                    if isinstance(el1, str) and isinstance(el2, dict):
                        my_str = el1
                        my_dict = el2
                    elif isinstance(el2, str) and isinstance(el1, dict):
                        my_str = el2
                        my_dict = el1
                    if my_str is not None:
                        if len(my_dict) == 1 and my_dict.get("name") == my_str:
                            continue
                assert isinstance(el1, type(el2)), \
                    f"Type mismatch, path: {this_path}, types: {type(el1)}, {type(el2)}"
                if isinstance(el1, (dict, list)):
                    # Iterables: Recursion
                    assert_equal_jsons(
                        el1, el2, allow_none=allow_none, allow_empty=allow_empty,
                        ignore_datetime=ignore_datetime, ignore_id_value=ignore_id_value,
                        allow_name_dict=allow_name_dict,
                        path=this_path)
                    continue
                if not (ignore_id_value and key == "id"):
                    assert equals_with_casting(el1, el2, ignore_datetime=ignore_datetime), (
                        f"Values at path {this_path} are not equal:\n{el1}\n{el2}")
                continue
            # Case 2: exists only in one collection
            existing = json1.get(key, json2.get(key))
            assert ((allow_none and is_recursively_none(existing))
                    or (allow_empty and existing == [])), (
                f"Element at path {this_path} is None or empty in one json and does not exist in "
                "the other.")
        return
    assert isinstance(json1, list) and isinstance(json2, list), f"Is not a list, path: {path}"
    assert len(json1) == len(json2), (f"Lists must have equal length, path: {path}\n"
                                      f"{json1}\n ---\n{json2}")
    for idx, (el1, el2) in enumerate(zip(json1, json2)):
        this_path = path + [idx]
        if isinstance(el1, dict):
            assert_equal_jsons(el1, el2, allow_none=allow_none, allow_empty=allow_empty,
                               ignore_datetime=ignore_datetime, ignore_id_value=ignore_id_value,
                               allow_name_dict=allow_name_dict,
                               path=this_path)
        else:
            assert equals_with_casting(el1, el2, ignore_datetime=ignore_datetime), (
                f"Values at path {this_path} are not equal:\n{el1},\n{el2}")


def equals_with_casting(value1, value2, ignore_datetime: bool = False) -> bool:
    """Compare two values, return True if equal, False otherwise.  Try to cast to clever datatypes.
    """
    try:
        dt1 = datetime.fromisoformat(value1)
        dt2 = datetime.fromisoformat(value2)
        if ignore_datetime:
            return True
        return dt1 == dt2
    except (ValueError, TypeError):
        pass
    if isinstance(value1, datetime) and isinstance(value1, datetime) and ignore_datetime:
        return True
    return value1 == value2


def purge_from_json(data: Union[dict, list], remove_keys: list[str]) -> Union[dict, list]:
    """Remove matching entries from json data.


    Parameters
    ----------
    data : Union[dict, list]
      The json data to clean.

    remove_keys : list[str]
       Remove all keys that are in this list

    Returns
    -------
    out : Union[dict, list]
      The cleaned result.
    """

    # Remove only from dicts
    if isinstance(data, dict):
        keys = set(data.keys())
        for removable in remove_keys:
            if removable in keys:
                data.pop(removable)
        elements = list(data.values())
    else:
        if not isinstance(data, list):
            raise ValueError("Data must be a dict or list.")
        elements = data

    # Recurse for all elements
    for element in elements:
        if isinstance(element, dict) or isinstance(element, list):
            purge_from_json(element, remove_keys=remove_keys)

    return data


def compare_workbooks(wb1: Workbook, wb2: Workbook, hidden: bool = True):
    """Compare two workbooks for equal content.

Raises an error if differences are found.

Parameters
----------

hidden: bool, optional
  Test if the "hidden" status of rows and columns is the same.
    """
    assert wb1.sheetnames == wb2.sheetnames, (
        f"Sheet names are different: \n{wb1.sheetnames}\n   !=\n{wb2.sheetnames}"
    )
    for sheetname in wb2.sheetnames:
        sheet_1 = wb1[sheetname]
        sheet_2 = wb2[sheetname]
        for irow, (row1, row2) in enumerate(zip(sheet_1.iter_rows(), sheet_2.iter_rows())):
            if hidden:
                assert (sheet_1.row_dimensions[irow].hidden
                        == sheet_2.row_dimensions[irow].hidden), f"hidden row: {sheetname}, {irow}"
            for icol, (cell1, cell2) in enumerate(zip(row1, row2)):
                if hidden:
                    assert (sheet_1.column_dimensions[cell1.column_letter].hidden
                            == sheet_2.column_dimensions[cell2.column_letter].hidden), (
                                f"hidden col: {sheetname}, {icol}")
                assert cell1.value == cell2.value, (
                    f"Sheet: {sheetname}, cell: {cell1.coordinate}, Values: \n"
                    f"{cell1.value}\n{cell2.value}"
                )
