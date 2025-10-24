"""
excel_tools.py
read_excel_preserve_decimals(): a drop-in replacement for pandas.read_excel
that *keeps the exact number of decimal places* a user sees in Excel.
"""

from collections.abc import Sequence
from pathlib import Path
from typing import IO, Any

import pandas as pd
from openpyxl import load_workbook

# ------------------------------------------------------------------ #
# robust import for format_cell (new path / old path / fallback)     #
# ------------------------------------------------------------------ #
try:  # official since openpyxl 3.1
    from openpyxl.utils.formatting import format_cell  # type: ignore
except ImportError:
    try:  # provisional path in some wheels
        from openpyxl.utils.cell import format_cell  # type: ignore
    except ImportError:
        # very small local fallback
        def format_cell(cell) -> str:  # type: ignore
            v = cell.value
            if v is None:
                return ""
            fmt = getattr(cell, "number_format", "")
            if not isinstance(v, (int, float)) or "." not in fmt:
                return str(v)
            decs = fmt.split(".", 1)[1].split(";")[0]
            n_dec = sum(ch == "0" for ch in decs)
            return f"{v:.{n_dec}f}"


# ------------------------------------------------------------------ #
# internal helper                                                    #
# ------------------------------------------------------------------ #
def _clean_cell(cell) -> Any:
    """Return a value that respects the cell’s displayed decimals."""
    if cell.data_type != "n":          # not numeric
        return cell.value

    shown = format_cell(cell)          # text Excel would display
    if "e" in shown.lower():           # scientific notation → leave as float
        return cell.value

    if "." in shown:                   # count decimal places and round
        n_dec = len(shown.split(".", 1)[1])
        return round(float(cell.value), n_dec)
    return cell.value                  # integer-like


# ------------------------------------------------------------------ #
# public API                                                         #
# ------------------------------------------------------------------ #
def read_excel_preserve_decimals(
    path: str | Path | IO[bytes],
    sheet_name: Any = 0,
    header: int | Sequence[int] | None = 0,
    **pd_kwargs,
) -> pd.DataFrame:
    """
    Load an Excel sheet while preserving user-visible decimals **and**
    reproduce pandas’ header logic (Unnamed columns + de-duplication).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]

    # 1 — read all rows, fixing numeric cells
    rows: list[list[Any]] = [[_clean_cell(c) for c in row] for row in ws.iter_rows()]

    # 2 — build DataFrame without headers first
    df = pd.DataFrame(rows, **pd_kwargs)

    # 3 — mimic pandas header behaviour
    if header is not None:
        hdr_row = df.iloc[header].tolist()

        # convert None → 'Unnamed: {i}', Decimal → str, then de-duplicate
        seen: dict[str, int] = {}
        clean_hdr: list[str] = []
        for i, col in enumerate(hdr_row):
            base = str(col) if col is not None else f"Unnamed: {i}"
            cnt = seen.get(base, 0)
            clean = base if cnt == 0 else f"{base}.{cnt}"
            seen[base] = cnt + 1
            clean_hdr.append(clean)

        df.columns = clean_hdr
        df = df.drop(index=list(range(header + 1))).reset_index(drop=True)

    return df

