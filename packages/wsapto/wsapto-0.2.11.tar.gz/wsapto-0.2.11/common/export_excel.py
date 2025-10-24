from django.conf import settings
import os
from datetime import datetime
from datetime import datetime, date 
import re
import json
import traceback
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
from urllib.parse import urlparse, parse_qs
from openpyxl import Workbook
import io

def export_excel(data, headers, title_config, subtitle_config, table_start, column_widths=None, left_aligned_headers=None):
    """
    Generate an Excel file in memory with customizable content.
 
    Parameters:
    - data (list of dict): The data to populate the Excel file.
    - headers (list of str): Column headers for the Excel sheet.
    - title_config (dict): Configuration for title with 'row', 'col', and 'text'.
    - subtitle_config (list of dict): List of subtitle configurations with 'row', 'col', and 'text'.
    - table_start (dict): Dictionary with 'row' and 'col' for table start position.
    - column_widths (dict, optional): Dictionary to set custom column widths.
    - left_aligned_headers (list of str, optional): Headers that should be left-aligned.
 
    Returns:
    - BytesIO: An in-memory Excel file.
    """
    if left_aligned_headers is None:
        left_aligned_headers = []
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
 
    # Set Title
    title_cell = ws.cell(row=title_config['row'], column=title_config['col'], value=title_config['text'])
    title_cell.font = Font(size=20, bold=True, color="000000")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=title_config['row'], start_column=title_config['col'], 
                   end_row=title_config['row'], end_column=title_config['col'] + len(headers) - 1)
 
    # Set Subtitles
    for subtitle in subtitle_config:
        subtitle_cell = ws.cell(row=subtitle['row'], column=subtitle['col'], value=subtitle['text'])
        subtitle_cell.font = Font(size=11, italic=True, color="555555")
        alignment = subtitle.get("alignment", "center")  # Default ke center jika tidak ada
        subtitle_cell.alignment = Alignment(horizontal=alignment, vertical="center")
        ws.merge_cells(start_row=subtitle['row'], start_column=subtitle['col'], 
                    end_row=subtitle['row'], end_column=subtitle['col'] + len(headers) - 1)
 
    # Set Table Headers
    header_row_index = table_start['row']
    for col_num, header in enumerate(headers, start=table_start['col']):
        cell = ws.cell(row=header_row_index, column=col_num, value=header)
        cell.font = Font(bold=True, size=12, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.border = Border(
            top=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
 
    # Populate Data Rows
    for row_index, row_data in enumerate(data, start=header_row_index + 1):
        for col_index, header in enumerate(headers, start=table_start['col']):
            cell_value = row_data.get(header, "")
            if isinstance(cell_value, (datetime, date)):
                cell_value = cell_value.strftime('%d-%m-%Y')
 
            cell = ws.cell(row=row_index, column=col_index, value=cell_value)
 
            if header in left_aligned_headers:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
 
            cell.border = Border(
                top=Side(border_style="thin", color="000000"),
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )
 
    # Adjust Column Widths
    for col_num, header in enumerate(headers, start=table_start['col']):
        if column_widths and header in column_widths:
            ws.column_dimensions[ws.cell(row=table_start['row'], column=col_num).column_letter].width = column_widths[header]
        else:
            max_length = max(len(str(header)), *(len(str(row_data.get(header, ""))) for row_data in data))
            ws.column_dimensions[ws.cell(row=table_start['row'], column=col_num).column_letter].width = max_length + 5
 
    # Save workbook to an in-memory buffer
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer

def export_excel_pandas(data_attendance, data_permit, headers_attendance, headers_permit, title_config, subtitle_config, table_start):
    """
    Generate an Excel file with multiple sheets using pandas.

    Parameters:
    - data_attendance (list of dict): Data for the Attendance sheet.
    - data_permit (list of dict): Data for the Permit sheet.
    - headers_attendance (list of str): Column headers for the Attendance sheet.
    - headers_permit (list of str): Column headers for the Permit sheet.
    - title_config (dict): Title configuration with 'row', 'col', and 'text'.
    - subtitle_config (list of dict): Subtitle configurations with 'row', 'col', and 'text'.
    - table_start (dict): Dictionary with 'row' and 'col' for table start.

    Returns:
    - BytesIO: An in-memory Excel file.
    """

    # Convert data to pandas DataFrame
    df_attendance = pd.DataFrame(data_attendance)
    df_permit = pd.DataFrame(data_permit)

    # Handle NaN and Inf values
    for df in [df_attendance, df_permit]:
        df.replace([float("inf"), -float("inf")], 0, inplace=True)  # Replace Inf with 0
        df.fillna("", inplace=True)  # Replace NaN with empty string

    # 🔹 Pastikan hanya kolom yang tersedia dimasukkan ke Excel
    df_attendance = df_attendance[[col for col in headers_attendance if col in df_attendance.columns]]
    df_permit = df_permit[[col for col in headers_permit if col in df_permit.columns]]

    # Save DataFrame to Excel with formatting
    output_buffer = io.BytesIO()
    
    with pd.ExcelWriter(output_buffer, engine="xlsxwriter") as writer:
        workbook = writer.book
        header_format = workbook.add_format({"bold": True, "bg_color": "#FFFF00", "align": "center", "border": 1})
        cell_format = workbook.add_format({"border": 1,'align':'Left'})   
        title_format = workbook.add_format({"bold": True, "font_size": 20, "align": "center"})
        subtitle_format = workbook.add_format({"italic": True, "font_size": 11, "align": "left", "font_color": "#555555"})

        # Function to write data into a sheet
        def write_sheet(df, sheet_name, headers):
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=table_start["row"] - 1)
            worksheet = writer.sheets[sheet_name]

            # Set title
            worksheet.merge_range(
                title_config["row"] - 1, title_config["col"] - 1, 
                title_config["row"] - 1, title_config["col"] - 1 + len(headers) - 1, 
                title_config["text"], title_format
            )

            # Set subtitles
            for subtitle in subtitle_config:
                worksheet.merge_range(
                    subtitle["row"] - 1, subtitle["col"] - 1,
                    subtitle["row"] - 1, subtitle["col"] - 1 + len(headers) - 1,
                    subtitle["text"], subtitle_format
                )

            # Apply header formatting
            for col_num, header in enumerate(headers):
                if header in df.columns:  # Pastikan hanya menulis header yang ada
                    worksheet.write(table_start["row"] - 1, col_num, header, header_format)

            # Apply border to all data cells
            for row_num in range(len(df)):
                for col_num, col_name in enumerate(df.columns):
                    worksheet.write(table_start["row"] + row_num, col_num, df.iloc[row_num, col_num], cell_format)

            # AutoFit columns
            for col_num, col_name in enumerate(df.columns):
                max_length = max(df[col_name].astype(str).map(len).max(), len(col_name)) + 2
                worksheet.set_column(col_num, col_num, max_length)

        # Write both sheets if they are not empty
        if not df_attendance.empty:
            write_sheet(df_attendance, "Attendance", headers_attendance)
        if not df_permit.empty:
            write_sheet(df_permit, "Permit", headers_permit)

    output_buffer.seek(0)
    return output_buffer