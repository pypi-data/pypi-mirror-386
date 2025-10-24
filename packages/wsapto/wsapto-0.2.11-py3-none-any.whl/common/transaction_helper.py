from django.db import connection, transaction
from django.db import connections
from django.conf import settings
import os
from datetime import datetime,timedelta
from datetime import datetime, date 
import re
import json
import traceback
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
from urllib.parse import urlparse, parse_qs
from openpyxl import Workbook
import io
import time
import hashlib
import base64
from collections import deque
import threading
import hmac
from typing import Any
from typing import Optional
import mimetypes
from urllib.parse import urlencode
import requests

# Global lock untuk sinkronisasi akses ke cache
cache_lock = threading.Lock()

# Cache untuk rate limiting per method dan IP
RATE_LIMIT_CACHE = {}

# Cache untuk duplicate request berdasarkan signature per IP
REQUEST_SIGNATURES = {}

# Maksimal duplicate request yang diperbolehkan dalam time_window
duplicate_limit = 5
duplicate_time_window = 10  # dalam detik

# Secret Key untuk API Key
SECRET_KEY = "MySatnusa"

# Konfigurasi batas ukuran
MAX_JSON_SIZE_MB = 4
MAX_FILE_SIZE_MB = 6
MAX_TOTAL_SIZE_MB = 10
MAX_FILES_ALLOWED = 20
ALLOWED_FILE_EXTENSIONS = {"jpg", "jpeg", "png", "pdf", "txt"}
ALLOWED_MIME_TYPES = {
    "jpg": "image/jpeg",
    "png": "image/png",
    "pdf": "application/pdf",
    "txt": "text/plain"
}
# Daftar User-Agent yang dikenal sebagai bot sah (opsional)
ALLOWED_BOTS = {
    "Googlebot": "Googlebot",
    "Bingbot": "bingbot",
    "YandexBot": "YandexBot",
    # Tambahkan bot lain yang diizinkan sesuai kebutuhan
}

# Daftar pola User-Agent mencurigakan
SUSPICIOUS_UA_PATTERNS = [
    r"(?i)\b(bot|crawl|spider|slurp|http|scrap|curl|wget|python-requests|java|node|phantomjs)\b",  # Bot/scraper umum
    r"(?i)\b(headless|automation|test|fake|unknown)\b",  # Indikator headless atau tes
    r"^[0-9]+$",  # User-Agent hanya angka (sangat tidak wajar)
    r"^\s*$",  # Kosong atau hanya whitespace
]
 
# Helper: Execute query 
def execute_query(sql_query, params=None, db_alias='default'):
    """
    Executes a raw SQL query on the specified database and returns the results as a list of dictionaries.
 
    Args:
        sql_query (str): The SQL query to execute.
        params (tuple or list, optional): Parameters to safely pass to the query.
        db_alias (str): The alias of the database to use.
 
    Returns:
        list of dict: A list of rows as dictionaries.
    """
    try:
        with connections[db_alias].cursor() as cursor:
            cursor.execute(sql_query, params or [])
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
 
        return [dict(zip(columns, row)) for row in rows]
 
    except Exception as e:
        raise Exception(f"Error executing query on database '{db_alias}': {e}")
 
# Helper: Insert data into table
def insert_data(table_name, data, db_alias='default'):
    """
    Helper to insert data into a table.
 
    Args:
        table_name (str): The name of the table where data will be inserted.
        data (dict): A dictionary containing the column names as keys and the values to be inserted as values.
 
    Returns:
        int or None: The inserted row's ID if successful, None if failed.
    """
    try:
        columns = ', '.join(data.keys())
        placeholders = ', '.join(['%s'] * len(data))
        sql_query = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"
 
        # Get the database connection using the db_alias parameter
        with transaction.atomic(using=db_alias):
            with connections[db_alias].cursor() as cursor:
                cursor.execute(sql_query, list(data.values()))
 
        return True  # Return True if the insert was successful
    except Exception as e:
        raise Exception(f"Error in insert: {e}")
 
# Helper: Read data from table  
def get_data(table_name, filters=None, search=None, search_columns=None, columns='*', 
             limit=None, offset=None, order_by=None, db_alias='default'):
    """
    Helper to read data from a table with support for operators like !=, IN, NOT IN, >, <, >=, <=, OR, and NULL handling.
 
    Args:
        table_name (str): The name of the table.
        filters (dict or list, optional): Conditions to filter data, supports OR with {"OR": [filter1, filter2]}.
        search (str, optional): Search keyword.
        search_columns (list, optional): Columns to perform search on.
        columns (str or list, optional): Columns to select, default is '*'.
        limit (int, optional): Limit the number of rows.
        offset (int, optional): Offset for pagination.
        order_by (str or list, optional): Column(s) to order by.
        db_alias (str, optional): The database alias to use (default is 'default').
 
    Returns:
        list of dict: Retrieved rows as a list of dictionaries.
    """
    try:
        if isinstance(columns, list):
            columns = ', '.join(columns)
 
        sql_query = f"SELECT {columns} FROM {table_name}"
 
        # Adding filters (WHERE clause)
        conditions = []
        values = []

        def process_filter(filter_dict):
            """Helper to process a single filter dictionary."""
            local_conditions = []
            local_values = []
            for key, value in filter_dict.items():
                if "__ne" in key:  # Not equal operator
                    column = key.replace("__ne", "")
                    if value is None:
                        local_conditions.append(f"{column} IS NOT NULL")
                    else:
                        local_conditions.append(f"{column} <> %s")
                        local_values.append(value)
                elif "__in" in key:  # IN operator
                    column = key.replace("__in", "")
                    if isinstance(value, (list, tuple)):
                        local_conditions.append(f"{column} IN ({', '.join(['%s'] * len(value))})")
                        local_values.extend(value)
                    else:
                        local_conditions.append(f"{column} = %s")
                        local_values.append(value)
                elif "__not_in" in key:  # NOT IN operator
                    column = key.replace("__not_in", "")
                    if isinstance(value, (list, tuple)):
                        local_conditions.append(f"{column} NOT IN ({', '.join(['%s'] * len(value))})")
                        local_values.extend(value)
                    else:
                        local_conditions.append(f"{column} <> %s")
                        local_values.append(value)
                elif "__between" in key:  # BETWEEN operator
                    column = key.replace("__between", "")
                    if isinstance(value, (list, tuple)) and len(value) == 2:
                        local_conditions.append(f"{column} BETWEEN %s AND %s")
                        local_values.extend(value)
                    else:
                        raise ValueError(f"Filter '{key}' requires a tuple/list with exactly two values for BETWEEN")
                elif "__gte" in key:  # Greater than or equal operator
                    column = key.replace("__gte", "")
                    local_conditions.append(f"{column} >= %s")
                    local_values.append(value)
                elif "__lte" in key:  # Less than or equal operator
                    column = key.replace("__lte", "")
                    local_conditions.append(f"{column} <= %s")
                    local_values.append(value)
                elif "__gt" in key:  # Greater than operator
                    column = key.replace("__gt", "")
                    local_conditions.append(f"{column} > %s")
                    local_values.append(value)
                elif "__lt" in key:  # Less than operator
                    column = key.replace("__lt", "")
                    local_conditions.append(f"{column} < %s")
                    local_values.append(value)
                else:
                    if value is None:
                        local_conditions.append(f"{key} IS NULL")
                    else:
                        local_conditions.append(f"{key} = %s")
                        local_values.append(value)
            return local_conditions, local_values

        if filters:
            if isinstance(filters, dict) and "OR" in filters:
                # Handle OR conditions
                or_conditions = []
                for filter_dict in filters["OR"]:
                    sub_conditions, sub_values = process_filter(filter_dict)
                    if sub_conditions:
                        or_conditions.append(f"({' AND '.join(sub_conditions)})")
                        values.extend(sub_values)
                if or_conditions:
                    conditions.append(f"({' OR '.join(or_conditions)})")
            else:
                # Handle standard AND conditions
                filter_conditions, filter_values = process_filter(filters if isinstance(filters, dict) else {})
                conditions.extend(filter_conditions)
                values.extend(filter_values)
 
        # Adding search condition (LIKE) for the given columns
        if search and search_columns:
            search_conditions = [f"{col}::text ILIKE %s" for col in search_columns]
            conditions.append(f"({' OR '.join(search_conditions)})")
            values += [f"%{search}%"] * len(search_columns)
 
        # Combine all conditions
        if conditions:
            where_clause = ' WHERE ' + ' AND '.join(conditions)
            sql_query += where_clause
 
        # Adding ORDER BY clause
        if order_by:
            if isinstance(order_by, list):
                order_by_clause = ', '.join(order_by)
            else:
                order_by_clause = order_by
            sql_query += f" ORDER BY {order_by_clause}"
 
        # Adding limit and offset
        if limit:
            sql_query += f" LIMIT {limit}"
 
        if offset:
            sql_query += f" OFFSET {offset}"
 
        # Executing query
        with connections[db_alias].cursor() as cursor:
            cursor.execute(sql_query, values)
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
 
        return [dict(zip(columns, row)) for row in rows]
 
    except Exception as e:
        raise Exception(f"Error in read from database '{db_alias}': {e}")
 
def update_data(table_name, data, filters, db_alias='default'):
    """
    Helper to update data in a table.
 
    Args:
        table_name (str): The name of the table where data will be updated.
        data (dict): The data to update, with column names as keys.
        filters (dict): Filters for the WHERE clause to specify which rows to update.
                        Supports '__ne' for 'not equal' conditions (e.g., {'column__ne': value}).
        db_alias (str, optional): The database alias to use (default is 'default').
 
    Returns:
        bool: True if update successful, False otherwise.
    """
    try:
        if not data:
            raise ValueError("Update data cannot be empty.")
        if not filters:
            raise ValueError("Filters cannot be empty to prevent updating all rows.")
 
        set_clause = ', '.join([f"{key}=%s" for key in data.keys()])
        
        # Prepare WHERE clause with support for __ne
        where_parts = []
        filter_values = []
        for key, value in filters.items():
            if key.endswith('__ne'):
                # Handle 'not equal' condition
                column = key[:-4]  # Remove '__ne' suffix
                where_parts.append(f"{column}!=%s")
                filter_values.append(value)
            else:
                # Handle standard equality condition
                where_parts.append(f"{key}=%s")
                filter_values.append(value)
        
        where_clause = ' AND '.join(where_parts)
        sql_query = f"UPDATE {table_name} SET {set_clause} WHERE {where_clause}"
 
        with transaction.atomic(using=db_alias):
            with connections[db_alias].cursor() as cursor:
                cursor.execute(sql_query, list(data.values()) + filter_values)
 
        return True
    except Exception as e:
        raise Exception(f"Error in update on database '{db_alias}': {e}")
 
# Helper: Delete data from table
def delete_data(table_name, filters, db_alias='default'):
    """
    Helper to delete data from a table from the specified database.
 
    Args:
        table_name (str): The name of the table to delete from.
        filters (dict): Filters for the WHERE clause to specify which rows to delete.
        db_alias (str): The alias for the database connection (default is 'default').
 
    Returns:
        bool: True if deletion successful, False otherwise.
    """
    try:
        where_clause = ' AND '.join([f"{key}=%s" for key in filters.keys()])
        sql_query = f"DELETE FROM {table_name} WHERE {where_clause}"
 
        # Get the database connection using the db_alias parameter
        with transaction.atomic(using=db_alias):
            with connections[db_alias].cursor() as cursor:
                cursor.execute(sql_query, list(filters.values()))
 
        return True  # Return True if deletion was successful
    except Exception as e:
        raise Exception(f"Error in delete: {e}")
 
# Helper: Insert data into table and get the inserted row's ID
def insert_get_id_data(table_name, data, column_id, db_alias='default'):
    """
    Helper to insert data into a table and return the inserted row's ID.
 
    Args:
        table_name (str): The name of the table where data will be inserted.
        data (dict): A dictionary containing the column names as keys and the values to be inserted.
        column_id (str): The name of the ID column (primary key).
        db_alias (str, optional): The database alias to use (default is 'default').
 
    Returns:
        int or None: The inserted row's ID if successful, None if failed.
    """
    try:
        if not data:
            raise ValueError("Insert data cannot be empty.")
 
        columns = ', '.join(data.keys())
        placeholders = ', '.join(['%s'] * len(data))
        sql_query = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders}) RETURNING {column_id}"
 
        with transaction.atomic(using=db_alias):
            with connections[db_alias].cursor() as cursor:
                cursor.execute(sql_query, list(data.values()))
                inserted_id = cursor.fetchone()[0]  # Get the inserted ID
 
        return inserted_id  # Return inserted ID
    except Exception as e:
        raise Exception(f"Error in insert_get_id on database '{db_alias}': {e}")
 
def first_data(table_name, filters=None, columns='*', order_by=None, db_alias='default'):
    """
    Helper to get the first row of data from a table.

    Args:
        table_name (str): The name of the table to read from.
        filters (dict, optional): Filters for the WHERE clause. 
            Supports equality (=) by default, not equal (__ne suffix, e.g., {'field__ne': value}),
            and IN (__in suffix, e.g., {'field__in': [value1, value2, ...]}).
        columns (str or list, optional): Columns to select, '*' by default.
        order_by (str or list, optional): Column or columns to order the results by.
        db_alias (str, optional): The database alias to use (default is 'default').

    Returns:
        dict or None: A dictionary representing the first row, or None if no data found.
    """
    try:
        if isinstance(columns, list):
            columns = ', '.join(columns)

        sql_query = f"SELECT {columns} FROM {table_name}"

        # Adding filters (WHERE clause)
        values = []
        if filters:
            where_conditions = []
            for key, value in filters.items():
                if '__ne' in key:       # not equal
                    field = key.split('__ne')[0]
                    operator = '<>'
                    where_conditions.append(f"{field} {operator} %s")
                    values.append(value)
                elif '__in' in key:     # IN clause
                    field = key.split('__in')[0]
                    if not isinstance(value, (list, tuple)) or not value:
                        raise ValueError(f"Value for {key} must be a non-empty list or tuple")
                    placeholders = ', '.join(['%s'] * len(value))
                    where_conditions.append(f"{field} IN ({placeholders})")
                    values.extend(value)
                else:                   # default equal
                    field = key
                    operator = '='
                    where_conditions.append(f"{field} {operator} %s")
                    values.append(value)

            sql_query += ' WHERE ' + ' AND '.join(where_conditions)

        # Adding order by clause
        if order_by:
            if isinstance(order_by, list):
                order_by = ', '.join(order_by)
            sql_query += f" ORDER BY {order_by}"

        sql_query += " LIMIT 1"  # Limit to the first row

        with connections[db_alias].cursor() as cursor:
            cursor.execute(sql_query, values)
            columns = [col[0] for col in cursor.description]
            row = cursor.fetchone()  # Get the first row

        return dict(zip(columns, row)) if row else None
    except Exception as e:
        raise Exception(f"Error in first_data on database '{db_alias}': {e}")
 
# Helper: Get the count of rows in a table
def count_data(table_name, filters=None, db_alias='default'):
    """
    Helper to get the count of rows in a table from the specified database.
 
    Args:
        table_name (str): The name of the table to count rows from.
        filters (dict, optional): Filters for the WHERE clause.
        db_alias (str): The alias for the database connection (default is 'default').
 
    Returns:
        int: The count of rows that match the filters, or 0 if no data found.
    """
    try:
        sql_query = f"SELECT COUNT(*) FROM {table_name}"
 
        # Adding filters (WHERE clause)
        if filters:
            where_clause = ' WHERE ' + ' AND '.join([f"{key}=%s" for key in filters.keys()])
            sql_query += where_clause
 
        # Get the database connection using the db_alias parameter
        with connections[db_alias].cursor() as cursor:
            cursor.execute(sql_query, list(filters.values()) if filters else [])
            count = cursor.fetchone()[0]  # Get the count
 
        return count
    except Exception as e:
        raise Exception(f"Error in count_data: {e}")
 
# Helper: Pluck data from a specific column in a table
def pluck_data(table_name, column_name, filters=None, db_alias='default'):
    """
    Helper to pluck values from a specific column in a table.
 
    Args:
        table_name (str): The name of the table to read from.
        column_name (str): The column to pluck data from.
        filters (dict, optional): Filters for the WHERE clause.
        db_alias (str, optional): The database alias to use (default is 'default').
 
    Returns:
        list: A list of values from the specified column.
    """
    try:
        sql_query = f"SELECT {column_name} FROM {table_name}"
 
        # Adding filters (WHERE clause)
        values = []
        if filters:
            where_conditions = [f"{key}=%s" for key in filters.keys()]
            sql_query += ' WHERE ' + ' AND '.join(where_conditions)
            values.extend(filters.values())
 
        with connections[db_alias].cursor() as cursor:
            cursor.execute(sql_query, values)
            rows = cursor.fetchall()  # Get all matching rows
 
        # Extract the specified column values into a list
        return [row[0] for row in rows]
    except Exception as e:
        raise Exception(f"Error in pluck_data on database '{db_alias}': {e}")
 
# Helper: Get distinct values from a specific column in a table
def distinct_data(table_name, column_name, filters=None, db_alias='default'):
    """
    Helper to get distinct values from a specific column in a table.
 
    Args:
        table_name (str): The name of the table to read from.
        column_name (str): The column to get distinct values from.
        filters (dict, optional): Filters for the WHERE clause.
        db_alias (str, optional): The database alias to use (default is 'default').
 
    Returns:
        list: A list of distinct values from the specified column.
    """
    try:
        sql_query = f"SELECT DISTINCT {column_name} FROM {table_name}"
 
        # Adding filters (WHERE clause)
        values = []
        if filters:
            where_conditions = [f"{key}=%s" for key in filters.keys()]
            sql_query += ' WHERE ' + ' AND '.join(where_conditions)
            values.extend(filters.values())
 
        with connections[db_alias].cursor() as cursor:
            cursor.execute(sql_query, values)
            rows = cursor.fetchall()  # Get all matching rows
 
        # Extract the specified column values into a list
        return [row[0] for row in rows]
    except Exception as e:
        raise Exception(f"Error in distinct_data on database '{db_alias}': {e}")
 
# Helper: Get ordered data from a table
def order_by_data(table_name, order_column, ascending=True, filters=None, limit=None, offset=None, db_alias='default'):
    """
    Helper to get ordered data from a table.
 
    Args:
        table_name (str): The name of the table to read from.
        order_column (str): The column to order by.
        ascending (bool): True for ascending order, False for descending. Default is True.
        filters (dict, optional): Filters for the WHERE clause.
        limit (int, optional): Limit for the number of records.
        offset (int, optional): Offset for records.
        db_alias (str, optional): The database alias to use (default is 'default').
 
    Returns:
        list of dict: A list of dictionaries representing the ordered rows.
    """
    try:
        sql_query = f"SELECT * FROM {table_name}"
 
        # Adding filters (WHERE clause)
        values = []
        if filters:
            where_conditions = [f"{key}=%s" for key in filters.keys()]
            sql_query += ' WHERE ' + ' AND '.join(where_conditions)
            values.extend(filters.values())
 
        # Adding ORDER BY clause
        order_direction = 'ASC' if ascending else 'DESC'
        sql_query += f" ORDER BY {order_column} {order_direction}"
 
        # Adding LIMIT and OFFSET
        if limit:
            sql_query += f" LIMIT {limit}"
        if offset:
            sql_query += f" OFFSET {offset}"
 
        with connections[db_alias].cursor() as cursor:
            cursor.execute(sql_query, values)
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
 
        return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        raise Exception(f"Error in order_by_data on database '{db_alias}': {e}")
 
# Helper: Check if data exists in a table
def exists_data(table_name, filters, id_column='id', exclude_id=None, db_alias='default'):
    """
    Helper to check if data exists in a table, with an option to exclude a specific ID.
 
    Args:
        table_name (str): The name of the table to check.
        filters (dict): Filters for the WHERE clause to specify which rows to check.
                        Supports '__ne', '__in', '__not_in', '__between', '__gt', '__lt', '__gte', '__lte'.
        id_column (str): The name of the ID column to check against.
        exclude_id (int, optional): The ID to exclude from the check.
        db_alias (str, optional): The database alias to use (default is 'default').
 
    Returns:
        bool: True if at least one row matches the filters, False otherwise.
    """
    try:
        where_clause = []
        params = []

        # Process each filter
        for key, value in filters.items():
            if key.endswith('__between'):
                # Handle BETWEEN operator
                column_name = key.replace('__between', '')
                if isinstance(value, list) and len(value) == 2:
                    where_clause.append(f"{column_name} BETWEEN %s AND %s")
                    params.extend(value)  # Add start and end values
                else:
                    raise ValueError(f"Invalid value for {key}: expected a list with two values")
            elif key.endswith('__ne'):
                # Handle not equal operator
                column_name = key.replace('__ne', '')
                where_clause.append(f"{column_name} != %s")
                params.append(value)
            elif key.endswith('__in'):
                # Handle IN operator
                column_name = key.replace('__in', '')
                if isinstance(value, (list, tuple)) and value:
                    placeholders = ', '.join(['%s'] * len(value))
                    where_clause.append(f"{column_name} IN ({placeholders})")
                    params.extend(value)
                else:
                    raise ValueError(f"Invalid value for {key}: expected a non-empty list or tuple")
            elif key.endswith('__not_in'):
                # Handle NOT IN operator
                column_name = key.replace('__not_in', '')
                if isinstance(value, (list, tuple)) and value:
                    placeholders = ', '.join(['%s'] * len(value))
                    where_clause.append(f"{column_name} NOT IN ({placeholders})")
                    params.extend(value)
                else:
                    raise ValueError(f"Invalid value for {key}: expected a non-empty list or tuple")
            elif key.endswith('__gte'):
                # Handle greater than or equal operator
                column_name = key.replace('__gte', '')
                where_clause.append(f"{column_name} >= %s")
                params.append(value)
            elif key.endswith('__lte'):
                # Handle less than or equal operator
                column_name = key.replace('__lte', '')
                where_clause.append(f"{column_name} <= %s")
                params.append(value)
            elif key.endswith('__gt'):
                # Handle greater than operator
                column_name = key.replace('__gt', '')
                where_clause.append(f"{column_name} > %s")
                params.append(value)
            elif key.endswith('__lt'):
                # Handle less than operator
                column_name = key.replace('__lt', '')
                where_clause.append(f"{column_name} < %s")
                params.append(value)
            else:
                # Standard equality filter
                where_clause.append(f"{key}=%s")
                params.append(value)

        # If exclude_id is provided, add the condition to exclude it
        if exclude_id is not None:
            where_clause.append(f"{id_column} != %s")
            params.append(exclude_id)

        # Combine all conditions with AND
        where_clause = ' AND '.join(where_clause)

        # Build and execute the query
        sql_query = f"SELECT EXISTS(SELECT 1 FROM {table_name} WHERE {where_clause})"

        with connections[db_alias].cursor() as cursor:
            cursor.execute(sql_query, params)
            exists = cursor.fetchone()[0]

        return exists
    except Exception as e:
        raise Exception(f"Error in exists_data on database '{db_alias}': {e}")
 
def fetch_records_with_conditions(table_name, null_column=None, not_null_column=None, additional_filters=None, db_alias='default'):
    """
    Fetch records from a specified table based on null or not null conditions and additional filters.
 
    :param table_name: The name of the table to query.
    :param null_column: The name of the column to check for NULL (use for IS NULL).
    :param not_null_column: The name of the column to check for NOT NULL.
    :param additional_filters: A dictionary of additional conditions (column: value).
    :param db_alias: The database alias to use (default is 'default').
    :return: A list of records with the specified column(s) conditions met.
    """
    try:
        with transaction.atomic(using=db_alias):
            with connections[db_alias].cursor() as cursor:
                # Start building the SQL query
                query = f"SELECT * FROM {table_name} WHERE 1=1"  # Start with a valid base
 
                params = []
 
                # Add IS NULL condition if specified
                if null_column:
                    query += f" AND {null_column} IS NULL"
 
                # Add IS NOT NULL condition if specified
                if not_null_column:
                    query += f" AND {not_null_column} IS NOT NULL"
 
                # Add additional filters if provided
                if additional_filters:
                    conditions = []
                    for column, value in additional_filters.items():
                        conditions.append(f"{column} = %s")
                        params.append(value)
                    # Join the conditions with 'AND'
                    query += " AND " + " AND ".join(conditions)
 
                # Execute the query
                cursor.execute(query, params)
 
                # Fetch all records and convert to a list of dictionaries
                rows = cursor.fetchall()
                columns = [col[0] for col in cursor.description]  # Get column names
                result = [dict(zip(columns, row)) for row in rows]
 
                return result
 
    except Exception as e:
        raise Exception(f"Error fetching records on database '{db_alias}': {str(e)}")   
 
# Helper: Get a single value from a table
def get_value(table_name, column_name, filters=None, type=None, db_alias='default'):
    """
    Helper to get a single value from a specified column in a table.
 
    Args:
        table_name (str): The name of the table to read from.
        column_name (str): The column from which to retrieve the value.
        filters (dict, optional): Filters for the WHERE clause. Supports equality (=) by default and not equal (__ne suffix, e.g., {'field__ne': value} for field != value).
        type (str, optional): Type of the value, e.g., "UUID" for UUID-specific error handling.
        db_alias (str, optional): The database alias to use (default is 'default').
 
    Returns:
        The value from the specified column, or raises an error if UUID is not found and type is "UUID".
    """
    try:
        sql_query = f"SELECT {column_name} FROM {table_name}"
 
        # Adding filters (WHERE clause)
        params = []
        if filters:
            where_parts = []
            for key, value in filters.items():
                if '__ne' in key:
                    field = key.split('__ne')[0]
                    operator = '<>'
                else:
                    field = key
                    operator = '='
                where_parts.append(f"{field} {operator} %s")
                params.append(value)
            
            where_clause = ' WHERE ' + ' AND '.join(where_parts)
            sql_query += where_clause
 
        sql_query += " LIMIT 1"  # Limit to the first row
 
        # Use the db_alias to get the correct connection for the desired database
        with connections[db_alias].cursor() as cursor:
            cursor.execute(sql_query, params)
            row = cursor.fetchone()  # Get the first row
 
        # If type is UUID and no row is found, raise specific UUID not found error
        if type == "UUID" and row is None:
            raise Exception("UUID tidak ditemukan")
 
        return row[0] if row else None  # Return the value of the specified column
    except Exception as e:
        if type != "UUID":
            raise Exception(f"Error in get_value: {e}")
        else:
            raise Exception("Error in get_value: UUID tidak ditemukan")  # Specific message for UUID not found
 
# Helper: Save an uploaded file to a specific directory
def save_uploaded_file(file, upload_dir='uploads'):
    """
    Save an uploaded file to a specific directory.
 
    Args:
        file: The file object to save.
        upload_dir (str): The directory to save the uploaded files.
 
    Returns:
        str: The file path of the saved file.
    """
    # Buat direktori jika belum ada
    upload_path = os.path.join(settings.MEDIA_ROOT, upload_dir)
    os.makedirs(upload_path, exist_ok=True)
 
    # Ambil nama file asli
    original_name = os.path.splitext(file.name)[0]
    extension = os.path.splitext(file.name)[1]
 
    # Buat timestamp dalam format yang diinginkan
    timestamp = datetime.now().strftime("%d%m%y_%H%M%S_%f")
 
    # Buat nama file baru
    new_file_name = f"{original_name}_{timestamp}{extension}"
    file_path = os.path.join(upload_path, new_file_name)
 
    # Cek jika file sudah ada, jika ada tambahkan angka
    base, ext = os.path.splitext(new_file_name)
    counter = 1
    while os.path.exists(file_path):
        new_file_name = f"{base}_{counter}{ext}"
        file_path = os.path.join(upload_path, new_file_name)
        counter += 1
 
    # Simpan file
    with open(file_path, 'wb+') as destination:
        for chunk in file.chunks():
            destination.write(chunk)
 
    return file_path
 
# Helper: sum data from a column in a table
def sum_data(table_name, column_name, filters=None, db_alias='default'):
    """
    Helper to sum data from a column in a table.
 
    Args:
        table_name (str): The name of the table.
        column_name (str): The name of the column to sum.
        filters (dict, optional): Filters for the WHERE clause.
        db_alias (str, optional): The database alias to use (default is 'default').
 
    Returns:
        float: The sum of the column data that matches the filters, or 0 if no data found.
    """
    try:
        sql_query = f"SELECT SUM({column_name}) FROM {table_name}"
 
        # Adding filters (WHERE clause)
        if filters:
            where_clause = ' WHERE ' + ' AND '.join([f"{key}=%s" for key in filters.keys()])
            sql_query += where_clause
 
        # Using the db_alias to connect to the correct database
        with connections[db_alias].cursor() as cursor:
            cursor.execute(sql_query, list(filters.values()) if filters else [])
            result = cursor.fetchone()[0]  # Get the sum result
 
        return result if result is not None else 0  # Return 0 if result is None (no rows)
 
    except Exception as e:
        raise Exception(f"Error in sum_data: {e}")
 
# Helper: Get the last row of data from a table
def last_data(table_name, filters=None, order_by_column='id', columns='*', db_alias='default'):
    """
    Helper to get the last row of data from a table.
 
    Args:
        table_name (str): The name of the table to read from.
        filters (dict, optional): Filters for the WHERE clause.
        order_by_column (str, optional): The column to order the results by (default is 'id').
        columns (str or list, optional): Columns to select, '*' by default.
        db_alias (str, optional): The database alias to use (default is 'default').
 
    Returns:
        dict or None: A dictionary representing the last row, or None if no data found.
    """
    try:
        if isinstance(columns, list):
            columns = ', '.join(columns)
 
        sql_query = f"SELECT {columns} FROM {table_name}"
 
        # Adding filters (WHERE clause)
        where_clause = ''
        if filters:
            where_clause = ' WHERE ' + ' AND '.join([f"{key}=%s" for key in filters.keys()])
            sql_query += where_clause
 
        # Order by the specified column in descending order
        sql_query += f" ORDER BY {order_by_column} DESC LIMIT 1"
 
        # Using db_alias to connect to the correct database
        with connections[db_alias].cursor() as cursor:
            cursor.execute(sql_query, list(filters.values()) if filters else [])
            columns = [col[0] for col in cursor.description]
            row = cursor.fetchone()  # Get the last row
 
        return dict(zip(columns, row)) if row else None
    except Exception as e:
        raise Exception(f"Error in last_data: {e}")
 
# Helper: Get the value user_id from uuid jwt
def jwt_uuid_conveter(uuid):
    user_id = get_value(
        table_name='sso.users',
        filters={
            'user_uuid': uuid
        },
        column_name="user_id"
    )
 
    if not user_id:
        return None
    return user_id
 
# Helper: Validate Request json from FE
def validate_request(data, rules):
    errors = {}
 
    for field, validations in rules.items():
        field_validations = validations.split('|')
 
        for validation in field_validations:
            # Required rule
            if validation == 'required' and field not in data:
                errors[field] = f"{field} is required."
                continue
 
            # Skip further checks if field is not present and it's not required
            if field not in data:
                continue
 
            value = data[field]
 
            # String rule
            if validation == 'string' and not isinstance(value, str):
                errors[field] = f"{field} must be a string."
 
            # Numeric rule
            if validation == 'numeric' and not isinstance(value, (int, float)):
                errors[field] = f"{field} must be numeric."
 
            # Boolean rule
            if validation == 'boolean' and not isinstance(value, bool):
                errors[field] = f"{field} must be true or false."
 
            # Email rule
            if validation == 'email':
                email_regex = r'^[\w\.-]+@[\w\.-]+\.\w+$'
                if not re.match(email_regex, value):
                    errors[field] = f"{field} must be a valid email."
 
            # Date rule
            if validation == 'date':
                try:
                    datetime.strptime(value, '%Y-%m-%d')
                except ValueError:
                    errors[field] = f"{field} must be a valid date (YYYY-MM-DD)."
 
            # Min length for strings or min value for numbers
            if validation.startswith('min:'):
                min_value = int(validation.split(':')[1])
                if isinstance(value, (int, float)) and value < min_value:
                    errors[field] = f"{field} must be at least {min_value}."
                if isinstance(value, str) and len(value) < min_value:
                    errors[field] = f"{field} must be at least {min_value} characters long."
 
            # Max length for strings or max value for numbers
            if validation.startswith('max:'):
                max_value = int(validation.split(':')[1])
                if isinstance(value, (int, float)) and value > max_value:
                    errors[field] = f"{field} must be no more than {max_value}."
                if isinstance(value, str) and len(value) > max_value:
                    errors[field] = f"{field} must be no more than {max_value} characters long."
 
            # In rule (check if value is in a list of allowed values)
            if validation.startswith('in:'):
                allowed_values = validation.split(':')[1].split(',')
                if value not in allowed_values:
                    errors[field] = f"{field} must be one of {', '.join(allowed_values)}."
 
            # Array rule
            if validation == 'array' and not isinstance(value, list):
                errors[field] = f"{field} must be an array."
 
            # Integer rule
            if validation == 'integer' and not isinstance(value, int):
                errors[field] = f"{field} must be an integer."
 
            # Float rule
            if validation == 'float' and not isinstance(value, float):
                errors[field] = f"{field} must be a floating-point number."
 
            # Digits rule (e.g. digits:5 means must be a 5-digit number)
            if validation.startswith('digits:'):
                num_digits = int(validation.split(':')[1])
                if not (isinstance(value, int) and len(str(value)) == num_digits):
                    errors[field] = f"{field} must be a {num_digits}-digit number."
 
            # URL rule
            if validation == 'url':
                url_regex = r'^(https?|ftp)://[^\s/$.?#].[^\s]*$'
                if not re.match(url_regex, value):
                    errors[field] = f"{field} must be a valid URL."
 
            # Starts with rule
            if validation.startswith('starts_with:'):
                prefix = validation.split(':')[1]
                if not value.startswith(prefix):
                    errors[field] = f"{field} must start with {prefix}."
 
            # Ends with rule
            if validation.startswith('ends_with:'):
                suffix = validation.split(':')[1]
                if not value.endswith(suffix):
                    errors[field] = f"{field} must end with {suffix}."
 
            # Unique rule (simulate by checking a list of existing values)
            if validation == 'unique':
                # Normally, you'd check a database for uniqueness
                # Simulate by checking against a list of existing values
                existing_values = ['example1', 'example2']  # Replace with actual database check
                if value in existing_values:
                    errors[field] = f"{field} must be unique."
 
    # Return errors if any are found
    if errors:
        return errors
 
    # No errors, validation passed
    return None
 
def validate_method(
    request,
    required_method,
    require_api_key=False,
    rate_limit=60,
    time_window=60,
    method_specific=True,
    block_bots=True
):
    """
    Validasi method, rate limiting, duplicate request, dan keamanan API Key.
    Args:
        request: Request object Django.
        required_method (str): Method yang diizinkan ("GET", "POST", dll).
        rate_limit (int): Maksimal request dalam `time_window` detik per IP dan endpoint.
        time_window (int): Durasi dalam detik untuk batasan rate limit.
        method_specific (bool): Validasi rate limit untuk method tertentu (True = spesifik per method, False = per IP global).
        require_api_key (bool): Apakah perlu API key dalam request header.
        block_bots (bool): Apakah harus memblokir akses dari bot atau user-agent tidak dikenal.
    Raises:
        ValueError: Jika terjadi pelanggaran aturan.
    """
    client_ip = get_client_ip(request)
    method = request.method.upper()
    now = time.monotonic()
    endpoint = request.path  # Menggunakan path sebagai bagian dari cache key

    # 1. Validasi Method
    if method != required_method.upper():
        raise ValueError(f"Method {method} not allowed. Use {required_method}.")
    
    # VALIDATE REQUEST [START]
    # VALIDATE REQUEST [START]
    # VALIDATE REQUEST [START]

    # 2. Cek apakah Content-Type valid (hanya menerima JSON atau multipart/form-data)
    content_type = request.headers.get("Content-Type", "")
    if not any(ct in content_type for ct in ["application/json", "multipart/form-data"]):
        raise ValueError("Invalid Content-Type. Only application/json or multipart/form-data is allowed.")

    # 3. Batasi ukuran request body
    max_size_bytes = MAX_TOTAL_SIZE_MB * 1024 * 1024  # 10 MB untuk multipart
    if request.META.get('CONTENT_LENGTH') and int(request.META.get('CONTENT_LENGTH')) > max_size_bytes:
        raise ValueError("Payload too large. Maximum allowed size is 10MB.")

    # 4. Validasi header User-Agent
    validate_user_agent(request)

    # 5. Validasi request body berdasarkan tipe content
    if "application/json" in content_type:
        validate_json_payload(request)
    elif "multipart/form-data" in content_type:
        validate_file_upload(request)
        
    # VALIDATE REQUEST [END]
    # VALIDATE REQUEST [END]
    # VALIDATE REQUEST [END]

    # 2. Rate Limiting per Method dan IP berdasarkan endpoint
    with cache_lock:
        if method_specific:
            # Memperhitungkan endpoint dalam rate limiting
            endpoint_cache = RATE_LIMIT_CACHE.setdefault(endpoint, {})
            ip_cache = endpoint_cache.setdefault(client_ip, {})
            method_cache = ip_cache.setdefault(method, deque(maxlen=rate_limit))
            while method_cache and now - method_cache[0] > time_window:
                method_cache.popleft()
            if len(method_cache) >= rate_limit:
                raise ValueError(
                    f"Rate limit exceeded for {method} on {endpoint} ({rate_limit} requests per {time_window} seconds). Try again later."
                )
            method_cache.append(now)
        else:
            # Rate limiting per IP dan endpoint
            endpoint_cache = RATE_LIMIT_CACHE.setdefault(endpoint, {})
            ip_cache = endpoint_cache.setdefault(client_ip, deque(maxlen=rate_limit))
            while ip_cache and now - ip_cache[0] > time_window:
                ip_cache.popleft()
            if len(ip_cache) >= rate_limit:
                raise ValueError(
                    f"Rate limit exceeded for {endpoint} ({rate_limit} requests per {time_window} seconds). Try again later."
                )
            ip_cache.append(now)

    # 3. Duplicate Request Limiting
    # request_signature = generate_request_signature(request, endpoint)
    with cache_lock:
        endpoint_signatures = REQUEST_SIGNATURES.setdefault(endpoint, {})
        ip_signatures = endpoint_signatures.setdefault(client_ip, deque(maxlen=duplicate_limit))
        while ip_signatures and now - ip_signatures[0] > duplicate_time_window:
            ip_signatures.popleft()
        if len(ip_signatures) >= duplicate_limit:
            raise ValueError("Duplicate request limit exceeded for this endpoint. Try again later.")
        ip_signatures.append(now)

    # 4. Validasi API Key (Opsional)
    if require_api_key:
        api_key = request.headers.get("X-API-KEY")

        # Pastikan API key ada
        if not api_key:
            raise ValueError("401 Unauthorized: Missing API Key")
        
        # Menentukan periode 6 bulan
        current_month = datetime.utcnow().month
        year = datetime.utcnow().year
        period = f"{year}-{'01' if current_month <= 6 else '07'}"  # Januari-Juni -> periode 01, Juli-Desember -> periode 07

        # Membuat API key yang diharapkan
        expected_api_key = hmac.new(SECRET_KEY.encode(), period.encode(), hashlib.sha256).hexdigest()

        # Bandingkan API key yang diterima dengan yang diharapkan
        if api_key != expected_api_key:
            raise ValueError("Unauthorized: Invalid API Key")

    # 5. Blokir User-Agent yang Mencurigakan (Anti Bot & Scraper)
    if block_bots:
        user_agent = request.headers.get("User-Agent", "").lower()
        if not user_agent or "bot" in user_agent or "crawler" in user_agent or len(user_agent) < 10:
            raise ValueError("Access denied: Suspicious user-agent detected.")

# Helper untuk Mendapatkan IP Client
def get_client_ip(request):
    forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR", "0.0.0.0")

# Helper untuk Membuat Signature Unik dari Request berdasarkan Endpoint
def generate_request_signature(request, endpoint):
    """
    Membuat signature unik menggunakan hash SHA-256 untuk mencegah replay attack,
    memperhitungkan endpoint dalam signature.
    """
    data = (
        f"{request.method}-"
        f"{request.META.get('REMOTE_ADDR')}-"
        f"{request.headers.get('User-Agent', '')}-"
        f"{request.headers.get('Content-Length', '')}-"
        f"{str(request.GET)}-"
        f"{str(request.body)}-"
        f"{endpoint}"  # Menambahkan endpoint ke dalam signature
    )
    hash_digest = hashlib.sha256(data.encode()).digest()
    return base64.b64encode(hash_digest).decode()

# Fungsi untuk memvalidasi payload JSON
def validate_json_payload(request):
    """Validasi payload JSON dari request."""

    if request.body:
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            raise ValueError("Invalid JSON format.")
        except Exception:
            raise ValueError("Invalid request payload.")

        # Batasi ukuran JSON payload
        if len(request.body) > (MAX_JSON_SIZE_MB * 1024 * 1024):
            raise ValueError(f"JSON payload too large. Maximum allowed size is {MAX_JSON_SIZE_MB}MB.")

        # Deteksi input mencurigakan secara rekursif
        if contains_malicious_input(data):
            raise ValueError("Potential hacking detected.")

def validate_file_upload(request):
    """Validasi file upload dan field formulir dari request multipart/form-data."""

    # 1. Validasi field formulir (request.POST)
    form_data = request.POST
    if contains_malicious_input(form_data):
        raise ValueError("Potential hacking detected in form fields.")

    # 2. Validasi file yang diunggah
    files = []
    for key in request.FILES:
        files.extend(request.FILES.getlist(key))

    if len(files) > MAX_FILES_ALLOWED:
        raise ValueError(f"Too many files uploaded. Maximum {MAX_FILES_ALLOWED} files are allowed.")

    total_size = 0
    for uploaded_file in files:
        file_size_mb = uploaded_file.size / (1024 * 1024)
        total_size += file_size_mb

        # Validasi ukuran per file
        if file_size_mb > MAX_FILE_SIZE_MB:
            raise ValueError(f"File '{uploaded_file.name}' is too large. Maximum allowed size is {MAX_FILE_SIZE_MB}MB.")

        # Validasi ekstensi file
        file_extension = uploaded_file.name.split(".")[-1].lower()
        if file_extension not in ALLOWED_FILE_EXTENSIONS:
            raise ValueError(f"File '{uploaded_file.name}' has an invalid extension. Allowed: {', '.join(ALLOWED_FILE_EXTENSIONS)}.")

        # Validasi MIME type menggunakan mimetypes
        mime_type, _ = mimetypes.guess_type(uploaded_file.name, strict=True)
        expected_mime = ALLOWED_MIME_TYPES.get(file_extension)
        if mime_type != expected_mime:
            raise ValueError(f"File '{uploaded_file.name}' has an invalid MIME type: {mime_type or 'unknown'}. Expected: {expected_mime}.")

        # Validasi isi file untuk file teks (opsional)
        if file_extension == "txt":
            try:
                file_content = uploaded_file.read().decode("utf-8")
                if contains_malicious_input(file_content):
                    raise ValueError(f"File '{uploaded_file.name}' contains potentially malicious content.")
            except UnicodeDecodeError:
                raise ValueError(f"File '{uploaded_file.name}' is not a valid text file.")
            finally:
                uploaded_file.seek(0)  # Reset pointer

    # Validasi total ukuran
    if total_size > MAX_TOTAL_SIZE_MB:
        raise ValueError(f"Total file upload size exceeds {MAX_TOTAL_SIZE_MB}MB.")

# Pola berbahaya yang lebih spesifik (tidak termasuk 'id' yang terlalu generik)
suspicious_patterns = [
    # tetap cek script/iframe/meta/img javascript secara langsung pada HTML mentah
    r"<script\b[^>]*>.*?</script>",
    r"(?i)\b(on\w+\s*=\s*['\"][^'\"]*['\"])",  # attribute event handler
    r"(?i)<iframe\b[^>]*src=['\"][^'\"]*['\"][^>]*>.*?</iframe>",
    r"(?i)<meta\b[^>]*http-equiv=['\"]refresh['\"][^>]*content=['\"][^'\"]*['\"][^>]*>",
    r"(?i)<img\b[^>]*src=['\"]javascript:[^'\"]*['\"][^>]*>",
    # SQL injection — lebih ketat, jangan gunakan single-token yang umum
    r"(?i)\b(union\s+select|drop\s+table|--\s|/\*|\*/|;\s*xp_cmdshell|exec\s+xp_)\b",
    r"(?i)\b(select\s+.*?from|insert\s+into|update\s+.*?set|delete\s+from)\s*['\"]",
    r"(?i)\b(load_file|outfile|dumpfile|sleep\(\d+\)|benchmark\(\d+,\d+\))\b",
    # JS API/function calls — pastikan ada tanda () untuk meminimalkan false positive
    r"(?i)\b(alert|confirm|prompt|document\.cookie|eval|setTimeout|setInterval|Function)\s*\(",
    r"(?i)\b(fetch|XMLHttpRequest|webkitRequestFileSystem)\s*\(",
    r"(?i)\b(window\.location|history\.go|localStorage|sessionStorage)\s*[=\.]",
    # Dangerous URI schemes (cek pada teks setelah decode juga)
    r"(?i)%3Cscript%3E|%3Ciframe%3E|%3Cimg%20src=['\"]javascript:|%3Cbody%20onload|%3Conerror=",
    r"(?i)\b(data:text/html|javascript:|vbscript:|file://)\b",
    # Command injection: buat lebih spesifik (hapus token 'id')
    r"(?i)\b(cat\s+/etc/passwd|ls\s+-la|whoami\b|uname\s+-a|rm\s+-rf|chmod\s+777)\b",
    # Path traversal
    r"(?i)\b(\.\./|\.\.\\|/etc/passwd|c:\\windows\\system32)\b",
]

def contains_malicious_input(data: Any) -> bool:
    """
    Versi yang lebih robust:
    - Jika input adalah HTML, kita lihat pola-pola HTML (script/iframe/img javascript) pada teks mentah.
    - Untuk pola 'perintah' dan token lain yang raw, hapus tag HTML dulu (so '.go.id' tidak matched).
    """
    if isinstance(data, dict):
        return any(contains_malicious_input(v) for v in data.values())
    if isinstance(data, list):
        return any(contains_malicious_input(v) for v in data)
    if isinstance(data, str):
        s = data.strip()

        # 1) Cek pola HTML-berbahaya langsung terhadap teks mentah (jangan menghapus tag dulu)
        for pat in suspicious_patterns[:5]:  # pola-pola awal yang ingin deteksi dalam HTML mentah
            if re.search(pat, s, flags=re.IGNORECASE | re.DOTALL):
                return True

        # 2) Hapus tag HTML sebelum cek pola generik yang bisa terpicu oleh domain/atribut
        s_no_tags = re.sub(r"<[^>]+>", " ", s)

        # 3) decode percent-encoding untuk mendeteksi encoded attacks (opsional)
        try:
            from urllib.parse import unquote
            s_decoded = unquote(s_no_tags)
        except Exception:
            s_decoded = s_no_tags

        # 4) cek semua pola yang tersisa pada teks tanpa tag (dan setelah decode)
        for pat in suspicious_patterns[5:]:
            if re.search(pat, s_decoded, flags=re.IGNORECASE):
                return True

    return False

def validate_user_agent(request) -> None:
    """
    Validasi header User-Agent untuk mendeteksi permintaan mencurigakan.
    Args:
        request: Request object Django.
    Raises:
        ValueError: Jika User-Agent mencurigakan atau tidak valid.
    """
    user_agent: str = request.headers.get("User-Agent", "").strip()
    
    # 1. Cek apakah User-Agent ada dan memiliki panjang minimal
    if not user_agent:
        raise ValueError("Missing User-Agent. Request seems suspicious.")
    if len(user_agent) < 10:
        # Izinkan pengecualian untuk bot yang sah jika diperlukan
        if not any(bot.lower() in user_agent.lower() for bot in ALLOWED_BOTS.values()):
            raise ValueError(f"User-Agent too short ({len(user_agent)} chars). Request seems suspicious.")

    # 2. Cek pola User-Agent mencurigakan
    for pattern in SUSPICIOUS_UA_PATTERNS:
        if re.search(pattern, user_agent):
            # Izinkan bot yang dikenal (opsional)
            if any(bot.lower() in user_agent.lower() for bot in ALLOWED_BOTS.values()):
                return  # Bot sah, lanjutkan
            raise ValueError(f"Suspicious User-Agent detected: {user_agent}")

    # 3. Cek apakah User-Agent terlalu generik atau tidak wajar
    if user_agent.lower() in ["mozilla", "browser", "default", "user-agent"]:
        raise ValueError("Generic User-Agent detected. Request seems suspicious.")

    # 4. (Opsional) Cek header tambahan untuk konsistensi
    accept_header = request.headers.get("Accept", "")
    if not accept_header and "html" not in accept_header.lower():
        # Jika tidak ada Accept header atau tidak menerima HTML, kemungkinan bukan browser
        if not any(bot.lower() in user_agent.lower() for bot in ALLOWED_BOTS.values()):
            raise ValueError("Inconsistent headers with User-Agent. Request seems suspicious.")

def log_exception(request, exception):
    try:
        # Parse traceback to get the file and line where the error occurred
        tb = traceback.extract_tb(exception.__traceback__)
        if tb:
            file_location = f"{tb[0].filename} line {tb[0].lineno}"
            error_message = f"{tb[-1].filename} line {tb[-1].lineno}: {tb[-1].line} - {str(exception)}"
        else:
            file_location = "Unknown file"
            error_message = f"{tb[-1].filename} line {tb[-1].lineno}: {tb[-1].line} - {str(exception)}"
 
        # Convert QueryDict to a JSON string
        query_params = json.dumps(request.GET.dict())
 
        # Insert log entry
        insert_data(
            table_name="log.tbl_log",
            data={
                "method": request.method,
                "path": request.path,
                "query_params": query_params,
                "client_ip": request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR')),
                "user_agent": request.META.get('HTTP_USER_AGENT'),
                "employee_uuid": getattr(request, 'jwt_uuid', None),
                "employee_no": getattr(request, 'jwt_badge_no', None),
                "employee_name": getattr(request, 'jwt_fullname', None),
                "file_location": file_location,
                "message_error": error_message,
                "created_at": datetime.now()
            }
        )
    except Exception as e:
        print(f"Failed to log exception: {str(e)}")
 
 
def generate_custom_excel(data, headers, title,):
    """
    Generate an Excel file in memory with customizable content.
 
    Parameters:
    - data (list of dict): The data to populate the Excel file, where each dict represents a row.
    - headers (list of str): Column headers for the Excel sheet.
    - title (str): Title for the report.
 
    Returns:
    - BytesIO: An in-memory Excel file.
    """
    # Create a new workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = title
 
    # Title Row
    ws['A1'] = title
    ws['A1'].font = Font(size=20, bold=True, color="000000")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
 
    # Subtitle Row with timestamp
    timestamp = datetime.now().strftime('%d %b %Y %H:%M:%S')
    ws['A2'] = f"Generated at : {timestamp}"
    ws['A2'].font = Font(size=11, italic=True, color="555555")
    ws['A2'].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
 
    # Header Row (starts from W4)
    header_row_index = 4
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_index, column=col_num, value=header)
        cell.font = Font(bold=True, size=12, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.border = Border(
            top=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
 
    # Populate Data Rows
    for row_index, row_data in enumerate(data, start=header_row_index + 1):
        for col_index, header in enumerate(headers, start=1):
            cell_value = row_data.get(header)
 
            # Check if the value is a date (datetime or date object)
            if isinstance(cell_value, (datetime, date)):
                # Format the date to the desired format (e.g., 'dd-mm-yyyy')
                cell_value = cell_value.strftime('%d-%m-%Y')
 
            cell = ws.cell(row=row_index, column=col_index, value=cell_value)
            cell.alignment = Alignment(horizontal="left" if header in ["Meeting title","Meeting room", "Date Meeting", "Start Meeting", "Finish Meeting", "Participant", "Request", "Booking Date", "Status","Meeting Qty"] else "center", vertical="center")
            cell.border = Border(
                top=Side(border_style="thin", color="000000"),
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )
 
    # Adjust column widths
    column_widths = {
        "A": 10, "B": 40, "C": 20, "D": 20, "E": 20,"F": 20,"F": 20,"G": 30,"H": 30,"I": 30
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
 
    # Save workbook to an in-memory buffer
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)  # Move cursor to the start of the buffer
 
    return output_buffer
 
def generate_excel_from_template(data, url=None, output_file_name="dashboard_report.xlsx"):
    """
    Generate an Excel file from data, including dynamic filter and report title.
    :param data: Data to be written in the Excel report (list of sheets, each containing 'sheet_name', 'headers', 'data').
    :param url: URL string to extract filter parameters, default is None.
    :param output_file_name: Name of the output Excel file, default is "dashboard_report.xlsx".
    :return: The path to the generated Excel file.
    """
    # Function to extract parameters from URL
    def extract_url_params(url):
        if not url:
            return "", "", "", ""  # Default empty values if URL is None
 
        parsed_url = urlparse(url)
        query_params = parse_qs(parsed_url.query)
 
        # Extract individual query parameters
        year = query_params.get('year', [''])[0]
        month = query_params.get('month', [''])[0]
        department = query_params.get('department', [''])[0]
        floor = query_params.get('floor', [''])[0]
 
        return year, month, department, floor
 
    # Extract parameters from URL (if provided)
    year, month, department, floor = extract_url_params(url)
 
    # Set up output folder to be the user's Downloads folder
    home_dir = os.path.expanduser("~")
    downloads_folder = os.path.join(home_dir, "Downloads")
 
    # Ensure the Downloads directory exists
    if not os.path.exists(downloads_folder):
        os.makedirs(downloads_folder)
 
    output_file = os.path.join(downloads_folder, output_file_name)
 
    # Check if file already exists and add a number if needed
    base_name, ext = os.path.splitext(output_file_name)
    counter = 1
    while os.path.exists(output_file):
        output_file = os.path.join(downloads_folder, f"{base_name}({counter}){ext}")
        counter += 1
 
    try:
        with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
            for sheet in data:
                sheet_name = sheet['sheet_name']
                headers = sheet['headers']
                data_rows = sheet['data']
 
                # Reformat the data to match the required format
                formatted_data = []
                for row in data_rows:
                    formatted_row = [row['label']]  # The first column will contain values like "Owl", "Nowl", "Total"
                    for month in headers:
                        formatted_row.append(row.get(month, 0))  # Default to 0 if no value exists
                    formatted_data.append(formatted_row)
 
                # Create the DataFrame with the first column as 'label' and the months as headers
                df = pd.DataFrame(formatted_data, columns=['label'] + headers)
 
                # Write the data to the Excel file without the header (since we're writing it manually)
                df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=4, header=False)
 
                # Access the workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
 
                # Header format (for the months)
                header_format = workbook.add_format({
                    'bold': True,
                    'text_wrap': True,
                    'align': 'center',
                    'valign': 'vcenter',
                    'bg_color': '#FFFF00',
                    'border': 1
                })
 
                # Title format (for the sheet title)
                title_format = workbook.add_format({
                    'bold': True,
                    'font_size': 16,
                    'align': 'center',
                    'valign': 'vcenter',
                    'font_color': '#000000'
                })
 
                # Filter label format
                label_format = workbook.add_format({
                    'bold': True,
                    'font_size': 12,
                    'align': 'left',
                    'valign': 'vcenter',
                })
 
                # Dynamically calculate the column range based on headers length
                last_col = len(headers)
                title_range = f"A1:{chr(65 + last_col)}1"  # Adjust for the column length dynamically
                worksheet.merge_range(title_range, f"{sheet_name}", title_format)
 
                # Create dynamic filter string, excluding empty filters
                filters = []
                if year:
                    filters.append(f"{year}")
                if month:
                    filters.append(f"{month}")
                if department:
                    filters.append(f"{department}")
                if floor:
                    filters.append(f"{floor}")
 
                # Join the filters with commas if any are present
                filter_by_text = f"Filter by: {', '.join(filters)}" if filters else "Filter by: N/A"
                generate_on_text = f"Generate on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
 
                # Write the filter and generation timestamp
                worksheet.write('A2', filter_by_text, label_format)
                worksheet.write('A3', generate_on_text, label_format)
 
                # Write the headers in row 4 (just below the filter)
                for col_num, header in enumerate(headers, start=1):
                    worksheet.write(3, col_num, header, header_format)
 
                # Auto-adjust column widths based on the content
                for i, column in enumerate(df.columns):
                    max_width = max(df[column].astype(str).map(len).max(), len(column)) + 2
                    worksheet.set_column(i, i, max_width)
 
                # Apply border around the entire data range (excluding title and filter rows)
                last_row = len(formatted_data) + 4  # Data starts from row 4
                last_col = len(headers) + 1  # Including 'label' column
 
                # Set border around the data cells (A4:Z[last_row])
                border_format = workbook.add_format({'border': 1})
 
                # Apply border only to the data range (A4:Z[last_row])
                worksheet.conditional_format(f"A4:{chr(64 + last_col)}{last_row}", {
                    'type': 'no_blanks',
                    'format': border_format
                })
 
                # Apply specific color to cells in column A for rows 5-7 without overwriting data
                row_color_format = workbook.add_format({
                    'bg_color': '#dddedd'
                })
                worksheet.conditional_format("A5:A7", {
                    'type': 'no_blanks',
                    'format': row_color_format
                })
 
        return output_file
 
    except Exception as e:
        raise Exception(f"Error saving Excel file: {e}")
    
def execute_query_with_pagination(
    request,
    sql_query,
    params = None,
    db_alias = 'default',
    default_page = 1,
    default_per_page = 10,
    order_by = None
) -> dict:
    """
    Executes a raw SQL query with pagination on the specified database and returns paginated results.
 
    Args:
        request: HttpRequest object to get pagination parameters and build URLs.
        sql_query: The base SQL query to execute.
        params: List of parameters to safely pass to the query.
        db_alias: The alias of the database to use.
        default_page: Default page number if not specified in request.
        default_per_page: Default items per page if not specified in request.
        order_by: Optional ORDER BY clause for sorting results.
 
    Returns:
        dict: A dictionary containing paginated data and metadata:
            - rows: List of result rows
            - current_page: Current page number
            - page_size: Items per page
            - total_pages: Total number of pages
            - total_rows: Total number of rows
            - last_page: Last page number
            - next_page_url: URL for next page (or None)
            - prev_page_url: URL for previous page (or None)
    """
    try:
        # Initialize variables
        params = params or []
        pagination = {}
        
        # Get pagination parameters from request
        try:
            page = int(request.GET.get("page", default_page))
            per_page = int(request.GET.get("page_size", default_per_page))
        except ValueError:
            page = default_page
            per_page = default_per_page
        
        offset = (page - 1) * per_page
        
        # Count total rows using execute_query
        count_query = f"SELECT COUNT(*) as total FROM ({sql_query}) AS total_data"
        total_rows = execute_query(count_query, params, db_alias)[0]["total"]
        
        # Calculate pagination metadata
        total_pages = (total_rows + per_page - 1) // per_page
        last_page = total_pages
        
        # Prepare paginated query
        paginated_params = params.copy()
        paginated_query = sql_query
        if order_by:
            paginated_query += f" {order_by}"
        paginated_query += " LIMIT %s OFFSET %s"
        paginated_params.extend([per_page, offset])
        
        # Execute paginated query using execute_query
        data = execute_query(paginated_query, paginated_params, db_alias)
        
        # Build pagination URLs
        base_url = request.build_absolute_uri(request.path)
        query_params = request.GET.copy()
        
        next_page_url = None
        prev_page_url = None
        
        if page < total_pages:
            query_params['current_page'] = str(page + 1)
            next_page_url = f"{base_url}?{urlencode(query_params)}"
        
        if page > 1:
            query_params['current_page'] = str(page - 1)
            prev_page_url = f"{base_url}?{urlencode(query_params)}"
        
        # Construct response
        pagination = {
            "current_page": page,
            "page_size": per_page,
            "total_pages": total_pages,
            "total_rows": total_rows,
            "last_page": last_page,
            "next_page_url": next_page_url,
            "prev_page_url": prev_page_url,
            "rows": data,
        }
        
        return pagination
    
    except Exception as e:
        raise Exception(f"Error executing paginated query on database '{db_alias}': {e}")
    
def get_data_with_pagination(
    request,
    table_name,
    filters=None,
    search=None,
    search_columns=None,
    columns='*',
    order_by=None,
    db_alias='default',
    default_page=1,
    default_per_page=10
) -> dict:
    """
    Helper to read data from a table with pagination, supporting operators like =, !=, IN, NOT IN, >, <, >=, <=, OR, NULL handling, and BETWEEN for dates.
 
    Args:
        request: HttpRequest object to get pagination parameters and build URLs.
        table_name: The name of the table.
        filters: Conditions to filter data (supports =, !=, IN, NOT IN, >, <, >=, <=, OR, NULL, and BETWEEN).
        search: Search keyword.
        search_columns: Columns to perform search on.
        columns: Columns to select, default is '*'.
        order_by: Column(s) to order by.
        db_alias: The database alias to use (default is 'default').
        default_page: Default page number if not specified in request.
        default_per_page: Default items per page if not specified in request.
 
    Returns:
        dict: A dictionary containing paginated data and metadata:
            - rows: List of result rows
            - current_page: Current page number
            - page_size: Items per page
            - total_pages: Total number of pages
            - total_rows: Total number of rows
            - last_page: Last page number
            - next_page_url: URL for next page (or None)
            - prev_page_url: URL for previous page (or None)
    """
    try:
        # Get pagination parameters from request
        try:
            page = int(request.GET.get("page", default_page))
            per_page = int(request.GET.get("page_size", default_per_page))
        except ValueError:
            page = default_page
            per_page = default_per_page
        
        offset = (page - 1) * per_page
        
        # Execute count query using get_data
        total_rows = get_data(
            table_name=table_name,
            filters=filters,
            search=search,
            search_columns=search_columns,
            columns=['COUNT(*) as total'],
            db_alias=db_alias
        )[0]["total"]
        
        # Calculate pagination metadata
        total_pages = (total_rows + per_page - 1) // per_page
        last_page = total_pages
        
        # Execute paginated query using get_data
        data = get_data(
            table_name=table_name,
            filters=filters,
            search=search,
            search_columns=search_columns,
            columns=columns,
            limit=per_page,
            offset=offset,
            order_by=order_by,
            db_alias=db_alias
        )
        
        # Build pagination URLs
        base_url = request.build_absolute_uri(request.path)
        query_params = request.GET.copy()
        
        next_page_url = None
        prev_page_url = None
        
        if page < total_pages:
            query_params['current_page'] = str(page + 1)
            next_page_url = f"{base_url}?{urlencode(query_params)}"
        
        if page > 1:
            query_params['current_page'] = str(page - 1)
            prev_page_url = f"{base_url}?{urlencode(query_params)}"
        
        # Construct response
        pagination = {
            "current_page": page,
            "page_size": per_page,
            "total_pages": total_pages,
            "total_rows": total_rows,
            "last_page": last_page,
            "next_page_url": next_page_url,
            "prev_page_url": prev_page_url,
            "rows": data,
        }
        
        return pagination
    
    except Exception as e:
        raise Exception(f"Error in read with pagination from database '{db_alias}': {e}")

def upload_file_wsapto(file_obj, folder_name="", api_key="mysatnusa-file", base_url="http://192.168.88.60:40020/api/file-management"):
    """
    Helper untuk upload file ke file-management API.

    Args:
        file_obj (str | file-like): Path string atau InMemoryUploadedFile/File.
        folder_name (str): Folder tujuan di server (opsional).
        api_key (str): API Key (default: 'mysatnusa-file').
        base_url (str): Base URL API file management.

    Returns:
        dict: Response JSON dari API (berisi file_url, relative_path, dll).
    """
    try:
        url = f"{base_url}/upload_file"

        headers = {
            "x-api-key-mysatnusa": api_key,
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        }

        # kalau input berupa string → anggap path file
        if isinstance(file_obj, str):
            files = {"file": open(file_obj, "rb")}
        else:
            # kalau input sudah file-like object (misalnya InMemoryUploadedFile)
            files = {"file": (getattr(file_obj, "name", "upload.bin"), file_obj)}

        data = {"folder_name": folder_name} if folder_name else {}

        response = requests.post(url, headers=headers, files=files, data=data)

        if response.status_code == 200:
            return response.json()
        else:
            return {"status_code": response.status_code, "message": response.text, "messagetype": "E", "data": []}

    except Exception as e:
        return {"status_code": 400, "message": f"Error in upload_file_api: {e}", "messagetype": "E", "data": []}
