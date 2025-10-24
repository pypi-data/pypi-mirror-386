import logging
import re

from openpyxl import Workbook
from openpyxl.styles import PatternFill

logger = logging.getLogger('__main__')


class stylized_excel:
    """
    Class to streamline openpxyl stylizing when saving Excel sheets

    """

    origin_point = (0, 0)

    def __init__(self, data,
                 style_func=None,
                 origin_offset=None,
                 sheetnames=None,
                 ignore_values=False,
                 ignore_fill=True):

        """
        data : Pandas Dataframe or list of dataframes
        style_func : callable(data) or list of callable(data) functions corresponding to data list that returns hex color code or ARGB color code from dataframe index values
        origin_offset : positive value tuple pair for offsetting of origin point from 0,0 {col#1, row#1}
                        eg. origin_offset = (2,5) offsets origin to {col#3, row#6}

        sheetname : list of sheetnames corresponding to given data parameter, must be same length as the data parameter

        ignore_values : Do not write dataframe values into cells. Useful when only coloring is desired
        ignore_fill : Skip coloring of cells, defaults to white background

        """

        self._data = [data] if not isinstance(data, list) else data
        self._styling_func = [style_func] if not isinstance(style_func, list) else style_func

        if origin_offset is not None:
            self.origin_point = origin_offset

        if sheetnames is not None:
            self.sheetnames = sheetnames
        else:
            self.sheetnames = [f'Sheet{i + 1}' for i, _ in enumerate(self._data)]

        self._ignore_fill = ignore_fill
        self._ignore_values = ignore_values

    def create_file(self):

        self.workbook = Workbook()

        for sheet in self.sheetnames:
            self.workbook.create_sheet(sheet)

            self.write_to_sheet(sheet)

        return self

    def write_to_sheet(self, sheetname):

        if sheetname in self.workbook.sheetnames:
            target_sheet = self.workbook[sheetname]

            data = self._data[self.sheetnames.index(sheetname)]
            try:
                style_callable = self._styling_func[self.sheetnames.index(sheetname)]
            except IndexError:
                style_callable = self._styling_func[0]

            # write index first
            for r, index_str in enumerate(data.index):
                # offset 1 row to make room for columns
                target_sheet.cell(column=1, row=r + 2, value=index_str)

            for c, col_str in enumerate(data.columns):
                # offset 1 col to make room for rows
                target_sheet.cell(column=c + 2, row=1, value=col_str)

            rows, cols = data.shape

            for row_ind in range(rows):
                for col_ind in range(cols):
                    check_val = data.iloc[row_ind, col_ind]
                    if not self._ignore_values:
                        target_sheet.cell(column=col_ind + 2, row=row_ind + 2, value=check_val)

                    if not self._ignore_fill:
                        target_sheet.cell(column=col_ind + 2, row=row_ind + 2).fill = PatternFill("solid", fgColor=self.verify_color_format(
                            style_callable(check_val)))
        else:
            return

    def save_file(self, savepath):
        self.workbook.save(savepath)

    @staticmethod
    def verify_color_format(color_string):
        if re.search("^#[A-Za-z0-9]{6}$", color_string) is not None:
            color_string = color_string.replace('#', 'ff')  # default to 100% opacity
        return color_string
