import asyncio
import csv
import os
from io import StringIO
from typing import Optional, Tuple

from openpyxl import load_workbook

from ..common.logger import logger
from .base_handler import BaseHandler

_BOM_UTF8 = b"\xef\xbb\xbf"
_BOM_UTF16_LE = b"\xff\xfe"
_BOM_UTF16_BE = b"\xfe\xff"


def _escape_md_cell(s: str) -> str:
    return s.replace("|", r"\|").replace("\n", " ").replace("\r", " ").strip()


def _strip_bom_char(s: str) -> str:
    return s.lstrip("\ufeff")


def _detect_utf_encoding(path: str) -> Tuple[str, bool]:
    with open(path, "rb") as f:
        head = f.read(4)
    if head.startswith(_BOM_UTF8):
        return ("utf-8-sig", True)
    if head.startswith(_BOM_UTF16_LE) or head.startswith(_BOM_UTF16_BE):
        # Use 'utf-16' umbrella codec so Python handles BOM and endianness correctly.
        return ("utf-16", True)
    return ("utf-8", False)


def _csv_to_markdown(path: str, delimiter: str, encoding: Optional[str] = None) -> str:
    chosen_enc = encoding or _detect_utf_encoding(path)[0]
    out = StringIO()
    with open(path, "r", newline="", encoding=chosen_enc, errors="replace") as f:
        reader = csv.reader(f, delimiter=delimiter)
        try:
            header = next(reader)
        except StopIteration:
            return ""
        header = ["" if h is None else _strip_bom_char(str(h)) for h in header]
        out.write("| " + " | ".join(_escape_md_cell(h) for h in header) + " |\n")
        out.write("| " + " | ".join(["---"] * len(header)) + " |\n")
        for row in reader:
            row = ["" if c is None else str(c) for c in row]
            # If a stray BOM slipped into the first cell of first data row (rare), strip it too:
            if row:
                row[0] = _strip_bom_char(row[0])
            out.write("| " + " | ".join(_escape_md_cell(c) for c in row) + " |\n")
    return out.getvalue()


def _xlsx_to_markdown(path: str) -> str:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    out = StringIO()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out.write(f"\n\n### Sheet: {sheet_name}\n\n")
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            out.write("_(empty sheet)_\n")
            continue
        if header_row is None:
            out.write("_(empty sheet)_\n")
            continue
        header_vals = ["" if v is None else str(v) for v in header_row]
        out.write("| " + " | ".join(_escape_md_cell(v) for v in header_vals) + " |\n")
        out.write("| " + " | ".join(["---"] * len(header_vals)) + " |\n")
        for row in rows_iter:
            if row is None:
                out.write("| " + " | ".join("" for _ in header_vals) + " |\n")
                continue
            cells = ["" if v is None else str(v) for v in row]
            if len(cells) < len(header_vals):
                cells += [""] * (len(header_vals) - len(cells))
            elif len(cells) > len(header_vals):
                cells = cells[: len(header_vals)]
            out.write("| " + " | ".join(_escape_md_cell(v) for v in cells) + " |\n")
    return out.getvalue()


class TabularHandler(BaseHandler):
    """Handler for .csv, .tsv, .xls, .xlsx files."""

    SUPPORTED_EXTENSIONS = frozenset([".csv", ".tsv", ".xls", ".xlsx"])

    async def handle(self, file_path: str, *args, **kwargs) -> Optional[str]:
        logger.info(f"TabularHandler: Processing tabular file: {file_path}")
        try:
            ext = os.path.splitext(file_path)[1].lower()
            # Optional user override, e.g., encoding='utf-16'
            user_enc = kwargs.get("encoding")
            if ext == ".csv":
                md = await asyncio.to_thread(_csv_to_markdown, file_path, ",", user_enc)
            elif ext == ".tsv":
                md = await asyncio.to_thread(_csv_to_markdown, file_path, "\t", user_enc)
            elif ext in [".xlsx"]:
                md = await asyncio.to_thread(_xlsx_to_markdown, file_path)
            elif ext in [".xls"]:
                import pandas as pd

                def _xls_to_md(p: str) -> str:
                    df = pd.read_excel(p)
                    return df.to_markdown(index=False)

                md = await asyncio.to_thread(_xls_to_md, file_path)
            else:
                raise RuntimeError("Unsupported tabular format")
            return md or ""
        except Exception as e:
            logger.error(f"TabularHandler: Error processing tabular file {file_path}: {e}")
            return None
