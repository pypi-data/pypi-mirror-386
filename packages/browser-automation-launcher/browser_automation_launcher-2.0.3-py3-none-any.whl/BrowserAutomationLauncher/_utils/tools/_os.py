"""
Copyright (c) 2025-now Martian Bugs All rights reserved.
Build Date: 2025-05-19
Author: Martian Bugs
Description: 系统工具类, 例如文件操作、路径操作等
"""

import csv
from os import path, remove
from socket import AF_INET, SOCK_STREAM, socket
from typing import Any
from warnings import filterwarnings

from openpyxl import load_workbook

from ._common import CommonTools

filterwarnings('ignore', category=UserWarning, module='openpyxl.styles.stylesheet')


class OsTools:
    @staticmethod
    def file_remove(file_path: str):
        """
        删除文件

        Args:
            file_path: 文件路径
        Returns:
            是否删除成功, 如果文件不存在或不是文件, 则返回False
        """

        if not path.exists(file_path) or not path.isfile(file_path):
            return False

        remove(file_path)
        return True

    @staticmethod
    def xlsx_read(file_path: str, sheet_name: str = None, titles: list[str] = None):
        """
        读取 xlsx 文件内容

        Args:
            file_path: xlsx 文件路径
            sheet_name: 工作表名称，默认激活的工作表
            titles: 标题列表，如果不指定，则读取第一行作为标题
        Returns:
            xlsx 内容列表
        """

        wb = load_workbook(file_path)
        sheet = (
            wb.active
            if not sheet_name or not isinstance(sheet_name, str)
            else wb[sheet_name]
        )

        titles_dict = {}
        title_row = next(sheet.rows)
        if not titles or not isinstance(titles, list):
            for ci, col in enumerate(title_row):
                titles_dict[col.value] = ci
        else:
            titles_dict = {k: -1 for k in titles}
            for ci, col in enumerate(title_row):
                if col.value in titles_dict:
                    titles_dict[col.value] = ci

        records: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2):
            record = {k: v for k, ci in titles_dict.items() for v in [row[ci].value]}
            records.append(record)

        return records

    @staticmethod
    def csv_read(file_path: str):
        """
        CSV 文件读取

        Args:
            file_path: 文件路径
        Returns:
            读取解析后的数据列表
        """

        if not path.exists(file_path):
            raise FileNotFoundError(f'文件不存在: {file_path}')

        records: list[dict] = []

        try:
            with open(file_path, encoding='utf-8', newline='') as f:
                header = [CommonTools.str_cleanup(h) for h in next(csv.reader(f))]

                csv_reader = csv.DictReader(f, fieldnames=header)
                for row in csv_reader:
                    cleanup_row = {}
                    for k, v in row.items():
                        cleanup_row[CommonTools.str_cleanup(k)] = (
                            CommonTools.str_cleanup(v)
                        )
                    records.append(cleanup_row)
        except Exception as err:
            raise RuntimeError(f'读取 CSV 文件出错: {file_path}', err) from err

        return records

    @staticmethod
    def port_up_check(port: int, timeout: int = 3):
        """
        检查端口号是否启用

        Args:
            port: 端口号
            timeout: 超时时间
        Returns:
            是否启用
        """

        with socket(AF_INET, SOCK_STREAM) as s:
            s.settimeout(timeout)
            return s.connect_ex(('127.0.0.1', port)) == 0
