"""
Common Database Benchmark Base Class

Provides reusable functionality for all database benchmarks:
- Command-line argument parsing
- Excel output generation with unified format
- Standard test execution flow
- Result formatting and aggregation

Company: eXonware.com
Author: Eng. Muhammad AlShehri
Email: connect@exonware.com
Version: 0.0.1
Generation Date: October 16, 2025
"""

import sys
import argparse
import random
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional
from abc import ABC, abstractmethod

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("ERROR: openpyxl is required for Excel output. Install with: pip install openpyxl")
    sys.exit(1)


class BaseBenchmarkRunner(ABC):
    """Base class for all database benchmarks"""
    
    def __init__(self, benchmark_name: str, default_test_sizes: List[int] = None, random_enabled: bool = True):
        """
        Initialize benchmark runner.
        
        Args:
            benchmark_name: Name of the benchmark (e.g., "Basic Database Benchmark (Node-Only)")
            default_test_sizes: Default test sizes if not provided via command line
            random_enabled: Whether to randomize configuration execution order (default: True)
        """
        self.benchmark_name = benchmark_name
        self.default_test_sizes = default_test_sizes or [1000, 10000, 100000]
        self.test_sizes = self.default_test_sizes.copy()
        self.random_enabled = random_enabled
        self.all_results = {}
        self.run_datetime = None
        self.timestamp_str = None
        self.random_seed = None
    
    @abstractmethod
    def run_single_benchmark(self, total_entities: int) -> Dict[str, Any]:
        """
        Run benchmark for a single test size.
        Must return a dictionary of results keyed by configuration name.
        
        Args:
            total_entities: Number of entities to test
            
        Returns:
            Dict mapping configuration name to result data
        """
        pass
    
    def parse_arguments(self):
        """Parse command-line arguments for test sizes and randomization"""
        parser = argparse.ArgumentParser(description=self.benchmark_name)
        parser.add_argument('test_sizes', type=int, nargs='*',
                          help=f'Test sizes (entity counts). Default: {self.default_test_sizes}')
        parser.add_argument('--no-random', dest='random_enabled', action='store_false',
                          help='Disable random execution order (run in defined order)')
        parser.add_argument('--seed', type=int, default=None,
                          help='Random seed for reproducible random ordering')
        parser.set_defaults(random_enabled=True)
        args = parser.parse_args()
        
        if args.test_sizes:
            # Validate minimum size
            if any(size < 1 for size in args.test_sizes):
                print(f"Error: Minimum 1 entity required")
                sys.exit(1)
            self.test_sizes = args.test_sizes
        
        # Update random settings from command line
        self.random_enabled = args.random_enabled
        self.random_seed = args.seed
        
        # Set random seed if provided
        if self.random_seed is not None:
            random.seed(self.random_seed)
            print(f"[RANDOM SEED] Using seed: {self.random_seed}")
        
        return args
    
    def shuffle_if_enabled(self, items: List[Any]) -> List[Any]:
        """
        Shuffle items if random execution is enabled.
        
        Args:
            items: List of items to potentially shuffle
            
        Returns:
            Shuffled list if random_enabled=True, original list otherwise
        """
        if self.random_enabled:
            shuffled = items.copy()
            random.shuffle(shuffled)
            return shuffled
        return items
    
    def run_all_tests(self):
        """Run all tests and generate Excel output"""
        print(f"\n{'='*80}")
        print(f"{self.benchmark_name.upper()} - AUTO RUN")
        print(f"{'='*80}")
        print(f"\nTests to run: {', '.join([f'{t:,}' for t in self.test_sizes])} entities")
        print(f"Random execution order: {'ENABLED' if self.random_enabled else 'DISABLED'}")
        if self.random_seed is not None:
            print(f"Random seed: {self.random_seed}")
        print(f"{'='*80}\n")
        
        # Store timestamp for this entire run
        self.run_datetime = datetime.now()
        self.timestamp_str = self.run_datetime.strftime("%Y-%m-%d %H:%M:%S")
        
        # Run each test
        for i, entities in enumerate(self.test_sizes, 1):
            print(f"\n{'*'*80}")
            print(f"TEST {i}/{len(self.test_sizes)}: {entities:,} entities")
            print(f"{'*'*80}")
            
            results = self.run_single_benchmark(entities)
            self.all_results[str(entities)] = results
        
        # Show overall top performers across all test sizes
        self.show_top_performers()
        
        # Generate Excel output
        self.generate_excel_output()
        
        print(f"\n{'='*80}")
        print(f"ALL BENCHMARKS COMPLETE")
        print(f"{'='*80}")
        print(f"\nTotal benchmarks completed: {sum(len(r) for r in self.all_results.values()):,}")
    
    def show_top_performers(self):
        """Show top performers across all test sizes"""
        print(f"\n{'#'*80}")
        print(f"# OVERALL TOP PERFORMERS - ALL TEST SIZES")
        print(f"{'#'*80}\n")
        
        # Aggregate results across all test sizes
        for entities_str in sorted(self.all_results.keys(), key=lambda x: int(x)):
            results = self.all_results[entities_str]
            successful_results = {k: v for k, v in results.items() if v.get('success', True)}
            
            if successful_results:
                sorted_results = sorted(successful_results.items(), 
                                       key=lambda x: x[1].get('total_time_ms', float('inf')))
                
                print(f"{'='*80}")
                print(f"TOP 10 - {entities_str} ENTITIES")
                print(f"{'='*80}")
                
                for rank, (name, data) in enumerate(sorted_results[:10], 1):
                    time_ms = data.get('total_time_ms', 0)
                    memory = data.get('peak_memory_mb', 0)
                    config_info = f"{data.get('node_mode', 'N/A')} + {data.get('edge_mode', 'N/A')}"
                    
                    # Add special indicators
                    medal = "[1]" if rank == 1 else "[2]" if rank == 2 else "[3]" if rank == 3 else "   "
                    print(f"  {medal} {rank:2}. {name[:60]:60} | {time_ms:8.2f}ms | {memory:6.1f}MB")
                
                # Show fastest performer
                winner_name, winner_data = sorted_results[0]
                print(f"\n>> FASTEST: {winner_name}")
                print(f"   Time: {winner_data.get('total_time_ms', 0):.2f}ms")
                print(f"   Memory: {winner_data.get('peak_memory_mb', 0):.1f}MB\n")
    
    def format_result_row(self, entities_str: str, name: str, data: Dict[str, Any],
                         additional_data: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        Format a result row with standard columns.
        
        Args:
            entities_str: Entity count as string
            name: Configuration name
            data: Result data from benchmark
            additional_data: Additional fields to include (e.g., group classification)
            
        Returns:
            Dictionary with standardized row format
        """
        metrics = data.get('metrics', {})
        insert_metrics = metrics.get('insert', {})
        read_metrics = metrics.get('read', {})
        update_metrics = metrics.get('update', {})
        delete_metrics = metrics.get('delete', {})
        
        node = data.get('node_mode', 'N/A')
        time_ms = data.get('total_time_ms', 0)
        memory = data.get('peak_memory_mb', 0)
        ops_sec = (50000 / (time_ms / 1000)) if time_ms > 0 else 0
        edge_mode = data.get('edge_mode', 'None')
        
        row_data = {
            'test_type': self.benchmark_name,
            'operations_size': int(entities_str),
            'datetime': self.timestamp_str,
            'name': name,
            'node': node,
            'edge_mode': edge_mode if edge_mode else 'None',
            'graph_manager': data.get('graph_manager', 'N/A'),
            'storage_format': data.get('storage_format', 'N/A'),
            'storage_smart_format': data.get('storage_smart_format', 'N/A'),
            'group': data.get('group', 'N/A'),
            'time_ms': time_ms,
            'ram_memory': memory,
            'file_size_kb': data.get('file_size_kb', 0),
            'ops_sec': ops_sec,
            'insert_time': insert_metrics.get("total_time_ms", 0),
            'insert_memory': insert_metrics.get("peak_memory_mb", 0),
            'read_time': read_metrics.get("total_time_ms", 0),
            'read_memory': read_metrics.get("peak_memory_mb", 0),
            'update_time': update_metrics.get("total_time_ms", 0),
            'update_memory': update_metrics.get("peak_memory_mb", 0),
            'delete_time': delete_metrics.get("total_time_ms", 0),
            'delete_memory': delete_metrics.get("peak_memory_mb", 0)
        }
        
        # Merge additional data if provided
        if additional_data:
            row_data.update(additional_data)
        
        return row_data
    
    def generate_excel_output(self):
        """Generate unified Excel output file"""
        print(f"\n{'='*80}")
        print(f"GENERATING EXCEL OUTPUT")
        print(f"{'='*80}")
        
        # Excel file in parent folder (unified across all benchmarks)
        output_dir = Path(__file__).parent.parent
        excel_file = output_dir / "results.xlsx"
        
        try:
            # Define standard headers
            headers = [
                'Test Type - Benchmark Name',
                'Operations Size',
                'Date & Time',
                'Rank',
                'Configuration',
                'Node Mode',
                'Edge Mode',
                'Graph Manager (ON/OFF)',
                'Storage Format',
                'Storage Smart Format (ON/OFF)',
                'Group',
                'Total Time (ms)',
                'Peak RAM Memory (MB)',
                'File Size (KB)',
                'Operations/sec',
                'Insert Time (ms)',
                'Insert Memory (MB)',
                'Read Time (ms)',
                'Read Memory (MB)',
                'Update Time (ms)',
                'Update Memory (MB)',
                'Delete Time (ms)',
                'Delete Memory (MB)'
            ]
            
            # Load existing workbook if file exists, or create new one
            existing_rows = []
            wb = None
            
            if excel_file.exists():
                try:
                    from openpyxl import load_workbook
                    print(f"  Loading existing results from {excel_file.name}...")
                    wb = load_workbook(excel_file)
                    
                    # CRITICAL FIX: Explicitly get "Benchmark Results" sheet, not wb.active
                    # Root cause: wb.active gets last active sheet (could be "Modes", "Dashboard", etc.)
                    # Solution: Explicitly select "Benchmark Results" sheet by name
                    # Priority: Usability #2 - Save data to correct sheet
                    if "Benchmark Results" in wb.sheetnames:
                        ws = wb["Benchmark Results"]
                    else:
                        # Create "Benchmark Results" sheet if it doesn't exist
                        ws = wb.create_sheet("Benchmark Results", 0)
                    
                    # Read existing data from "Benchmark Results" sheet (skip header row)
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:  # If first column has data
                            row_data = list(row)
                            # Pad with None if row is shorter than expected
                            while len(row_data) < len(headers):
                                row_data.append(None)
                            existing_rows.append(row_data[:3] + row_data[4:])  # Skip rank column
                    
                    print(f"  Found {len(existing_rows)} existing records in 'Benchmark Results' sheet")
                    print(f"  Preserving {len(wb.sheetnames)} existing sheet(s): {', '.join(wb.sheetnames)}")
                    
                    # Clear the "Benchmark Results" sheet content (we'll rewrite everything)
                    ws.delete_rows(1, ws.max_row)
                    
                except Exception as e:
                    print(f"  Warning: Could not load existing file ({e}), will create new file")
                    existing_rows = []
                    wb = None
            
            # Create new workbook if loading failed or file doesn't exist
            if wb is None:
                wb = Workbook()
                ws = wb.active
                ws.title = "Benchmark Results"
            
            # Write headers
            ws.append(headers)
            
            # Write existing rows first
            for row_data in existing_rows:
                ws.append(row_data[:3] + [''] + row_data[3:])
            
            # Collect new data rows
            all_rows = []
            for entities_str, results in self.all_results.items():
                successful_results = {k: v for k, v in results.items() if v.get('success', True)}
                
                for name, data in successful_results.items():
                    row_data = self.format_result_row(entities_str, name, data)
                    all_rows.append(row_data)
            
            # Write new data rows
            for row_data in all_rows:
                ws.append([
                    row_data['test_type'],
                    row_data['operations_size'],
                    row_data['datetime'],
                    '',  # Rank - will be filled by formula
                    row_data['name'],
                    row_data['node'],
                    row_data['edge_mode'],
                    row_data['graph_manager'],
                    row_data['storage_format'],
                    row_data['storage_smart_format'],
                    row_data['group'],
                    row_data['time_ms'],
                    row_data['ram_memory'],
                    row_data['file_size_kb'],
                    row_data['ops_sec'],
                    row_data['insert_time'],
                    row_data['insert_memory'],
                    row_data['read_time'],
                    row_data['read_memory'],
                    row_data['update_time'],
                    row_data['update_memory'],
                    row_data['delete_time'],
                    row_data['delete_memory']
                ])
            
            # Calculate total rows
            total_rows = len(existing_rows) + len(all_rows)
            
            # Add ranking formulas for ALL rows
            for row_idx in range(2, total_rows + 2):
                formula = f'=COUNTIFS($A:$A,$A{row_idx},$B:$B,$B{row_idx},$C:$C,$C{row_idx},$L:$L,"<"&$L{row_idx})+1'
                ws[f'D{row_idx}'] = formula
            
            # Create Excel Table (remove existing table if present)
            # Remove existing table if it exists
            # Remove existing tables to prevent duplicate error
            # Root cause: Table reference persists even after deletion
            # Solution: Clear tables dict completely and use unique timestamp-based name
            if ws.tables:
                # Clear all existing tables
                ws.tables.clear()
            
            # Use unique table name to prevent conflicts
            # Priority: Maintainability #3 - Robust table management
            import time
            table_name = f"BenchmarkResults_{int(time.time())}"
            table_ref = f"A1:{get_column_letter(len(headers))}{total_rows + 1}"
            
            try:
                table = Table(displayName=table_name, ref=table_ref)
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                table.tableStyleInfo = style
                ws.add_table(table)
            except ValueError as e:
                # If still fails, skip table formatting (data is already saved)
                # Priority: Usability #2 - Don't fail entire export for formatting issue
                print(f"[WARN] Could not add table formatting: {e}")
                print(f"[OK] Data saved successfully without table formatting")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Hide columns: Edge Mode (G), Graph Manager (H), Storage Format (I), Storage Smart Format (J)
            for col_letter in ['G', 'H', 'I', 'J']:
                ws.column_dimensions[col_letter].hidden = True
            
            # Save workbook
            wb.save(excel_file)
            
            # Print summary
            if len(existing_rows) > 0:
                print(f"[OK] Excel File: {excel_file.name} (UPDATED)")
                print(f"  - Existing records: {len(existing_rows)}")
                print(f"  - New records added: {len(all_rows)}")
                print(f"  - Total records: {total_rows}")
                print(f"  - Total sheets preserved: {len(wb.sheetnames)}")
            else:
                print(f"[OK] Excel File: {excel_file.name} (NEW)")
                print(f"  - New records: {len(all_rows)}")
            print(f"  - Excel Table 'BenchmarkResults' with filtering enabled")
            print(f"  - Rank calculated by formula (grouped by Test Type, Operations Size, Date & Time)")
            print(f"\n  All existing sheets have been preserved!")
        
        except PermissionError:
            print(f"[WARNING] Could not write {excel_file.name} - file is open in another program")
        except Exception as e:
            print(f"[ERROR] generating Excel file: {e}")
            import traceback
            traceback.print_exc()
    
    def main(self):
        """Main entry point for benchmark"""
        self.parse_arguments()
        self.run_all_tests()

