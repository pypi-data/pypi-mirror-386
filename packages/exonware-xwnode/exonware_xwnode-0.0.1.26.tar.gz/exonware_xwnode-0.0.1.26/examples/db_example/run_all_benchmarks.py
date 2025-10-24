#!/usr/bin/env python3
"""
Run All Benchmarks - Master Script

Executes all database benchmarks in sequence:
1. x1_basic_db - Node-only testing
2. x2_classic_db - Predefined classic configurations
3. x3_extensive_db - All NodeMode × EdgeMode combinations
4. x4_db_graph_on_off - Graph Manager ON/OFF comparison
5. x5_file_db - File serialization testing (xwsystem)
6. x6_file_advance_db - Advanced file operations (xwsystem)

Each benchmark automatically appends its results to: db_example/results.xlsx
The Excel file maintains all historical results with timestamps.

Usage:
    python run_all_benchmarks.py [test_sizes...]
    
Examples:
    python run_all_benchmarks.py              # Use defaults: 1, 10, 100
    python run_all_benchmarks.py 100 1000     # Test with 100 and 1000 entities
    python run_all_benchmarks.py 5000         # Single test with 5000 entities

Company: eXonware.com
Author: Eng. Muhammad AlShehri
Email: connect@exonware.com
Version: 0.0.1
Generation Date: October 16, 2025
"""

import sys
import subprocess
from pathlib import Path
from datetime import datetime
from typing import List

def run_benchmark(script_path: Path, name: str, test_sizes: List[int]) -> bool:
    """
    Run a single benchmark script.
    
    Args:
        script_path: Path to the benchmark script
        name: Benchmark name for display
        test_sizes: List of entity counts to test
        
    Returns:
        True if successful, False otherwise
    """
    print(f"\n{'='*80}")
    print(f"RUNNING: {name}")
    print(f"Script: {script_path}")
    print(f"Test Sizes: {', '.join([f'{s:,}' for s in test_sizes])}")
    print(f"{'='*80}\n")
    
    try:
        # Build command with test sizes as arguments
        cmd = [sys.executable, str(script_path)] + [str(size) for size in test_sizes]
        
        print(f"Command: {' '.join(cmd)}")
        print(f"Working directory: {script_path.parent}\n")
        
        result = subprocess.run(
            cmd,
            cwd=str(script_path.parent),
            capture_output=False,  # Show output in real-time
            text=True,
            check=False
        )
        
        if result.returncode == 0:
            print(f"\n{'='*80}")
            print(f"[OK] {name} completed successfully")
            print(f"{'='*80}")
            return True
        else:
            print(f"\n{'='*80}")
            print(f"[FAIL] {name} failed with return code {result.returncode}")
            print(f"{'='*80}")
            return False
            
    except Exception as e:
        print(f"\n{'='*80}")
        print(f"[ERROR] {name} failed with error: {e}")
        print(f"{'='*80}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Run all database benchmarks with specified test sizes',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python run_all_benchmarks.py              # Use defaults: 1, 10, 100
  python run_all_benchmarks.py 100 1000     # Test with 100 and 1000 entities
  python run_all_benchmarks.py 5000         # Single test with 5000 entities
        """
    )
    parser.add_argument(
        'test_sizes',
        nargs='*',
        type=int,
        help='Test sizes (entity counts). Default: [1, 10, 100]'
    )
    
    args = parser.parse_args()
    
    # Use provided test sizes or defaults
    test_sizes = args.test_sizes if args.test_sizes else [1, 10, 100]
    
    print(f"\n{'#'*80}")
    print(f"# DATABASE BENCHMARK SUITE - COMPLETE RUN")
    print(f"# Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"# Test Sizes: {', '.join([f'{s:,}' for s in test_sizes])}")
    print(f"{'#'*80}\n")
    
    base_dir = Path(__file__).parent
    results_file = base_dir / "results.xlsx"
    
    # Define all benchmarks in order
    benchmarks = [
        (base_dir / "x1_basic_db" / "benchmark.py", "x1 - Basic Database (Node-Only)"),
        (base_dir / "x2_classic_db" / "benchmark.py", "x2 - Classic Configurations"),
        (base_dir / "x3_extensive_db" / "benchmark.py", "x3 - Extensive Combinations"),
        (base_dir / "x4_db_graph_on_off" / "benchmark.py", "x4 - Graph Manager ON/OFF"),
        #(base_dir / "x5_file_db" / "benchmark.py", "x5 - File Serialization"),
        #(base_dir / "x6_file_advance_db" / "benchmark.py", "x6 - Advanced File Operations"),
    ]
    
    # Check which benchmarks exist
    available_benchmarks = []
    for script_path, name in benchmarks:
        if script_path.exists():
            available_benchmarks.append((script_path, name))
        else:
            print(f"[SKIP] {name} - script not found: {script_path}")
    
    print(f"\nFound {len(available_benchmarks)} available benchmarks")
    print(f"Results will be saved to: {results_file}\n")
    
    # Run all benchmarks
    results = {}
    for i, (script_path, name) in enumerate(available_benchmarks, 1):
        print(f"\n{'*'*80}")
        print(f"Progress: {i}/{len(available_benchmarks)}")
        print(f"{'*'*80}")
        
        success = run_benchmark(script_path, name, test_sizes)
        results[name] = success
    
    # Print summary
    print(f"\n{'#'*80}")
    print(f"# BENCHMARK SUITE COMPLETE")
    print(f"# Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'#'*80}\n")
    
    print("Summary:")
    successful = sum(1 for v in results.values() if v)
    failed = len(results) - successful
    
    for name, success in results.items():
        status = "[PASS]" if success else "[FAIL]"
        print(f"  {status} - {name}")
    
    print(f"\nTotal: {successful} passed, {failed} failed")
    
    # Check if results file exists
    if results_file.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(results_file)
            ws = wb.active
            row_count = ws.max_row - 1  # Subtract header row
            wb.close()
            print(f"\n[OK] Results file exists: {results_file}")
            print(f"  Total records in Excel: {row_count:,}")
        except Exception as e:
            print(f"\n[WARNING] Results file exists but couldn't read: {e}")
    else:
        print(f"\n[WARNING] Results file not found: {results_file}")
        print("   Individual benchmarks should have created this file.")
        print("   Check for errors in the benchmark output above.")
    
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())

