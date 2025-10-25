from abc import ABC, abstractmethod
from pathlib import Path
import pandas as pd
import numpy as np
from typing import Optional, Dict, Any, List
from dataclasses import dataclass, field
import csv
import re

@dataclass
class ParseOptions:
    encoding: str = 'utf-8'
    skip_rows: int = 0
    header_rows: int = 1
    skip_footer: int = 0
    na_values: List[str] = field(default_factory=list)
    parse_dates: bool = True
    chunk_size: Optional[int] = None
    data_only: bool = True
    preserve_formatting: bool = False
    handle_merged_cells: str = 'unmerge'
    ignore_hidden: bool = True
    max_rows: Optional[int] = None

class FormatHandler(ABC):
    @abstractmethod
    def can_handle(self, format_type: str) -> bool:
        pass

    @abstractmethod
    def parse(self, file_path: Path, sheet: Optional[str], options: ParseOptions) -> pd.DataFrame:
        pass

    @abstractmethod
    def get_sheets(self, file_path: Path) -> List[str]:
        pass

    @abstractmethod
    def validate(self, file_path: Path) -> tuple[bool, Optional[str]]:
        pass

class XLSXHandler(FormatHandler):
    def can_handle(self, format_type: str) -> bool:
        return format_type in ['xlsx', 'xlsm']

    def parse(self, file_path: Path, sheet: Optional[str], options: ParseOptions) -> pd.DataFrame:
        try:
            import openpyxl
        except ImportError:
            # Fall back to pandas
            return pd.read_excel(
                file_path,
                sheet_name=sheet,
                skiprows=options.skip_rows,
                skipfooter=options.skip_footer,
                na_values=options.na_values,
                engine='openpyxl'
            )

        wb = openpyxl.load_workbook(
            file_path,
            read_only=True,
            data_only=options.data_only,
            keep_links=False
        )

        ws = wb[sheet] if sheet else wb.active

        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))

        wb.close()

        if not data:
            return pd.DataFrame()

        # Handle skip_rows
        if options.skip_rows > 0:
            data = data[options.skip_rows:]

        # Handle skip_footer
        if options.skip_footer > 0:
            data = data[:-options.skip_footer]

        # Handle headers
        if options.header_rows > 0 and len(data) > options.header_rows:
            headers = data[:options.header_rows]
            data = data[options.header_rows:]

            if options.header_rows == 1:
                columns = [str(h) if h else f'col_{i}' for i, h in enumerate(headers[0])]
            else:
                # Multi-row headers - combine them
                columns = []
                for col_idx in range(len(headers[0])):
                    col_parts = []
                    for row in headers:
                        if col_idx < len(row) and row[col_idx]:
                            col_parts.append(str(row[col_idx]))
                    columns.append('_'.join(col_parts) if col_parts else f'col_{col_idx}')
        else:
            columns = [f'col_{i}' for i in range(len(data[0]) if data else 0)]

        df = pd.DataFrame(data, columns=columns)

        # Clean data
        df = self._clean_excel_data(df, options)

        return df

    def _clean_excel_data(self, df: pd.DataFrame, options: ParseOptions) -> pd.DataFrame:
        excel_errors = ['#DIV/0!', '#N/A', '#NAME?', '#NULL!', '#NUM!', '#REF!', '#VALUE!']
        df = df.replace(excel_errors, np.nan)

        # Handle NA values
        if options.na_values:
            df = df.replace(options.na_values, np.nan)

        return df

    def get_sheets(self, file_path: Path) -> List[str]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except:
            # Fallback
            xl_file = pd.ExcelFile(file_path, engine='openpyxl')
            sheets = xl_file.sheet_names
            xl_file.close()
            return sheets

    def validate(self, file_path: Path) -> tuple[bool, Optional[str]]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            wb.close()
            return True, None
        except Exception as e:
            return False, str(e)

class XLSHandler(FormatHandler):
    def can_handle(self, format_type: str) -> bool:
        return format_type == 'xls'

    def parse(self, file_path: Path, sheet: Optional[str], options: ParseOptions) -> pd.DataFrame:
        try:
            # Try using pandas with xlrd
            df = pd.read_excel(
                file_path,
                sheet_name=sheet if sheet else 0,
                skiprows=options.skip_rows,
                skipfooter=options.skip_footer,
                na_values=options.na_values,
                engine='xlrd'
            )
            return df
        except ImportError:
            # If xlrd not available, try openpyxl
            return pd.read_excel(
                file_path,
                sheet_name=sheet if sheet else 0,
                skiprows=options.skip_rows,
                skipfooter=options.skip_footer,
                na_values=options.na_values
            )

    def get_sheets(self, file_path: Path) -> List[str]:
        try:
            xl_file = pd.ExcelFile(file_path)
            sheets = xl_file.sheet_names
            xl_file.close()
            return sheets
        except:
            return ['Sheet1']

    def validate(self, file_path: Path) -> tuple[bool, Optional[str]]:
        try:
            xl_file = pd.ExcelFile(file_path)
            xl_file.close()
            return True, None
        except Exception as e:
            return False, str(e)

class CSVHandler(FormatHandler):
    def can_handle(self, format_type: str) -> bool:
        return format_type in ['csv', 'tsv']

    def parse(self, file_path: Path, sheet: Optional[str], options: ParseOptions) -> pd.DataFrame:
        # Detect delimiter
        delimiter = self._detect_delimiter(file_path, options.encoding)

        # Parse with pandas
        try:
            df = pd.read_csv(
                file_path,
                encoding=options.encoding,
                delimiter=delimiter,
                skiprows=options.skip_rows,
                skipfooter=options.skip_footer,
                nrows=options.max_rows,
                na_values=options.na_values or ['', 'NA', 'N/A', 'null', 'NULL', '#N/A'],
                parse_dates=options.parse_dates,
                chunksize=options.chunk_size,
                engine='python' if options.skip_footer > 0 else 'c'
            )
            return df
        except Exception as e:
            # Try with different encoding
            try:
                df = pd.read_csv(
                    file_path,
                    encoding='latin-1',
                    delimiter=delimiter,
                    skiprows=options.skip_rows,
                    na_values=options.na_values,
                    error_bad_lines=False
                )
                return df
            except:
                raise e

    def _detect_delimiter(self, file_path: Path, encoding: str) -> str:
        try:
            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                sample = f.read(8192)

            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample)
            return dialect.delimiter
        except:
            # Fallback to common delimiters
            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                lines = f.readlines()[:10]

            delimiters = [',', '\t', ';', '|']
            delimiter_scores = {}

            for delim in delimiters:
                counts = [line.count(delim) for line in lines if line.strip()]
                if counts and counts[0] > 0:
                    avg_count = sum(counts) / len(counts)
                    variance = sum((c - avg_count) ** 2 for c in counts) / len(counts) if len(counts) > 1 else 0
                    delimiter_scores[delim] = avg_count / (variance + 1)

            if delimiter_scores:
                return max(delimiter_scores, key=delimiter_scores.get)

            return ','

    def get_sheets(self, file_path: Path) -> List[str]:
        return ['Sheet1']

    def validate(self, file_path: Path) -> tuple[bool, Optional[str]]:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                f.read(1024)
            return True, None
        except Exception as e:
            return False, str(e)