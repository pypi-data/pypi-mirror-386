import os
import csv
import pytest
from pathlib import Path
import pandas as pd
import openpyxl

import mcp_excel.server as server
from mcp_excel.formats.detector import FormatDetector, FormatInfo
from mcp_excel.formats.handlers import XLSXHandler, CSVHandler, ParseOptions
from mcp_excel.formats.normalizer import DataNormalizer
from mcp_excel.formats.manager import FormatManager

pytestmark = pytest.mark.unit


@pytest.fixture(autouse=True)
def reset_server_state():
    server.catalog.clear()
    server.load_configs.clear()
    server.init_server()
    yield
    server.catalog.clear()
    server.load_configs.clear()


class TestFormatDetector:
    def test_detect_xlsx(self, temp_dir):
        # Create a simple XLSX file
        xlsx_file = temp_dir / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B", "C"])
        ws.append([1, 2, 3])
        wb.save(xlsx_file)

        detector = FormatDetector()
        info = detector.detect(xlsx_file)

        assert info.format_type == 'xlsx'
        assert info.confidence >= 0.9
        assert not info.has_macros
        assert not info.is_encrypted

    def test_detect_csv(self, temp_dir):
        # Create a CSV file
        csv_file = temp_dir / "test.csv"
        with open(csv_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Name', 'Age', 'City'])
            writer.writerow(['Alice', 30, 'NYC'])
            writer.writerow(['Bob', 25, 'LA'])

        detector = FormatDetector()
        info = detector.detect(csv_file)

        assert info.format_type == 'csv'
        assert info.confidence >= 0.5

    def test_detect_tsv(self, temp_dir):
        # Create a TSV file
        tsv_file = temp_dir / "test.tsv"
        with open(tsv_file, 'w', newline='') as f:
            writer = csv.writer(f, delimiter='\t')
            writer.writerow(['Name', 'Age', 'City'])
            writer.writerow(['Alice', 30, 'NYC'])
            writer.writerow(['Bob', 25, 'LA'])

        detector = FormatDetector()
        info = detector.detect(tsv_file)

        assert info.format_type == 'tsv'
        assert info.confidence >= 0.5

    def test_detect_unknown(self, temp_dir):
        # Create a random binary file
        bin_file = temp_dir / "test.bin"
        with open(bin_file, 'wb') as f:
            f.write(b'\x00\x01\x02\x03\x04\x05')

        detector = FormatDetector()
        info = detector.detect(bin_file)

        assert info.format_type == 'unknown'
        assert info.confidence == 0.0


class TestFormatHandlers:
    def test_xlsx_handler(self, temp_dir):
        # Create XLSX file with multiple sheets
        xlsx_file = temp_dir / "test.xlsx"
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["Name", "Value"])
        ws1.append(["A", 100])
        ws1.append(["B", 200])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["ID", "Amount"])
        ws2.append([1, 50.5])
        ws2.append([2, 75.25])

        wb.save(xlsx_file)

        handler = XLSXHandler()
        assert handler.can_handle('xlsx')

        # Get sheets
        sheets = handler.get_sheets(xlsx_file)
        assert len(sheets) == 2
        assert "Sheet1" in sheets
        assert "Sheet2" in sheets

        # Parse Sheet1
        options = ParseOptions(header_rows=1)
        df = handler.parse(xlsx_file, "Sheet1", options)
        assert len(df) == 2
        assert list(df.columns) == ["Name", "Value"]

        # Validate
        is_valid, error = handler.validate(xlsx_file)
        assert is_valid
        assert error is None

    def test_csv_handler(self, temp_dir):
        # Create CSV file
        csv_file = temp_dir / "test.csv"
        with open(csv_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Product', 'Price', 'Quantity'])
            writer.writerow(['Apple', '1.50', '10'])
            writer.writerow(['Banana', '0.75', '20'])
            writer.writerow(['Orange', '2.00', '15'])

        handler = CSVHandler()
        assert handler.can_handle('csv')

        # Parse CSV
        options = ParseOptions()
        df = handler.parse(csv_file, None, options)
        assert len(df) == 3
        assert list(df.columns) == ['Product', 'Price', 'Quantity']

        # Validate
        is_valid, error = handler.validate(csv_file)
        assert is_valid


class TestDataNormalizer:
    def test_normalize_whitespace(self):
        df = pd.DataFrame({
            'A': ['  hello  ', 'world\n\n', '\xa0test'],
            'B': [1, 2, 3]
        })

        normalizer = DataNormalizer()
        df_clean = normalizer.clean_whitespace(df, {})

        assert df_clean['A'].iloc[0] == 'hello'
        assert df_clean['A'].iloc[1] == 'world'
        assert df_clean['A'].iloc[2] == 'test'

    def test_normalize_numbers(self):
        df = pd.DataFrame({
            'Amount': ['$1,234.56', '(500.00)', '€2.500,50', '  100  '],
            'Text': ['abc', 'def', 'ghi', 'jkl']
        })

        normalizer = DataNormalizer()
        df_clean = normalizer.normalize_numbers(df, {})

        assert pd.api.types.is_numeric_dtype(df_clean['Amount'])
        assert df_clean['Amount'].iloc[0] == 1234.56
        assert df_clean['Amount'].iloc[1] == -500.00
        assert df_clean['Text'].dtype == object

    def test_normalize_dates(self):
        df = pd.DataFrame({
            'Date1': [44927, 44928, 44929],  # Excel serial dates
            'Date2': ['2023-01-15', '2023-02-20', '2023-03-25'],
            'Text': ['abc', 'def', 'ghi']
        })

        normalizer = DataNormalizer()
        df_clean = normalizer.normalize_dates(df, {})

        # Check if Date1 was converted from Excel serial
        assert pd.api.types.is_datetime64_any_dtype(df_clean['Date1'])

        # Check if Date2 was parsed
        assert pd.api.types.is_datetime64_any_dtype(df_clean['Date2'])

        # Text should remain as object
        assert df_clean['Text'].dtype == object

    def test_handle_missing_values(self):
        df = pd.DataFrame({
            'A': ['value', 'NA', 'N/A', 'null', '-', ''],
            'B': [1, 2, 3, 4, 5, 6]
        })

        normalizer = DataNormalizer()
        df_clean = normalizer.handle_missing_values(df, {'empty_string_as_na': True})

        # Check that missing value representations became NaN
        assert pd.isna(df_clean['A'].iloc[1])
        assert pd.isna(df_clean['A'].iloc[2])
        assert pd.isna(df_clean['A'].iloc[3])
        assert pd.isna(df_clean['A'].iloc[4])
        assert pd.isna(df_clean['A'].iloc[5])
        assert df_clean['A'].iloc[0] == 'value'


class TestFormatManager:
    def test_load_xlsx_file(self, temp_dir):
        # Create XLSX file
        xlsx_file = temp_dir / "data.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Amount", "Date"])
        ws.append(["Alice", "$1,234.56", "2023-01-15"])
        ws.append(["Bob", "$2,345.67", "2023-02-20"])
        wb.save(xlsx_file)

        manager = FormatManager()
        df = manager.load_file(xlsx_file, options={'normalize': True})

        assert len(df) == 2
        assert list(df.columns) == ["Name", "Amount", "Date"]

        # Check normalization worked
        assert pd.api.types.is_numeric_dtype(df['Amount'])
        assert df['Amount'].iloc[0] == 1234.56

    def test_load_csv_file(self, temp_dir):
        # Create CSV file with messy data
        csv_file = temp_dir / "data.csv"
        with open(csv_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Name', 'Value', 'Status'])
            writer.writerow(['  Item1  ', '100', 'true'])
            writer.writerow(['Item2', 'N/A', 'false'])
            writer.writerow(['Item3', '200', 'yes'])

        manager = FormatManager()
        df = manager.load_file(csv_file, options={'normalize': True})

        assert len(df) == 3
        assert df['Name'].iloc[0] == 'Item1'  # Whitespace cleaned
        assert pd.isna(df['Value'].iloc[1])  # N/A converted to NaN

    def test_get_sheets_various_formats(self, temp_dir):
        # Create XLSX file
        xlsx_file = temp_dir / "data.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Data"
        ws1.append(["A", "B"])
        ws2 = wb.create_sheet("Summary")
        ws2.append(["C", "D"])
        wb.save(xlsx_file)

        # Create CSV file
        csv_file = temp_dir / "data.csv"
        with open(csv_file, 'w') as f:
            f.write("A,B\n1,2\n")

        manager = FormatManager()

        # Test XLSX sheets
        xlsx_sheets = manager.get_sheets(xlsx_file)
        assert len(xlsx_sheets) == 2
        assert "Data" in xlsx_sheets
        assert "Summary" in xlsx_sheets

        # Test CSV sheets (always returns Sheet1)
        csv_sheets = manager.get_sheets(csv_file)
        assert csv_sheets == ['Sheet1']


class TestIntegrationWithServer:
    def test_load_csv_through_server(self, temp_dir):
        # Create CSV file
        csv_file = temp_dir / "sales.csv"
        with open(csv_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Product', 'Revenue', 'Quarter'])
            writer.writerow(['Widget', '1000', 'Q1'])
            writer.writerow(['Gadget', '2000', 'Q2'])
            writer.writerow(['Gizmo', '1500', 'Q1'])

        # Load through server
        result = server.load_dir(str(temp_dir), alias="test")

        assert result['files_count'] >= 1
        assert result['sheets_count'] >= 1
        assert result['tables_count'] >= 1

        # Get the actual table name from catalog
        table_names = list(server.catalog.keys())
        sales_table = [t for t in table_names if 'sales' in t.lower() and 'sheet1' in t.lower()][0]

        # Query the data
        query_result = server.query(f'SELECT * FROM "{sales_table}"')
        assert query_result['row_count'] == 3

    def test_load_mixed_formats(self, temp_dir):
        # Create XLSX file
        xlsx_file = temp_dir / "data.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name"])
        ws.append([1, "Alice"])
        ws.append([2, "Bob"])
        wb.save(xlsx_file)

        # Create CSV file
        csv_file = temp_dir / "extra.csv"
        with open(csv_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['ID', 'Score'])
            writer.writerow([1, 85])
            writer.writerow([2, 92])

        # Load directory with mixed formats
        result = server.load_dir(str(temp_dir), alias="test")

        assert result['files_count'] == 2
        assert result['sheets_count'] == 2
        assert result['tables_count'] == 2

        # Get actual table names from catalog
        table_names = list(server.catalog.keys())
        data_table = [t for t in table_names if 'data' in t.lower()][0]
        extra_table = [t for t in table_names if 'extra' in t.lower()][0]

        # In RAW mode, Excel files don't have headers, so columns are col_0, col_1, etc.
        # CSV files in RAW mode also don't parse headers
        # Let's just verify the data is loaded
        query_result = server.query(f'SELECT COUNT(*) as cnt FROM "{data_table}"')
        assert query_result['rows'][0][0] >= 2

        query_result = server.query(f'SELECT COUNT(*) as cnt FROM "{extra_table}"')
        assert query_result['rows'][0][0] >= 2

        # Verify we can select from both tables
        query_result = server.query(f'SELECT * FROM "{data_table}"')
        assert query_result['row_count'] >= 2

        query_result = server.query(f'SELECT * FROM "{extra_table}"')
        assert query_result['row_count'] >= 2