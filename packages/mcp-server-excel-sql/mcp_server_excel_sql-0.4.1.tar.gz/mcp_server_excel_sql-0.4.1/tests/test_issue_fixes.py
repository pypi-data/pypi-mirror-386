import os
import time
import pytest
from pathlib import Path
import openpyxl

import mcp_excel.server as server
from mcp_excel.server import (
    load_dir,
    refresh,
    catalog,
    load_configs,
)

pytestmark = pytest.mark.regression


@pytest.fixture(autouse=True)
def reset_server_state():
    catalog.clear()
    server.load_configs.clear()
    if server.conn:
        try:
            for table_name in list(catalog.keys()):
                try:
                    server.conn.execute(f'DROP VIEW IF EXISTS "{table_name}"')
                except:
                    pass
        except:
            pass
    server.init_server()
    yield
    catalog.clear()
    server.load_configs.clear()


class TestIssue3RefreshAliasExtraction:
    """
    Issue: refresh() uses meta.table_name.split("__")[0] to get alias,
    but table names use dots (.) not double underscores (__).

    This causes config lookup to fail, preventing refresh from working.
    """

    def test_refresh_finds_config_with_regular_table_name(self, temp_excel_dir):
        """
        Test that refresh can find the config for tables with dots in name.

        Before fix: meta.table_name.split("__")[0] returns full table name
        After fix: meta.alias is used directly
        """
        excel_file = temp_excel_dir / "sales_data.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Q1"
        ws.append(["Product", "Revenue"])
        ws.append(["Widget", "1000"])
        wb.save(excel_file)

        result = load_dir(str(temp_excel_dir), alias="sales")
        assert result["tables_count"] == 1

        table_name = list(catalog.keys())[0]
        assert "sales" in table_name
        assert "sales_data" in table_name
        assert "q1" in table_name
        assert "__" not in table_name

        original_meta = catalog[table_name]
        assert original_meta.alias == "sales"

        time.sleep(0.01)

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append(["Gadget", "2000"])
        wb.save(excel_file)

        refresh_result = refresh(alias="sales", full=False)

        assert refresh_result["changed"] == 1, \
            "Refresh should detect and update the modified file"

        updated_meta = catalog[table_name]
        assert updated_meta.mtime > original_meta.mtime, \
            "Modified time should be updated after refresh"

    def test_refresh_handles_underscores_in_filename(self, temp_excel_dir):
        """
        Test that files with underscores don't confuse the alias extraction.

        Filename: Q1_2024_final.xlsx
        Table name: sales.q1_2024_final.sheet1
        Should extract alias: sales (not sales.q1_2024_final.sheet1)
        """
        excel_file = temp_excel_dir / "Q1_2024_final.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append([1, 2])
        wb.save(excel_file)

        load_dir(str(temp_excel_dir), alias="sales")

        table_name = list(catalog.keys())[0]
        meta = catalog[table_name]

        assert meta.alias == "sales"
        assert "q1_2024_final" in table_name or "q12024final" in table_name
        assert "_" in table_name or "q1" in table_name

        time.sleep(0.01)
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append([3, 4])
        wb.save(excel_file)

        refresh_result = refresh(alias="sales", full=False)
        assert refresh_result["changed"] == 1

    def test_system_views_still_work_with_double_underscores(self, temp_excel_dir):
        """
        System views use __ (e.g., sales.__files), ensure they still work.
        """
        excel_file = temp_excel_dir / "data.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Col1"])
        ws.append(["Val1"])
        wb.save(excel_file)

        load_dir(str(temp_excel_dir), alias="myalias")

        files_view = f"myalias.__files"
        tables_view = f"myalias.__tables"

        result = server.conn.execute(f'SELECT * FROM "{files_view}"').fetchall()
        assert len(result) >= 1, "System view __files should be queryable"

        result = server.conn.execute(f'SELECT * FROM "{tables_view}"').fetchall()
        assert len(result) >= 1, "System view __tables should be queryable"

    def test_meta_alias_matches_load_config_key(self, temp_excel_dir):
        """
        Verify that TableMeta.alias matches the key in load_configs.
        """
        excel_file = temp_excel_dir / "report.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Data"])
        wb.save(excel_file)

        load_dir(str(temp_excel_dir), alias="reports")

        assert "reports" in load_configs

        for table_name, meta in catalog.items():
            if not table_name.startswith("reports."):
                continue

            assert meta.alias == "reports"
            assert meta.alias in load_configs

            config = load_configs[meta.alias]
            assert config.alias == meta.alias


class TestIssue4SystemViewMemoryLeak:
    """
    Issue: _create_system_views() registers temporary DataFrames but doesn't
    unregister them, potentially causing memory accumulation.
    """

    def test_repeated_system_view_creation_does_not_leak(self, temp_excel_dir):
        """
        Test that repeatedly calling _create_system_views doesn't accumulate temp tables.
        """
        excel_file = temp_excel_dir / "data.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A"])
        ws.append([1])
        wb.save(excel_file)

        load_dir(str(temp_excel_dir), alias="test")

        initial_tables = server.conn.execute("SHOW TABLES").fetchall()
        initial_count = len(initial_tables)

        for i in range(5):
            time.sleep(0.01)
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            ws.append([i + 2])
            wb.save(excel_file)

            refresh(alias="test", full=True)

        final_tables = server.conn.execute("SHOW TABLES").fetchall()
        final_count = len(final_tables)

        assert final_count <= initial_count + 2, \
            f"Expected at most {initial_count + 2} tables, got {final_count}. " \
            f"Temporary tables may be accumulating."

    def test_system_views_are_queryable_after_refresh(self, temp_excel_dir):
        """
        Test that system views remain queryable after multiple refreshes.
        """
        excel_file = temp_excel_dir / "data.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Col"])
        ws.append(["Val"])
        wb.save(excel_file)

        load_dir(str(temp_excel_dir), alias="test")

        for i in range(3):
            result = server.conn.execute('SELECT * FROM "test.__files"').fetchall()
            assert len(result) >= 1

            result = server.conn.execute('SELECT * FROM "test.__tables"').fetchall()
            assert len(result) >= 1

            if i < 2:
                refresh(alias="test", full=True)


class TestIssue5TransactionCleanup:
    """
    Issue: Query timeout via conn.interrupt() may leave transactions open
    or connection in inconsistent state.
    """

    def test_connection_state_after_successful_query(self, temp_excel_dir):
        """
        Verify connection is in clean state after successful query.
        """
        from mcp_excel.server import query

        excel_file = temp_excel_dir / "data.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append([1, 2])
        ws.append([3, 4])
        wb.save(excel_file)

        load_dir(str(temp_excel_dir), alias="test")

        table_name = [t for t in catalog.keys() if not "__" in t][0]

        result = query(f'SELECT * FROM "{table_name}"')
        assert result["row_count"] >= 2

        result2 = query(f'SELECT COUNT(*) as cnt FROM "{table_name}"')
        assert result2["row_count"] == 1

    def test_connection_usable_after_query_error(self, temp_excel_dir):
        """
        Verify connection is still usable after a failed query.
        """
        from mcp_excel.server import query

        excel_file = temp_excel_dir / "data.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A"])
        ws.append([1])
        wb.save(excel_file)

        load_dir(str(temp_excel_dir), alias="test")

        with pytest.raises(RuntimeError):
            query("SELECT * FROM nonexistent_table")

        table_name = [t for t in catalog.keys() if not "__" in t][0]
        result = query(f'SELECT * FROM "{table_name}"')
        assert result["row_count"] >= 1

    def test_query_timeout_doesnt_break_connection(self, temp_excel_dir):
        """
        Test that after a timeout, the connection is still usable.

        Note: This is a simplified test. A real timeout with conn.interrupt()
        is hard to test without a truly long-running query.
        """
        from mcp_excel.server import query

        excel_file = temp_excel_dir / "data.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A"])
        for i in range(10):
            ws.append([i])
        wb.save(excel_file)

        load_dir(str(temp_excel_dir), alias="test")

        table_name = [t for t in catalog.keys() if not "__" in t][0]

        result = query(f'SELECT * FROM "{table_name}"', timeout_ms=5000)
        assert result["row_count"] >= 1

        result2 = query(f'SELECT COUNT(*) FROM "{table_name}"')
        assert result2["row_count"] == 1

    def test_ddl_operations_in_refresh_are_safe(self, temp_excel_dir):
        """
        Test that DDL operations (DROP VIEW, CREATE VIEW) in refresh don't
        interfere with subsequent operations.
        """
        excel_file = temp_excel_dir / "data.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A"])
        ws.append([1])
        wb.save(excel_file)

        load_dir(str(temp_excel_dir), alias="test")

        table_name = [t for t in catalog.keys() if not "__" in t][0]

        result_before = server.conn.execute(f'SELECT * FROM "{table_name}"').fetchall()
        assert len(result_before) >= 1

        refresh(alias="test", full=True)

        table_name_after = [t for t in catalog.keys() if not "__" in t][0]

        result_after = server.conn.execute(f'SELECT * FROM "{table_name_after}"').fetchall()
        assert len(result_after) >= 1
        assert len(result_after) == len(result_before)

        from mcp_excel.server import query
        query_result = query(f'SELECT * FROM "{table_name_after}"')
        assert query_result["row_count"] >= 1
