import json
import logging
import os
from pathlib import Path
import re
import openpyxl
import pandas as pd

from nemo_library.features.nemo_persistence_api import getColumns
from nemo_library.features.nemo_persistence_api import getProjects
from nemo_library.model.migman import MigMan
from nemo_library.utils.config import Config
from nemo_library.utils.utils import get_internal_name

SYNONYM_FIELDS = {
    "S_Mengeneinheit.Mengeneinheit": [
        "M_LTPos.MengenEinheit",
        "SBM_PartCustQuant.Mengeneinheit",
        "S_ArtME.MengenEinheit",
        "S_EAN.MengenEinheit",
        "VS_AuftragPos.MengenEinheit",
        "V_BELEGPOS.mengeneinheit",
        "S_Artikel.LagerME",
        "S_Artikel.GewichtME",
        "S_Artikel.StkME",
    ],
    "S_Artikel.Artikel": [
        "E_ArtLief.Artikel",
        "E_ArtLiefBei.Artikel",
        "E_ArtLiefDatum.Artikel",
        "E_ArtLiefRabStaffel.Artikel",
        "E_ArtLiefZeich.Artikel",
        "E_BelegPos.Artikel",
        "E_RA_Pos.Artikel",
        "MC_Charge.Artikel",
        "MD_Artikel.Artikel",
        "MS_ProdAktePos.Artikel",
        "MS_SNR.ArtikelAktuell",
        "M_Aktivitaet.Artikel",
        "M_LTPos.Artikel",
        "M_Ressource.Artikel",
        "P_ZeichnungArtikel.Artikel",
        "SBM_PartCustQuant.Artikel",
        "SBM_PartCustQuant.Artikel",
        "S_Anlage.Artikel",
        "S_ArtAlternativ.Artikel",
        "S_ArtGruppe.ArtikelGruppe",
        "S_ArtHerst.Artikel",
        "S_ArtKunde.Artikel",
        "S_ArtKunde.ArtikelNr",
        "S_ArtKundeDatum.Artikel",
        "S_ArtKundeRabDatum.Artikel",
        "S_ArtME.Artikel",
        "S_ArtPreiseStaffel.Artikel",
        "S_Artikel.ArtikelArt",
        "S_Artikel.ArtikelGruppe",
        "S_ArtikelSpr.Artikel",
        "S_EAN.Artikel",
        "S_Kunde.Artikelstatistik",
        "VS_AuftragMKT.Artikel",
        "VS_AuftragPos.Artikel",
        "VS_AuftragPos.ArtikelWartung",
        "V_BELEGPOS.artikel",
        "V_BELEGPOS.artikel",
        "V_BELEGPOS.artikel",
        "V_BELEGPOS.artikel",
        "V_BelegPos.Artikel",
        "V_BelegPos.Artikel",
    ],
    "S_Kunde.Kunde": [
        "JBT_Project.Kunde",
        "MMM_Transport.Kunde",
        "MS_SNR.Kunde",
        "MS_ServiceObjekt.Kunde",
        "SBM_EDL.Kunde",
        "SBM_PartCustQuant.Kunde",
        "S_Abladestelle.Kunde",
        "S_ArtKunde.Kunde",
        "S_ArtKundeDatum.Kunde",
        "S_ArtKundeRabDatum.Kunde",
        "S_KunARabDatum.Kunde",
        "S_KundeVertreter.Kunde",
        "S_LieferAdresse.Kunde",
        "S_RechnungsAdresse.Kunde",
        "VC_Interessent.Kunde",
        "VS_Auftrag.Kunde",
        "V_BELEGKOPF.kunde",
        "V_BelegKopf.Kunde",
        "S_Ansprech.Konto",
        "S_UStID.Konto",
    ],
    "S_Lieferant.Lieferant": [
        "E_ArtLief.Lieferant",
        "E_ArtLiefBei.Lieferant",
        "E_ArtLiefDatum.Lieferant",
        "E_ArtLiefRabStaffel.Lieferant",
        "E_ArtLiefZeich.Lieferant",
        "E_BelegKopf.Lieferant",
        "E_RA_Kopf.Lieferant",
        "MMM_Transport.Lieferant",
        "M_Aktivitaet.Lieferant",
        "M_Ressource.Lieferant",
        "SBM_EDL.Lieferant",
        "S_Anlage.Lieferant",
        "S_BestellAdresse.Lieferant",
        "S_EAN.Lieferant",
        "S_Vertreter.Lieferant",
        "V_BelegKopf.Lieferant",
        "S_Ansprech.Konto",
        "S_UStID.Konto",
    ],
}

DUPLICATE_CHECK_FIELDS = {
    "Parts": [
        "S_ArtikelSpr.Bezeichnung1",
        "S_ArtikelSpr.Bezeichnung2",
        "S_ArtikelSpr.Bezeichnung3",
        "S_ArtikelSpr.Bezeichnung4",
    ],
    "Suppliers": [
        "S_Adresse.Name1",
        "S_Adresse.Staat",
        "S_Adresse.Ort",
        "S_Adresse.Name2",
        "S_Adresse.Name3",
        "S_Adresse.PLZ",
        "S_Adresse.Strasse",
        "S_Adresse.Hausnummer",
    ],
    "Customers": [
        "S_Adresse.Name1",
        "S_Adresse.Staat",
        "S_Adresse.Ort",
        "S_Adresse.Name2",
        "S_Adresse.Name3",
        "S_Adresse.PLZ",
        "S_Adresse.Strasse",
        "S_Adresse.Hausnummer",
    ],
    "Prospects": [
        "S_Adresse.Name1",
        "S_Adresse.Staat",
        "S_Adresse.Ort",
        "S_Adresse.Name2",
        "S_Adresse.Name3",
        "S_Adresse.PLZ",
        "S_Adresse.Strasse",
        "S_Adresse.Hausnummer",
    ],
}

DATA_CLASSIFICATIONS = {
    "S_Adresse.HomePage": "url",
    "S_Adresse.EMail": "email",
    "S_Adresse.Handy": "phone_number",
    "S_Adresse.Telefax": "phone_number",
    "S_Adresse.Telefon": "phone_number",
    "S_Adresse.Telefon2": "phone_number",
    "S_Adresse.Name1": "address",
    "S_Adresse.Name2": "address",
    "S_Adresse.Name3": "address",
    "S_Adresse.Staat": "address",
    "S_Adresse.Ort": "address",
    "S_Adresse.PLZ": "address",
    "S_Adresse.Strasse": "address",
    "S_Adresse.Hausnummer": "address",
}


def get_migman_project_list(config: Config) -> list[str] | None:
    proALPHA_project_status_file = config.get_migman_proALPHA_project_status_file()
    if proALPHA_project_status_file:
        return getNEMOStepsFrompAMigrationStatusFile(proALPHA_project_status_file)
    else:
        return config.get_migman_projects()


def initializeFolderStructure(
    project_path: str,
) -> None:
    """
    Initialize the folder structure for a given project path.

    Args:
        project_path (str): The path to the project directory.
    """
    folders = [
        "templates",
        "mappings",
        "srcdata",
        "other",
        "to_proalpha",
        "to_customer",
    ]
    for folder in folders:
        os.makedirs(os.path.join(project_path, folder), exist_ok=True)


def getMappingFilePath(projectname: str, local_project_path: str) -> str:
    """
    Get the file path for the mapping file of a given project.

    Args:
        projectname (str): The name of the project.
        local_project_path (str): The local path to the project directory.

    Returns:
        str: The file path to the mapping file.
    """
    return os.path.join(local_project_path, "mappings", f"{projectname}.csv")


def getProjectName(project: str, addon: str, postfix: str) -> str:
    """
    Generate a project name based on the given parameters.

    Args:
        project (str): The base project name.
        addon (str): An optional addon to the project name.
        postfix (str): An optional postfix to the project name.

    Returns:
        str: The generated project name.
    """
    return f"{project}{" " + addon if addon else ""}{(" (" + postfix + ")") if postfix else ""}"


def getNEMOStepsFrompAMigrationStatusFile(file: str) -> list[str]:
    """
    Extract NEMO steps from a pA Migration Status file.

    Args:
        file (str): The path to the pA Migration Status file.

    Returns:
        list[str]: A list of NEMO steps.
    """
    path = Path(file)
    if not path.exists():
        raise FileNotFoundError(f"The file {file} does not exist.")
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook["Status Datenübernahme"]

    data = []
    for row in worksheet.iter_rows(
        min_row=10, max_row=300, min_col=1, max_col=10, values_only=True
    ):
        data.append(row)

    # Create a DataFrame from the extracted data
    columns = [
        worksheet.cell(row=9, column=i).value for i in range(1, 11)
    ]  # Headers in row 9
    dataframe = pd.DataFrame(data, columns=columns)

    # Drop rows where "Importreihenfolge" is NaN or empty
    if "Importreihenfolge" in dataframe.columns:
        dataframe = dataframe.dropna(subset=["Importreihenfolge"])
    else:
        raise ValueError("The column 'Importreihenfolge' does not exist in the data.")

    if "Name des Importprograms / Name der Erfassungsmaske" in dataframe.columns:
        nemosteps = dataframe[dataframe["Migrationsart"] == "NEMO"][
            "Name des Importprograms / Name der Erfassungsmaske"
        ].to_list()

        nemosteps = [x.title().strip() for x in nemosteps]
        replacements = {
            "European Article Numbers": "Global Trade Item Numbers",
            "Part-Storage Areas Relationship": "Part-Storage Areas Relationships",
            "Sales Tax Id": "Sales Tax ID",
            "Mrp Parameters": "MRP Parameters",
            "Sales Units Of Measure": "Sales Units of Measure",
            "Standard Boms (Header Data)": "Standard BOMs (Header Data)",
            "Standard Boms (Line Data)": "Standard BOMs (Line Data)",
            "Routings (Standard Boms)": "Routings (Standard BOMs)",
            "Bills Of Materials For Operations (Routings Production)": "Bills of Materials for Operations (Routings Production)",
        }

        nemosteps = [
            replacements[item] if item in replacements else item for item in nemosteps
        ]

        return nemosteps
    else:
        raise ValueError(
            "The column 'Name des Importprograms / Name der Erfassungsmaske' does not exist in the data."
        )


def getMappingRelations(config: Config) -> pd.DataFrame:
    """
    Get mapping relations based on the given configuration.

    Args:
        config (Config): The configuration object.

    Returns:
        pd.DataFrame: A DataFrame containing the mapping relations.
    """
    # get configuration
    mapping_fields = config.get_migman_mapping_fields()
    if not mapping_fields:
        logging.info(f"no mapping fields defined")
        return pd.DataFrame()
    additional_fields = config.get_migman_additional_fields()
    migman_projects = get_migman_project_list(config)
    if not migman_projects:
        raise ValueError("No migman projects defined.")

    # get data projects
    projects_display_name_migman = [
        project.displayName
        for project in getProjects(config)
        if project.displayName in migman_projects
    ]

    # scan projects for fields
    data = []
    for project in projects_display_name_migman:

        logging.info(f"scan project '{project}' for mapping fields...")

        # remove (xxx)
        def remove_brackets_if_present(name):
            pattern = r"\(\d{3}\)$"
            if re.search(pattern, name):
                return re.sub(pattern, "", name).strip()
            return name

        # get list of fields
        cols = getColumns(config=config, projectname=project)
        cols_cleaned = {
            remove_brackets_if_present(col.displayName): col for col in cols
        }

        # let's search the fields now
        for field in mapping_fields:
            # Check if the mapping field or any of its synonyms exists in imported_columns
            matching_field = None
            if field in cols_cleaned:
                matching_field = field
            else:
                for synonym in SYNONYM_FIELDS.get(field, []):
                    if synonym in cols_cleaned:
                        matching_field = synonym
                        break

            # If the mapping field or one of its synonyms is found
            if matching_field:

                # Check if all additional fields are also present
                additional_fields_present = additional_fields is None or all(
                    additional_field in cols_cleaned
                    for additional_field in additional_fields.get(field, [])
                )
                if additional_fields_present:

                    # we have cut of the (...) for easier handling. Now we have add them back again and add useful information for further processing
                    matching_field_display_name = cols_cleaned[
                        matching_field
                    ].displayName
                    matching_field_internal_name = cols_cleaned[
                        matching_field
                    ].internalName
                    matching_field_import_name = cols_cleaned[matching_field].importName

                    additional_field_information = []
                    for additional_field in additional_fields.get(field, []):
                        additional_field_display_name = cols_cleaned[
                            additional_field
                        ].displayName
                        additional_field_internal_name = cols_cleaned[
                            additional_field
                        ].internalName
                        additional_field_import_name = cols_cleaned[
                            additional_field
                        ].importName
                        additional_field_information.append(
                            (
                                additional_field_display_name,
                                additional_field_internal_name,
                                additional_field_import_name,
                            )
                        )

                    # Save the data for this mapping field
                    data.append(
                        {
                            "project": project,
                            "mapping_field": field,
                            "matching_field_display_name": matching_field_display_name,
                            "matching_field_internal_name": matching_field_internal_name,
                            "matching_field_import_name": matching_field_import_name,
                            "additional_fields": additional_field_information,
                        }
                    )

    logging.info(f"mapping related fields found: {json.dumps(data,indent=2)}")
    return pd.DataFrame(data)


def sqlQueryInMappingTable(
    config: Config,
    field: str,
    newProject: bool,
    mappingrelationsdf: pd.DataFrame,
) -> str:
    """
    Generate an SQL query for the mapping table based on the given parameters.

    Args:
        config (Config): The configuration object.
        field (str): The field to be queried.
        newProject (bool): Whether the project is new.
        mappingrelationsdf (pd.DataFrame): The DataFrame containing mapping relations.

    Returns:
        str: The generated SQL query.
    """
    projects = mappingrelationsdf["project"].to_list()
    display_names = mappingrelationsdf["matching_field_display_name"].to_list()
    internal_names = mappingrelationsdf["matching_field_internal_name"].to_list()
    additional_fields = mappingrelationsdf["additional_fields"].to_list()
    additional_fields_defined = config.get_migman_additional_fields()
    additional_field_global_definition = (
        additional_fields_defined.get(field, []) if additional_fields_defined else []
    )

    # setup CTEs to load data from source projects
    ctes = []
    for project, display_name, internal_name, additional_fields in zip(
        projects, display_names, internal_names, additional_fields
    ):

        subselect = [f'{internal_name} AS "source {field}"']
        if any(additional_fields):
            for (
                additional_field_label,
                additional_field_internal_name,
                additional_field_import_name,
            ), additional_field_definition in zip(
                additional_fields, additional_field_global_definition
            ):
                subselect.extend(
                    [
                        f'{additional_field_internal_name} AS "source {additional_field_definition}"'
                    ]
                )

        ctes.append(
            f"""CTE_{get_internal_name(project)} AS (
    SELECT DISTINCT
        {"\n\t,".join(subselect)}
    FROM 
        $schema.PROJECT_{get_internal_name(project)}
)"""
        )

    # global CTE to UNION ALL everything
    source_fields = [f'"source {field}"']
    for additional_field in additional_field_global_definition:
        source_fields.append(f'"source {additional_field}"')

    cteallfrags = [
        f"""SELECT
        {"\n\t, ".join(source_fields)} from CTE_{get_internal_name(project)} """
        for project in projects
    ]

    joinfrags = [
        f'mapping.{get_internal_name(field.strip('"'))} = cte.{field}'
        for field in source_fields
    ]

    # build the final query
    query = f""" WITH {"\n, ".join(ctes)}
, CTEALL AS (
    {"\nUNION ALL\n\t".join(cteallfrags)}
)
, CTEALLDISTINCT AS (
    SELECT DISTINCT
        {"\n\t, ".join(source_fields)}
    FROM   
        CTEALL  
)
SELECT
    cte.{"\n\t, cte.".join(source_fields)}
    , {"NULL" if newProject else f"mapping.TARGET_{get_internal_name(field)}"} AS "target {field}"
FROM    
    CTEALLDISTINCT cte
"""
    if not newProject:
        query += f"""LEFT JOIN
    $schema.$table mapping
ON  
    {"\n\tAND ".join(joinfrags)}"""

    return query


def is_migman_project_existing(
    database: list[MigMan],
    project: str,
) -> bool:
    return len([x.nemo_import_name for x in database if x.project == project]) > 0


def get_migman_postfixes(
    database: list[MigMan],
    project: str,
) -> list[str]:
    return list(set([x.postfix for x in database if x.project == project]))


def get_migman_fields(
    database: list[MigMan],
    project: str,
    postfix: str,
) -> list[str]:
    return [
        x.nemo_import_name
        for x in database
        if x.project == project and x.postfix == postfix
    ]


def get_migman_mandatory_fields(
    database: list[MigMan],
    project: str,
    postfix: str,
) -> list[str]:
    return [
        x.nemo_import_name
        for x in database
        if x.project == project and x.postfix == postfix and x.snow_mandatory == True
    ]


def get_mig_man_field(
    database: list[MigMan], project: str, postfix: str, index: int
) -> MigMan | None:
    columns = [
        x
        for x in database
        if x.project == project and x.postfix == postfix and x.index == index
    ]
    return columns[0] if columns else None
