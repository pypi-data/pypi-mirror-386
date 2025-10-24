import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

logger = logging.getLogger(__name__)

def load_input_file(input_file, sheets_list=None):
    '''
    Load an input file which can be a .csv, .xls, .xlsx, or .xlsm file.
    Returns a dictionary of dataframes, with the key as the sheet name (or 'Sheet' for csv)
    and the value as the dataframe.
    '''
    file_extension = os.path.splitext(input_file)[1].lower()
    sheets = {}
    
    if file_extension == '.csv':
        df = pd.read_csv(input_file)
        sheets['Sheet'] = df
    elif file_extension in ['.xls', '.xlsx', '.xlsm']:
        xls = pd.ExcelFile(input_file)
        if not sheets_list:
            sheets_list = xls.sheet_names
        for sheet_name in sheets_list:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            sheets[sheet_name] = df
    else:
        raise ValueError("Unsupported file format. Please use .csv, .xls, .xlsx, or .xlsm files.")
    
    logger.debug(f"Sheets loaded: {sheets.keys()}")
    
    return sheets


def load_template(template_name, sheets):
    '''
    Load a template workbook if provided, otherwise create a new workbook.
    Ensures that all sheets in the data are present in the workbook.
    '''
    if template_name:
        wb = load_workbook(template_name)
        logger.debug("Template loaded.")
        include_header = False  # if there is a template, do not include headers
    else:
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name in sheets.keys():
            wb.create_sheet(sheet_name)
        logger.debug(f"Loaded new workbook with sheets: {sheets.keys()}.")
        include_header = True
    return wb, include_header


def split_files(output_folder, splitter_column, template_name, naming_suffix, sheets, current_time):
    '''
    The "sheets" input is a dictionary with the key as the sheetname and the value as a dataframe
    of the data to be added into that sheet. "current_time" is a timestamp.
    "splitter_column" is a column in each sheet that will be used to split the template into 
    separate files, based on the unique values of the column.
    '''
    newpath = f'{output_folder}_{current_time}'  # create a new output folder name with current time
    os.makedirs(newpath)  # create output folder

    # create a file with all properties in it:
    
    wb, include_header = load_template(template_name, sheets)

    for sheet_name, query_results in sheets.items(): 
        logger.debug(f"Sheet: {sheet_name}")
        logger.debug(f"Data:\n{query_results}")
        for r in dataframe_to_rows(query_results, index=False, header=include_header):
            wb[sheet_name].append(r)

    wb.save(f'{newpath}/_COMBINED_LIST.xlsx')
    wb.close()

    logger.info(f"Combined List workbook saved to {output_folder}_{current_time}")

    list_of_dfs_for_split_column = []
    for _, query_results in sheets.items():
        try:
            list_of_dfs_for_split_column.append(query_results[splitter_column])
        except KeyError:
            raise KeyError(f"Column '{splitter_column}' not found in one of the sheets.")
    unique_split_series = pd.concat(list_of_dfs_for_split_column).drop_duplicates().dropna()

    output_log = pd.DataFrame()

    for each in unique_split_series:  # iterate over each servicer in the unique servicer list
        logger.debug(f"Loop: {each}")
        row_counter = 0
        print_output = {'Segment': each,}
        
        wb, include_header = load_template(template_name, sheets)  # load the template workbook

        for sheet_name, query_results in sheets.items():
            logger.debug(f"Sheet: {sheet_name}")
            df_filtered = query_results[query_results[splitter_column] == each]
            logger.debug(f"Filtered:\n{df_filtered}")
            row_counter += df_filtered.shape[0]
            for r in dataframe_to_rows(df_filtered, index=False, header=include_header):
                wb[sheet_name].append(r)
            wb[sheet_name].title = f'{sheet_name} ({df_filtered.shape[0]})'  # rename the sheet with the count
            print_output[sheet_name] = df_filtered.shape[0]

        output_filename = f"{each}_{naming_suffix}({row_counter}).xlsx"

        print_output['Total'] = row_counter
        
        wb.save(f'{newpath}/{output_filename}')  # save the file
        wb.close()

        data = pd.DataFrame(print_output, index=[0])
        output_log = pd.concat([output_log, data], ignore_index=True)

    logger.info(f"All workbooks exported to {output_folder}_{current_time}")
    logger.info("*************************\nOUTPUT:")
    logger.info(output_log)
    logger.info("\n")

    return newpath, output_log