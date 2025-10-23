import re

import openpyxl
import pytest

from bam_masterdata.excel import MasterdataExcelExtractor
from bam_masterdata.metadata.definitions import DataType


@pytest.fixture
def excel_extractor(tmp_path):
    """Fixture to create an instance of MasterdataExcelExtractor with a dummy Excel file."""
    # Create a dummy Excel file in the temporary test directory
    dummy_excel = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add some dummy data
    ws.append(["Header1", "Header2", "Header3"])
    ws.append(["Data1", "Data2", "Data3"])
    ws.append(["", "", ""])  # Empty row
    ws.append(["Last", "Row", "Data"])

    wb.save(dummy_excel)

    # Create an instance with the dummy file
    return MasterdataExcelExtractor(excel_path=str(dummy_excel))


# Tests for `index_to_excel_column`
@pytest.mark.parametrize(
    "index, expected_column",
    [
        (1, "A"),
        (2, "B"),
        (26, "Z"),
        (27, "AA"),
        (28, "AB"),
        (52, "AZ"),
        (53, "BA"),
        (702, "ZZ"),
        (703, "AAA"),
        (704, "AAB"),
        (1378, "AZZ"),
        (1379, "BAA"),
        (18278, "ZZZ"),
        (18279, "AAAA"),
        (18280, "AAAB"),
    ],
)
def test_index_to_excel_column(excel_extractor, index, expected_column):
    """Tests the index_to_excel_column method."""
    result = excel_extractor.index_to_excel_column(index)
    assert result == expected_column


# Failing cases for `index_to_excel_column`
@pytest.mark.parametrize(
    "invalid_index",
    [-1, 0],  # Invalid cases
)
def test_index_to_excel_column_invalid(excel_extractor, invalid_index):
    """Tests that index_to_excel_column fails with invalid input."""
    with pytest.raises(ValueError):
        excel_extractor.index_to_excel_column(invalid_index)


# Tests for `get_last_non_empty_row`
def test_get_last_non_empty_row(excel_extractor):
    """Tests finding the last non-empty row."""
    sheet = excel_extractor.workbook["Sheet1"]

    result = excel_extractor.get_last_non_empty_row(sheet, 1)

    assert result == 2  # Last non-empty row should be 4


# Failing cases for `get_last_non_empty_row`
@pytest.mark.parametrize(
    "invalid_sheet, start_index, expected_exception",
    [
        (None, 1, AttributeError),  # No sheet provided
        ("Sheet1", -1, ValueError),  # Negative index
        ("Sheet1", 999, ValueError),  # Index out of bounds
    ],
)
def test_get_last_non_empty_row_invalid(
    excel_extractor, invalid_sheet, start_index, expected_exception
):
    """Tests that get_last_non_empty_row fails with invalid input."""
    try:
        sheet = (
            excel_extractor.workbook[invalid_sheet]
            if isinstance(invalid_sheet, str)
            else invalid_sheet
        )
    except KeyError:
        if expected_exception is KeyError:
            return  # Test passes if KeyError was expected
        raise  # Otherwise, raise the error

    with pytest.raises(expected_exception):
        excel_extractor.get_last_non_empty_row(sheet, start_index)


@pytest.mark.parametrize(
    "value, expected_result, log_message, log_message_level",
    [
        ("true", True, None, None),
        ("false", False, None, None),
        (True, True, None, None),
        (False, False, None, None),
        ("TrUe", True, None, None),
        ("FaLsE", False, None, None),
        ("", False, None, None),
        (None, False, None, None),
        (
            "yes",
            False,
            "Invalid boolean term value found in the Boolean Term column at position A1 in Sheet1. Accepted values: TRUE or FALSE.",
            "error",
        ),
        (
            "123",
            False,
            "Invalid boolean term value found in the Boolean Term column at position A1 in Sheet1. Accepted values: TRUE or FALSE.",
            "error",
        ),
    ],
)
def test_str_to_bool(
    cleared_log_storage,
    excel_extractor,
    value,
    expected_result,
    log_message,
    log_message_level,
):
    """Tests conversion of string values to boolean with `str_to_bool`."""
    result = excel_extractor.str_to_bool(value, "Boolean Term", "A1", "Sheet1")
    assert result == expected_result

    if log_message:
        # Check if the expected error message appears in logs
        assert any(
            log["event"] == log_message and log["level"] == log_message_level
            for log in cleared_log_storage
        )
    else:
        # Ensure no error messages are logged
        assert not any(log["level"] == "error" for log in cleared_log_storage)


@pytest.mark.parametrize(
    "value, term, coordinate, sheet_title, is_description, is_code, is_data, is_url, expected_result, log_message, log_message_level",
    [
        # Valid cases (No errors should be logged)
        (
            "Valid Code",
            "Code",
            "A1",
            "Sheet1",
            False,
            False,
            False,
            False,
            "Valid Code",
            None,
            None,
        ),
        (
            "Valid Description // Gültige Beschreibung",
            "Description",
            "A1",
            "Sheet1",
            True,
            False,
            False,
            False,
            "Valid Description // Gültige Beschreibung",
            None,
            None,
        ),
        (
            "VALID_CODE",
            "Code",
            "B2",
            "Sheet2",
            False,
            True,
            False,
            False,
            "VALID_CODE",
            None,
            None,
        ),
        (
            "INTEGER",
            "Data Type",
            "C3",
            "Sheet3",
            False,
            False,
            True,
            False,
            "INTEGER",
            None,
            None,
        ),
        (
            "https://valid-url.com",
            "URL",
            "D4",
            "Sheet4",
            False,
            False,
            False,
            True,
            "https://valid-url.com",
            None,
            None,
        ),
        # Invalid cases (Errors should be logged)
        (
            "Missing Separator",
            "Description",
            "A5",
            "Sheet5",
            True,
            False,
            False,
            False,
            "Missing Separator",
            "Invalid description value found in the Description column at position A5 in Sheet5. "
            "Description should follow the schema: English Description + '//' + German Description. ",
            "error",
        ),
        (
            "Invalid Code!",
            "Code",
            "B6",
            "Sheet6",
            False,
            True,
            False,
            False,
            "Invalid Code!",
            "Invalid code value found in the Code column at position B6 in Sheet6.",
            "error",
        ),
        (
            "unknown_data_type",
            "Data Type",
            "C7",
            "Sheet7",
            False,
            False,
            True,
            False,
            "UNKNOWN_DATA_TYPE",
            f"Invalid data type value found in the Data Type column at position C7 in Sheet7. "
            f"The Data Type should be one of the following: {list(dt.value for dt in DataType)}",
            "error",
        ),
        (
            "invalid-url",
            "URL",
            "D8",
            "Sheet8",
            False,
            False,
            False,
            True,
            "invalid-url",
            "Invalid url value found in the URL column at position D8 in Sheet8.",
            "error",
        ),
    ],
)
def test_get_and_check_property(
    cleared_log_storage,
    excel_extractor,
    value,
    term,
    coordinate,
    sheet_title,
    is_description,
    is_code,
    is_data,
    is_url,
    expected_result,
    log_message,
    log_message_level,
):
    """Tests property validation and formatting with `get_and_check_property`."""
    result = excel_extractor.get_and_check_property(
        value, term, coordinate, sheet_title, is_description, is_code, is_data, is_url
    )
    assert result == expected_result

    if log_message:
        # Normalize log message formatting to avoid strict mismatches
        def normalize_log_message(msg):
            return re.sub(r"\s+", " ", msg.replace(". ", ".")).strip()

        cleaned_logs = [
            normalize_log_message(log["event"]) for log in cleared_log_storage
        ]
        expected_cleaned_message = normalize_log_message(log_message)

        # Generalized check (avoids strict spacing mismatches)
        assert any(expected_cleaned_message in log for log in cleaned_logs)
    else:
        # Ensure no error messages are logged
        assert not any(log["level"] == "error" for log in cleared_log_storage)


@pytest.mark.parametrize(
    "term, cell_value, coordinate, sheet_title, expected_result, log_message, log_message_level",
    [
        # Valid cases (No errors should be logged)
        ("Mandatory", "true", "A1", "Sheet1", True, None, None),
        ("Show in edit views", "False", "B2", "Sheet2", False, None, None),
        ("Code", "VALID_CODE", "C3", "Sheet3", "VALID_CODE", None, None),
        ("Data type", "INTEGER", "D4", "Sheet4", "INTEGER", None, None),
        # Invalid cases (Errors should be logged)
        (
            "Mandatory",
            "invalid_boolean",
            "A7",
            "Sheet7",
            False,
            "Invalid mandatory value found in the Mandatory column at position A7 in Sheet7. Accepted values: TRUE or FALSE.",
            "error",
        ),
        (
            "Show in edit views",
            123,
            "B8",
            "Sheet8",
            False,
            "Invalid show in edit views value found in the Show in edit views column at position B8 in Sheet8. Accepted values: TRUE or FALSE.",
            "error",
        ),
        (
            "Code",
            "Invalid Code!",
            "C9",
            "Sheet9",
            "Invalid Code!",
            "Invalid code value found in the Code column at position C9 in Sheet9.",
            "error",
        ),
        (
            "Data type",
            "UNKNOWN_TYPE",
            "D10",
            "Sheet10",
            "UNKNOWN_TYPE",
            f"Invalid data type value found in the Data type column at position D10 in Sheet10. The Data Type should be one of the following: {list(dt.value for dt in DataType)}",
            "error",
        ),
    ],
)
def test_process_term(
    cleared_log_storage,
    excel_extractor,
    term,
    cell_value,
    coordinate,
    sheet_title,
    expected_result,
    log_message,
    log_message_level,
):
    """Tests `process_term` to validate and process terms correctly."""

    result = excel_extractor.process_term(term, cell_value, coordinate, sheet_title)

    assert result == expected_result

    if log_message:
        # Normalize log message formatting to avoid strict mismatches
        def normalize_log_message(msg):
            return re.sub(r"\s+", " ", msg.replace(". ", ".")).strip()

        cleaned_logs = [
            normalize_log_message(log["event"]) for log in cleared_log_storage
        ]
        expected_cleaned_message = normalize_log_message(log_message)

        # Ensure log appears
        assert any(expected_cleaned_message in log for log in cleaned_logs), (
            f"Expected log message was not found. Logs: {cleaned_logs}"
        )
    else:
        # Ensure no error messages are logged
        assert not any(log["level"] == "error" for log in cleared_log_storage)


@pytest.mark.parametrize(
    "cell_value, row, column, validation_pattern, is_description, is_data, is_url, expected_result, log_message, log_message_level",
    [
        # Valid cases (No errors should be logged)
        (
            "VALID_CODE",
            1,
            1,
            r"^\$?[A-Z0-9_.]+$",
            False,
            False,
            False,
            "VALID_CODE",
            None,
            None,
        ),
        (
            "Valid Description // Gültige Beschreibung",
            2,
            2,
            r".*//.*",
            True,
            False,
            False,
            "Valid Description // Gültige Beschreibung",
            None,
            None,
        ),
        ("INTEGER", 3, 3, None, False, True, False, "INTEGER", None, None),
        (
            "https://valid-url.com",
            4,
            4,
            r"https?://(?:www\.)?[a-zA-Z0-9-._~:/?#@!$&'()*+,;=%]+",
            False,
            False,
            True,
            "https://valid-url.com",
            None,
            None,
        ),
        # Invalid cases (Errors should be logged)
        (
            "Missing Separator",
            6,
            1,
            r".*//.*",
            True,
            False,
            False,
            "Missing Separator",
            "Invalid value 'Missing Separator' at row 6, column 1 in sheet TestSheet Description should follow the schema: English Description + '//' + German Description.",
            "error",
        ),
        (
            "unknown_data_type",
            7,
            2,
            None,
            False,
            True,
            False,
            "unknown_data_type",
            f"Invalid value 'unknown_data_type' at row 7, column 2 in sheet TestSheet The Data Type should be one of the following: {list(dt.value for dt in DataType)}",
            "error",
        ),
        (
            "invalid-url",
            8,
            3,
            r"https?://(?:www\.)?[a-zA-Z0-9-._~:/?#@!$&'()*+,;=%]+",
            False,
            False,
            True,
            "invalid-url",
            "Invalid value 'invalid-url' at row 8, column 3 in sheet TestSheet It should be an URL or empty",
            "error",
        ),
        (
            "123INVALID",
            9,
            4,
            r"^[A-Z]+_\d+$",
            False,
            False,
            False,
            "123INVALID",
            "Invalid value '123INVALID' at row 9, column 4 in sheet TestSheet",
            "error",
        ),
        (
            "invalid code",
            9,
            4,
            r"^\$?[A-Z0-9_.]+$",
            False,
            False,
            False,
            "invalid code",
            "Invalid value 'invalid code' at row 9, column 4 in sheet TestSheet",
            "error",
        ),
        # Empty cell (should return empty string, no error)
        (None, 10, 5, None, False, False, False, "", None, None),
        ("", 11, 6, None, False, False, False, "", None, None),
    ],
)
def test_extract_value(
    cleared_log_storage,
    excel_extractor,
    cell_value,
    row,
    column,
    validation_pattern,
    is_description,
    is_data,
    is_url,
    expected_result,
    log_message,
    log_message_level,
):
    """Tests `extract_value` function for extracting and validating cell values."""

    # Create a dummy worksheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "TestSheet"

    # Set the cell value
    sheet.cell(row=row, column=column, value=cell_value)

    # Call function
    result = excel_extractor.extract_value(
        sheet, row, column, validation_pattern, is_description, is_data, is_url
    )

    # Assert value extraction result
    assert result == expected_result

    if log_message:
        # Normalize log message formatting
        def normalize_log_message(msg):
            return re.sub(r"\s+", " ", msg.replace(". ", ".")).strip()

        cleaned_logs = [
            (normalize_log_message(log["event"]), log["level"])
            for log in cleared_log_storage
        ]
        expected_cleaned_message = normalize_log_message(log_message)

        # Ensure log appears with the correct level
        assert any(
            expected_cleaned_message in log_text and log_level == log_message_level
            for log_text, log_level in cleaned_logs
        ), (
            f"Expected log message '{expected_cleaned_message}' with level '{log_message_level}' was not found. Logs: {cleaned_logs}"
        )

    else:
        # Ensure no error messages are logged
        assert not any(log["level"] == "error" for log in cleared_log_storage), (
            "Unexpected error log found!"
        )


@pytest.mark.parametrize(
    "header_terms, expected_terms, entity_type, cell_values, expected_attributes, log_message, log_message_level",
    [
        # Valid Case: SAMPLE_TYPE (No errors expected)
        (
            ["Code", "Description", "Generated code prefix"],  # Header terms
            [
                "Code",
                "Description",
                "Generated code prefix",
            ],  # Expected extracted terms
            "SAMPLE_TYPE",
            [
                "ABC123",
                "A valid description // Eine gültige Beschreibung",
                "ABC",
            ],  # Cell values
            {
                "code": "ABC123",
                "description": "A valid description // Eine gültige Beschreibung",
                "generatedCodePrefix": "ABC",
            },  # Expected attributes
            None,  # No error
            None,
        ),
        # Invalid Case: Unknown Data Type (PROPERTY_TYPE)
        (
            ["Code", "Data type"],
            ["Code", "Data type"],
            "PROPERTY_TYPE",
            ["ABC123", "UNKNOWN_TYPE"],
            {"code": "ABC123", "dataType": "UNKNOWN_TYPE"},
            "Invalid Data Type: UNKNOWN_TYPE in B3 (Sheet: TestSheet). Should be one of the following: "
            f"{[dt.value for dt in DataType]}",
            "error",
        ),
        # Invalid Case: Boolean Error (SAMPLE_TYPE)
        (
            ["Code", "Auto generate codes"],
            ["Code", "Auto generate codes"],
            "SAMPLE_TYPE",
            ["ABC123", "maybe"],  # Invalid boolean
            {"code": "ABC123", "autoGeneratedCode": False},  # Expected processed value
            "Invalid auto generate codes value found in the Auto generate codes column at position B3 in TestSheet. Accepted values: TRUE or FALSE.",
            "error",
        ),
        # Invalid URL Case (VOCABULARY_TYPE)
        (
            ["Code", "Url template"],
            ["Code", "Url template"],
            "VOCABULARY_TYPE",
            ["ABC123", "invalid-url"],
            {"code": "ABC123", "url_template": "invalid-url"},
            "Invalid URL format: invalid-url in B3 (Sheet: TestSheet)",
            "error",
        ),
        # Invalid Generated Code Prefix (SAMPLE_TYPE)
        (
            ["Code", "Generated code prefix"],
            ["Code", "Generated code prefix"],
            "SAMPLE_TYPE",
            ["ABC123", "ABC_123"],
            {"code": "ABC123", "generatedCodePrefix": "ABC_123"},
            "Invalid Generated code prefix value 'ABC_123' in B3 (Sheet: TestSheet). Generated code prefix should be part of the 'Code' ABC123.",
            "warning",
        ),
    ],
)
def test_process_entity(
    cleared_log_storage,
    excel_extractor,
    header_terms,
    expected_terms,
    entity_type,
    cell_values,
    expected_attributes,
    log_message,
    log_message_level,
):
    """Tests processing entity attributes and validation logging in `process_entity`."""

    # Create dummy Excel sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "TestSheet"

    # Insert headers in row 1
    for col, header in enumerate(header_terms, start=1):
        sheet.cell(row=1, column=col, value=header)

    # Insert test values in row 3 (since `start_index_row + 2` is used)
    for col, value in enumerate(cell_values, start=1):
        sheet.cell(row=3, column=col, value=value)

    # Call the function
    result = excel_extractor.process_entity(
        sheet,
        start_index_row=1,
        header_terms=header_terms,
        expected_terms=expected_terms,
        entity_type=entity_type,
    )

    # Assert entity attributes
    assert result == expected_attributes, (
        f"Expected: {expected_attributes}, but got: {result}"
    )

    # Log message checking
    if log_message:
        cleaned_logs = [
            re.sub(r"\s+", " ", log["event"]).strip() for log in cleared_log_storage
        ]
        expected_cleaned_message = re.sub(r"\s+", " ", log_message).strip()

        assert any(expected_cleaned_message in log for log in cleaned_logs), (
            "Expected log message was not found!"
        )

        # Ensure the correct log level is used
        assert any(log["level"] == log_message_level for log in cleared_log_storage), (
            f"Expected log level '{log_message_level}' not found."
        )
    else:
        assert not any(log["level"] == "error" for log in cleared_log_storage), (
            "Unexpected error logs found!"
        )


@pytest.mark.parametrize(
    "header_terms, cell_values, last_non_empty_row, expected_properties, log_message, log_message_level",
    [
        # Valid Case: Extract properties correctly
        (
            [
                "Code",
                "Description",
                "Mandatory",
                "Show in edit views",
                "Section",
                "Property label",
                "Data type",
                "Vocabulary code",
                "Metadata",
                "Dynamic script",
            ],
            [
                [
                    "PROP_001",
                    "Sample description",
                    "True",
                    "False",
                    "General",
                    "Property Label",
                    "INTEGER",
                    "VOCAB_ABC",
                    "Meta info",
                    "Script example",
                ],
                [
                    "PROP_002",
                    "Another description",
                    "False",
                    "True",
                    "Advanced",
                    "Another Label",
                    "BOOLEAN",
                    "VOCAB_DEF",
                    None,  # Metadata not provided
                    None,  # Dynamic script not provided
                ],
            ],
            6,
            {
                "PROP_001": {
                    "permId": "PROP_001",
                    "code": "PROP_001",
                    "description": "Sample description",
                    "mandatory": True,
                    "show_in_edit_views": False,
                    "section": "General",
                    "label": "Property Label",
                    "dataType": "INTEGER",
                    "vocabularyCode": "VOCAB_ABC",
                    "metadata": "Meta info",
                    "dynamic_script": "Script example",
                },
                "PROP_002": {
                    "permId": "PROP_002",
                    "code": "PROP_002",
                    "description": "Another description",
                    "mandatory": False,
                    "show_in_edit_views": True,
                    "section": "Advanced",
                    "label": "Another Label",
                    "dataType": "BOOLEAN",
                    "vocabularyCode": "VOCAB_DEF",
                    "metadata": None,
                    "dynamic_script": None,
                },
            },
            None,
            None,
        ),
        # Missing Code (Should be ignored)
        (
            [
                "Description",
                "Mandatory",
                "Show in edit views",
                "Section",
                "Property label",
                "Data type",
                "Vocabulary code",
                "Metadata",
                "Dynamic script",
            ],  # Missing "Code"
            [
                [
                    "Description only",
                    "True",
                    "False",
                    "General",
                    "Label",
                    "INTEGER",
                    "VOCAB_ABC",
                    "Meta info",
                    "Script example",
                ]
            ],
            5,
            {},
            "'Code' not found in the properties headers for sheet TestSheet.",
            "error",
        ),
    ],
)
def test_properties_to_dict(
    cleared_log_storage,
    excel_extractor,
    header_terms,
    cell_values,
    last_non_empty_row,
    expected_properties,
    log_message,
    log_message_level,
):
    """Tests `properties_to_dict` function to validate properties extraction."""

    # Create dummy Excel sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "TestSheet"

    # Insert headers at row 4 (header_index = start_index_row + 3)
    for col, header in enumerate(header_terms, start=1):
        sheet.cell(row=4, column=col, value=header)

    # Insert test values starting from row 5
    for row_index, row_values in enumerate(cell_values, start=5):
        for col_index, value in enumerate(row_values, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

    # Call function
    result = excel_extractor.properties_to_dict(
        sheet, start_index_row=1, last_non_empty_row=last_non_empty_row
    )

    # Assert dictionary structure
    assert result == expected_properties, (
        f"Expected: {expected_properties}, but got: {result}"
    )

    # Log message checking
    if log_message:
        cleaned_logs = [
            re.sub(r"\s+", " ", log["event"]).strip() for log in cleared_log_storage
        ]
        expected_cleaned_message = re.sub(r"\s+", " ", log_message).strip()

        # Ensure the expected log message appears
        assert any(expected_cleaned_message in log for log in cleaned_logs), (
            "Expected log message was not found!"
        )

        # Ensure the correct log level is used
        assert any(log["level"] == log_message_level for log in cleared_log_storage), (
            f"Expected log level '{log_message_level}' not found."
        )
    else:
        assert not any(log["level"] == "error" for log in cleared_log_storage), (
            "Unexpected error logs found!"
        )


@pytest.mark.parametrize(
    "header_terms, cell_values, last_non_empty_row, expected_terms, log_message, log_message_level",
    [
        # Valid Case: Extract vocabulary terms correctly
        (
            ["Code", "Description", "Url template", "Label", "Official"],
            [
                [
                    "TERM_001",
                    "Sample description",
                    "https://example.com",
                    "Label1",
                    "TRUE",
                ],
                [
                    "TERM_002",
                    "Another description",
                    "https://another.com",
                    "Label2",
                    "FALSE",
                ],
            ],
            6,
            {
                "TERM_001": {
                    "permId": "TERM_001",
                    "code": "TERM_001",
                    "descriptions": "Sample description",
                    "url_template": "https://example.com",
                    "label": "Label1",
                    "official": True,
                },
                "TERM_002": {
                    "permId": "TERM_002",
                    "code": "TERM_002",
                    "descriptions": "Another description",
                    "url_template": "https://another.com",
                    "label": "Label2",
                    "official": False,
                },
            },
            None,
            None,
        ),
        # Missing Header: "Url template"
        (
            ["Code", "Description", "Label", "Official"],  # Missing "Url template"
            [["TERM_003", "Description", "Label3", "TRUE"]],
            4,
            {},
            "Url template not found in the properties headers.",
            "warning",
        ),
        # Invalid "Code" (should log an error)
        (
            ["Code", "Description", "Url template", "Label", "Official"],
            [
                [
                    "Invalid Code!",
                    "Description",
                    "https://example.com",
                    "Label4",
                    "FALSE",
                ]
            ],
            5,
            {
                "Invalid Code!": {
                    "permId": "Invalid Code!",
                    "code": "Invalid Code!",
                    "descriptions": "Description",
                    "url_template": "https://example.com",
                    "label": "Label4",
                    "official": False,
                }
            },
            "Invalid code value found in the Code column",
            "error",
        ),
        # Invalid "Url template" (should log an error)
        (
            ["Code", "Description", "Url template", "Label", "Official"],
            [["TERM_004", "Description", "invalid-url", "Label5", "TRUE"]],
            5,
            {
                "TERM_004": {
                    "permId": "TERM_004",
                    "code": "TERM_004",
                    "descriptions": "Description",
                    "url_template": "invalid-url",  # Still included even if invalid
                    "label": "Label5",
                    "official": True,
                }
            },
            "Invalid url template value found in the Url template column at position C5 in TestSheet.",
            "error",
        ),
        # Invalid "Official" value (should log an error)
        (
            ["Code", "Description", "Url template", "Label", "Official"],
            [["TERM_005", "Description", "https://valid.com", "Label6", "INVALID"]],
            5,
            {
                "TERM_005": {
                    "permId": "TERM_005",
                    "code": "TERM_005",
                    "descriptions": "Description",
                    "url_template": "https://valid.com",
                    "label": "Label6",
                    "official": False,  # Should be False because is default value when criteria not meet
                }
            },
            "Invalid official value found in the Official column",
            "error",
        ),
    ],
)
def test_terms_to_dict(
    cleared_log_storage,
    excel_extractor,
    header_terms,
    cell_values,
    last_non_empty_row,
    expected_terms,
    log_message,
    log_message_level,
):
    """Tests `terms_to_dict` function to validate vocabulary extraction."""

    # Create a dummy Excel sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "TestSheet"

    # Insert headers at row 4 (header_index = start_index_row + 3)
    for col, header in enumerate(header_terms, start=1):
        sheet.cell(row=4, column=col, value=header)

    # Insert test values starting from row 5
    for row_index, row_values in enumerate(cell_values, start=5):
        for col_index, value in enumerate(row_values, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

    # Call function
    result = excel_extractor.terms_to_dict(
        sheet, start_index_row=1, last_non_empty_row=last_non_empty_row
    )

    # Assert dictionary structure
    assert result == expected_terms, f"Expected: {expected_terms}, but got: {result}"

    # Log message checking
    if log_message:
        cleaned_logs = [
            re.sub(r"\s+", " ", log["event"]).strip() for log in cleared_log_storage
        ]
        expected_cleaned_message = re.sub(r"\s+", " ", log_message).strip()

        # Ensure the expected log message appears
        assert any(expected_cleaned_message in log for log in cleaned_logs), (
            "Expected log message was not found!"
        )

        # Ensure the correct log level is used
        assert any(log["level"] == log_message_level for log in cleared_log_storage), (
            f"Expected log level '{log_message_level}' not found."
        )
    else:
        assert not any(log["level"] == "error" for log in cleared_log_storage), (
            "Unexpected error logs found!"
        )


import pytest


@pytest.mark.parametrize(
    "entity_type, header_terms, cell_values, last_non_empty_row, expected_entities, expected_exception",
    [
        # Valid Case: Extract SAMPLE_TYPE
        (
            "SAMPLE_TYPE",
            [
                "Code",
                "Description",
                "Validation script",
                "Generated code prefix",
                "Auto generate codes",
            ],
            [
                [
                    "SAMPLE_001",
                    "Sample Description",
                    "val.py",
                    "SAM_001",
                    "True",
                ],
            ],
            6,  # Last non-empty row index (including the blank row)
            {
                "SAMPLE_001": {
                    "code": "SAMPLE_001",
                    "description": "Sample Description",
                    "validationPlugin": "val.py",  # Expected extracted field
                    "generatedCodePrefix": "SAM_001",  # Extracted field
                    "autoGeneratedCode": True,  # Boolean value converted
                    "properties": {},  # Empty properties extracted
                },
            },
            None,
        ),
        # Valid Case: Extract VOCABULARY_TYPE
        (
            "VOCABULARY_TYPE",
            ["Code", "Description", "Url template"],
            [
                ["VOC_001", "Description 1", "https://example.com"],
            ],
            5,
            {
                "VOC_001": {
                    "code": "VOC_001",
                    "description": "Description 1",
                    "url_template": "https://example.com",
                    "terms": {},
                },
            },
            None,
        ),
        # Invalid Entity Type
        (
            "INVALID_TYPE",
            [
                "Code",
                "Description",
                "Validation script",
                "Generated code prefix",
                "Auto generate codes",
            ],
            [["SAMPLE_003", "Invalid Sample", "True", "False"]],
            4,
            None,
            pytest.raises(ValueError, match="Invalid entity type: INVALID_TYPE"),
        ),
    ],
)
def test_block_to_entity_dict(
    excel_extractor,
    entity_type,
    header_terms,
    cell_values,
    last_non_empty_row,
    expected_entities,
    expected_exception,
):
    """Tests `block_to_entity_dict` function for extracting entities from Excel blocks."""

    # Create a dummy Excel sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "TestSheet"

    # Insert entity type in A1
    sheet["A1"] = entity_type

    # Insert headers at row 2 (header_index = start_index_row + 1)
    for col, header in enumerate(header_terms, start=1):
        sheet.cell(row=2, column=col, value=header)

    # Insert test values starting from row 3
    for row_index, row_values in enumerate(cell_values, start=3):
        for col_index, value in enumerate(row_values, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

    # Call function
    complete_dict = {}
    if expected_exception:
        with expected_exception:  # Properly assert the exception
            result = excel_extractor.block_to_entity_dict(
                sheet,
                start_index_row=1,
                last_non_empty_row=last_non_empty_row,
                complete_dict=complete_dict,
            )
    else:
        result = excel_extractor.block_to_entity_dict(
            sheet,
            start_index_row=1,
            last_non_empty_row=last_non_empty_row,
            complete_dict=complete_dict,
        )

        # Assert dictionary structure
        assert result == expected_entities, (
            f"Expected: {expected_entities}, but got: {result}"
        )


@pytest.mark.parametrize(
    "sheets_data, expected_result",
    [
        # Case 1: Single sheet with one valid entity block
        (
            {
                "Object Types": [
                    ["SAMPLE_TYPE"],
                    [
                        "Code",
                        "Description",
                        "Validation script",
                        "Generated code prefix",
                        "Auto generate codes",
                    ],
                    [
                        "SAMPLE_001",
                        "Sample Description",
                        "val.py",
                        "SAM_001",
                        "True",
                    ],
                ]
            },
            {
                "object_types": {
                    "SAMPLE_001": {
                        "code": "SAMPLE_001",
                        "description": "Sample Description",
                        "validationPlugin": "val.py",
                        "generatedCodePrefix": "SAM_001",
                        "autoGeneratedCode": True,
                        "properties": {},
                    }
                }
            },
        ),
        # Case 2: Multiple sheets with multiple entity blocks (Object Types + Collection Types) - **Now includes additional entity types per sheet**
        (
            {
                "Object Types": [
                    ["SAMPLE_TYPE"],
                    [
                        "Code",
                        "Description",
                        "Validation script",
                        "Generated code prefix",
                        "Auto generate codes",
                    ],
                    [
                        "SAMPLE_001",
                        "Sample Description",
                        "val.py",
                        "SAM_001",
                        "True",
                    ],
                    [],
                    ["SAMPLE_TYPE"],  # Second entity block in the same sheet
                    [
                        "Code",
                        "Description",
                        "Validation script",
                        "Generated code prefix",
                        "Auto generate codes",
                    ],
                    [
                        "SAMPLE_002",
                        "Another Sample",
                        None,
                        "SAM_002",
                        "False",
                    ],
                ],
                "Collection Types": [
                    ["EXPERIMENT_TYPE"],
                    [
                        "Code",
                        "Description",
                        "Validation script",
                        "Generated code prefix",
                        "Auto generate codes",
                    ],
                    [
                        "COLL_001",
                        "Collection Example",
                        "val_coll.py",
                        "COLL_ABC",
                        "False",
                    ],
                    [],
                    ["EXPERIMENT_TYPE"],  # Second entity block in the same sheet
                    [
                        "Code",
                        "Description",
                        "Validation script",
                        "Generated code prefix",
                        "Auto generate codes",
                    ],
                    [
                        "COLL_002",
                        "Another Collection",
                        None,
                        "COLL_DEF",
                        "True",
                    ],
                ],
            },
            {
                "object_types": {
                    "SAMPLE_001": {
                        "code": "SAMPLE_001",
                        "description": "Sample Description",
                        "validationPlugin": "val.py",
                        "generatedCodePrefix": "SAM_001",
                        "autoGeneratedCode": True,
                        "properties": {},
                    },
                    "SAMPLE_002": {
                        "code": "SAMPLE_002",
                        "description": "Another Sample",
                        "validationPlugin": None,
                        "generatedCodePrefix": "SAM_002",
                        "autoGeneratedCode": False,
                        "properties": {},
                    },
                },
                "collection_types": {
                    "COLL_001": {
                        "code": "COLL_001",
                        "description": "Collection Example",
                        "validationPlugin": "val_coll.py",
                        "properties": {},
                    },
                    "COLL_002": {
                        "code": "COLL_002",
                        "description": "Another Collection",
                        "validationPlugin": None,
                        "properties": {},
                    },
                },
            },
        ),
        # Case 4: Sheet with two consecutive empty rows (end detection, so second sample not readed)
        (
            {
                "Object Types": [
                    ["SAMPLE_TYPE"],
                    [
                        "Code",
                        "Description",
                        "Validation script",
                        "Generated code prefix",
                        "Auto generate codes",
                    ],
                    [
                        "SAMPLE_001",
                        "Sample Description",
                        "val.py",
                        "SAM_001",
                        "True",
                    ],
                    [],
                    [],
                    [
                        "SAMPLE_TYPE"
                    ],  # Second entity block in the same sheet that should be ignored
                    [
                        "Code",
                        "Description",
                        "Validation script",
                        "Generated code prefix",
                        "Auto generate codes",
                    ],
                    [
                        "SAMPLE_002",
                        "Another Sample",
                        None,
                        "SAM_002",
                        "False",
                    ],
                ],
            },
            {
                "object_types": {
                    "SAMPLE_001": {
                        "code": "SAMPLE_001",
                        "description": "Sample Description",
                        "validationPlugin": "val.py",
                        "generatedCodePrefix": "SAM_001",
                        "autoGeneratedCode": True,
                        "properties": {},
                    }
                }
            },
        ),
        # Case 5: Fully empty sheet
        (
            {
                "EmptySheet": [
                    ["", ""],  # Empty row
                    ["", ""],  # Empty row
                ]
            },
            {},  # Expected empty dictionary since no entities exist
        ),
    ],
)
def test_excel_to_entities(excel_extractor, sheets_data, expected_result):
    """Tests `excel_to_entities` function for extracting entities from Excel sheets."""

    # Create a dummy workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Populate workbook with test data
    for sheet_name, rows in sheets_data.items():
        sheet = wb.create_sheet(title=sheet_name)
        for row_index, row_values in enumerate(rows, start=1):
            for col_index, value in enumerate(row_values, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)

    # Assign workbook to extractor
    excel_extractor.workbook = wb

    # Call function
    result = excel_extractor.excel_to_entities()

    # Assert the extracted entities match the expected structure
    assert result == expected_result, f"Expected: {expected_result}, but got: {result}"
