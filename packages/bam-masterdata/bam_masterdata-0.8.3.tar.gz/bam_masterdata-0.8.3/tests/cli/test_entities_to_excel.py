import os

import pytest
from openpyxl import Workbook

from bam_masterdata.cli.entities_to_excel import entities_to_excel
from bam_masterdata.utils import import_module


@pytest.mark.parametrize(
    "module_name, entity_name, entity_headers",
    [
        (
            "collection_types",
            "EXPERIMENT_TYPE",
            ["Code", "Description", "Validation script"],
        ),
        # ('dataset_types', 'DATASET_TYPE', ['Code', 'Description', 'Validation script']),  # ! this module does not have classes yet
        (
            "object_types",
            "SAMPLE_TYPE",
            [
                "Code",
                "Description",
                "Validation script",
                "Generated code prefix",
                "Auto generate codes",
            ],
        ),
        (
            "vocabulary_types",
            "VOCABULARY_TYPE",
            ["Code", "Description", "Url template"],
        ),
    ],
)
def test_entities_to_excel(
    module_name: str, entity_name: str, entity_headers: list[str]
):
    """Test the `entities_to_excel` function."""
    definitions_module = import_module(
        module_path="./bam_masterdata/metadata/definitions.py"
    )

    # Get the Python modules to process the datamodel
    module_path = os.path.join("./bam_masterdata/datamodel", f"{module_name}.py")
    wb = Workbook()
    ws = wb.active
    ws.title = (
        os.path.basename(module_path).capitalize().replace(".py", "").replace("_", " ")
    )

    entities_to_excel(
        worksheet=ws,
        module_path=module_path,
        definitions_module=definitions_module,
    )

    assert len(wb.worksheets) == 1
    # Header for type of entity
    assert ws.cell(row=1, column=1).value == entity_name
    # Headers for definitions
    col = []
    for c in range(1, 101):
        if not ws.cell(row=2, column=c).value:
            break
        col.append(ws.cell(row=2, column=c).value)
    assert col == entity_headers
