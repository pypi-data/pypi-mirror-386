# -*- coding: UTF-8 -*-
from urllib.parse import urlparse

import openpyxl
import requests
import urllib3
from openpyxl.utils import get_column_letter

from kyutil.config import BUILD_PATH_LOGGER_FILE
from kyutil.log import zero_log

urllib3.disable_warnings()
logger = zero_log(__file__, BUILD_PATH_LOGGER_FILE)


def send_request(url, **kwargs):
    method = str(kwargs["method"]).lower() if "method" in kwargs.keys() else "get"
    data = kwargs["data"] if "data" in kwargs.keys() else None
    headers = kwargs["headers"] if "headers" in kwargs.keys() else None
    cookies = kwargs["cookies"] if "cookies" in kwargs.keys() else None
    proxies = kwargs["proxies"] if "proxies" in kwargs.keys() else None
    timeout = kwargs["timeout"] if "timeout" in kwargs.keys() else None
    verify = kwargs["verify"] if "verify" in kwargs.keys() else False
    json_ = kwargs["json"] if "json" in kwargs.keys() else None
    allow_redirects = kwargs["allow_redirects"] if "allow_redirects" in kwargs.keys() else True
    if method.lower() not in ["get", "post", "head", "delete", "put", "options", 'patch']:
        return
    func = getattr(requests, method.lower())
    resp = func(url, data=data, headers=headers, cookies=cookies, proxies=proxies,
                timeout=timeout, verify=verify, allow_redirects=allow_redirects, json=json_)
    return resp


def is_url(url):
    try:
        r = urlparse(url)
        return all([r.scheme, r.netloc])
    except ValueError:
        return


def is_merged_cell(cell, sheet):
    """
    检查单元格是否是合并单元格的一部分
    :param cell: 要检查的单元格
    :param sheet: 工作表对象
    :return: 如果是合并单元格的一部分，返回 True，否则返回 False
    """
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return True
    return False


def is_top_left_merged_cell(cell, merged_range):
    """
    检查单元格是否是合并单元格的左上角单元格
    :param cell: 要检查的单元格
    :param merged_range: 合并单元格范围对象
    :return: 如果是合并单元格的左上角单元格，返回 True，否则返回 False
    """
    return cell.coordinate == str(merged_range.start_cell.column_letter) + str(merged_range.min_row)


def get_merged_cell_info(merged_range):
    """
    获取合并单元格的 rowspan 和 colspan
    :param merged_range: 合并单元格范围对象
    :return: 包含 rowspan 和 colspan 的元组
    """
    rowspan = merged_range.max_row - merged_range.min_row + 1
    colspan = merged_range.max_col - merged_range.min_col + 1
    return rowspan, colspan


def process_merged_cell(cell, merged_range, html, merged_cells):
    """
    处理合并单元格
    :param cell: 合并单元格的左上角单元格
    :param merged_range: 合并单元格范围对象
    :param html: 当前的 HTML 字符串
    :param merged_cells: 已处理的合并单元格列表
    :param get_column_letter: 获取列字母的函数
    :return: 更新后的 HTML 字符串
    """
    rowspan, colspan = get_merged_cell_info(merged_range)
    value = cell.value if cell.value is not None else ""
    html += f'<td rowspan="{rowspan}" colspan="{colspan}">{value}</td>'
    merged_cells.extend([f'{get_column_letter(col)}{row}' for row in
                         range(merged_range.min_row, merged_range.max_row + 1) for col in
                         range(merged_range.min_col, merged_range.max_col + 1)])
    return html


def process_normal_cell(cell, html):
    """
    处理普通单元格
    :param cell: 普通单元格
    :param html: 当前的 HTML 字符串
    :return: 更新后的 HTML 字符串
    """
    value = cell.value if cell.value is not None else ""
    html += f'<td>{value}</td>'
    return html


def deal_rows(row, html, sheet, merged_cells):
    """
    处理每行单元格
    """
    for cell in row:
        if is_merged_cell(cell, sheet):
            for merged_range in sheet.merged_cells.ranges:
                if is_top_left_merged_cell(cell, merged_range):
                    html = process_merged_cell(cell, merged_range, html, merged_cells)
                    break
        elif cell.coordinate not in merged_cells:
            html = process_normal_cell(cell, html)
    return html


def excel_to_html(file_path):
    # 加载 Excel 文件
    workbook = openpyxl.load_workbook(file_path)
    # 获取第一个工作表
    sheet = workbook.active

    # 开始构建 HTML 表格
    html = '<table border="2">'

    # 用于跟踪已经处理过的合并单元格
    merged_cells = []

    # 遍历工作表的每一行
    for row in sheet.iter_rows():
        html += '<tr>'
        html = deal_rows(row, html, sheet, merged_cells)
        html += '</tr>'

    html += '</table>'
    return html


def result_post_server(_url, _data, logger_=logger):
    """回传数据至服务端，10次连接尝试"""
    for _ in range(10):
        try:
            r = send_request(url=str(_url), method="POST", data=_data, verify=False, timeout=600)
            logger_.info(f"ISO回传成功，状态：{r.status_code}，数据：{_data}")
            return True
        except requests.exceptions.ConnectTimeout:
            logger_.info(f"ISO回传失败，数据  ===== ：{_data}")
            return False
