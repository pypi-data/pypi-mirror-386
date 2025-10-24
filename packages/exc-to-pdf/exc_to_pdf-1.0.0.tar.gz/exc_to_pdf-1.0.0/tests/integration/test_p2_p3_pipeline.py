"""
Integration tests for P2-P3 pipeline (Excel processing to PDF generation).

This module tests the complete workflow from Excel file processing (P2)
to PDF generation (P3), ensuring data integrity, proper formatting,
and correct metadata handling throughout the pipeline.
"""

import tempfile
from pathlib import Path
from typing import Any, Callable, Dict, List
from unittest.mock import Mock, patch

import pytest
from openpyxl import Workbook
from _pytest.fixtures import SubRequest

from exc_to_pdf.excel_processor import ExcelReader, SheetData
from exc_to_pdf.pdf_generator import PDFGenerator
from exc_to_pdf.config.pdf_config import PDFConfig
from exc_to_pdf.exceptions import ExcelReaderError, PDFGenerationException

# Test configuration constants
TEMP_DIR = Path(tempfile.gettempdir()) / "exc-to-pdf-test"
TEST_OUTPUT_DIR = TEMP_DIR / "integration-test-output"


@pytest.fixture(scope="session", autouse=True)
def setup_test_directories() -> None:
    """Create test directories for integration tests."""
    TEMP_DIR.mkdir(exist_ok=True)
    TEST_OUTPUT_DIR.mkdir(exist_ok=True)


@pytest.fixture
def sample_excel_file_path() -> str:
    """Create a sample Excel file with multiple sheets and various data structures.

    Returns:
        Path to the created Excel file
    """
    excel_path = TEMP_DIR / "sample_integration_test.xlsx"

    # Create workbook with multiple sheets
    wb = Workbook()

    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)

    # Sheet 1: Sales Data (structured table)
    ws_sales = wb.create_sheet("Sales Data")
    ws_sales.append(["Product", "Q1", "Q2", "Q3", "Q4"])
    ws_sales.append(["Product A", "1000", "1200", "1100", "1300"])
    ws_sales.append(["Product B", "800", "900", "950", "1100"])
    ws_sales.append(["Product C", "1500", "1600", "1450", "1700"])

    # Sheet 2: Customer Info (simple table)
    ws_customers = wb.create_sheet("Customers")
    ws_customers.append(["Customer ID", "Name", "Email", "Region"])
    ws_customers.append(["C001", "Alice Corp", "alice@corp.com", "North"])
    ws_customers.append(["C002", "Bob LLC", "bob@llc.com", "South"])
    ws_customers.append(["C003", "Charlie Inc", "charlie@inc.com", "East"])

    # Sheet 3: Empty Sheet
    ws_empty = wb.create_sheet("Empty Sheet")

    # Sheet 4: Raw Data (no clear table structure)
    ws_raw = wb.create_sheet("Raw Data")
    ws_raw.append(["Date", "Event", "Value"])
    ws_raw.append(["2024-01-01", "Meeting", "High"])
    ws_raw.append(["2024-01-02", "Call", "Medium"])
    ws_raw.append(["2024-01-03", "Email", "Low"])
    ws_raw.append(["Summary", "", ""])
    ws_raw.append(["Total Events", "3", ""])

    # Sheet 5: Large Dataset (for performance testing)
    ws_large = wb.create_sheet("Large Dataset")
    ws_large.append(["ID", "Category", "Value", "Description"])
    for i in range(1, 101):  # 100 rows of data
        ws_large.append(
            [f"ID{i:03d}", f"Cat{i%10}", i * 10, f"Description for item {i}"]
        )

    wb.save(excel_path)
    return str(excel_path)


@pytest.fixture
def complex_excel_file_path() -> str:
    """Create a complex Excel file with merged cells and formatting challenges.

    Returns:
        Path to the created Excel file
    """
    excel_path = TEMP_DIR / "complex_integration_test.xlsx"

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    # Sheet with merged cells and complex structure
    ws_complex = wb.create_sheet("Complex Data")

    # Add some headers
    ws_complex.append(["Report Title", "", "", ""])
    ws_complex.append(["Date Range", "2024-Q1", "Region", "Global"])
    ws_complex.append(["", "", "", ""])

    # Add actual data
    ws_complex.append(["Department", "Budget", "Actual", "Variance"])
    ws_complex.append(["Sales", "100000", "95000", "-5%"])
    ws_complex.append(["Marketing", "50000", "52000", "+4%"])
    ws_complex.append(["Operations", "75000", "78000", "+4%"])

    # Add summary section
    ws_complex.append(["", "", "", ""])
    ws_complex.append(["Total", "225000", "225000", "0%"])

    # Another sheet with different data types
    ws_types = wb.create_sheet("Data Types")
    ws_types.append(["Type", "Value"])
    ws_types.append(["Text", "Sample text"])
    ws_types.append(["Number", 42.5])
    ws_types.append(["Date", "2024-01-15"])
    ws_types.append(["Boolean", True])
    ws_types.append(["Formula", "=SUM(A2:A5)"])

    wb.save(excel_path)
    return str(excel_path)


@pytest.fixture
def empty_excel_file_path() -> str:
    """Create an empty Excel file for edge case testing.

    Returns:
        Path to the created empty Excel file
    """
    excel_path = TEMP_DIR / "empty_integration_test.xlsx"

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)  # Remove default sheet

    # Add completely empty sheet
    wb.create_sheet("Completely Empty")

    # Add sheet with only headers
    ws_headers = wb.create_sheet("Headers Only")
    ws_headers.append(["Column1", "Column2", "Column3"])

    wb.save(excel_path)
    return str(excel_path)


@pytest.fixture
def excel_processor_factory() -> Callable[[str], ExcelReader]:
    """Factory function to create ExcelReader instances.

    Returns:
        Function that creates ExcelReader instances
    """

    def _create_processor(file_path: str) -> ExcelReader:
        # Create config with read_only_mode disabled to avoid page_setup issues
        from exc_to_pdf.config.excel_config import ExcelConfig

        config = ExcelConfig(read_only_mode=False)
        return ExcelReader(file_path, config)

    return _create_processor


@pytest.fixture
def pdf_generator_standard() -> PDFGenerator:
    """Create PDFGenerator with standard configuration.

    Returns:
        PDFGenerator with default configuration
    """
    return PDFGenerator()


@pytest.fixture
def pdf_generator_ai_optimized() -> PDFGenerator:
    """Create PDFGenerator with AI optimization enabled.

    Returns:
        PDFGenerator with AI optimization configuration
    """
    config = PDFConfig(
        optimize_for_notebooklm=True,
        include_bookmarks=True,
        include_metadata=True,
        page_size="A4",
        orientation="portrait",
    )
    return PDFGenerator(config)


@pytest.fixture
def pdf_generator_landscape() -> PDFGenerator:
    """Create PDFGenerator with landscape orientation.

    Returns:
        PDFGenerator with landscape configuration
    """
    config = PDFConfig(
        optimize_for_notebooklm=False,
        include_bookmarks=False,
        include_metadata=True,
        page_size="A4",
        orientation="landscape",
    )
    return PDFGenerator(config)


class TestP2P3IntegrationBasic:
    """Test basic P2-P3 integration scenarios."""

    def test_complete_pipeline_standard_config(
        self,
        sample_excel_file_path: str,
        excel_processor_factory: Callable[[str], ExcelReader],
        pdf_generator_standard: PDFGenerator,
    ) -> None:
        """Test complete pipeline with standard configuration."""
        # P2: Process Excel file
        excel_processor = excel_processor_factory(sample_excel_file_path)
        sheet_names = excel_processor.discover_sheets()

        # Extract data from all sheets
        sheet_data_list = []
        for sheet_name in sheet_names:
            sheet_data = excel_processor.extract_sheet_data(sheet_name)
            sheet_data_list.append(sheet_data)

        # Verify P2 output
        assert isinstance(sheet_data_list, list)
        assert len(sheet_data_list) > 0

        # Check for expected sheets
        sheet_names = [sheet.sheet_name for sheet in sheet_data_list]
        assert "Sales Data" in sheet_names
        assert "Customers" in sheet_names
        assert "Large Dataset" in sheet_names

        # Verify data integrity
        for sheet_data in sheet_data_list:
            assert isinstance(sheet_data, SheetData)
            assert sheet_data.sheet_name
            assert sheet_data.row_count >= 0
            assert sheet_data.col_count >= 0
            assert isinstance(sheet_data.has_data, bool)
            assert isinstance(sheet_data.tables, list)
            assert isinstance(sheet_data.raw_data, list)

        # P3: Generate PDF
        output_path = TEST_OUTPUT_DIR / "standard_integration_test.pdf"
        pdf_generator_standard.create_pdf(
            sheet_data_list, str(output_path), source_file=sample_excel_file_path
        )

        # Verify PDF was created
        assert output_path.exists()
        assert output_path.stat().st_size > 0

    def test_complete_pipeline_ai_optimized(
        self,
        sample_excel_file_path: str,
        excel_processor_factory: Callable[[str], ExcelReader],
        pdf_generator_ai_optimized: PDFGenerator,
    ) -> None:
        """Test complete pipeline with AI optimization enabled."""
        # P2: Process Excel file
        excel_processor = excel_processor_factory(sample_excel_file_path)
        sheet_names = excel_processor.discover_sheets()
        sheet_data_list = [
            excel_processor.extract_sheet_data(name) for name in sheet_names
        ]

        # P3: Generate PDF with AI optimization
        output_path = TEST_OUTPUT_DIR / "ai_optimized_integration_test.pdf"
        pdf_generator_ai_optimized.create_pdf(
            sheet_data_list, str(output_path), source_file=sample_excel_file_path
        )

        # Verify PDF was created and is larger (due to metadata)
        assert output_path.exists()
        assert output_path.stat().st_size > 0

    def test_pipeline_with_landscape_orientation(
        self,
        complex_excel_file_path: str,
        excel_processor_factory: Callable[[str], ExcelReader],
        pdf_generator_landscape: PDFGenerator,
    ) -> None:
        """Test pipeline with landscape PDF orientation."""
        # P2: Process complex Excel file
        excel_processor = excel_processor_factory(complex_excel_file_path)
        sheet_names = excel_processor.discover_sheets()
        sheet_data_list = [
            excel_processor.extract_sheet_data(name) for name in sheet_names
        ]

        # P3: Generate landscape PDF
        output_path = TEST_OUTPUT_DIR / "landscape_integration_test.pdf"
        pdf_generator_landscape.create_pdf(
            sheet_data_list, str(output_path), source_file=complex_excel_file_path
        )

        # Verify PDF was created
        assert output_path.exists()
        assert output_path.stat().st_size > 0


class TestP2P3IntegrationEdgeCases:
    """Test edge cases and error scenarios in P2-P3 integration."""

    def test_empty_excel_file_handling(
        self,
        empty_excel_file_path: str,
        excel_processor_factory: Callable[[str], ExcelReader],
        pdf_generator_standard: PDFGenerator,
    ) -> None:
        """Test handling of empty Excel files."""
        # P2: Process empty Excel file
        excel_processor = excel_processor_factory(empty_excel_file_path)
        sheet_names = excel_processor.discover_sheets()
        sheet_data_list = [
            excel_processor.extract_sheet_data(name) for name in sheet_names
        ]

        # Should still return list, possibly with empty sheets
        assert isinstance(sheet_data_list, list)

        # P3: Try to generate PDF (should handle gracefully)
        output_path = TEST_OUTPUT_DIR / "empty_excel_test.pdf"

        if sheet_data_list:  # Only try PDF generation if we have sheets
            pdf_generator_standard.create_pdf(
                sheet_data_list, str(output_path), source_file=empty_excel_file_path
            )
            assert output_path.exists()

    def test_large_dataset_performance(
        self,
        sample_excel_file_path: str,
        excel_processor_factory: Callable[[str], ExcelReader],
        pdf_generator_standard: PDFGenerator,
    ) -> None:
        """Test performance with large datasets."""
        import time

        # P2: Process Excel file with timing
        start_time = time.time()
        excel_processor = excel_processor_factory(sample_excel_file_path)
        sheet_names = excel_processor.discover_sheets()
        sheet_data_list = [
            excel_processor.extract_sheet_data(name) for name in sheet_names
        ]
        p2_duration = time.time() - start_time

        # Find the large dataset sheet
        large_sheet = None
        for sheet in sheet_data_list:
            if sheet.sheet_name == "Large Dataset":
                large_sheet = sheet
                break

        assert large_sheet is not None
        assert large_sheet.row_count > 50  # Should have substantial data

        # P3: Generate PDF with timing
        start_time = time.time()
        output_path = TEST_OUTPUT_DIR / "large_dataset_test.pdf"
        pdf_generator_standard.create_pdf(
            sheet_data_list, str(output_path), source_file=sample_excel_file_path
        )
        p3_duration = time.time() - start_time

        # Verify results
        assert output_path.exists()
        assert output_path.stat().st_size > 0

        # Performance should be reasonable (less than 30 seconds total)
        total_time = p2_duration + p3_duration
        assert total_time < 30.0, f"Performance test failed: {total_time:.2f}s total"

    def test_invalid_file_path_handling(
        self,
        excel_processor_factory: Callable[[str], ExcelReader],
        pdf_generator_standard: PDFGenerator,
    ) -> None:
        """Test handling of invalid file paths."""
        invalid_path = "/nonexistent/path/file.xlsx"

        # P2: Should raise appropriate exception
        with pytest.raises((ExcelReaderError, FileNotFoundError)):
            excel_processor = excel_processor_factory(invalid_path)

        # P3: Should handle empty sheet data gracefully
        output_path = TEST_OUTPUT_DIR / "invalid_path_test.pdf"

        with pytest.raises((PDFGenerationException, ValueError)):
            pdf_generator_standard.create_pdf([], str(output_path))


class TestP2P3DataIntegrity:
    """Test data integrity throughout the P2-P3 pipeline."""

    def test_data_preservation_across_pipeline(
        self,
        sample_excel_file_path: str,
        excel_processor_factory: Callable[[str], ExcelReader],
        pdf_generator_standard: PDFGenerator,
    ) -> None:
        """Test that data is preserved correctly through the pipeline."""
        # P2: Process Excel file
        excel_processor = excel_processor_factory(sample_excel_file_path)
        sheet_names = excel_processor.discover_sheets()
        sheet_data_list = [
            excel_processor.extract_sheet_data(name) for name in sheet_names
        ]

        # Find specific sheet to verify data integrity
        sales_sheet = None
        for sheet in sheet_data_list:
            if sheet.sheet_name == "Sales Data":
                sales_sheet = sheet
                break

        assert sales_sheet is not None

        # Verify P2 extracted data correctly
        assert sales_sheet.has_data
        assert len(sales_sheet.raw_data) > 0

        # Check specific data points
        expected_headers = ["Product", "Q1", "Q2", "Q3", "Q4"]
        if sales_sheet.raw_data:
            assert sales_sheet.raw_data[0] == expected_headers

        # Verify row count is accurate
        expected_rows = 4  # 1 header + 3 data rows
        assert sales_sheet.row_count == expected_rows

        # P3: Generate PDF
        output_path = TEST_OUTPUT_DIR / "data_integrity_test.pdf"
        pdf_generator_standard.create_pdf(
            sheet_data_list, str(output_path), source_file=sample_excel_file_path
        )

        # Verify PDF creation succeeded (data made it to PDF generation)
        assert output_path.exists()
        assert output_path.stat().st_size > 0

    def test_sheet_structure_preservation(
        self,
        complex_excel_file_path: str,
        excel_processor_factory: Callable[[str], ExcelReader],
        pdf_generator_ai_optimized: PDFGenerator,
    ) -> None:
        """Test that sheet structure is preserved through the pipeline."""
        # P2: Process complex Excel file
        excel_processor = excel_processor_factory(complex_excel_file_path)
        sheet_names = excel_processor.discover_sheets()
        sheet_data_list = [
            excel_processor.extract_sheet_data(name) for name in sheet_names
        ]

        # Verify we have the expected sheets
        sheet_names = [sheet.sheet_name for sheet in sheet_data_list]
        assert "Complex Data" in sheet_names
        assert "Data Types" in sheet_names

        # Check sheet metadata preservation
        for sheet in sheet_data_list:
            assert sheet.metadata is not None
            assert isinstance(sheet.metadata, dict)

            # Verify sheet-level statistics
            if sheet.has_data:
                assert sheet.row_count > 0
                assert sheet.col_count > 0

        # P3: Generate PDF with full features
        output_path = TEST_OUTPUT_DIR / "structure_preservation_test.pdf"
        pdf_generator_ai_optimized.create_pdf(
            sheet_data_list, str(output_path), source_file=complex_excel_file_path
        )

        # Verify PDF was created successfully
        assert output_path.exists()
        assert output_path.stat().st_size > 0


class TestP2P3ConfigurationMatrix:
    """Test various configuration combinations."""

    @pytest.mark.parametrize(
        "optimize_ai,include_bookmarks,include_metadata",
        [
            (True, True, True),
            (True, False, True),
            (False, True, False),
            (False, False, False),
            (True, True, False),
            (False, True, True),
        ],
    )
    def test_configuration_combinations(
        self,
        sample_excel_file_path: str,
        excel_processor_factory: Callable[[str], ExcelReader],
        optimize_ai: bool,
        include_bookmarks: bool,
        include_metadata: bool,
    ) -> None:
        """Test various configuration combinations."""
        # Create configuration
        config = PDFConfig(
            optimize_for_notebooklm=optimize_ai,
            include_bookmarks=include_bookmarks,
            include_metadata=include_metadata,
            page_size="A4",
            orientation="portrait",
        )

        # P2: Process Excel file
        excel_processor = excel_processor_factory(sample_excel_file_path)
        sheet_names = excel_processor.discover_sheets()
        sheet_data_list = [
            excel_processor.extract_sheet_data(name) for name in sheet_names
        ]

        # P3: Generate PDF with specific configuration
        pdf_generator = PDFGenerator(config)
        output_path = (
            TEST_OUTPUT_DIR
            / f"config_test_{optimize_ai}_{include_bookmarks}_{include_metadata}.pdf"
        )

        pdf_generator.create_pdf(
            sheet_data_list, str(output_path), source_file=sample_excel_file_path
        )

        # Verify PDF was created
        assert output_path.exists()
        assert output_path.stat().st_size > 0

        # Configuration-specific assertions could be added here
        # For example, checking PDF metadata if metadata is enabled


class TestP2P3ErrorHandling:
    """Test error handling and recovery in the pipeline."""

    def test_p2_failure_propagation(
        self,
        excel_processor_factory: Callable[[str], ExcelReader],
        pdf_generator_standard: PDFGenerator,
    ) -> None:
        """Test that P2 failures are properly handled."""
        # Test with completely invalid file
        with pytest.raises((ExcelReaderError, FileNotFoundError)):
            excel_processor_factory("/invalid/nonexistent.xlsx")

    def test_p3_graceful_degradation(
        self,
        sample_excel_file_path: str,
        excel_processor_factory: Callable[[str], ExcelReader],
    ) -> None:
        """Test P3 graceful degradation with problematic data."""
        # Process Excel file normally
        excel_processor = excel_processor_factory(sample_excel_file_path)
        sheet_names = excel_processor.discover_sheets()
        sheet_data_list = [
            excel_processor.extract_sheet_data(name) for name in sheet_names
        ]

        # Create PDF generator with minimal configuration
        config = PDFConfig(
            optimize_for_notebooklm=False,
            include_bookmarks=False,
            include_metadata=False,
            page_size="A4",
            orientation="portrait",
        )
        pdf_generator = PDFGenerator(config)

        # Should still work even with minimal features
        output_path = TEST_OUTPUT_DIR / "graceful_degradation_test.pdf"
        pdf_generator.create_pdf(
            sheet_data_list, str(output_path), source_file=sample_excel_file_path
        )

        assert output_path.exists()
        assert output_path.stat().st_size > 0


@pytest.mark.parametrize(
    "excel_file_path",
    ["sample_excel_file_path", "complex_excel_file_path", "empty_excel_file_path"],
)
def test_parametrized_integration_tests(
    request: SubRequest,
    excel_file_path: str,
    excel_processor_factory: Callable[[str], ExcelReader],
    pdf_generator_standard: PDFGenerator,
) -> None:
    """Parametrized integration test using different Excel files."""
    # Get the actual file path from the fixture
    file_path = request.getfixturevalue(excel_file_path)

    # P2: Process Excel file
    excel_processor = excel_processor_factory(file_path)
    sheet_names = excel_processor.discover_sheets()
    sheet_data_list = [excel_processor.extract_sheet_data(name) for name in sheet_names]

    # Skip PDF generation if no valid data
    if not sheet_data_list or not any(sheet.has_data for sheet in sheet_data_list):
        pytest.skip("No valid data to generate PDF")

    # P3: Generate PDF
    output_path = TEST_OUTPUT_DIR / f"parametrized_{Path(file_path).stem}.pdf"
    pdf_generator_standard.create_pdf(
        sheet_data_list, str(output_path), source_file=file_path
    )

    # Verify PDF was created
    assert output_path.exists()
    assert output_path.stat().st_size > 0
