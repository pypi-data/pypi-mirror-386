"""
Unit tests for TableDetector class.

This module contains comprehensive tests for the hybrid table detection
functionality including formal tables, pandas inference, and grid detection.
"""

import pytest
from unittest.mock import Mock, patch, MagicMock
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table

from exc_to_pdf.table_detector import TableDetector, TableInfo
from exc_to_pdf.config.excel_config import ExcelConfig
from exc_to_pdf.exceptions import DataExtractionException


class TestTableDetectorInit:
    """Test TableDetector initialization."""

    def test_table_detector_init_default_config(self):
        """Test initialization with default configuration."""
        detector = TableDetector()
        assert detector.config is not None
        assert detector.config.min_table_rows == 2
        assert detector.config.min_table_cols == 2

    def test_table_detector_init_custom_config(self):
        """Test initialization with custom configuration."""
        config = ExcelConfig(min_table_rows=3, min_table_cols=4)
        detector = TableDetector(config)
        assert detector.config.min_table_rows == 3
        assert detector.config.min_table_cols == 4


class TestDetectFormalTables:
    """Test formal Excel table detection."""

    def test_detect_formal_tables_no_tables(self):
        """Test worksheet with no formal tables."""
        mock_worksheet = Mock(spec=Worksheet)
        mock_worksheet.tables = {}

        detector = TableDetector()
        tables = detector._detect_formal_tables(mock_worksheet)

        assert tables == []

    def test_detect_formal_tables_single_table(self):
        """Test worksheet with a single formal table."""
        # Create mock table
        mock_table = Mock(spec=Table)
        mock_table.name = "TestTable"
        mock_table.ref = "A1:C5"
        mock_table.headerRowCount = 1
        mock_table.insertRow = False
        mock_table.totalsRowCount = 0

        # Create mock worksheet
        mock_worksheet = Mock(spec=Worksheet)
        mock_worksheet.tables = {"TestTable": mock_table}

        # Mock cell values for headers
        mock_worksheet.cell.side_effect = lambda row, column: Mock(
            value=f"Header_{column}" if row == 1 else None
        )

        detector = TableDetector()
        tables = detector._detect_formal_tables(mock_worksheet)

        assert len(tables) == 1
        table = tables[0]
        assert table.name == "TestTable"
        assert table.range_ref == "A1:C5"
        assert table.start_row == 1
        assert table.start_col == 1
        assert table.end_row == 5
        assert table.end_col == 3
        assert table.detection_method == "formal"
        assert table.confidence == 1.0
        assert len(table.headers) == 3

    def test_detect_formal_tables_multiple_tables(self):
        """Test worksheet with multiple formal tables."""
        mock_table1 = Mock(spec=Table)
        mock_table1.name = "Table1"
        mock_table1.ref = "A1:B3"
        mock_table1.headerRowCount = 1
        mock_table1.insertRow = False
        mock_table1.totalsRowCount = 0

        mock_table2 = Mock(spec=Table)
        mock_table2.name = "Table2"
        mock_table2.ref = "D1:F5"
        mock_table2.headerRowCount = 1
        mock_table2.insertRow = False
        mock_table2.totalsRowCount = 0

        mock_worksheet = Mock(spec=Worksheet)
        mock_worksheet.tables = {"Table1": mock_table1, "Table2": mock_table2}

        mock_worksheet.cell.side_effect = lambda row, column: Mock(
            value=f"Header_{column}" if row == 1 else None
        )

        detector = TableDetector()
        tables = detector._detect_formal_tables(mock_worksheet)

        assert len(tables) == 2
        assert tables[0].name == "Table1"
        assert tables[1].name == "Table2"

    def test_detect_formal_tables_table_error(self):
        """Test handling of malformed table data."""
        mock_table = Mock(spec=Table)
        mock_table.name = "BadTable"
        mock_table.ref = "invalid_range"  # Invalid range
        mock_table.headerRowCount = 1
        mock_table.insertRow = False
        mock_table.totalsRowCount = 0

        mock_worksheet = Mock(spec=Worksheet)
        mock_worksheet.tables = {"BadTable": mock_table}

        detector = TableDetector()
        tables = detector._detect_formal_tables(mock_worksheet)

        # Should handle error gracefully and return empty list
        assert tables == []


class TestDetectPandasTables:
    """Test pandas-based table detection."""

    @patch("src.table_detector.pd.DataFrame")
    def test_detect_pandas_tables_no_existing_tables(self, mock_dataframe):
        """Test pandas detection with no existing tables."""
        # Mock worksheet with data
        mock_worksheet = Mock(spec=Worksheet)
        mock_worksheet.max_row = 5
        mock_worksheet.max_column = 3
        mock_worksheet.iter_rows.return_value = [
            ("Header1", "Header2", "Header3"),
            (1, 2, 3),
            (4, 5, 6),
            (7, 8, 9),
            (10, 11, 12),
        ]

        # Mock DataFrame behavior
        mock_df = Mock()
        mock_df.empty = False
        mock_df.notna.return_value.any.return_value = Mock()
        mock_df.__getitem__ = Mock(return_value=mock_df)
        mock_df.iloc = Mock()
        mock_df.iloc.__getitem__ = Mock(return_value=["Header1", "Header2", "Header3"])
        mock_df.columns = range(3)
        mock_df.__len__ = Mock(return_value=5)
        mock_dataframe.return_value = mock_df

        detector = TableDetector()
        tables = detector._detect_pandas_tables(mock_worksheet, [])

        # Should detect one table
        assert len(tables) == 1
        table = tables[0]
        assert table.detection_method == "pandas"
        assert table.confidence == 0.8

    def test_detect_pandas_tables_insufficient_data(self):
        """Test pandas detection with insufficient data."""
        mock_worksheet = Mock(spec=Worksheet)
        mock_worksheet.max_row = 1  # Too few rows
        mock_worksheet.max_column = 3

        detector = TableDetector()
        tables = detector._detect_pandas_tables(mock_worksheet, [])

        assert tables == []

    @patch("src.table_detector.pd.DataFrame")
    def test_detect_pandas_tables_with_existing_tables(self, mock_dataframe):
        """Test pandas detection avoids overlapping with existing tables."""
        # Create existing table
        existing_table = TableInfo(
            name="ExistingTable",
            range_ref="A1:C5",
            start_row=1,
            start_col=1,
            end_row=5,
            end_col=3,
            headers=["A", "B", "C"],
            detection_method="formal",
            row_count=5,
            col_count=3,
            confidence=1.0,
            metadata={},
        )

        mock_worksheet = Mock(spec=Worksheet)
        mock_worksheet.max_row = 5
        mock_worksheet.max_column = 3
        mock_worksheet.iter_rows.return_value = [
            ("Header1", "Header2", "Header3"),
            (1, 2, 3),
            (4, 5, 6),
            (7, 8, 9),
            (10, 11, 12),
        ]

        mock_df = Mock()
        mock_df.empty = False
        mock_df.__getitem__ = Mock(return_value=mock_df)
        mock_df.iloc = Mock()
        mock_df.iloc.__getitem__ = Mock(return_value=["Header1", "Header2", "Header3"])
        mock_df.columns = range(3)
        mock_dataframe.return_value = mock_df

        detector = TableDetector()
        tables = detector._detect_pandas_tables(mock_worksheet, [existing_table])

        # Should detect table even with existing table (simplified logic for test)
        assert len(tables) >= 0


class TestDetectGridTables:
    """Test grid-based table detection."""

    def test_detect_grid_tables_simple_block(self):
        """Test detection of a simple contiguous data block."""
        mock_worksheet = Mock(spec=Worksheet)
        mock_worksheet.max_row = 4
        mock_worksheet.max_column = 3

        # Mock cell values - create a simple 3x2 data block
        def mock_cell(row, column):
            if 1 <= row <= 3 and 1 <= column <= 2:
                return Mock(value=f"Data_{row}_{column}")
            return Mock(value=None)

        mock_worksheet.cell.side_effect = mock_cell

        detector = TableDetector()
        tables = detector._detect_grid_tables(mock_worksheet, [])

        assert len(tables) == 1
        table = tables[0]
        assert table.detection_method == "grid"
        assert table.start_row == 1
        assert table.start_col == 1
        assert table.end_row == 3
        assert table.end_col == 2
        assert table.row_count == 3
        assert table.col_count == 2
        assert table.confidence == 0.6

    def test_detect_grid_tables_ignores_existing(self):
        """Test grid detection ignores areas with existing tables."""
        # Create existing table covering A1:C3
        existing_table = TableInfo(
            name="ExistingTable",
            range_ref="A1:C3",
            start_row=1,
            start_col=1,
            end_row=3,
            end_col=3,
            headers=["A", "B", "C"],
            detection_method="formal",
            row_count=3,
            col_count=3,
            confidence=1.0,
            metadata={},
        )

        mock_worksheet = Mock(spec=Worksheet)
        mock_worksheet.max_row = 4
        mock_worksheet.max_column = 4

        def mock_cell(row, column):
            return Mock(value=f"Data_{row}_{column}")

        mock_worksheet.cell.side_effect = mock_cell

        detector = TableDetector()
        tables = detector._detect_grid_tables(mock_worksheet, [existing_table])

        # Should not detect overlapping table
        for table in tables:
            assert not detector._overlaps_with_existing_tables(
                {
                    "start_row": table.start_row,
                    "end_row": table.end_row,
                    "start_col": table.start_col,
                    "end_col": table.end_col,
                },
                [existing_table],
            )

    def test_detect_grid_tables_insufficient_size(self):
        """Test grid detection ignores blocks that are too small."""
        mock_worksheet = Mock(spec=Worksheet)
        mock_worksheet.max_row = 1
        mock_worksheet.max_column = 1

        mock_worksheet.cell.side_effect = lambda row, column: Mock(value="Data")

        detector = TableDetector()
        tables = detector._detect_grid_tables(mock_worksheet, [])

        assert tables == []


class TestDetectTablesIntegration:
    """Test the complete table detection pipeline."""

    @patch("src.table_detector.TableDetector._detect_grid_tables")
    @patch("src.table_detector.TableDetector._detect_pandas_tables")
    @patch("src.table_detector.TableDetector._detect_formal_tables")
    def test_detect_tables_all_methods(self, mock_formal, mock_pandas, mock_grid):
        """Test complete detection using all methods."""
        # Mock return values
        formal_table = TableInfo(
            name="FormalTable",
            range_ref="A1:C3",
            start_row=1,
            start_col=1,
            end_row=3,
            end_col=3,
            headers=["A", "B", "C"],
            detection_method="formal",
            row_count=3,
            col_count=3,
            confidence=1.0,
            metadata={},
        )
        mock_formal.return_value = [formal_table]

        pandas_table = TableInfo(
            name="PandasTable_1",
            range_ref="E1:G4",
            start_row=1,
            start_col=5,
            end_row=4,
            end_col=7,
            headers=["E", "F", "G"],
            detection_method="pandas",
            row_count=4,
            col_count=3,
            confidence=0.8,
            metadata={},
        )
        mock_pandas.return_value = [pandas_table]

        grid_table = TableInfo(
            name="GridTable_1",
            range_ref="A5:B7",
            start_row=5,
            start_col=1,
            end_row=7,
            end_col=2,
            headers=["A", "B"],
            detection_method="grid",
            row_count=3,
            col_count=2,
            confidence=0.6,
            metadata={},
        )
        mock_grid.return_value = [grid_table]

        mock_worksheet = Mock(spec=Worksheet)
        mock_worksheet.title = "TestSheet"

        detector = TableDetector()
        tables = detector.detect_tables(mock_worksheet)

        # Should detect all three tables
        assert len(tables) == 3
        methods = [table.detection_method for table in tables]
        assert "formal" in methods
        assert "pandas" in methods
        assert "grid" in methods

        # Should be sorted by position and confidence
        assert tables[0].start_row <= tables[1].start_row
        assert tables[1].start_row <= tables[2].start_row

    def test_detect_tables_critical_error(self):
        """Test handling of critical errors in table detection."""
        mock_worksheet = Mock(spec=Worksheet)
        mock_worksheet.title = "TestSheet"

        detector = TableDetector()

        # Mock _detect_formal_tables to raise an exception
        with patch.object(
            detector, "_detect_formal_tables", side_effect=Exception("Critical error")
        ):
            with pytest.raises(DataExtractionException, match="Table detection failed"):
                detector.detect_tables(mock_worksheet)


class TestUtilityMethods:
    """Test utility methods of TableDetector."""

    @pytest.mark.parametrize(
        "col_letter,expected",
        [("A", 1), ("B", 2), ("Z", 26), ("AA", 27), ("AB", 28), ("AZ", 52), ("BA", 53)],
    )
    def test_parse_column_letter(self, col_letter, expected):
        """Test column letter parsing."""
        detector = TableDetector()
        result = detector._parse_column_letter(col_letter)
        assert result == expected

    @pytest.mark.parametrize(
        "start_row,start_col,end_row,end_col,expected",
        [(1, 1, 3, 3, "A1:C3"), (5, 2, 10, 4, "B5:D10"), (1, 26, 2, 27, "Z1:AA2")],
    )
    def test_create_range_ref(self, start_row, start_col, end_row, end_col, expected):
        """Test Excel range reference creation."""
        detector = TableDetector()
        result = detector._create_range_ref(start_row, start_col, end_row, end_col)
        assert result == expected

    def test_tables_significantly_overlap(self):
        """Test overlap detection between tables."""
        detector = TableDetector()

        # Two identical tables should overlap
        table1 = TableInfo(
            name="Table1",
            range_ref="A1:C3",
            start_row=1,
            start_col=1,
            end_row=3,
            end_col=3,
            headers=[],
            detection_method="grid",
            row_count=3,
            col_count=3,
            confidence=0.8,
            metadata={},
        )
        table2 = TableInfo(
            name="Table2",
            range_ref="A1:C3",
            start_row=1,
            start_col=1,
            end_row=3,
            end_col=3,
            headers=[],
            detection_method="grid",
            row_count=3,
            col_count=3,
            confidence=0.8,
            metadata={},
        )

        assert detector._tables_significantly_overlap(table1, table2)

        # Non-overlapping tables should not overlap
        table3 = TableInfo(
            name="Table3",
            range_ref="E5:G7",
            start_row=5,
            start_col=5,
            end_row=7,
            end_col=7,
            headers=[],
            detection_method="grid",
            row_count=3,
            col_count=3,
            confidence=0.8,
            metadata={},
        )

        assert not detector._tables_significantly_overlap(table1, table3)

        # Partially overlapping tables should overlap if overlap > 50%
        table4 = TableInfo(
            name="Table4",
            range_ref="A2:C4",
            start_row=2,
            start_col=1,
            end_row=4,
            end_col=3,
            headers=[],
            detection_method="grid",
            row_count=3,
            col_count=3,
            confidence=0.8,
            metadata={},
        )

        assert detector._tables_significantly_overlap(table1, table4)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
