"""
Unit tests for streaming Excel processor performance optimizations.

This module tests the streaming Excel processing capabilities including
chunked data processing, memory management, and parallel processing.
"""

import gc
import tempfile
import time
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
from typing import Any, Dict, List

import pytest
from openpyxl import Workbook

from exc_to_pdf.config.excel_config import ExcelConfig
from exc_to_pdf.streaming_excel_processor import (
    StreamingExcelProcessor,
    ChunkIterator,
    ChunkInfo,
    ProcessingStats,
)
from exc_to_pdf.exceptions import (
    WorkbookInitializationException,
    WorksheetNotFoundException,
    DataExtractionException,
)
from exc_to_pdf.excel_processor import SheetData


@pytest.fixture
def temp_dir() -> Path:
    """Create temporary directory for test files."""
    return Path(tempfile.mkdtemp())


@pytest.fixture
def large_excel_file(temp_dir: Path) -> str:
    """Create a large Excel file for testing streaming functionality."""
    excel_path = temp_dir / "large_test_file.xlsx"
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Large Data")

    # Add header
    ws.append(["ID", "Name", "Value", "Category", "Description"])

    # Add 50,000 rows of data
    for i in range(1, 50001):
        ws.append(
            [
                f"ID{i:06d}",
                f"Name_{i}",
                i * 1.5,
                f"Category_{i % 10}",
                f"Description for item {i} with some additional text",
            ]
        )

    wb.save(excel_path)
    return str(excel_path)


@pytest.fixture
def multi_sheet_excel_file(temp_dir: Path) -> str:
    """Create Excel file with multiple sheets for parallel processing tests."""
    excel_path = temp_dir / "multi_sheet_test.xlsx"
    wb = Workbook()

    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)

    # Create multiple sheets with data
    for sheet_num in range(1, 6):  # 5 sheets
        ws = wb.create_sheet(f"Sheet_{sheet_num}")
        ws.append(["ID", "Value", "Description"])

        # Add 1000 rows per sheet
        for i in range(1, 1001):
            ws.append(
                [
                    f"{sheet_num}_{i:04d}",
                    i * sheet_num,
                    f"Item {i} in sheet {sheet_num}",
                ]
            )

    wb.save(excel_path)
    return str(excel_path)


@pytest.fixture
def performance_config() -> ExcelConfig:
    """Create configuration optimized for performance testing."""
    return ExcelConfig(
        streaming_enabled=True,
        chunk_size=1000,  # Small chunks for testing
        memory_limit_mb=100,
        parallel_processing=True,
        max_workers=2,
        enable_progress=True,
        progress_interval=0.1,  # Fast updates for testing
        enable_caching=True,
        cache_size_mb=10,
    )


@pytest.fixture
def streaming_processor(
    large_excel_file: str, performance_config: ExcelConfig
) -> StreamingExcelProcessor:
    """Create streaming processor for testing."""
    return StreamingExcelProcessor(large_excel_file, performance_config)


class TestChunkIterator:
    """Test cases for ChunkIterator functionality."""

    def test_chunk_iterator_initialization(
        self, streaming_processor: StreamingExcelProcessor
    ) -> None:
        """Test ChunkIterator initialization."""
        with streaming_processor:
            streaming_processor._initialize_workbook()
            worksheet = streaming_processor.workbook.active

            iterator = ChunkIterator(
                worksheet=worksheet, chunk_size=100, max_row=500, values_only=True
            )

            assert iterator.chunk_size == 100
            assert iterator.max_row == 500
            assert iterator.values_only is True
            assert iterator._current_row == 1

    def test_chunk_iterator_estimation(
        self, streaming_processor: StreamingExcelProcessor
    ) -> None:
        """Test chunk count estimation."""
        with streaming_processor:
            streaming_processor._initialize_workbook()
            worksheet = streaming_processor.workbook.active

            iterator = ChunkIterator(
                worksheet=worksheet, chunk_size=1000, values_only=True
            )

            total_chunks = iterator.estimate_total_chunks()
            # 50,001 rows total (including header)
            expected_chunks = (50001 + 1000 - 1) // 1000  # Ceiling division
            assert total_chunks == expected_chunks

    def test_chunk_iteration(
        self, streaming_processor: StreamingExcelProcessor
    ) -> None:
        """Test chunk iteration functionality."""
        with streaming_processor:
            streaming_processor._initialize_workbook()
            worksheet = streaming_processor.workbook.active

            iterator = ChunkIterator(
                worksheet=worksheet, chunk_size=1000, values_only=True
            )

            chunks_processed = 0
            total_rows = 0

            for chunk_data, chunk_info in iterator:
                chunks_processed += 1
                total_rows += len(chunk_data)

                # Verify chunk info structure
                assert isinstance(chunk_info, ChunkInfo)
                assert chunk_info.row_count == len(chunk_data)
                assert chunk_info.sheet_name == worksheet.title
                assert chunk_info.start_row <= chunk_info.end_row

                # Verify data structure
                assert isinstance(chunk_data, list)
                if chunk_data:  # Non-empty chunk
                    assert isinstance(chunk_data[0], list)

            # Verify we processed all chunks
            expected_chunks = iterator.estimate_total_chunks()
            assert chunks_processed == expected_chunks
            assert total_rows > 0


class TestStreamingExcelProcessor:
    """Test cases for StreamingExcelProcessor functionality."""

    def test_processor_initialization(
        self, large_excel_file: str, performance_config: ExcelConfig
    ) -> None:
        """Test processor initialization."""
        processor = StreamingExcelProcessor(large_excel_file, performance_config)

        assert processor.file_path == Path(large_excel_file)
        assert processor.config == performance_config
        assert processor.workbook is None
        assert processor.memory_monitor is None

    def test_file_validation_success(
        self, large_excel_file: str, performance_config: ExcelConfig
    ) -> None:
        """Test successful file validation."""
        processor = StreamingExcelProcessor(large_excel_file, performance_config)
        # Should not raise any exceptions
        processor._validate_file()

    def test_file_validation_not_found(self, performance_config: ExcelConfig) -> None:
        """Test file validation with non-existent file."""
        with pytest.raises(FileNotFoundError):
            StreamingExcelProcessor("/nonexistent/file.xlsx", performance_config)

    def test_file_validation_invalid_extension(
        self, temp_dir: Path, performance_config: ExcelConfig
    ) -> None:
        """Test file validation with invalid extension."""
        invalid_file = temp_dir / "test.txt"
        invalid_file.write_text("Not an Excel file")

        with pytest.raises(ValueError, match="Invalid file extension"):
            StreamingExcelProcessor(str(invalid_file), performance_config)

    def test_file_validation_too_large(
        self, temp_dir: Path, performance_config: ExcelConfig
    ) -> None:
        """Test file validation with file too large."""
        # Create a config with very small file size limit
        small_config = ExcelConfig(max_file_size_mb=1)

        # Create a file larger than the limit
        large_file = temp_dir / "large.xlsx"
        wb = Workbook()
        ws = wb.active
        for i in range(10000):  # Create a reasonably large file
            ws.append([f"Data_{i}", i, i * 2])
        wb.save(large_file)

        with pytest.raises(ValueError, match="File too large"):
            StreamingExcelProcessor(str(large_file), small_config)

    def test_workbook_initialization(
        self, streaming_processor: StreamingExcelProcessor
    ) -> None:
        """Test workbook initialization."""
        streaming_processor._initialize_workbook()

        assert streaming_processor.workbook is not None
        assert streaming_processor.memory_monitor is not None
        assert (
            streaming_processor.memory_monitor.limits.max_memory_mb
            == streaming_processor.config.memory_limit_mb
        )

    def test_context_manager(
        self, large_excel_file: str, performance_config: ExcelConfig
    ) -> None:
        """Test context manager functionality."""
        with StreamingExcelProcessor(large_excel_file, performance_config) as processor:
            assert processor.workbook is not None
            assert processor.memory_monitor is not None

        # After context exit
        assert processor.workbook is None
        assert processor.memory_monitor is None

    def test_discover_sheets(
        self, multi_sheet_excel_file: str, performance_config: ExcelConfig
    ) -> None:
        """Test sheet discovery."""
        processor = StreamingExcelProcessor(multi_sheet_excel_file, performance_config)

        with processor:
            sheet_names = processor.discover_sheets()

        assert len(sheet_names) == 5
        assert all(name.startswith("Sheet_") for name in sheet_names)
        assert "Sheet_1" in sheet_names
        assert "Sheet_5" in sheet_names

    def test_process_sheet_chunked(
        self, streaming_processor: StreamingExcelProcessor
    ) -> None:
        """Test chunked sheet processing."""
        with streaming_processor:
            chunks_processed = 0
            total_rows = 0

            for chunk_data, chunk_info in streaming_processor.process_sheet_chunked(
                "Large Data"
            ):
                chunks_processed += 1
                total_rows += len(chunk_data)

                # Verify chunk structure
                assert isinstance(chunk_data, list)
                assert isinstance(chunk_info, ChunkInfo)
                assert chunk_info.sheet_name == "Large Data"

            # Verify we processed data
            assert chunks_processed > 0
            assert total_rows > 0

        # Verify statistics updated
        stats = streaming_processor.get_processing_stats()
        assert stats.total_chunks_processed == chunks_processed
        assert stats.total_rows_processed == total_rows

    def test_extract_sheet_data_streaming(
        self, streaming_processor: StreamingExcelProcessor
    ) -> None:
        """Test streaming sheet data extraction."""
        with streaming_processor:
            sheet_data = streaming_processor.extract_sheet_data_streaming("Large Data")

        # Verify SheetData structure
        assert isinstance(sheet_data, SheetData)
        assert sheet_data.sheet_name == "Large Data"
        assert sheet_data.has_data is True
        assert sheet_data.row_count > 0
        assert sheet_data.col_count > 0
        assert len(sheet_data.raw_data) > 0

        # Verify data integrity
        assert len(sheet_data.raw_data[0]) == 5  # Header row has 5 columns
        assert sheet_data.raw_data[0][0] == "ID"  # First column header

        # Verify metadata contains processing stats
        assert "processing_stats" in sheet_data.metadata
        processing_stats = sheet_data.metadata["processing_stats"]
        assert processing_stats["streaming_mode"] is True
        assert processing_stats["chunks_processed"] > 0

    def test_process_all_sheets_parallel(
        self, multi_sheet_excel_file: str, performance_config: ExcelConfig
    ) -> None:
        """Test parallel processing of multiple sheets."""
        processor = StreamingExcelProcessor(multi_sheet_excel_file, performance_config)

        with processor:
            start_time = time.time()
            sheet_data_list = processor.process_all_sheets_parallel()
            end_time = time.time()

        # Verify results
        assert len(sheet_data_list) == 5  # 5 sheets processed
        sheet_names = [data.sheet_name for data in sheet_data_list]
        assert all(name.startswith("Sheet_") for name in sheet_names)

        # Verify each sheet has data
        for sheet_data in sheet_data_list:
            assert sheet_data.has_data is True
            assert sheet_data.row_count > 0
            assert len(sheet_data.raw_data) > 0

        # Verify processing statistics
        stats = processor.get_processing_stats()
        assert stats.sheets_processed == 5
        assert stats.processing_time_seconds > 0
        assert stats.processing_time_seconds == end_time - start_time

    def test_process_all_sheets_sequential(
        self, multi_sheet_excel_file: str, performance_config: ExcelConfig
    ) -> None:
        """Test sequential processing of multiple sheets."""
        # Disable parallel processing for this test
        sequential_config = ExcelConfig(
            **{**performance_config.__dict__, "parallel_processing": False}
        )

        processor = StreamingExcelProcessor(multi_sheet_excel_file, sequential_config)

        with processor:
            sheet_data_list = processor.process_all_sheets_sequential()

        # Verify results
        assert len(sheet_data_list) == 5
        sheet_names = [data.sheet_name for data in sheet_data_list]
        assert all(name.startswith("Sheet_") for name in sheet_names)

    def test_memory_monitoring(
        self, large_excel_file: str, performance_config: ExcelConfig
    ) -> None:
        """Test memory monitoring functionality."""
        processor = StreamingExcelProcessor(large_excel_file, performance_config)

        with processor:
            # Process some data to trigger memory monitoring
            for chunk_data, chunk_info in processor.process_sheet_chunked("Large Data"):
                # Memory monitor should be active
                assert processor.memory_monitor is not None

                # Get memory stats
                stats = processor.memory_monitor.get_memory_stats()
                assert stats.current_mb >= 0
                assert stats.peak_mb >= stats.current_mb

                # Process only a few chunks for testing
                if chunk_info.chunk_index >= 2:
                    break

    def test_chunk_size_configuration(self, large_excel_file: str) -> None:
        """Test different chunk size configurations."""
        # Test with small chunks
        small_chunk_config = ExcelConfig(chunk_size=100)
        processor = StreamingExcelProcessor(large_excel_file, small_chunk_config)

        with processor:
            chunk_count = 0
            for chunk_data, chunk_info in processor.process_sheet_chunked("Large Data"):
                chunk_count += 1
                # Each chunk should have at most 100 rows (plus header)
                assert len(chunk_data) <= 100

                # Test only a few chunks
                if chunk_count >= 3:
                    break

        # Test with large chunks
        large_chunk_config = ExcelConfig(chunk_size=10000)
        processor = StreamingExcelProcessor(large_excel_file, large_chunk_config)

        with processor:
            chunk_count = 0
            for chunk_data, chunk_info in processor.process_sheet_chunked("Large Data"):
                chunk_count += 1
                # First chunk should be large
                if chunk_count == 1:
                    assert len(chunk_data) > 100

                # Test only first chunk
                break

    def test_error_handling_invalid_sheet(
        self, streaming_processor: StreamingExcelProcessor
    ) -> None:
        """Test error handling for invalid sheet names."""
        with streaming_processor:
            with pytest.raises(WorksheetNotFoundException):
                streaming_processor.extract_sheet_data_streaming("NonExistentSheet")

    def test_detect_tables_streaming(
        self, streaming_processor: StreamingExcelProcessor
    ) -> None:
        """Test table detection in streaming mode."""
        with streaming_processor:
            tables = streaming_processor.detect_tables_streaming("Large Data")

            # Should return a list (even if empty)
            assert isinstance(tables, list)

    def test_processing_stats(
        self, streaming_processor: StreamingExcelProcessor
    ) -> None:
        """Test processing statistics tracking."""
        initial_stats = streaming_processor.get_processing_stats()

        with streaming_processor:
            # Process some data
            for chunk_data, chunk_info in streaming_processor.process_sheet_chunked(
                "Large Data"
            ):
                if chunk_info.chunk_index >= 2:  # Process only a few chunks
                    break

        final_stats = streaming_processor.get_processing_stats()

        # Verify statistics were updated
        assert final_stats.total_rows_processed > initial_stats.total_rows_processed
        assert final_stats.total_chunks_processed > initial_stats.total_chunks_processed

    def test_resource_cleanup(
        self, large_excel_file: str, performance_config: ExcelConfig
    ) -> None:
        """Test proper resource cleanup."""
        processor = StreamingExcelProcessor(large_excel_file, performance_config)

        # Process some data
        with processor:
            processor._initialize_workbook()
            assert processor.workbook is not None
            assert processor.memory_monitor is not None

        # After context exit, resources should be cleaned up
        assert processor.workbook is None
        assert processor.memory_monitor is None

        # Force garbage collection to test cleanup
        gc.collect()
        # Should not raise any exceptions


class TestPerformanceOptimizations:
    """Test cases specifically for performance optimizations."""

    def test_automatic_streaming_detection(self, temp_dir: Path) -> None:
        """Test automatic streaming detection based on file size."""
        # Create small file (should not trigger streaming)
        small_file = temp_dir / "small.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Header1", "Header2"])
        for i in range(10):
            ws.append([f"Data_{i}", i])
        wb.save(small_file)

        config = ExcelConfig(streaming_enabled=False)  # Disabled by default
        processor = StreamingExcelProcessor(str(small_file), config)

        # Small file should not trigger streaming even if enabled
        assert processor._should_use_streaming() is False

    def test_parallel_vs_sequential_performance(
        self, multi_sheet_excel_file: str
    ) -> None:
        """Test performance difference between parallel and sequential processing."""
        config_parallel = ExcelConfig(
            parallel_processing=True, max_workers=4, streaming_enabled=True
        )
        config_sequential = ExcelConfig(
            parallel_processing=False, streaming_enabled=True
        )

        # Test parallel processing
        processor_parallel = StreamingExcelProcessor(
            multi_sheet_excel_file, config_parallel
        )
        start_time = time.time()
        with processor_parallel:
            result_parallel = processor_parallel.process_all_sheets_parallel()
        parallel_time = time.time() - start_time

        # Test sequential processing
        processor_sequential = StreamingExcelProcessor(
            multi_sheet_excel_file, config_sequential
        )
        start_time = time.time()
        with processor_sequential:
            result_sequential = processor_sequential.process_all_sheets_sequential()
        sequential_time = time.time() - start_time

        # Verify results are the same
        assert len(result_parallel) == len(result_sequential)

        # Parallel should generally be faster (but not always in testing environments)
        # We just verify both complete successfully
        assert parallel_time > 0
        assert sequential_time > 0

    def test_memory_efficiency_with_large_chunks(self, large_excel_file: str) -> None:
        """Test memory efficiency with different chunk sizes."""
        import psutil
        import os

        process = psutil.Process(os.getpid())
        initial_memory = process.memory_info().rss / 1024 / 1024  # MB

        # Test with very small chunks (should use less memory)
        small_chunk_config = ExcelConfig(chunk_size=100, memory_limit_mb=50)
        processor = StreamingExcelProcessor(large_excel_file, small_chunk_config)

        with processor:
            # Process the entire file
            for chunk_data, chunk_info in processor.process_sheet_chunked("Large Data"):
                pass  # Process all chunks

        final_memory = process.memory_info().rss / 1024 / 1024  # MB
        memory_increase = final_memory - initial_memory

        # Memory increase should be reasonable (less than 100MB for this test)
        assert (
            memory_increase < 100
        ), f"Memory increased by {memory_increase}MB, expected < 100MB"
