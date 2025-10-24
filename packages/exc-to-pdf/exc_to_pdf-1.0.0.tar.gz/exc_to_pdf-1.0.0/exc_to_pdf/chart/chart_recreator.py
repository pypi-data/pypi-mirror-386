"""
Chart recreation module using matplotlib.

This module provides capabilities to recreate Excel charts using matplotlib
based on extracted chart information and data.
"""

from typing import List, Dict, Any, Optional, Tuple
import logging
import re
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
import pandas as pd
from matplotlib.figure import Figure
from matplotlib.axes import Axes

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from .chart_extractor import ChartInfo, ChartType

logger = logging.getLogger(__name__)


class ChartRecreator:
    """Recreates Excel charts using matplotlib."""

    def __init__(self, excel_file_path: str):
        """Initialize chart recreator.

        Args:
            excel_file_path: Path to the Excel file containing chart data
        """
        self.excel_file_path = excel_file_path
        self.workbook = load_workbook(excel_file_path, data_only=True)

        # Set matplotlib style for professional appearance
        plt.style.use(
            "seaborn-v0_8" if "seaborn-v0_8" in plt.style.available else "default"
        )

    def recreate_chart(self, chart_info: ChartInfo) -> Optional[Figure]:
        """Recreate a chart based on ChartInfo.

        Args:
            chart_info: Chart information from extractor

        Returns:
            matplotlib Figure object or None if recreation failed
        """
        try:
            # Extract data from Excel ranges
            data = self._extract_chart_data(chart_info)
            if not data:
                logger.warning(f"No data extracted for chart: {chart_info.title}")
                return None

            # Create figure and recreate chart
            fig, ax = plt.subplots(figsize=(10, 6))

            if chart_info.chart_type == ChartType.BAR:
                self._recreate_bar_chart(ax, data, chart_info)
            elif chart_info.chart_type == ChartType.LINE:
                self._recreate_line_chart(ax, data, chart_info)
            elif chart_info.chart_type == ChartType.PIE:
                self._recreate_pie_chart(ax, data, chart_info)
            elif chart_info.chart_type == ChartType.SCATTER:
                self._recreate_scatter_chart(ax, data, chart_info)
            elif chart_info.chart_type == ChartType.AREA:
                self._recreate_area_chart(ax, data, chart_info)
            else:
                logger.warning(f"Unsupported chart type: {chart_info.chart_type}")
                plt.close(fig)
                return None

            # Apply styling
            self._apply_styling(ax, chart_info)

            # Set title
            if chart_info.title:
                ax.set_title(chart_info.title, fontsize=14, fontweight="bold", pad=20)

            # Adjust layout
            plt.tight_layout()

            logger.info(f"Successfully recreated chart: {chart_info.title}")
            return fig

        except Exception as e:
            logger.error(f"Error recreating chart {chart_info.title}: {e}")
            if "fig" in locals():
                plt.close(fig)
            return None

    def _extract_chart_data(self, chart_info: ChartInfo) -> Optional[pd.DataFrame]:
        """Extract data from Excel ranges for the chart.

        Args:
            chart_info: Chart information containing data ranges

        Returns:
            pandas DataFrame with chart data or None if extraction failed
        """
        try:
            all_data = []

            for data_range in chart_info.data_ranges:
                # Parse range like "Sheet1!$A$1:$B$10" or "A1:B10"
                range_parts = self._parse_range(data_range)
                if not range_parts:
                    continue

                worksheet_name, start_col, start_row, end_col, end_row = range_parts

                # Get worksheet
                if worksheet_name:
                    worksheet = self.workbook[worksheet_name]
                else:
                    # Use the chart's worksheet if no specific worksheet mentioned
                    worksheet = self.workbook[chart_info.worksheet_name]

                # Extract data
                range_data = []
                for row in worksheet.iter_rows(
                    min_row=start_row,
                    max_row=end_row,
                    min_col=start_col,
                    max_col=end_col,
                    values_only=True,
                ):
                    range_data.append(row)

                # Convert to DataFrame
                if range_data:
                    df = pd.DataFrame(range_data)
                    all_data.append(df)

            if not all_data:
                return None

            # Combine all data
            combined_data = pd.concat(all_data, axis=1)

            # Use first row as header if it contains strings
            if combined_data.shape[0] > 0:
                first_row = combined_data.iloc[0]
                if any(isinstance(val, str) for val in first_row if val is not None):
                    combined_data.columns = [
                        str(val) if val is not None else f"Col_{i}"
                        for i, val in enumerate(first_row)
                    ]
                    combined_data = combined_data.iloc[1:]

            return combined_data

        except Exception as e:
            logger.error(f"Error extracting chart data: {e}")
            return None

    def _parse_range(self, range_str: str) -> Optional[Tuple[str, int, int, int, int]]:
        """Parse Excel range string into components.

        Args:
            range_str: Excel range string like "Sheet1!$A$1:$B$10"

        Returns:
            Tuple of (worksheet_name, start_col, start_row, end_col, end_row)
        """
        try:
            # Remove $ symbols and split
            clean_range = range_str.replace("$", "")

            # Check if worksheet name is included
            if "!" in clean_range:
                worksheet_name, range_part = clean_range.split("!", 1)
            else:
                worksheet_name = None
                range_part = clean_range

            # Parse range part
            if ":" not in range_part:
                return None

            start_cell, end_cell = range_part.split(":", 1)

            # Convert column letters to numbers
            start_col = self._col_letter_to_number(start_cell.rstrip("0123456789"))
            start_row = int(start_cell.lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
            end_col = self._col_letter_to_number(end_cell.rstrip("0123456789"))
            end_row = int(end_cell.lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))

            return worksheet_name, start_col, start_row, end_col, end_row

        except Exception as e:
            logger.error(f"Error parsing range '{range_str}': {e}")
            return None

    def _col_letter_to_number(self, col_letter: str) -> int:
        """Convert Excel column letter to number.

        Args:
            col_letter: Column letter like 'A', 'B', 'AA'

        Returns:
            Column number (1-based)
        """
        result = 0
        for char in col_letter:
            result = result * 26 + (ord(char.upper()) - ord("A") + 1)
        return result

    def _recreate_bar_chart(self, ax: Axes, data: pd.DataFrame, chart_info: ChartInfo):
        """Recreate a bar chart.

        Args:
            ax: matplotlib Axes object
            data: Chart data
            chart_info: Chart information
        """
        if data.shape[1] < 2:
            logger.warning("Insufficient data for bar chart")
            return

        # Use first column as categories, remaining as series
        categories = data.iloc[:, 0]
        series_data = data.iloc[:, 1:]

        # Create bar chart
        x_pos = np.arange(len(categories))
        width = 0.8 / len(series_data.columns)

        for i, (col_name, series) in enumerate(series_data.items()):
            ax.bar(x_pos + i * width, series, width, label=str(col_name))

        ax.set_xlabel(str(data.columns[0]))
        ax.set_ylabel("Values")
        ax.set_xticks(x_pos + width * (len(series_data.columns) - 1) / 2)
        ax.set_xticklabels([str(val) for val in categories], rotation=45)

        if len(series_data.columns) > 1:
            ax.legend()

    def _recreate_line_chart(self, ax: Axes, data: pd.DataFrame, chart_info: ChartInfo):
        """Recreate a line chart.

        Args:
            ax: matplotlib Axes object
            data: Chart data
            chart_info: Chart information
        """
        if data.shape[1] < 2:
            logger.warning("Insufficient data for line chart")
            return

        # Use first column as x-axis, remaining as series
        x_data = data.iloc[:, 0]
        series_data = data.iloc[:, 1:]

        # Create line chart
        for col_name, series in series_data.items():
            ax.plot(x_data, series, marker="o", label=str(col_name))

        ax.set_xlabel(str(data.columns[0]))
        ax.set_ylabel("Values")

        if len(series_data.columns) > 1:
            ax.legend()

    def _recreate_pie_chart(self, ax: Axes, data: pd.DataFrame, chart_info: ChartInfo):
        """Recreate a pie chart.

        Args:
            ax: matplotlib Axes object
            data: Chart data
            chart_info: Chart information
        """
        if data.shape[1] < 2:
            logger.warning("Insufficient data for pie chart")
            return

        # Use first column as labels, second as values
        labels = data.iloc[:, 0]
        values = data.iloc[:, 1]

        # Create pie chart
        ax.pie(
            values,
            labels=[str(val) for val in labels],
            autopct="%1.1f%%",
            startangle=90,
        )
        ax.axis("equal")  # Equal aspect ratio ensures that pie is drawn as a circle

    def _recreate_scatter_chart(
        self, ax: Axes, data: pd.DataFrame, chart_info: ChartInfo
    ):
        """Recreate a scatter chart.

        Args:
            ax: matplotlib Axes object
            data: Chart data
            chart_info: Chart information
        """
        if data.shape[1] < 2:
            logger.warning("Insufficient data for scatter chart")
            return

        # Use first two columns as x and y
        x_data = data.iloc[:, 0]
        y_data = data.iloc[:, 1]

        # Create scatter chart
        ax.scatter(x_data, y_data, alpha=0.6)

        ax.set_xlabel(str(data.columns[0]))
        ax.set_ylabel(str(data.columns[1]))

        # Add third dimension as color if available
        if data.shape[1] > 2:
            colors = data.iloc[:, 2]
            scatter = ax.scatter(x_data, y_data, c=colors, alpha=0.6, cmap="viridis")
            plt.colorbar(scatter, ax=ax, label=str(data.columns[2]))

    def _recreate_area_chart(self, ax: Axes, data: pd.DataFrame, chart_info: ChartInfo):
        """Recreate an area chart.

        Args:
            ax: matplotlib Axes object
            data: Chart data
            chart_info: Chart information
        """
        if data.shape[1] < 2:
            logger.warning("Insufficient data for area chart")
            return

        # Use first column as x-axis, remaining as series
        x_data = data.iloc[:, 0]
        series_data = data.iloc[:, 1:]

        # Create area chart (stacked if multiple series)
        ax.stackplot(
            x_data,
            [series for _, series in series_data.items()],
            labels=[str(col) for col in series_data.columns],
            alpha=0.7,
        )

        ax.set_xlabel(str(data.columns[0]))
        ax.set_ylabel("Values")

        if len(series_data.columns) > 1:
            ax.legend()

    def _apply_styling(self, ax: Axes, chart_info: ChartInfo):
        """Apply styling to the chart based on chart_info.

        Args:
            ax: matplotlib Axes object
            chart_info: Chart information containing styling
        """
        try:
            styling = chart_info.styling

            # Apply legend settings
            if "legend" in styling:
                legend_settings = styling["legend"]
                if ax.get_legend():
                    ax.get_legend().set_visible(legend_settings.get("visible", True))

            # Apply axis titles
            if "x_axis" in styling:
                x_axis = styling["x_axis"]
                if "title" in x_axis and x_axis["title"]:
                    ax.set_xlabel(x_axis["title"])

            if "y_axis" in styling:
                y_axis = styling["y_axis"]
                if "title" in y_axis and y_axis["title"]:
                    ax.set_ylabel(y_axis["title"])

            # Apply grid
            ax.grid(True, alpha=0.3)

            # Set background color
            ax.set_facecolor("#ffffff")

        except Exception as e:
            logger.error(f"Error applying styling: {e}")

    def recreate_all_charts(self, charts_info: List[ChartInfo]) -> List[Figure]:
        """Recreate multiple charts.

        Args:
            charts_info: List of ChartInfo objects

        Returns:
            List of matplotlib Figure objects
        """
        figures = []

        for chart_info in charts_info:
            fig = self.recreate_chart(chart_info)
            if fig:
                figures.append(fig)

        logger.info(
            f"Successfully recreated {len(figures)} out of {len(charts_info)} charts"
        )
        return figures
