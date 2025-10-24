"""
Hybrid table detection system for Excel worksheets.

This module provides a comprehensive table detection system that combines
formal Excel tables, pandas inference, and grid pattern detection to
identify data tables in worksheets.
"""

import logging
from typing import List, Dict, Any, Optional, Tuple, Iterator
from dataclasses import dataclass
import pandas as pd

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table

from .exceptions import DataExtractionException
from .config.excel_config import ExcelConfig, DEFAULT_CONFIG

# Configure logging
logger = logging.getLogger(__name__)


@dataclass
class TableInfo:
    """
    Information about a detected table.

    Contains metadata about the table including location, dimensions,
    headers, and detection method used.
    """

    name: str
    range_ref: str
    start_row: int
    start_col: int
    end_row: int
    end_col: int
    headers: List[str]
    detection_method: str  # "formal", "pandas", or "grid"
    row_count: int
    col_count: int
    confidence: float  # 0.0 to 1.0
    metadata: Dict[str, Any]


class TableDetector:
    """
    Hybrid table detection system for Excel worksheets.

    Detects tables using three methods in order of preference:
    1. Formal Excel tables (defined in the worksheet)
    2. Pandas DataFrame inference (pattern recognition)
    3. Grid pattern detection (contiguous data blocks)
    """

    def __init__(self, config: Optional[ExcelConfig] = None):
        """
        Initialize TableDetector with configuration.

        Args:
            config: Configuration object for detection settings
        """
        self.config = config or DEFAULT_CONFIG

    def detect_tables(self, worksheet: Worksheet) -> List[TableInfo]:
        """
        Detect all tables in the worksheet using hybrid approach.

        Args:
            worksheet: The worksheet to analyze

        Returns:
            List of detected tables ordered by confidence and position

        Raises:
            DataExtractionException: If table detection fails critically
        """
        try:
            logger.debug(f"Starting table detection for worksheet: {worksheet.title}")

            # Method 1: Detect formal Excel tables
            formal_tables = self._detect_formal_tables(worksheet)
            logger.debug(f"Found {len(formal_tables)} formal tables")

            # Method 2: Detect tables using pandas inference
            pandas_tables = self._detect_pandas_tables(worksheet, formal_tables)
            logger.debug(f"Found {len(pandas_tables)} pandas-inferred tables")

            # Method 3: Detect tables using grid patterns (fallback)
            grid_tables = self._detect_grid_tables(
                worksheet, formal_tables + pandas_tables
            )
            logger.debug(f"Found {len(grid_tables)} grid-based tables")

            # Combine and deduplicate tables
            all_tables = self._combine_and_deduplicate_tables(
                formal_tables, pandas_tables, grid_tables
            )

            # Sort by confidence and position
            all_tables.sort(key=lambda t: (t.start_row, t.start_col, -t.confidence))

            logger.info(
                f"Detected {len(all_tables)} total tables in worksheet '{worksheet.title}'"
            )
            return all_tables

        except Exception as e:
            error_msg = f"Table detection failed for worksheet '{worksheet.title}': {e}"
            logger.error(
                error_msg, extra={"worksheet": worksheet.title, "error": str(e)}
            )
            raise DataExtractionException(
                error_msg,
                context={"worksheet": worksheet.title, "original_error": str(e)},
            ) from e

    def _detect_formal_tables(self, worksheet: Worksheet) -> List[TableInfo]:
        """
        Detect formal Excel tables defined in the worksheet.

        Args:
            worksheet: The worksheet to analyze

        Returns:
            List of formal table information
        """
        formal_tables: List[TableInfo] = []

        if not hasattr(worksheet, "tables") or not worksheet.tables:
            return formal_tables

        for table in worksheet.tables.values():
            try:
                # Parse table range reference
                start_cell, end_cell = table.ref.split(":")
                start_col = self._parse_column_letter(start_cell.rstrip("0123456789"))
                start_row = int(start_cell.lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
                end_col = self._parse_column_letter(end_cell.rstrip("0123456789"))
                end_row = int(end_cell.lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))

                # Extract headers from first row of table
                headers = []
                if table.headerRowCount > 0:
                    for col_idx in range(start_col, end_col + 1):
                        cell = worksheet.cell(row=start_row, column=col_idx)
                        headers.append(
                            str(cell.value)
                            if cell.value is not None
                            else f"Column_{col_idx}"
                        )

                # Calculate table dimensions
                row_count = end_row - start_row + 1
                col_count = end_col - start_col + 1

                table_info = TableInfo(
                    name=table.name,
                    range_ref=table.ref,
                    start_row=start_row,
                    start_col=start_col,
                    end_row=end_row,
                    end_col=end_col,
                    headers=headers,
                    detection_method="formal",
                    row_count=row_count,
                    col_count=col_count,
                    confidence=1.0,  # Formal tables have highest confidence
                    metadata={
                        "header_row_count": table.headerRowCount,
                        "insert_row": table.insertRow,
                        "totals_row_count": table.totalsRowCount,
                    },
                )
                formal_tables.append(table_info)

            except Exception as e:
                logger.warning(f"Error processing formal table '{table.name}': {e}")
                continue

        return formal_tables

    def _detect_pandas_tables(
        self, worksheet: Worksheet, existing_tables: List[TableInfo]
    ) -> List[TableInfo]:
        """
        Detect tables using pandas DataFrame inference.

        Args:
            worksheet: The worksheet to analyze
            existing_tables: Tables already detected (to avoid duplicates)

        Returns:
            List of pandas-inferred table information
        """
        pandas_tables: List[TableInfo] = []

        try:
            # Convert worksheet to DataFrame in chunks
            max_row = worksheet.max_row
            max_col = worksheet.max_column

            if (
                max_row < self.config.min_table_rows
                or max_col < self.config.min_table_cols
            ):
                return pandas_tables

            # Read data into DataFrame
            data = []
            for row in worksheet.iter_rows(values_only=True):
                data.append(row)

            if not data:
                return pandas_tables

            # Create DataFrame and analyze structure
            df = pd.DataFrame(data)

            # Look for table-like patterns in the DataFrame
            potential_tables = self._find_dataframe_tables(df, existing_tables)

            for table_data in potential_tables:
                try:
                    table_info = TableInfo(
                        name=f"PandasTable_{len(pandas_tables) + 1}",
                        range_ref=self._create_range_ref(
                            table_data["start_row"],
                            table_data["start_col"],
                            table_data["end_row"],
                            table_data["end_col"],
                        ),
                        start_row=table_data["start_row"],
                        start_col=table_data["start_col"],
                        end_row=table_data["end_row"],
                        end_col=table_data["end_col"],
                        headers=table_data["headers"],
                        detection_method="pandas",
                        row_count=table_data["row_count"],
                        col_count=table_data["col_count"],
                        confidence=table_data["confidence"],
                        metadata={
                            "pattern_type": table_data.get("pattern_type", "unknown")
                        },
                    )
                    pandas_tables.append(table_info)

                except Exception as e:
                    logger.warning(f"Error creating pandas table info: {e}")
                    continue

        except Exception as e:
            logger.warning(f"Error in pandas table detection: {e}")

        return pandas_tables

    def _detect_grid_tables(
        self, worksheet: Worksheet, existing_tables: List[TableInfo]
    ) -> List[TableInfo]:
        """
        Detect tables using grid pattern analysis.

        Args:
            worksheet: The worksheet to analyze
            existing_tables: Tables already detected (to avoid duplicates)

        Returns:
            List of grid-detected table information
        """
        grid_tables: List[TableInfo] = []

        try:
            max_row = worksheet.max_row
            max_col = worksheet.max_column

            # Create a grid representation of the worksheet
            data_grid = []
            for row_idx in range(1, max_row + 1):
                row_data = []
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    row_data.append(cell.value is not None)
                data_grid.append(row_data)

            # Find contiguous data blocks
            data_blocks = self._find_contiguous_data_blocks(data_grid)

            for block in data_blocks:
                # Skip if too small
                if (
                    block["height"] < self.config.min_table_rows
                    or block["width"] < self.config.min_table_cols
                ):
                    continue

                # Check if this block overlaps with existing tables
                if self._overlaps_with_existing_tables(block, existing_tables):
                    continue

                try:
                    # Extract headers from first row
                    headers = []
                    for col_idx in range(block["start_col"], block["end_col"] + 1):
                        cell = worksheet.cell(row=block["start_row"], column=col_idx)
                        headers.append(
                            str(cell.value)
                            if cell.value is not None
                            else f"Column_{col_idx}"
                        )

                    table_info = TableInfo(
                        name=f"GridTable_{len(grid_tables) + 1}",
                        range_ref=self._create_range_ref(
                            block["start_row"],
                            block["start_col"],
                            block["end_row"],
                            block["end_col"],
                        ),
                        start_row=block["start_row"],
                        start_col=block["start_col"],
                        end_row=block["end_row"],
                        end_col=block["end_col"],
                        headers=headers,
                        detection_method="grid",
                        row_count=block["height"],
                        col_count=block["width"],
                        confidence=0.6,  # Grid detection has lower confidence
                        metadata={"block_type": "contiguous_data"},
                    )
                    grid_tables.append(table_info)

                except Exception as e:
                    logger.warning(f"Error creating grid table info: {e}")
                    continue

        except Exception as e:
            logger.warning(f"Error in grid table detection: {e}")

        return grid_tables

    def _find_dataframe_tables(
        self, df: pd.DataFrame, existing_tables: List[TableInfo]
    ) -> List[Dict[str, Any]]:
        """
        Find table patterns in a DataFrame.

        Args:
            df: DataFrame to analyze
            existing_tables: Already detected tables to avoid duplicates

        Returns:
            List of potential table information
        """
        potential_tables: List[Dict[str, Any]] = []

        # Simple pattern: look for regions with consistent data
        if df.empty:
            return potential_tables

        # Find non-empty regions
        non_empty_mask = df.notna().any(axis=1)
        non_empty_rows = df[non_empty_mask]

        if len(non_empty_rows) < self.config.min_table_rows:
            return potential_tables

        # Basic table detection: assume first row contains headers
        if len(non_empty_rows.columns) >= self.config.min_table_cols:
            headers = [
                str(col) if pd.notna(col) else f"Column_{i}"
                for i, col in enumerate(non_empty_rows.iloc[0])
            ]

            potential_tables.append(
                {
                    "start_row": 1,
                    "start_col": 1,
                    "end_row": len(non_empty_rows),
                    "end_col": len(non_empty_rows.columns),
                    "headers": headers,
                    "row_count": len(non_empty_rows),
                    "col_count": len(non_empty_rows.columns),
                    "confidence": 0.8,
                    "pattern_type": "header_data",
                }
            )

        return potential_tables

    def _find_contiguous_data_blocks(
        self, data_grid: List[List[bool]]
    ) -> List[Dict[str, int]]:
        """
        Find contiguous blocks of data in the grid.

        Args:
            data_grid: 2D boolean grid representing data presence

        Returns:
            List of data block information
        """
        blocks = []
        visited = [[False for _ in row] for row in data_grid]

        for row_idx in range(len(data_grid)):
            for col_idx in range(len(data_grid[0])):
                if data_grid[row_idx][col_idx] and not visited[row_idx][col_idx]:
                    block = self._explore_block(data_grid, visited, row_idx, col_idx)
                    if block:
                        blocks.append(block)

        return blocks

    def _explore_block(
        self,
        data_grid: List[List[bool]],
        visited: List[List[bool]],
        start_row: int,
        start_col: int,
    ) -> Optional[Dict[str, int]]:
        """
        Explore a contiguous block starting from given position.

        Args:
            data_grid: 2D boolean grid
            visited: Grid marking visited cells
            start_row: Starting row index
            start_col: Starting column index

        Returns:
            Block information or None if invalid
        """
        if not data_grid or not data_grid[0]:
            return None

        rows = len(data_grid)
        cols = len(data_grid[0])

        # Simple approach: find the bounding rectangle of the contiguous region
        min_row, max_row = start_row, start_row
        min_col, max_col = start_col, start_col

        # BFS/DFS to explore the contiguous region
        stack = [(start_row, start_col)]

        while stack:
            row, col = stack.pop()

            if (
                row < 0
                or row >= rows
                or col < 0
                or col >= cols
                or not data_grid[row][col]
                or visited[row][col]
            ):
                continue

            visited[row][col] = True

            min_row = min(min_row, row)
            max_row = max(max_row, row)
            min_col = min(min_col, col)
            max_col = max(max_col, col)

            # Add neighbors
            for dr, dc in [(0, 1), (1, 0), (0, -1), (-1, 0)]:
                stack.append((row + dr, col + dc))

        return {
            "start_row": min_row + 1,  # Convert to 1-based indexing
            "start_col": min_col + 1,
            "end_row": max_row + 1,
            "end_col": max_col + 1,
            "height": max_row - min_row + 1,
            "width": max_col - min_col + 1,
        }

    def _overlaps_with_existing_tables(
        self, block: Dict[str, int], existing_tables: List[TableInfo]
    ) -> bool:
        """
        Check if a block overlaps with existing tables.

        Args:
            block: Data block information
            existing_tables: List of existing tables

        Returns:
            True if block overlaps with any existing table
        """
        for table in existing_tables:
            # Check if blocks overlap
            if not (
                block["end_row"] < table.start_row
                or block["start_row"] > table.end_row
                or block["end_col"] < table.start_col
                or block["start_col"] > table.end_col
            ):
                return True
        return False

    def _combine_and_deduplicate_tables(
        self,
        formal_tables: List[TableInfo],
        pandas_tables: List[TableInfo],
        grid_tables: List[TableInfo],
    ) -> List[TableInfo]:
        """
        Combine tables from different detection methods and remove duplicates.

        Args:
            formal_tables: Tables from formal Excel tables
            pandas_tables: Tables from pandas inference
            grid_tables: Tables from grid detection

        Returns:
            Deduplicated list of tables
        """
        all_tables = formal_tables + pandas_tables + grid_tables
        deduplicated_tables: List[TableInfo] = []

        for table in all_tables:
            is_duplicate = False

            for existing in deduplicated_tables:
                # Check if tables significantly overlap
                if self._tables_significantly_overlap(table, existing):
                    # Keep the table with higher confidence
                    if table.confidence > existing.confidence:
                        deduplicated_tables.remove(existing)
                        deduplicated_tables.append(table)
                    is_duplicate = True
                    break

            if not is_duplicate:
                deduplicated_tables.append(table)

        return deduplicated_tables

    def _tables_significantly_overlap(
        self, table1: TableInfo, table2: TableInfo
    ) -> bool:
        """
        Check if two tables significantly overlap.

        Args:
            table1: First table
            table2: Second table

        Returns:
            True if tables significantly overlap
        """
        # Calculate overlap area
        overlap_rows = max(
            0,
            min(table1.end_row, table2.end_row)
            - max(table1.start_row, table2.start_row)
            + 1,
        )
        overlap_cols = max(
            0,
            min(table1.end_col, table2.end_col)
            - max(table1.start_col, table2.start_col)
            + 1,
        )

        if overlap_rows <= 0 or overlap_cols <= 0:
            return False

        overlap_area = overlap_rows * overlap_cols
        area1 = table1.row_count * table1.col_count
        area2 = table2.row_count * table2.col_count

        # Consider overlapping if more than 50% of the smaller table overlaps
        overlap_ratio = overlap_area / min(area1, area2)
        return overlap_ratio > 0.5

    def _parse_column_letter(self, col_letter: str) -> int:
        """
        Convert column letter (e.g., 'A', 'B', 'AA') to column number.

        Args:
            col_letter: Column letter(s)

        Returns:
            Column number (1-based)
        """
        col_num = 0
        for char in col_letter.upper():
            col_num = col_num * 26 + (ord(char) - ord("A") + 1)
        return col_num

    def _create_range_ref(
        self, start_row: int, start_col: int, end_row: int, end_col: int
    ) -> str:
        """
        Create Excel range reference from row and column numbers.

        Args:
            start_row: Starting row (1-based)
            start_col: Starting column (1-based)
            end_row: Ending row (1-based)
            end_col: Ending column (1-based)

        Returns:
            Excel range reference (e.g., 'A1:C5')
        """

        def col_num_to_letter(col_num: int) -> str:
            result = ""
            while col_num > 0:
                col_num -= 1
                result = chr(ord("A") + col_num % 26) + result
                col_num //= 26
            return result

        start_letter = col_num_to_letter(start_col)
        end_letter = col_num_to_letter(end_col)

        return f"{start_letter}{start_row}:{end_letter}{end_row}"
