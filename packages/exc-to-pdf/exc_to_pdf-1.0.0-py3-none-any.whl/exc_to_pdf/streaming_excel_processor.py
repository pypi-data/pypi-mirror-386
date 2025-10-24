"""
Streaming Excel file processor for large file handling.

This module provides memory-efficient Excel processing with streaming,
chunked data processing, and parallel sheet processing capabilities.
"""

import gc
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from itertools import islice
from pathlib import Path
from typing import Iterator, List, Optional, Dict, Any, Tuple, Union

import structlog
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .config.excel_config import ExcelConfig
from .excel_processor import SheetData
from .exceptions import (
    ExcelReaderError,
    WorkbookInitializationException,
    WorksheetNotFoundException,
    DataExtractionException,
)
from .memory_monitor import MemoryMonitor, ResourceLimits
from .progress_tracker import ProgressTracker, MultiProgressTracker
from .table_detector import TableDetector, TableInfo

logger = structlog.get_logger()


@dataclass
class ChunkInfo:
    """Information about a data chunk."""

    chunk_index: int
    start_row: int
    end_row: int
    row_count: int
    sheet_name: str


@dataclass
class ProcessingStats:
    """Processing performance statistics."""

    total_rows_processed: int = 0
    total_chunks_processed: int = 0
    processing_time_seconds: float = 0.0
    peak_memory_mb: float = 0.0
    sheets_processed: int = 0
    errors_encountered: int = 0


class ChunkIterator:
    """
    Iterator for processing Excel data in chunks.

    Provides memory-efficient chunked access to Excel worksheet data,
    with configurable chunk sizes and automatic resource management.
    """

    def __init__(
        self,
        worksheet: Worksheet,
        chunk_size: int,
        max_row: Optional[int] = None,
        values_only: bool = True,
    ):
        """
        Initialize chunk iterator.

        Args:
            worksheet: OpenPyXL worksheet object
            chunk_size: Number of rows per chunk
            max_row: Maximum row to process (None for all rows)
            values_only: Whether to return only values (not cell objects)
        """
        self.worksheet = worksheet
        self.chunk_size = chunk_size
        self.max_row = max_row or worksheet.max_row
        self.values_only = values_only
        self._current_row = 1

    def __iter__(self) -> Iterator[Tuple[List[List[Any]], ChunkInfo]]:
        """Iterate over data chunks."""
        while self._current_row <= self.max_row:
            chunk_data = []
            start_row = self._current_row
            end_row = min(self._current_row + self.chunk_size - 1, self.max_row)

            # Extract chunk data
            for row in islice(
                self.worksheet.iter_rows(values_only=self.values_only),
                start_row - 1,
                end_row,
            ):
                # Convert row to list and filter None values at the end
                row_data = list(row) if row else []
                while row_data and row_data[-1] is None:
                    row_data.pop()
                chunk_data.append(row_data)

            # Create chunk info
            chunk_info = ChunkInfo(
                chunk_index=(start_row - 1) // self.chunk_size,
                start_row=start_row,
                end_row=end_row,
                row_count=len(chunk_data),
                sheet_name=self.worksheet.title,
            )

            # Update position for next iteration
            self._current_row = end_row + 1

            # Only yield non-empty chunks
            if chunk_data:
                yield chunk_data, chunk_info

    def estimate_total_chunks(self) -> int:
        """
        Estimate total number of chunks.

        Returns:
            Estimated total chunks
        """
        total_rows = self.max_row
        return (total_rows + self.chunk_size - 1) // self.chunk_size


class StreamingExcelProcessor:
    """
    High-performance streaming Excel file processor.

    Provides memory-efficient processing of large Excel files with
    chunked data access, parallel sheet processing, and comprehensive
    progress tracking.
    """

    def __init__(self, file_path: str, config: Optional[ExcelConfig] = None):
        """
        Initialize streaming Excel processor.

        Args:
            file_path: Path to Excel file
            config: Optional configuration object
        """
        self.file_path = Path(file_path)
        self.config = config or ExcelConfig()

        # Validate file
        self._validate_file()

        # Initialize components
        self.workbook: Optional[Workbook] = None
        self.memory_monitor: Optional[MemoryMonitor] = None
        self.table_detector = TableDetector(self.config)
        self.processing_stats = ProcessingStats()

        # Configure logging
        structlog.configure(
            processors=[
                structlog.stdlib.filter_by_level,
                structlog.stdlib.add_logger_name,
                structlog.stdlib.add_log_level,
                structlog.stdlib.PositionalArgumentsFormatter(),
                structlog.processors.TimeStamper(fmt="iso"),
                structlog.processors.StackInfoRenderer(),
                structlog.processors.format_exc_info,
                structlog.processors.UnicodeDecoder(),
                structlog.processors.JSONRenderer(),
            ],
            context_class=dict,
            logger_factory=structlog.stdlib.LoggerFactory(),
            wrapper_class=structlog.stdlib.BoundLogger,
            cache_logger_on_first_use=True,
        )

        logger.info(
            "Streaming Excel processor initialized",
            file_path=str(self.file_path),
            streaming_enabled=self.config.streaming_enabled,
            chunk_size=self.config.chunk_size,
            memory_limit_mb=self.config.memory_limit_mb,
        )

    def _validate_file(self) -> None:
        """Validate the Excel file."""
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {self.file_path}")

        if self.file_path.suffix.lower() not in self.config.allowed_extensions:
            raise ValueError(f"Invalid file extension: {self.file_path.suffix}")

        file_size_mb = self.file_path.stat().st_size / (1024 * 1024)
        if file_size_mb > self.config.max_file_size_mb:
            raise ValueError(
                f"File too large: {file_size_mb:.1f}MB. "
                f"Maximum allowed: {self.config.max_file_size_mb}MB"
            )

        logger.debug("File validation passed", file_path=str(self.file_path))

    def _initialize_workbook(self) -> None:
        """Initialize workbook in read-only mode."""
        try:
            logger.debug(f"Opening workbook: {self.file_path}")

            self.workbook = load_workbook(
                filename=str(self.file_path),
                read_only=True,
                data_only=True,
                keep_links=False,  # Improve performance
            )

            # Initialize memory monitoring
            if self.config.streaming_enabled:
                resource_limits = ResourceLimits(
                    max_memory_mb=self.config.memory_limit_mb,
                    gc_threshold_mb=self.config.chunk_size * 0.1,  # 10% of chunk size
                )
                self.memory_monitor = MemoryMonitor(resource_limits)
                self.memory_monitor.start_monitoring()

            logger.info(
                "Workbook opened successfully",
                sheets=len(self.workbook.sheetnames),
                file_size_mb=self.file_path.stat().st_size / (1024 * 1024),
            )

        except Exception as e:
            logger.error("Failed to initialize workbook", error=str(e))
            raise WorkbookInitializationException(
                f"Failed to initialize workbook: {e}", str(self.file_path)
            ) from e

    def close(self) -> None:
        """Close workbook and cleanup resources."""
        if self.workbook:
            try:
                self.workbook.close()
                logger.debug("Workbook closed")
            except Exception as e:
                logger.warning(f"Error closing workbook: {e}")
            finally:
                self.workbook = None

        if self.memory_monitor:
            self.memory_monitor.stop_monitoring()
            self.memory_monitor = None

        # Force garbage collection
        gc.collect()

    def __enter__(self) -> "StreamingExcelProcessor":
        """Context manager entry."""
        self._initialize_workbook()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        """Context manager exit."""
        self.close()

    def discover_sheets(self) -> List[str]:
        """
        Discover all worksheets in the workbook.

        Returns:
            List of worksheet names
        """
        if not self.workbook:
            self._initialize_workbook()

        if not self.workbook:
            raise WorkbookInitializationException("Workbook not initialized")

        sheet_names = list(self.workbook.sheetnames)
        logger.info(f"Discovered {len(sheet_names)} sheets", sheets=sheet_names)
        return sheet_names

    def process_sheet_chunked(
        self, sheet_name: str, progress_tracker: Optional[ProgressTracker] = None
    ) -> Iterator[Tuple[List[List[Any]], ChunkInfo]]:
        """
        Process a worksheet in chunks.

        Args:
            sheet_name: Name of worksheet to process
            progress_tracker: Optional progress tracker

        Yields:
            Tuples of (chunk_data, chunk_info)
        """
        if not self.workbook:
            self._initialize_workbook()

        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            raise WorksheetNotFoundException(
                f"Worksheet '{sheet_name}' not found", str(self.file_path)
            )

        worksheet = self.workbook[sheet_name]
        logger.info(f"Processing sheet '{sheet_name}' in chunks")

        # Create chunk iterator
        max_row = (
            worksheet.max_row
            if worksheet.max_row is not None
            else self.config.max_row_count
        )
        chunk_iterator = ChunkIterator(
            worksheet=worksheet,
            chunk_size=self.config.chunk_size,
            max_row=min(max_row, self.config.max_row_count),
            values_only=True,
        )

        # Initialize progress tracking
        if progress_tracker:
            total_chunks = chunk_iterator.estimate_total_chunks()
            progress_tracker.total_items = total_chunks
            progress_tracker.start()

        # Process chunks
        for chunk_data, chunk_info in chunk_iterator:
            # Check memory limits
            if self.memory_monitor and not self.memory_monitor.check_memory_limits():
                logger.error("Memory limits exceeded during chunk processing")
                raise MemoryError("Memory limits exceeded")

            # Update progress
            if progress_tracker:
                progress_tracker.update(
                    chunk_info.chunk_index + 1,
                    f"Processing chunk {chunk_info.chunk_index + 1} "
                    f"(rows {chunk_info.start_row}-{chunk_info.end_row})",
                )

            # Update statistics
            self.processing_stats.total_rows_processed += chunk_info.row_count
            self.processing_stats.total_chunks_processed += 1

            yield chunk_data, chunk_info

            # Force garbage collection after each chunk
            gc.collect()

        # Complete progress tracking
        if progress_tracker:
            progress_tracker.complete(f"Completed processing sheet '{sheet_name}'")

        self.processing_stats.sheets_processed += 1
        logger.info(
            f"Completed chunked processing of sheet '{sheet_name}'",
            total_rows=self.processing_stats.total_rows_processed,
            total_chunks=self.processing_stats.total_chunks_processed,
        )

    def extract_sheet_data_streaming(
        self, sheet_name: str, progress_tracker: Optional[ProgressTracker] = None
    ) -> SheetData:
        """
        Extract sheet data using streaming approach.

        Args:
            sheet_name: Name of worksheet to process
            progress_tracker: Optional progress tracker

        Returns:
            SheetData object with extracted information
        """
        if not self.workbook:
            self._initialize_workbook()

        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            raise WorksheetNotFoundException(
                f"Worksheet '{sheet_name}' not found", str(self.file_path)
            )

        worksheet = self.workbook[sheet_name]
        start_time = time.time()

        # Initialize result containers
        all_data: List[List[Any]] = []
        row_count = 0
        col_count = 0
        has_data = False

        # Process sheet in chunks
        for chunk_data, chunk_info in self.process_sheet_chunked(
            sheet_name, progress_tracker
        ):
            # Process chunk data
            for row in chunk_data:
                if row:  # Only add non-empty rows
                    all_data.append(row)
                    row_count = chunk_info.end_row
                    col_count = max(col_count, len(row))
                    has_data = True

            # Check memory limits
            if self.memory_monitor and not self.memory_monitor.check_memory_limits():
                logger.warning(
                    "Memory pressure detected, but continuing with streaming"
                )

        # Extract metadata
        metadata = {
            "title": worksheet.title,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "sheet_state": getattr(worksheet, "sheet_state", "visible"),
            "processing_stats": {
                "chunks_processed": self.processing_stats.total_chunks_processed,
                "processing_time_seconds": time.time() - start_time,
                "streaming_mode": True,
            },
        }

        # Extract formal tables (if any) - in read-only mode this may be limited
        tables = []
        if hasattr(worksheet, "tables") and worksheet.tables:
            for table in worksheet.tables.values():
                tables.append(
                    {
                        "name": table.name,
                        "ref": table.ref,
                        "headerRowCount": table.headerRowCount,
                        "insertRow": getattr(table, "insertRow", False),
                        "totalsRowCount": getattr(table, "totalsRowCount", 0),
                    }
                )

        # Create sheet data
        sheet_data = SheetData(
            sheet_name=sheet_name,
            tables=tables,
            metadata=metadata,
            raw_data=all_data,
            row_count=row_count,
            col_count=col_count,
            has_data=has_data,
        )

        logger.info(
            f"Extracted streaming data from sheet '{sheet_name}'",
            rows=row_count,
            cols=col_count,
            tables=len(tables),
            processing_time=time.time() - start_time,
        )

        return sheet_data

    def process_all_sheets_parallel(
        self, progress_tracker: Optional[ProgressTracker] = None
    ) -> List[SheetData]:
        """
        Process all sheets in parallel.

        Args:
            progress_tracker: Optional progress tracker

        Returns:
            List of SheetData objects for all sheets
        """
        if not self.config.parallel_processing:
            # Fallback to sequential processing
            return self.process_all_sheets_sequential(progress_tracker)

        if not self.workbook:
            self._initialize_workbook()

        sheet_names = self.discover_sheets()
        if not sheet_names:
            return []

        # Create multi-progress tracker
        multi_tracker = None
        if progress_tracker:
            multi_tracker = MultiProgressTracker(
                "Processing sheets", self.config.enable_progress
            )
            multi_progress = ProgressTracker(
                "Overall progress",
                len(sheet_names),
                self.config.progress_interval,
                self.config.enable_progress,
            )

        results: List[SheetData] = []
        errors: List[Exception] = []

        def process_single_sheet(sheet_name: str) -> SheetData:
            """Process a single sheet."""
            sheet_progress = None
            if multi_tracker:
                sheet_progress = multi_tracker.add_tracker(sheet_name, 1)

            try:
                return self.extract_sheet_data_streaming(sheet_name, sheet_progress)
            except Exception as e:
                logger.error(f"Error processing sheet '{sheet_name}': {e}")
                raise

        # Process sheets in parallel
        with ThreadPoolExecutor(max_workers=self.config.max_workers) as executor:
            # Submit all tasks
            future_to_sheet = {
                executor.submit(process_single_sheet, sheet_name): sheet_name
                for sheet_name in sheet_names
            }

            # Collect results as they complete
            for future in as_completed(future_to_sheet):
                sheet_name = future_to_sheet[future]
                try:
                    sheet_data = future.result()
                    results.append(sheet_data)

                    if multi_tracker:
                        multi_progress.increment(1, f"Completed sheet '{sheet_name}'")

                except Exception as e:
                    errors.append(e)
                    self.processing_stats.errors_encountered += 1
                    logger.error(f"Failed to process sheet '{sheet_name}': {e}")

        # Complete progress tracking
        if multi_tracker:
            multi_progress.complete(f"Processed {len(results)} sheets successfully")

        # Handle errors
        if errors and self.config.raise_on_validation_error:
            raise DataExtractionException(
                f"Failed to process {len(errors)} sheets",
                str(self.file_path),
                context={"errors": [str(e) for e in errors]},
            )

        logger.info(
            "Parallel sheet processing completed",
            sheets_processed=len(results),
            errors=len(errors),
            total_rows_processed=self.processing_stats.total_rows_processed,
        )

        return results

    def process_all_sheets_sequential(
        self, progress_tracker: Optional[ProgressTracker] = None
    ) -> List[SheetData]:
        """
        Process all sheets sequentially.

        Args:
            progress_tracker: Optional progress tracker

        Returns:
            List of SheetData objects for all sheets
        """
        sheet_names = self.discover_sheets()
        if not sheet_names:
            return []

        results: List[SheetData] = []

        # Create progress tracker for all sheets
        overall_progress = None
        if progress_tracker:
            overall_progress = ProgressTracker(
                "Processing all sheets",
                len(sheet_names),
                self.config.progress_interval,
                self.config.enable_progress,
            )
            overall_progress.start()

        # Process each sheet
        for i, sheet_name in enumerate(sheet_names):
            try:
                sheet_progress = None
                if overall_progress:
                    sheet_progress = ProgressTracker(
                        f"Sheet {sheet_name}",
                        1,
                        self.config.progress_interval,
                        False,  # Don't show individual sheet progress
                    )
                    sheet_progress.start()

                sheet_data = self.extract_sheet_data_streaming(
                    sheet_name, sheet_progress
                )
                results.append(sheet_data)

                if overall_progress:
                    overall_progress.increment(1, f"Completed '{sheet_name}'")

            except Exception as e:
                logger.error(f"Error processing sheet '{sheet_name}': {e}")
                self.processing_stats.errors_encountered += 1

                if self.config.raise_on_validation_error:
                    raise

        # Complete progress tracking
        if overall_progress:
            overall_progress.complete(f"Processed {len(results)} sheets")

        logger.info(
            "Sequential sheet processing completed",
            sheets_processed=len(results),
            total_rows_processed=self.processing_stats.total_rows_processed,
        )

        return results

    def get_processing_stats(self) -> ProcessingStats:
        """
        Get current processing statistics.

        Returns:
            ProcessingStats object with performance metrics
        """
        if self.memory_monitor:
            stats = self.memory_monitor.get_memory_stats()
            self.processing_stats.peak_memory_mb = stats.peak_mb

        return self.processing_stats

    def detect_tables_streaming(
        self, sheet_name: str, progress_tracker: Optional[ProgressTracker] = None
    ) -> List[TableInfo]:
        """
        Detect tables in a worksheet using streaming approach.

        Args:
            sheet_name: Name of worksheet to analyze
            progress_tracker: Optional progress tracker

        Returns:
            List of detected tables
        """
        # For streaming mode, we'll process the sheet and detect tables
        # This is a simplified approach since full table detection
        # requires more complex analysis
        sheet_data = self.extract_sheet_data_streaming(sheet_name, progress_tracker)

        # Use the table detector on the collected data
        # Note: This is not fully streaming yet, but provides the
        # same interface as the original processor
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return []

        worksheet = self.workbook[sheet_name]
        return self.table_detector.detect_tables(worksheet)
