"""
Excel file processing engine with hybrid table detection and multi-sheet support.

This module provides the main ExcelReader class for comprehensive Excel file
analysis including sheet discovery, table detection, and data extraction
optimized for PDF generation. Includes streaming capabilities for large files.
"""

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Dict, Any, Iterator, TYPE_CHECKING

if TYPE_CHECKING:
    from .streaming_excel_processor import StreamingExcelProcessor

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .exceptions import (
    ExcelReaderError,
    InvalidFileException,
    WorkbookException,
    WorksheetNotFoundException,
    DataExtractionException,
    WorkbookInitializationException,
)
from .config.excel_config import ExcelConfig, DEFAULT_CONFIG
from .table_detector import TableDetector, TableInfo
from .data_validator import DataValidator, ValidationResult, ValidationRule
from .progress_tracker import ProgressTracker
from .cache_manager import get_global_cache

# Configure logging
logger = logging.getLogger(__name__)


@dataclass
class SheetData:
    """
    Data structure containing extracted worksheet information.

    Contains tables, metadata, and raw data extracted from a worksheet.
    """

    sheet_name: str
    tables: List[Dict[str, Any]]
    metadata: Dict[str, Any]
    raw_data: List[List[Any]]
    row_count: int
    col_count: int
    has_data: bool


class ExcelReader:
    """
    Excel file processing engine with hybrid table detection and multi-sheet support.

    Provides comprehensive Excel file analysis capabilities including sheet discovery,
    table detection, and data extraction optimized for PDF generation.

    Attributes:
        file_path: Path to the Excel file
        config: Configuration object for processing options
        workbook: OpenPyXL workbook instance
        _is_read_only: Flag indicating read-only mode usage
    """

    def __init__(self, file_path: str, config: Optional[ExcelConfig] = None) -> None:
        """
        Initialize ExcelReader with file path and optional configuration.

        Args:
            file_path: Path to the Excel file (.xlsx format)
            config: Optional configuration object for processing settings

        Raises:
            FileNotFoundError: If the specified file does not exist
            InvalidFileException: If the file is not a valid Excel file

        Example:
            >>> reader = ExcelReader("data.xlsx")
            >>> sheets = reader.discover_sheets()
        """
        self.file_path = Path(file_path)
        self.config = config or DEFAULT_CONFIG

        # Validate file existence and format
        self._validate_file()

        # Initialize workbook
        self.workbook: Optional[Workbook] = None
        self._is_read_only: bool = False

        # Set up logging
        logging.basicConfig(level=getattr(logging, self.config.log_level))

        # Initialize table detector
        self.table_detector = TableDetector(self.config)

        # Initialize data validator
        self.data_validator = DataValidator(self.config)

        # Initialize streaming processor for large files
        self.streaming_processor: Optional["StreamingExcelProcessor"] = None

        # Initialize cache
        self.cache = get_global_cache() if self.config.enable_caching else None

        logger.info(f"Initialized ExcelReader for file: {self.file_path}")

    def _validate_file(self) -> None:
        """
        Validate the Excel file before processing.

        Raises:
            InvalidFileException: If file validation fails
        """
        # Check file existence
        if not self.file_path.exists():
            raise InvalidFileException(
                f"File not found: {self.file_path}", str(self.file_path)
            )

        # Check file extension
        if self.file_path.suffix.lower() not in self.config.allowed_extensions:
            raise InvalidFileException(
                f"Invalid file extension: {self.file_path.suffix}. "
                f"Allowed extensions: {self.config.allowed_extensions}",
                str(self.file_path),
            )

        # Check file size
        file_size_mb = self.file_path.stat().st_size / (1024 * 1024)
        if file_size_mb > self.config.max_file_size_mb:
            raise InvalidFileException(
                f"File too large: {file_size_mb:.1f}MB. "
                f"Maximum allowed: {self.config.max_file_size_mb}MB",
                str(self.file_path),
                context={
                    "file_size_mb": file_size_mb,
                    "max_size_mb": self.config.max_file_size_mb,
                },
            )

        # Check if file is empty
        if self.file_path.stat().st_size == 0:
            raise InvalidFileException(
                f"File is empty: {self.file_path}", str(self.file_path)
            )

        logger.debug(f"File validation passed: {self.file_path}")

    def _initialize_workbook(self, read_only: bool = True) -> None:
        """
        Initialize the workbook with optimal memory settings.

        Args:
            read_only: Whether to open workbook in read-only mode for memory efficiency

        Raises:
            WorkbookInitializationException: If unable to initialize workbook

        Example:
            >>> reader = ExcelReader("large_file.xlsx")
            >>> # Uses read-only mode by default for memory efficiency
        """
        try:
            logger.debug(f"Opening workbook: {self.file_path} (read_only={read_only})")

            self.workbook = load_workbook(
                filename=str(self.file_path),
                read_only=read_only,
                data_only=True,  # Get values, not formulas
            )
            self._is_read_only = read_only

            logger.info(
                f"Workbook opened successfully with {len(self.workbook.sheetnames)} sheets"
            )

        except Exception as e:
            error_msg = f"Failed to initialize workbook: {e}"
            logger.error(
                error_msg, extra={"file_path": str(self.file_path), "error": str(e)}
            )
            raise WorkbookInitializationException(
                error_msg,
                str(self.file_path),
                context={"read_only": read_only, "original_error": str(e)},
            ) from e

    def close(self) -> None:
        """
        Close the workbook and release resources.

        Important for memory management, especially in read-only mode.

        Example:
            >>> reader = ExcelReader("data.xlsx")
            >>> # ... process data ...
            >>> reader.close()  # Release resources
        """
        if self.workbook is not None:
            try:
                self.workbook.close()
                logger.debug("Workbook closed successfully")
            except Exception as e:
                logger.warning(f"Error closing workbook: {e}")
            finally:
                self.workbook = None
                self._is_read_only = False

    def __enter__(self) -> "ExcelReader":
        """Context manager entry."""
        self._initialize_workbook(self.config.read_only_mode)
        return self

    def __exit__(
        self,
        exc_type: Optional[type],
        exc_val: Optional[Exception],
        exc_tb: Optional[object],
    ) -> None:
        """Context manager exit."""
        self.close()

    def discover_sheets(self) -> List[str]:
        """
        Discover all worksheets in the Excel file.

        Returns:
            List of worksheet names in order of appearance

        Raises:
            WorkbookException: If unable to access workbook sheets

        Example:
            >>> reader = ExcelReader("data.xlsx")
            >>> sheets = reader.discover_sheets()
            >>> print(sheets)  # ['Sheet1', 'Sheet2', 'Data']
        """
        if self.workbook is None:
            self._initialize_workbook(self.config.read_only_mode)

        try:
            if self.workbook is None:
                raise WorkbookException("Workbook not initialized", str(self.file_path))

            sheet_names = list(self.workbook.sheetnames)
            logger.info(f"Discovered {len(sheet_names)} sheets: {sheet_names}")
            return sheet_names

        except Exception as e:
            error_msg = f"Failed to discover sheets: {e}"
            logger.error(
                error_msg, extra={"file_path": str(self.file_path), "error": str(e)}
            )
            raise WorkbookException(
                error_msg, str(self.file_path), context={"original_error": str(e)}
            ) from e

    def extract_sheet_data(self, sheet_name: str) -> SheetData:
        """
        Extract all data and table information from a specific worksheet.

        Args:
            sheet_name: Name of the worksheet to process

        Returns:
            SheetData object containing tables, metadata, and raw data

        Raises:
            WorksheetNotFoundException: If sheet_name does not exist
            DataExtractionException: If unable to extract sheet data

        Example:
            >>> reader = ExcelReader("data.xlsx")
            >>> sheet_data = reader.extract_sheet_data("Sheet1")
            >>> print(f"Found {len(sheet_data.tables)} tables")
        """
        if self.workbook is None:
            self._initialize_workbook(self.config.read_only_mode)

        try:
            if self.workbook is None:
                raise WorkbookException("Workbook not initialized", str(self.file_path))

            # Check if sheet exists
            if sheet_name not in self.workbook.sheetnames:
                raise WorksheetNotFoundException(
                    f"Worksheet '{sheet_name}' not found",
                    str(self.file_path),
                    context={"available_sheets": self.workbook.sheetnames},
                )

            # Get worksheet
            worksheet = self.workbook[sheet_name]

            # Extract basic metadata
            metadata = {
                "title": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "sheet_state": getattr(worksheet, "sheet_state", "visible"),
                "page_setup": {
                    "orientation": (
                        getattr(worksheet.page_setup, "orientation", None)
                        if hasattr(worksheet, "page_setup") and worksheet.page_setup
                        else None
                    ),
                    "paper_size": (
                        getattr(worksheet.page_setup, "paperSize", None)
                        if hasattr(worksheet, "page_setup") and worksheet.page_setup
                        else None
                    ),
                },
            }

            # Extract raw data (limited for memory efficiency)
            raw_data = []
            row_count = 0
            col_count = 0
            has_data = False

            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                if row_idx > self.config.max_row_count:
                    logger.warning(
                        f"Reached maximum row limit ({self.config.max_row_count})"
                    )
                    break

                # Convert row to list and filter None values at the end
                row_data = list(row)
                while row_data and row_data[-1] is None:
                    row_data.pop()

                if row_data:  # Only add non-empty rows
                    raw_data.append(row_data)
                    row_count = row_idx
                    col_count = max(col_count, len(row_data))
                    has_data = True

            # Extract formal tables (if any)
            tables = []
            if hasattr(worksheet, "tables") and worksheet.tables:
                for table in worksheet.tables.values():
                    tables.append(
                        {
                            "name": table.name,
                            "ref": table.ref,
                            "headerRowCount": table.headerRowCount,
                            "insertRow": table.insertRow,
                            "totalsRowCount": table.totalsRowCount,
                        }
                    )

            sheet_data = SheetData(
                sheet_name=sheet_name,
                tables=tables,
                metadata=metadata,
                raw_data=raw_data,
                row_count=row_count,
                col_count=col_count,
                has_data=has_data,
            )

            logger.info(
                f"Extracted data from sheet '{sheet_name}': "
                f"{row_count} rows, {col_count} cols, {len(tables)} tables"
            )

            return sheet_data

        except WorksheetNotFoundException:
            raise  # Re-raise as-is
        except Exception as e:
            error_msg = f"Failed to extract data from sheet '{sheet_name}': {e}"
            logger.error(
                error_msg,
                extra={
                    "file_path": str(self.file_path),
                    "sheet_name": sheet_name,
                    "error": str(e),
                },
            )
            raise DataExtractionException(
                error_msg,
                str(self.file_path),
                context={"sheet_name": sheet_name, "original_error": str(e)},
            ) from e

    def detect_tables(self, sheet_name: str) -> List[TableInfo]:
        """
        Detect all tables in a specific worksheet using hybrid detection.

        Args:
            sheet_name: Name of the worksheet to analyze

        Returns:
            List of detected tables with metadata

        Raises:
            WorksheetNotFoundException: If sheet_name does not exist
            DataExtractionException: If table detection fails

        Example:
            >>> reader = ExcelReader("data.xlsx")
            >>> tables = reader.detect_tables("Sheet1")
            >>> for table in tables:
            ...     print(f"Table '{table.name}' detected using {table.detection_method}")
        """
        if self.workbook is None:
            self._initialize_workbook(self.config.read_only_mode)

        try:
            if self.workbook is None:
                raise WorkbookException("Workbook not initialized", str(self.file_path))

            # Check if sheet exists
            if sheet_name not in self.workbook.sheetnames:
                raise WorksheetNotFoundException(
                    f"Worksheet '{sheet_name}' not found",
                    str(self.file_path),
                    context={"available_sheets": self.workbook.sheetnames},
                )

            # Get worksheet
            worksheet = self.workbook[sheet_name]

            # Detect tables using hybrid approach
            tables = self.table_detector.detect_tables(worksheet)

            logger.info(f"Detected {len(tables)} tables in sheet '{sheet_name}'")
            return tables

        except WorksheetNotFoundException:
            raise  # Re-raise as-is
        except Exception as e:
            error_msg = f"Failed to detect tables in sheet '{sheet_name}': {e}"
            logger.error(
                error_msg,
                extra={
                    "file_path": str(self.file_path),
                    "sheet_name": sheet_name,
                    "error": str(e),
                },
            )
            raise DataExtractionException(
                error_msg,
                str(self.file_path),
                context={"sheet_name": sheet_name, "original_error": str(e)},
            ) from e

    def validate_sheet_data(
        self, sheet_name: str, validation_rules: Optional[List[ValidationRule]] = None
    ) -> ValidationResult:
        """
        Validate data in a specific worksheet using validation rules.

        Args:
            sheet_name: Name of the worksheet to validate
            validation_rules: Optional validation rules for each column

        Returns:
            ValidationResult with validation details

        Raises:
            WorksheetNotFoundException: If sheet_name does not exist
            DataExtractionException: If validation fails critically

        Example:
            >>> reader = ExcelReader("data.xlsx")
            >>> result = reader.validate_sheet_data("Sheet1")
            >>> print(f"Validation result: {result.is_valid}, confidence: {result.confidence_score}")
        """
        # Extract sheet data first
        sheet_data = self.extract_sheet_data(sheet_name)

        # Validate the data
        return self.data_validator.validate_table_data(
            data=sheet_data.raw_data,
            headers=sheet_data.raw_data[0] if sheet_data.raw_data else [],
            rules=validation_rules,
        )

    def _should_use_streaming(self) -> bool:
        """
        Determine if streaming should be used for this file.

        Returns:
            True if streaming should be used
        """
        # Use streaming if explicitly enabled or file is large
        file_size_mb = self.file_path.stat().st_size / (1024 * 1024)
        return (
            self.config.streaming_enabled or file_size_mb > 50
        )  # Use streaming for files > 50MB

    def _get_streaming_processor(self) -> "StreamingExcelProcessor":
        """
        Get or create streaming processor.

        Returns:
            StreamingExcelProcessor instance
        """
        if self.streaming_processor is None:
            # Import here to avoid circular import
            from .streaming_excel_processor import StreamingExcelProcessor

            self.streaming_processor = StreamingExcelProcessor(
                str(self.file_path), self.config
            )
        return self.streaming_processor

    def extract_sheet_data_optimized(
        self, sheet_name: str, progress_tracker: Optional[ProgressTracker] = None
    ) -> SheetData:
        """
        Extract sheet data using optimal method (streaming or traditional).

        Args:
            sheet_name: Name of the worksheet to process
            progress_tracker: Optional progress tracker

        Returns:
            SheetData object with extracted information

        Raises:
            WorksheetNotFoundException: If sheet_name does not exist
            DataExtractionException: If unable to extract sheet data
        """
        # Check cache first
        if self.cache:
            cache_key = [
                "sheet_data",
                str(self.file_path),
                sheet_name,
                str(self.config),
            ]
            cached_result = self.cache.get(cache_key)
            if cached_result:
                logger.info(f"Using cached data for sheet '{sheet_name}'")
                return cached_result

        # Choose extraction method
        if self._should_use_streaming():
            logger.info(f"Using streaming extraction for sheet '{sheet_name}'")
            processor = self._get_streaming_processor()
            result = processor.extract_sheet_data_streaming(
                sheet_name, progress_tracker
            )
        else:
            logger.info(f"Using traditional extraction for sheet '{sheet_name}'")
            result = self.extract_sheet_data(sheet_name)

        # Cache result
        if self.cache and result:
            cache_key = [
                "sheet_data",
                str(self.file_path),
                sheet_name,
                str(self.config),
            ]
            self.cache.put(cache_key, result)

        return result

    def process_all_sheets_optimized(
        self, progress_tracker: Optional[ProgressTracker] = None
    ) -> List[SheetData]:
        """
        Process all sheets using optimal method.

        Args:
            progress_tracker: Optional progress tracker

        Returns:
            List of SheetData objects for all sheets
        """
        # Check cache first
        if self.cache:
            cache_key = ["all_sheets", str(self.file_path), str(self.config)]
            cached_result = self.cache.get(cache_key)
            if cached_result:
                logger.info("Using cached data for all sheets")
                return cached_result

        # Choose processing method
        if self._should_use_streaming():
            logger.info("Using streaming processing for all sheets")
            processor = self._get_streaming_processor()

            if self.config.parallel_processing:
                result = processor.process_all_sheets_parallel(progress_tracker)
            else:
                result = processor.process_all_sheets_sequential(progress_tracker)
        else:
            logger.info("Using traditional processing for all sheets")
            sheet_names = self.discover_sheets()
            result = []

            for sheet_name in sheet_names:
                sheet_data = self.extract_sheet_data(sheet_name)
                result.append(sheet_data)

        # Cache result
        if self.cache and result:
            cache_key = ["all_sheets", str(self.file_path), str(self.config)]
            self.cache.put(cache_key, result)

        return result

    def detect_tables_optimized(
        self, sheet_name: str, progress_tracker: Optional[ProgressTracker] = None
    ) -> List[TableInfo]:
        """
        Detect tables using optimal method.

        Args:
            sheet_name: Name of the worksheet to analyze
            progress_tracker: Optional progress tracker

        Returns:
            List of detected tables with metadata
        """
        # Check cache first
        if self.cache:
            cache_key = ["tables", str(self.file_path), sheet_name, str(self.config)]
            cached_result = self.cache.get(cache_key)
            if cached_result:
                logger.info(f"Using cached table detection for sheet '{sheet_name}'")
                return cached_result

        # Choose detection method
        if self._should_use_streaming():
            logger.info(f"Using streaming table detection for sheet '{sheet_name}'")
            processor = self._get_streaming_processor()
            result = processor.detect_tables_streaming(sheet_name, progress_tracker)
        else:
            logger.info(f"Using traditional table detection for sheet '{sheet_name}'")
            result = self.detect_tables(sheet_name)

        # Cache result
        if self.cache and result:
            cache_key = ["tables", str(self.file_path), sheet_name, str(self.config)]
            self.cache.put(cache_key, result)

        return result

    def get_file_info(self) -> Dict[str, Any]:
        """
        Get comprehensive file information.

        Returns:
            Dictionary with file metadata and statistics
        """
        file_stats = self.file_path.stat()
        file_size_mb = file_stats.st_size / (1024 * 1024)

        # Get sheet information
        sheet_names = self.discover_sheets()

        info = {
            "file_path": str(self.file_path),
            "file_size_mb": file_size_mb,
            "file_size_bytes": file_stats.st_size,
            "last_modified": file_stats.st_mtime,
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
            "streaming_enabled": self._should_use_streaming(),
            "parallel_processing": self.config.parallel_processing,
            "chunk_size": self.config.chunk_size,
            "memory_limit_mb": self.config.memory_limit_mb,
            "cache_enabled": self.config.enable_caching,
        }

        # Add processing statistics if available
        if self.streaming_processor:
            processing_stats = self.streaming_processor.get_processing_stats()
            info["processing_stats"] = {
                "total_rows_processed": processing_stats.total_rows_processed,
                "total_chunks_processed": processing_stats.total_chunks_processed,
                "processing_time_seconds": processing_stats.processing_time_seconds,
                "peak_memory_mb": processing_stats.peak_memory_mb,
                "sheets_processed": processing_stats.sheets_processed,
                "errors_encountered": processing_stats.errors_encountered,
            }

        return info

    def clear_cache(self) -> None:
        """Clear cached data for this file."""
        if self.cache:
            # Invalidate all cache entries for this file
            cache_keys = [
                ["sheet_data", str(self.file_path)],
                ["all_sheets", str(self.file_path)],
                ["tables", str(self.file_path)],
            ]

            for base_key in cache_keys:
                # Note: This is a simplified cache invalidation
                # In a production system, you might want more sophisticated invalidation
                pass

            logger.info(f"Cache cleared for file: {self.file_path}")

    def close(self) -> None:
        """
        Close the workbook and release resources.

        Important for memory management, especially in read-only mode.
        """
        # Close streaming processor if active
        if self.streaming_processor:
            self.streaming_processor.close()
            self.streaming_processor = None

        # Close traditional workbook if active
        if self.workbook is not None:
            try:
                self.workbook.close()
                logger.debug("Workbook closed successfully")
            except Exception as e:
                logger.warning(f"Error closing workbook: {e}")
            finally:
                self.workbook = None
                self._is_read_only = False

        logger.info("ExcelReader resources released")
