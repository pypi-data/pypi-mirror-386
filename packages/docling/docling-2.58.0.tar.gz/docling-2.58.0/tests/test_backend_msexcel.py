import logging
from pathlib import Path

import pytest
from openpyxl import load_workbook

from docling.backend.msexcel_backend import MsExcelDocumentBackend
from docling.datamodel.base_models import InputFormat
from docling.datamodel.document import ConversionResult, DoclingDocument, InputDocument
from docling.document_converter import DocumentConverter

from .test_data_gen_flag import GEN_TEST_DATA
from .verify_utils import verify_document, verify_export

_log = logging.getLogger(__name__)

GENERATE = GEN_TEST_DATA


def get_excel_paths():
    # Define the directory you want to search
    directory = Path("./tests/data/xlsx/")

    # List all Excel files in the directory and its subdirectories
    excel_files = sorted(directory.rglob("*.xlsx")) + sorted(directory.rglob("*.xlsm"))
    return excel_files


def get_converter():
    converter = DocumentConverter(allowed_formats=[InputFormat.XLSX])

    return converter


@pytest.fixture(scope="module")
def documents() -> list[tuple[Path, DoclingDocument]]:
    documents: list[dict[Path, DoclingDocument]] = []

    excel_paths = get_excel_paths()
    converter = get_converter()

    for excel_path in excel_paths:
        _log.debug(f"converting {excel_path}")

        gt_path = (
            excel_path.parent.parent / "groundtruth" / "docling_v2" / excel_path.name
        )

        conv_result: ConversionResult = converter.convert(excel_path)

        doc: DoclingDocument = conv_result.document

        assert doc, f"Failed to convert document from file {gt_path}"
        documents.append((gt_path, doc))

    return documents


def test_e2e_excel_conversions(documents) -> None:
    for gt_path, doc in documents:
        pred_md: str = doc.export_to_markdown()
        assert verify_export(pred_md, str(gt_path) + ".md"), "export to md"

        pred_itxt: str = doc._export_to_indented_text(
            max_text_len=70, explicit_tables=False
        )
        assert verify_export(pred_itxt, str(gt_path) + ".itxt"), (
            "export to indented-text"
        )

        assert verify_document(doc, str(gt_path) + ".json", GENERATE), (
            "document document"
        )


def test_pages(documents) -> None:
    """Test the page count and page size of converted documents.

    Args:
        documents: The paths and converted documents.
    """
    # number of pages from the backend method
    path = next(item for item in get_excel_paths() if item.stem == "xlsx_01")
    in_doc = InputDocument(
        path_or_stream=path,
        format=InputFormat.XLSX,
        filename=path.stem,
        backend=MsExcelDocumentBackend,
    )
    backend = MsExcelDocumentBackend(in_doc=in_doc, path_or_stream=path)
    assert backend.page_count() == 4

    # number of pages from the converted document
    doc = next(item for path, item in documents if path.stem == "xlsx_01")
    assert len(doc.pages) == 4

    # page sizes as number of cells
    assert doc.pages.get(1).size.as_tuple() == (3.0, 7.0)
    assert doc.pages.get(2).size.as_tuple() == (9.0, 18.0)
    assert doc.pages.get(3).size.as_tuple() == (13.0, 36.0)
    assert doc.pages.get(4).size.as_tuple() == (0.0, 0.0)


def test_chartsheet(documents) -> None:
    """Test the conversion of Chartsheets.

    Args:
        documents: The paths and converted documents.
    """
    doc = next(item for path, item in documents if path.stem == "xlsx_03_chartsheet")
    assert len(doc.pages) == 2

    # Chartseet content is for now ignored
    assert doc.groups[1].name == "sheet: Duck Chart"
    assert doc.pages[2].size.height == 0
    assert doc.pages[2].size.width == 0


def test_inflated_rows_handling(documents) -> None:
    """Test that files with inflated max_row are handled correctly.

    xlsx_04_inflated.xlsx has inflated max_row (1,048,496) but only 7 rows of actual data.
    This test verifies that our backend correctly identifies true data bounds.
    """
    # First, verify the file has inflated max_row using openpyxl directly
    path = next(item for item in get_excel_paths() if item.stem == "xlsx_04_inflated")

    wb = load_workbook(path)
    ws = wb.active
    reported_max_row = ws.max_row

    # Assert that openpyxl reports inflated max_row
    assert reported_max_row > 100000, (
        f"xlsx_04_inflated.xlsx should have inflated max_row (expected >100k, got {reported_max_row:,}). "
        f"This test file is designed to verify proper handling of Excel files with inflated row counts."
    )

    _log.info(
        f"xlsx_04_inflated.xlsx - Openpyxl reported max_row: {reported_max_row:,}"
    )

    # Now test that our backend handles it correctly
    in_doc = InputDocument(
        path_or_stream=path,
        format=InputFormat.XLSX,
        filename=path.stem,
        backend=MsExcelDocumentBackend,
    )
    backend = MsExcelDocumentBackend(in_doc=in_doc, path_or_stream=path)

    # Verify backend detects correct number of pages (should be 4, like test-01)
    page_count = backend.page_count()
    assert page_count == 4, (
        f"Backend should detect 4 pages (same as test-01), got {page_count}"
    )

    # Verify converted document has correct pages
    doc = next(item for path, item in documents if path.stem == "xlsx_04_inflated")
    assert len(doc.pages) == 4, f"Document should have 4 pages, got {len(doc.pages)}"

    # Verify page sizes match expected dimensions (same as test-01)
    # These should reflect actual data, not inflated row counts
    assert doc.pages.get(1).size.as_tuple() == (3.0, 7.0), (
        f"Page 1 should be 3x7 cells, got {doc.pages.get(1).size.as_tuple()}"
    )
    assert doc.pages.get(2).size.as_tuple() == (9.0, 18.0), (
        f"Page 2 should be 9x18 cells, got {doc.pages.get(2).size.as_tuple()}"
    )
    assert doc.pages.get(3).size.as_tuple() == (13.0, 36.0), (
        f"Page 3 should be 13x36 cells, got {doc.pages.get(3).size.as_tuple()}"
    )
    assert doc.pages.get(4).size.as_tuple() == (0.0, 0.0), (
        f"Page 4 should be 0x0 cells (empty), got {doc.pages.get(4).size.as_tuple()}"
    )

    _log.info(
        f"✓ Successfully handled inflated max_row: "
        f"reported {reported_max_row:,} rows, "
        f"correctly processed as {page_count} pages with proper dimensions"
    )
