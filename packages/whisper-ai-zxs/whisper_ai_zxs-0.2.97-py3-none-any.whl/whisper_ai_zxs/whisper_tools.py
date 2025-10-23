from .whisper_db import WhisperDB
from .whisper_db_pool import WhisperDB_Pool
from datetime import datetime
import requests
from PIL import ImageGrab
import base64
import hashlib
import json
import time
import csv

import zipfile
import os
from openpyxl import load_workbook, Workbook
from ftplib import FTP
import glob

import logging
logger = logging.getLogger("whisper_ai")

class WhisperTools_ChatList:
    def __init__(self, name, before_time):
        """ 初始化一个空字典用于存储函数 """
        self._chat_list = []
        self._chat_name = name
        self._before_chat_time = before_time

    def add(self, chat_list):
        for item in chat_list:
            if (self._chat_name != item["name"]):
                ##if (len(self._chat_list) == 0 and self._before_chat_time == ""):
                ##    logger.error(f"第一条消息接收时，发现用户名不一致: {item['name']}， {self._chat_name}")
                if (self._before_chat_time == ""):
                    return False
                else:
                    return True
            if (item["time"] == self._before_chat_time):
                return False
            self._chat_list.append(item)
        return True

    def get(self):
        return self._chat_list
    

class WhisperTools_ChatRecord:
    @staticmethod
    def record_user_chat(kf_name, user_name, content):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        insert_query = """
            INSERT INTO openai_chat_list (chat_time, shop_name, chat_name, sender, act, content)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        
        with WhisperDB() as myDB:
            myDB.query(insert_query, (current_time, kf_name, user_name, user_name, "ask", content))
            myDB.commit()
    @staticmethod
    def record_chatGPT_action(kf_name, user_name, act, content):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        insert_query = """
            INSERT INTO openai_chat_list (chat_time, shop_name, chat_name, sender, act, content)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        
        with WhisperDB() as myDB:
            myDB.query(insert_query, (current_time, kf_name, user_name, "chatGPT", act, content))
            myDB.commit()

class WhisperTools_Qywx:
    @staticmethod
    def send_to_error_robot(msg):
        # 截取整个屏幕
        screenshot = ImageGrab.grab()
        screenshot_path = r"D:\WhisperAgent\异常截图\screenshot.png"  # 确保路径存在
        screenshot.save(screenshot_path)

        # 读取图片内容
        with open(screenshot_path, "rb") as f:
            image_data = f.read()

        # 计算 Base64 编码和 MD5 值
        image_base64 = base64.b64encode(image_data).decode("utf-8")
        image_md5 = hashlib.md5(image_data).hexdigest()

        # 企业微信 Webhook URL
        webhook_url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=f2113e16-d190-42b8-a386-6032ae7def7f'

        # 发送文本消息
        text_data = {
            "msgtype": "text",
            "text": {"content": msg}
        }
        text_response = requests.post(webhook_url, json=text_data)

        # 发送图片消息
        image_data = {
            "msgtype": "image",
            "image": {
                "base64": image_base64,
                "md5": image_md5
            }
        }
        image_response = requests.post(webhook_url, json=image_data)

        return {
            "text_response": text_response.json(),
            "image_response": image_response.json()
        }
    @staticmethod
    def send_to_kf_robot(agent, msg):
        webhook_url = WhisperTools_Qywx.get_robot_hook(agent)
        if (webhook_url != ""):
            data = {
                'msgtype': 'text',
                'text': {'content': msg}
            }
            response = requests.post(webhook_url, json=data)
            return response.text
        else:
            return "webhook_url is blank!"
    @staticmethod
    def send_to_kf_robot_report(agent, id, shop_name, user_id, type, detail, ):
        webhook_url = WhisperTools_Qywx.get_robot_hook(agent)
        if (webhook_url != ""):
            data = {
                'msgtype': 'markdown',
                "markdown": {
                    "content": f"""收到一条客户反馈的售后工单，请相关同事处理。\n
                    >编号:<font color='comment'>{id}</font>
                    >类型:<font color='warning'>{type}</font>
                    >店铺:<font color='comment'>{shop_name}</font>
                    >客户:<font color='comment'>{user_id}</font>
                    >说明:<font color='comment'>{detail}</font>\n
                    [现在处理](https://zxslife.com/qywxServer/bdManage_app/kf_service_manage/{agent.get_company_name()}?company={agent.get_company_name()}&menu=false&id={id})"""
                }
            }
            if os.getenv('TEST_ENV') != 'true':
                response = requests.post(webhook_url, json=data)
                return response.text
            else:
                return "测试环境，跳过关键功能执行。"
        else:
            return "webhook_url is blank!"
    @staticmethod
    def get_robot_hook(agent):
        # 使用 with 语句管理数据库连接
        with WhisperDB() as db:  # 自动管理连接
            # 使用参数化查询来防止 SQL 注入
            query = """
                SELECT robot_hook
                FROM openai_kf_manage
                JOIN openai_company ON openai_kf_manage.company = openai_company.name
                WHERE shop_name = %s;
            """
            # 获取查询结果
            result = db.query(query, (agent.get_kf_name(),))

        # 如果查询结果存在，则返回第一行
        return result[0][0] if result else ""

#用于解析通用的在售商品信息的工具
class WhisperTools_UploadSellingProduct:
    def __init__(self, shop_name):
        self._shop_name = shop_name

    def _save_to_csv(self, data, output_file):
        """将数据写入 CSV 文件"""
        with open(output_file, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            for row in data:
                processed_row = []
                for cell in row:
                    if isinstance(cell, list):  # 如果单元格是列表
                        processed_row.append(json.dumps(cell, ensure_ascii=False))  # 将列表转换为 JSON 字符串
                    else:
                        processed_row.append(cell)
                writer.writerow(processed_row)
        #logging.debug(f"CSV 文件已保存: {output_file}")

    def _upload_file(self, local_file_path, remote_file_path):
        # 连接到 FTP 服务器
        host = os.getenv("WHISPER_SERVER_HOST", "172.27.0.4")  # 默认用172.27.0.4，pytest环境可以修改DB_HOST
        ftp = FTP(host)
        ftp.login("ftpuser", "Zxs123")
        
        # 打开本地文件
        with open(local_file_path, 'rb') as file:
            # 使用 STOR 命令上传文件
            ftp.storbinary(f'STOR {remote_file_path}', file)
        
        # 关闭 FTP 连接
        ftp.quit()

    def _update_database(self):

        # 单独执行 USE 语句
        sql_use = "USE zxs_order;"
        sql_del = f"DELETE FROM `tm_selling_product` WHERE `shop_name` ='{self._shop_name}'"

        # 加载数据的 SQL 语句
        sql_load_data = """
            LOAD DATA INFILE 'd:\\\\ftp_server\\\\unified_selling_product.csv'
            INTO TABLE tm_selling_product
            FIELDS TERMINATED BY ','
            OPTIONALLY ENCLOSED BY '\"'
            ESCAPED BY '\\\\'
            LINES TERMINATED BY '\\r\\n'
            IGNORE 1 LINES
            (product_code, product_name, discount_plan, on_sale_sku, off_sale_sku, purchase_link, shop_name);
        """
        #logging.debug(f"sql: {sql_load_data}")

        with WhisperDB() as myDB:                
            # 执行 USE 语句
            myDB.query(sql_use)
            myDB.query(sql_del)
            # 执行加载数据的 SQL 语句
            myDB.query(sql_load_data)
            # 提交更改
            myDB.commit()

    def analyze(self, file_name):
        pass

#用于解析小红书的在售商品信息的工具
class WhisperTools_UploadSellingProduct_Red(WhisperTools_UploadSellingProduct):
    def _unzip_file(self, zip_path, extract_to='.'):  
        """解压 ZIP 文件并返回 Excel 文件路径"""
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_to)
            for file_name in zip_ref.namelist():
                if file_name.endswith('.xlsx'):
                    return os.path.join(extract_to, file_name)
        return None
        
    def _process_excel(self, file_path):
        """读取 Excel 文件，从第二行开始处理数据"""
        wb = load_workbook(file_path)
        ws = wb.active
        data = []

        head_row = [
            "商品编码","商品名称","优惠活动","在售规格","停售规格","购买链接","shop_name"
        ]
        data.append(head_row)
        
        for row in ws.iter_rows(min_row=2, values_only=True):  # 遍历第二行到最后一行
            found_flag = False
            for index, value in enumerate(data):
                if (value[0] == row[3]):
                    found_flag = True
                    if (row[8] > 0):
                        data[index][3].append(row[4])
                    else:
                        data[index][4].append(row[4])
                    break

            if (not found_flag):
                processed_row = []
                if (row[8] > 0):
                    processed_row = [
                        str(row[3]).strip() if row[3] is not None else '',
                        str(row[3]).strip() if row[3] is not None else '',
                        "",
                        [str(row[4]).strip() if row[4] is not None else ''],
                        [],
                        "",
                        self._shop_name
                    ]
                else:
                    processed_row = [
                        str(row[3]).strip() if row[3] is not None else '',
                        str(row[3]).strip() if row[3] is not None else '',
                        "",
                        [],
                        [str(row[4]).strip() if row[4] is not None else ''],
                        "",
                        self._shop_name
                    ]
                data.append(processed_row)
        
        return data
    def analyze(self, file_name):
        zip_path = file_name  # 需要解压的 ZIP 文件路径
        extract_folder = os.path.join(os.path.dirname(zip_path), "extracted_files")  # 生成解压目录
        output_file = os.path.join(os.path.dirname(zip_path), "processed_data.csv")  # 处理后生成的新 Excel 文件
        
        os.makedirs(extract_folder, exist_ok=True)
        excel_file = self._unzip_file(zip_path, extract_folder)
        
        if excel_file:
            #logging.debug(f"解压成功，找到 Excel 文件: {excel_file}")
            processed_data = self._process_excel(excel_file)
            self._save_to_csv(processed_data, output_file)
        else:
            logger.error("未找到 Excel 文件，请检查 ZIP 包内容。")
        
        self._upload_file(output_file, "unified_selling_product.csv")

        if os.getenv('TEST_ENV') != 'true':
            self._update_database()
        else:
            logger.debug("测试环境，跳过关键功能执行。")
            
#用于解析天猫的在售商品信息的工具
class WhisperTools_UploadSellingProduct_Taobao(WhisperTools_UploadSellingProduct):     
    def _get_xlsx_files(self, directory):
        """获取指定目录下的所有有效 .xlsx 文件，排除 Excel 生成的临时文件"""
        os.chdir(directory)  # 切换到指定目录
        xlsx_files = [f for f in glob.glob('*.xlsx') if not f.startswith("~$")]  # 过滤掉临时文件
        return xlsx_files

    def _process_excel(self, file_path):
        """读取 Excel 文件，从第二行开始处理数据"""
        wb = load_workbook(file_path)
        ws = wb.active
        data_count = {}
        for row in ws.iter_rows(min_row=2, values_only=True):  # 遍历第二行到最后一行
            key = row[1]
            if key in data_count:
                data_count[key] += 1
            else:
                data_count[key] = 1
        
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # 遍历第二行到最后一行
            if data_count[row[1]] == 1:
                #如果只有一条记录，直接判断库存
                processed_row = []
                selling = []
                unselling = []
                if (int(row[14]) > 0):
                    selling = [str(row[3]).strip() if row[3] is not None else '默认规格']
                else:
                    unselling = [str(row[3]).strip() if row[3] is not None else '默认规格']

                processed_row = [
                    str(row[1]).strip() if row[1] is not None else '',
                    str(row[0]).strip() if row[0] is not None else '',
                    "",
                    selling,
                    unselling,
                    f"https://detail.tmall.com/item.htm?id={row[1]}",
                    self._shop_name
                ]
                data.append(processed_row)
            else:
                if row[3] is None:
                    continue
                found_flag = False
                for index, value in enumerate(data):
                    if (value[0] == row[1]):
                        found_flag = True
                        if (int(row[14]) > 0):
                            data[index][3].append(row[3])
                        else:
                            data[index][4].append(row[3])
                        break
                if found_flag == False :
                    processed_row = []
                    selling = []
                    unselling = []
                    if (int(row[14]) > 0):
                        selling = [str(row[3]).strip() if row[3] is not None else '默认规格']
                    else:
                        unselling = [str(row[3]).strip() if row[3] is not None else '默认规格']

                    processed_row = [
                        str(row[1]).strip() if row[1] is not None else '',
                        str(row[0]).strip() if row[0] is not None else '',
                        "",
                        selling,
                        unselling,
                        f"https://detail.tmall.com/item.htm?id={row[1]}",
                        self._shop_name
                    ]
                    data.append(processed_row)

        return data
    def analyze(self, path_name):
        xlsx_files = self._get_xlsx_files(path_name)
        output_file = os.path.join(os.path.dirname(path_name), "processed_data.csv")  # 处理后生成的新 Excel 文件
        processed_data = []
        head_row = [
            "商品编码","商品名称","优惠活动","在售规格","停售规格","购买链接","shop_name"
        ]
        processed_data.append(head_row)

        for file in xlsx_files:
            logger.debug(f"开始解析文件: {file}")
            processed_data = processed_data + self._process_excel(file)

        self._save_to_csv(processed_data, output_file)
        self._upload_file(output_file, "unified_selling_product.csv")

        if os.getenv('TEST_ENV') != 'true':
            self._update_database()
        else:
            logger.debug("测试环境，跳过关键功能执行。")
            
class WhisperTools_QYWX:
    def __init__(self):
        self.corpid = 'ww068f7775fc823616'
        self.corpsecret = 'g0fEQ_QFtQ55G6Jw4lJmyTVWZ_xk-SbV8ld7l34Qz_w'
        self.corpsecret_archive = 'VEocLpZYY9QFecBrXWctmIfULGUbwLbIIrCJWkADcYc'
        self.group_map = {}
        self.token_cache = None
        self.token_expires = 0
        self.token_cache_archive = None
        self.token_expires_archive = 0
        # 优先从数据库获取群名称
        with WhisperDB_Pool() as myDB:
            sql = "SELECT roomid, group_name FROM wx_chat_groupid_map WHERE 1"
            result = myDB.query(sql)
            if result:
                for row in result:
                    roomid = row[0]
                    group_name = row[1]
                    self.group_map[roomid] = group_name

    def get_reply_msg(self):
        reply_msg = []
        with WhisperDB_Pool() as myDB:
            select_query = """
                SELECT id, `to`, msgtype, msgcontent
                FROM wx_chat_reply 
                WHERE send_time IS NULL 
                    AND create_time >= DATE_SUB(NOW(), INTERVAL 10 MINUTE)
                    AND create_time <= NOW()
                    AND action = 'reply'
                ORDER BY create_time ASC 
            """
            result = myDB.query(select_query)
            to_name = ""
            for row in result:
                to_name = self._get_group_name(row[1])
                if (to_name == row[1]):
                    to_name = self._get_user_name(row[1])

                reply_msg.append({
                    "id": row[0],
                    "to": to_name,
                    "msgtype": row[2],
                    "msgcontent": row[3]
                })
        return reply_msg
    
    def add_reply_msg(self, to, msgtype, msgcontent, action="notify"):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        insert_query = """
            INSERT INTO wx_chat_reply (`to`, msgtype, msgcontent, action, create_time, msgid)
            VALUES (%s, %s, %s, %s, %s, "")
        """
        
        with WhisperDB_Pool() as myDB:
            myDB.query(insert_query, (to, msgtype, msgcontent, action, current_time))
            myDB.commit()
        return
    
    def add_action_msg(self, msgcontent, actiontype):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        insert_query = """
            INSERT INTO wx_chat_reply (`to`, msgtype, msgcontent, action, create_time, msgid)
            VALUES (%s, %s, %s, %s, %s, "")
        """
        
        with WhisperDB_Pool() as myDB:
            myDB.query(insert_query, ("phone", "text", msgcontent, f"action:{actiontype}", current_time))
            myDB.commit()
        return
    
    def get_action_msg(self):
        reply_msg = []
        with WhisperDB_Pool() as myDB:
            select_query = """
                SELECT id, action, `to`, msgtype, msgcontent
                FROM wx_chat_reply 
                WHERE send_time IS NULL 
                    AND create_time >= DATE_SUB(NOW(), INTERVAL 10 MINUTE)
                    AND create_time <= NOW()
                    AND action LIKE %s
                ORDER BY create_time ASC 
            """
            result = myDB.query(select_query, ('action:%',))
            to_name = ""
            for row in result:
                to_name = self._get_group_name(row[2])
                if (to_name == row[2]):
                    to_name = self._get_user_name(row[2])

                reply_msg.append({
                    "id": row[0],
                    "action": row[1],
                    "to": to_name,
                    "msgtype": row[3],
                    "msgcontent": row[4]
                })
        return reply_msg
    
    def _get_group_name(self, group_id):
        # 优先从数据库获取群名称
        #with WhisperDB() as myDB:
        #    sql = "SELECT group_name FROM wx_chat_groupid_map WHERE roomid = %s"
        #    result = myDB.query(sql, (group_id,))
        #    if result:
        #        to_name = result[0][0]
        #        logging.info(f"本地数据库获取群名称成功，group_id: {group_id}, 群名称: {to_name}")
        #        return to_name
        if group_id in self.group_map:
            to_name = self.group_map[group_id]
            logging.info(f"本地缓存获取群名称成功，group_id: {group_id}, 群名称: {to_name}")
            return to_name

        # 数据库没有则调用 groupchat 接口
        url = f'https://qyapi.weixin.qq.com/cgi-bin/externalcontact/groupchat/get?access_token={self.getToken()}'
        payload = {
            "chat_id": group_id,
            "need_name": 0
        }
        response = requests.post(url, json=payload)
        response = response.json()
        if response.get("errcode", 0) == 0:
            to_name = response.get("group_chat", "").get("name", group_id)
            logging.info(f"groupchat接口获取群名称成功，group_id: {group_id}, 群名称: {to_name}")
            self.group_map[group_id] = to_name
            return to_name

        # groupchat接口失败则尝试 msgaudit 接口
        url2 = f'https://qyapi.weixin.qq.com/cgi-bin/msgaudit/groupchat/get?access_token={self.getToken_Archive()}'
        payload2 = {
            "roomid": group_id
        }
        response2 = requests.post(url2, json=payload2)
        response_data2 = response2.json()
        if response_data2.get('errcode', -1) == 0:
            to_name = response_data2.get('roomname', group_id)
            logging.info(f"msgaudit接口获取群名称成功，group_id: {group_id}, 群名称: {to_name}")
            self.group_map[group_id] = to_name
            return to_name

        # 所有接口失败则返回原始 group_id
        logging.info(f"所有接口获取群名称失败，group_id: {group_id}, 该用原始id代替")
        return group_id

    def _get_user_name(self, userid):
        user_name = ""
        if (len(userid) == 32):
            url = f'https://qyapi.weixin.qq.com/cgi-bin/externalcontact/get?access_token={self.getToken()}&external_userid={userid}'
            response = requests.get(url)
            response = response.json()
            if (response.get("errcode", 0) == 0):
                user_name = response.get("external_contact", {}).get("name", userid)
            else:
                user_name = userid
        else:
            url = f'https://qyapi.weixin.qq.com/cgi-bin/user/get?access_token={self.getToken()}&userid={userid}'
            response = requests.get(url)
            response = response.json()
            if (response.get("errcode", 0) == 0):
                user_name = response.get("name", userid)
            else:
                user_name = userid
        return user_name

    def set_reply_sended(self, id):
        with WhisperDB_Pool() as myDB:
            update_query = """
                UPDATE wx_chat_reply
                SET send_time = NOW()
                WHERE id = %s
            """
            myDB.query(update_query, (id,))
            myDB.commit()
        return
    
    def set_action_result(self, id, result):
        with WhisperDB_Pool() as myDB:
            update_query = """
                UPDATE wx_chat_reply
                SET send_time = NOW(), msgcontent = %s
                WHERE id = %s
            """
            myDB.query(update_query, (result, id))
            myDB.commit()
        return

    def getToken(self):
        if self.token_cache is None or time.time() >= self.token_expires - 10:
            sql_select = "SELECT `access_token`, `expires_time` FROM `wx_info` WHERE `token_type` = %s"
            with WhisperDB_Pool() as db:
                result = db.query(sql_select, (self.corpsecret,))
                if not result:
                    access_token = self._getToken()
                    self.token_cache = access_token
                    self.token_expires = int(time.time()) + 7200
                    sql_insert = "INSERT INTO `wx_info` (`access_token`, `expires_time`, `token_type`) VALUES (%s, %s, %s)"
                    db.query(sql_insert, (access_token, self.token_expires, self.corpsecret))
                    db.commit()
                elif result[0][1] <= (time.time() - 10):
                    access_token = self._getToken()
                    self.token_cache = access_token
                    self.token_expires = int(time.time()) + 7200
                    sql_update = "UPDATE `wx_info` SET `access_token`=%s, `expires_time`=%s WHERE `token_type` = %s"
                    db.query(sql_update, (access_token, self.token_expires, self.corpsecret))
                    db.commit()
                else:
                    self.token_cache = result[0][0]
                    self.token_expires = result[0][1]
        return self.token_cache
    
    def _getToken(self):
        url = f'https://qyapi.weixin.qq.com/cgi-bin/gettoken?corpid={self.corpid}&corpsecret={self.corpsecret}'
        response = requests.get(url)
        data = response.json()
        return data.get('access_token', '')
        
    def getToken_Archive(self):
        if not hasattr(self, 'token_cache_archive') or self.token_cache_archive is None or time.time() >= self.token_expires_archive - 10:
            sql_select = "SELECT `access_token`, `expires_time` FROM `wx_info` WHERE `token_type` = %s"
            with WhisperDB_Pool() as db:
                result = db.query(sql_select, (self.corpsecret_archive,))
                if not result:
                    access_token = self._getToken_Archive()
                    self.token_cache_archive = access_token
                    self.token_expires_archive = int(time.time()) + 7200
                    sql_insert = "INSERT INTO `wx_info` (`access_token`, `expires_time`, `token_type`) VALUES (%s, %s, %s)"
                    db.query(sql_insert, (access_token, self.token_expires_archive, self.corpsecret_archive))
                    db.commit()
                elif result[0][1] <= (time.time() - 10):
                    access_token = self._getToken_Archive()
                    self.token_cache_archive = access_token
                    self.token_expires_archive = int(time.time()) + 7200
                    sql_update = "UPDATE `wx_info` SET `access_token`=%s, `expires_time`=%s WHERE `token_type` = %s"
                    db.query(sql_update, (access_token, self.token_expires_archive, self.corpsecret_archive))
                    db.commit()
                else:
                    self.token_cache_archive = result[0][0]
                    self.token_expires_archive = result[0][1]
        return self.token_cache_archive
    
    def _getToken_Archive(self):
        url = f'https://qyapi.weixin.qq.com/cgi-bin/gettoken?corpid={self.corpid}&corpsecret={self.corpsecret_archive}'
        response = requests.get(url)
        data = response.json()
        return data.get('access_token', '')

class WhisperTools_OrderManageTools:
    # 上传文件格式到数据库中
    @staticmethod
    def upload_file_format(file_path):
        # 打开file_path这个excel文件，读取“渠道订单文件配置”sheet页面。将这个页面的内容写入数据库。
        # 这个页面的格式是固定的，第一行是表头，后续行是数据。一个行为：渠道名称	微信群名称	字段名称1	字段名称2	字段名称3	字段名称4	字段名称5	字段名称6	字段名称7	字段名称8	字段名称9	字段名称10	字段名称11	字段名称12	字段名称13	字段名称14	字段名称15	字段名称16	字段名称17	字段名称18	字段名称19	字段名称20	字段名称21	字段名称22	字段名称23	字段名称24
        wb = load_workbook(file_path)
        ws = wb["渠道订单文件配置"]
        with WhisperDB() as myDB:
            # 先清空表
            sql_delete = "DELETE FROM bd_order_file_format WHERE 1"
            myDB.query(sql_delete)
            # 然后插入数据
            for row in ws.iter_rows(min_row=2, values_only=True):  # 从第二行开始读取
                channel_name = row[0] if row[0] is not None else ""
                group_name = row[1] if row[1] is not None else ""
                fields = row[2:26]  # 获取字段名称1到字段名称24
                # 将字段名称列表转换为JSON字符串存储
                fields_json = json.dumps(fields, ensure_ascii=False)
                sql_insert = "INSERT INTO bd_order_file_format (channel_name, group_name, fields) VALUES (%s, %s, %s)"
                myDB.query(sql_insert, (channel_name, group_name, fields_json))
            myDB.commit()
        return