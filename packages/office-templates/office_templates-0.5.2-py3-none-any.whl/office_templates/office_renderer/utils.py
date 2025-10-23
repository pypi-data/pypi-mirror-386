"""
Utility functions for handling Office-related documents (Excel, PowerPoint, etc.).

This module provides helper functions for working with Microsoft Office document formats,
including file type detection and workbook loading utilities.
"""

from typing import IO
import zipfile

from .exceptions import UnsupportedFileType


def get_load_workbook():
    """
    Returns the load_workbook function from openpyxl.

    This function dynamically imports openpyxl's load_workbook function,
    allowing for lazy loading of the dependency only when needed.

    Returns:
        function: The load_workbook function from openpyxl

    Raises:
        ImportError: If openpyxl is not installed in the environment
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl is not installed. Please install it to use this feature."
        )

    return load_workbook


def identify_file_type(file_input: str | IO[bytes]):
    """
    Identifies the type of Office file by examining its internal structure.

    Office files (xlsx, pptx, etc.) are actually ZIP archives with specific
    directory structures. This function examines those structures to determine
    the file type.

    Args:
        file_input: Either a string path to the file or a file-like object (opened in binary mode).


    Returns:
        str: Identified file type ('xlsx', 'pptx', 'unknown') or 'not a valid zip file'
             if the file cannot be read as a ZIP archive
    """

    # Determine if file_input is a filepath or a file object.
    if isinstance(file_input, str):
        file_obj = open(file_input, "rb")
        should_close = True
    else:
        file_obj = file_input
        should_close = False

    # Remember the current position to restore later if needed.
    try:
        current_position = file_obj.tell()
    except (AttributeError, IOError):
        current_position = None

    try:
        # Rewind to the beginning
        file_obj.seek(0)

        with zipfile.ZipFile(file_obj, "r") as z:
            names = z.namelist()
            if any(name.startswith("xl/") for name in names):
                return "xlsx"
            elif any(name.startswith("ppt/") for name in names):
                return "pptx"

    except zipfile.BadZipFile:
        pass

    finally:
        # Restore the file position if it was provided by the caller
        if current_position is not None and not should_close:
            file_obj.seek(current_position)

        # Close the file if it was opened in this function
        if should_close:
            file_obj.close()

    raise UnsupportedFileType(
        "Unsupported file type. Please provide a valid XLSX or PPTX file."
    )
