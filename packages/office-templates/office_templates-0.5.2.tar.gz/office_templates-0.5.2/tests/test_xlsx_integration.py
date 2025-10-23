import os
import tempfile
import unittest
from unittest.mock import patch
from openpyxl import Workbook, load_workbook

from office_templates.office_renderer import render_xlsx

from tests.utils import has_view_permission


# Dummy objects for integration testing
class DummyUser:
    def __init__(self, name, email, is_active=True):
        self.name = name
        self.email = email
        self.is_active = is_active
        self._meta = True  # Simulate a Django model

    def __str__(self):
        return self.name


class DummyDepartment:
    def __init__(self, name, budget):
        self.name = name
        self.budget = budget
        self._meta = True  # Make it look like a Django object

    def __str__(self):
        return self.name


class DummyRequestUser:
    def has_perm(self, perm, obj):
        # Deny permission if object's name contains "deny"
        if hasattr(obj, "name") and "deny" in obj.name.lower():
            return False
        return True


class TestXlsxIntegration(unittest.TestCase):
    def setUp(self):
        # Create a minimal XLSX workbook with sample data and templates
        self.wb = Workbook()

        # First worksheet - simple text placeholders
        ws1 = self.wb.active
        ws1.title = "Summary"
        ws1["A1"] = "User Information"
        ws1["A2"] = "Name: {{ user.name }}"
        ws1["A3"] = "Email: {{ user.email }}"
        ws1["A4"] = "Report Date: {{ date }}"
        ws1["B2"] = "Department: {{ department.name }}"
        ws1["B3"] = "Budget: ${{ department.budget }}"

        # Second worksheet - single placeholder that should expand to multiple rows
        ws2 = self.wb.create_sheet(title="Employees")
        # This is the key test - this should expand to multiple rows with employee names
        ws2["A1"] = "Employee names"
        ws2["A2"] = "{{ employees.name }}"

        # Third worksheet - numeric values as dictionary keys
        ws3 = self.wb.create_sheet(title="Numbers")
        ws3["A1"] = "{{ value_dict.val1 }}"
        ws3["A2"] = "{{ value_dict.val2 }}"
        ws3["A3"] = "{{ value_dict.val3 }}"

        # Save this XLSX to a temporary file
        self.temp_input = tempfile.mktemp(suffix=".xlsx")
        self.temp_output = tempfile.mktemp(suffix=".xlsx")
        self.wb.save(self.temp_input)

        # Set up the context
        self.employee1 = DummyUser("Alice", "alice@example.com", is_active=True)
        self.employee2 = DummyUser("Bob", "bob@example.com", is_active=True)
        self.employee3 = DummyUser("Carol", "carol@example.com", is_active=False)
        self.department = DummyDepartment("Engineering", 150000)

        self.context = {
            "user": self.employee1,
            "employees": [self.employee1, self.employee2, self.employee3],
            "department": self.department,
            "date": "2023-05-15",
            "value_dict": {
                "val1": 100.5,
                "val2": 200.75,
                "val3": 300.25,
            },
        }

        self.request_user = DummyRequestUser()

    def tearDown(self):
        # Clean up temporary files
        for temp_file in [self.temp_input, self.temp_output]:
            if os.path.exists(temp_file):
                os.remove(temp_file)

    def test_integration_renderer(self):
        """Test the full integration of the XLSX renderer."""
        # Run the renderer
        rendered, errors = render_xlsx(
            self.temp_input, self.context, self.temp_output, None
        )

        self.assertIsNone(errors)

        # Open the rendered XLSX and verify its contents
        output_wb = load_workbook(rendered)

        # Check Summary sheet
        ws_summary = output_wb["Summary"]
        self.assertEqual(ws_summary["A2"].value, "Name: Alice")
        self.assertEqual(ws_summary["A3"].value, "Email: alice@example.com")
        self.assertEqual(ws_summary["A4"].value, "Report Date: 2023-05-15")
        self.assertEqual(ws_summary["B2"].value, "Department: Engineering")
        self.assertEqual(ws_summary["B3"].value, "Budget: $150000")

        # Check Employees sheet - this should have expanded to multiple rows for each employee
        ws_employees = output_wb["Employees"]

        # Check that the 3 cells in column A contain the employee names
        self.assertEqual(ws_employees["A1"].value, "Employee names")
        expected_names = ["Alice", "Bob", "Carol"]
        for i, name in enumerate(expected_names):
            # A2 should be Alice, A3 should be Bob, A4 should be Carol
            cell_pos = f"A{i+2}"
            self.assertEqual(ws_employees[cell_pos].value, name)

        # Check Numbers sheet - should have converted strings to floats
        ws_numbers = output_wb["Numbers"]
        self.assertEqual(ws_numbers["A1"].value, 100.5)
        self.assertEqual(ws_numbers["A2"].value, 200.75)
        self.assertEqual(ws_numbers["A3"].value, 300.25)

    def test_integration_with_file_objects(self):
        """Test the renderer with file-like objects instead of paths."""
        # Create file-like objects
        with open(self.temp_input, "rb") as input_file:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
                # Run the renderer with file objects
                check_permissions = lambda obj: has_view_permission(
                    obj, self.request_user
                )
                rendered, errors = render_xlsx(
                    input_file, self.context, output_file, check_permissions
                )

                self.assertIsNone(errors)
                output_path = output_file.name

        # Open the rendered XLSX and perform basic verification
        output_wb = load_workbook(output_path)
        ws_summary = output_wb["Summary"]
        self.assertEqual(ws_summary["A2"].value, "Name: Alice")

        # Clean up
        if os.path.exists(output_path):
            os.remove(output_path)

    def test_integration_permission_denied(self):
        """Test permission handling in integration."""
        # We need to force a permission exception by mocking
        # Create a simple test file
        wb = Workbook()
        ws = wb.active
        # Create a template that explicitly mentions to access a "deny" object
        ws["A1"] = "{{ deny_obj }}"

        test_input = tempfile.mktemp(suffix=".xlsx")
        test_output = tempfile.mktemp(suffix=".xlsx")
        wb.save(test_input)

        # Create a context with an object that will be denied
        deny_dept = DummyDepartment("DenyAccess", 50000)
        deny_context = {"deny_obj": deny_dept}

        # Run renderer - should naturally fail due to permission denied
        check_permissions = lambda obj: has_view_permission(obj, self.request_user)
        rendered, errors = render_xlsx(
            test_input, deny_context, test_output, check_permissions
        )

        # Should have errors due to permission denied
        self.assertIsNotNone(errors)
        self.assertTrue(len(errors) > 0)
        self.assertIn("Permission denied", str(errors[0]))

        # Clean up
        for temp_file in [test_input, test_output]:
            if os.path.exists(temp_file):
                os.remove(temp_file)

    def test_cell_overwrite_error(self):
        """Test that CellOverwriteError is raised when trying to expand a list into cells with content."""
        # Create a real workbook with a conflicting setup
        wb = Workbook()
        ws = wb.active
        ws.title = "Conflict"

        # Cell A1 has a placeholder that should expand to multiple rows
        ws["A1"] = "{{ employees }}"

        # Cell A2 already has content which would be overwritten when A1 expands
        ws["A2"] = "Existing content"

        # Save this workbook to a temporary file
        temp_input = tempfile.mktemp(suffix=".xlsx")
        temp_output = tempfile.mktemp(suffix=".xlsx")
        wb.save(temp_input)

        # Create context with a list that would expand
        context = {"employees": ["Alice", "Bob", "Carol"]}

        try:
            # This should raise a CellOverwriteError when rendering
            rendered, errors = render_xlsx(temp_input, context, temp_output, None)

            # Verify that rendering failed with appropriate error
            self.assertIsNone(rendered)
            self.assertIsNotNone(errors)

            # Check that the error message mentions the cell overwrite problem
            self.assertTrue(
                any(
                    "Cannot expand list into non-empty cell" in str(err) for err in errors
                )
            )
            self.assertTrue(any("A2" in str(err) for err in errors))

        finally:
            # Clean up temporary files
            for temp_file in [temp_input, temp_output]:
                if os.path.exists(temp_file):
                    os.remove(temp_file)

    def test_callable_math_expression_integration(self):
        """Test that math operations on callable results are evaluated in XLSX rendering."""

        class DummyCallable:
            def multiply(self, a, b):
                return a * b

            def add(self, a, b):
                return a + b

        # Add a new worksheet for this test
        ws = self.wb.create_sheet(title="CallableMath")
        ws["A1"] = "{{ dummy.multiply(3,4) * 10 }}"
        ws["A2"] = "{{ dummy.multiply(3,4) + 5 }}"
        ws["A3"] = "{{ dummy.add(10,5) * 2 }}"

        # Add dummy callable to context
        self.context["dummy"] = DummyCallable()
        self.wb.save(self.temp_input)

        rendered, errors = render_xlsx(
            self.temp_input, self.context, self.temp_output, None
        )
        self.assertIsNone(errors)
        output_wb = load_workbook(rendered)
        ws_callable = output_wb["CallableMath"]
        self.assertEqual(ws_callable["A1"].value, 120.0)  # 3*4=12, 12*10=120
        self.assertEqual(ws_callable["A2"].value, 17.0)  # 3*4=12, 12+5=17
        self.assertEqual(ws_callable["A3"].value, 30.0)  # 10+5=15, 15*2=30

    def test_image_cell_integration(self):
        """A cell starting with %imagesqueeze% should be replaced with a picture."""

        from PIL import Image

        img = Image.new("RGB", (2, 2), color="blue")
        img_file = tempfile.mktemp(suffix=".png")
        img.save(img_file)

        ws = self.wb.create_sheet(title="Images")
        ws["A1"] = f"%imagesqueeze% file://{img_file}"
        self.wb.save(self.temp_input)

        rendered, errors = render_xlsx(
            self.temp_input, self.context, self.temp_output, None
        )
        self.assertIsNone(errors)

        output_wb = load_workbook(rendered)
        ws_out = output_wb["Images"]
        self.assertEqual(len(ws_out._images), 1)
        anchor = ws_out._images[0].anchor._from.row, ws_out._images[0].anchor._from.col
        self.assertEqual(anchor, (0, 0))  # anchored at A1 (0-indexed)

        os.remove(img_file)


if __name__ == "__main__":
    unittest.main()
