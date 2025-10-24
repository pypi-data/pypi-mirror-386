"""
Balance updater for syncing Quicken database balances with Excel workbook.

Reads account balances from Quicken SQLite database and updates corresponding
entries in the Excel template using fuzzy matching.
"""

import shutil
import sqlite3
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from thefuzz import fuzz, process

    HAS_FUZZ = True
    FUZZ_LIBRARY = "thefuzz"
except ImportError:
    try:
        from rapidfuzz import fuzz, process

        HAS_FUZZ = True
        FUZZ_LIBRARY = "rapidfuzz"
    except ImportError:
        HAS_FUZZ = False
        FUZZ_LIBRARY = None

import openpyxl


class BalanceUpdaterError(Exception):
    """Base exception for balance updater errors."""

    pass


class BalanceUpdater:
    """Updates Excel workbook balances from Quicken database."""

    def __init__(
        self,
        db_path: Path,
        fuzzy_threshold: int = 80,
        bank_account_name: str = "PECU Checking",
        auto_backup: bool = True,
    ):
        """Initialize balance updater.

        Args:
            db_path: Path to Quicken SQLite database
            fuzzy_threshold: Minimum score for fuzzy matches (0-100)
            bank_account_name: Name of checking account to use for bank balance
            auto_backup: Whether to create backup before updating
        """
        if not HAS_FUZZ:
            raise ImportError(
                "Fuzzy matching library required for balance updates. "
                "Install one of: pip install 'thefuzz[speedup]' or pip install rapidfuzz"
            )

        self.db_path = Path(db_path)
        self.fuzzy_threshold = fuzzy_threshold
        self.bank_account_name = bank_account_name
        self.auto_backup = auto_backup

        if not self.db_path.exists():
            raise FileNotFoundError(f"Quicken database not found: {self.db_path}")

    def backup_excel(self, xlsx_path: Path) -> Path:
        """Create timestamped backup of Excel file.

        Args:
            xlsx_path: Path to Excel file

        Returns:
            Path to backup file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = xlsx_path.with_name(f"{xlsx_path.name}.backup.{timestamp}")
        shutil.copy2(xlsx_path, backup_path)
        return backup_path

    def connect_db(self) -> sqlite3.Connection:
        """Connect to Quicken database in read-only mode.

        Returns:
            SQLite connection
        """
        return sqlite3.connect(f"file:{self.db_path}?mode=ro", uri=True)

    def load_quicken_balances(
        self,
    ) -> Tuple[Dict[str, Dict], List[str], List[str], List[str]]:
        """Load account balances from Quicken database.

        Prefers ZONLINEBANKINGLEDGERBALANCEAMOUNT when available, otherwise sums
        transactions up to the current date. Uses Apple Cocoa timestamp format.

        Returns:
            Tuple of (accounts_by_name, credit_card_names, checking_names, savings_names)
        """
        conn = self.connect_db()
        conn.row_factory = sqlite3.Row

        try:
            cur = conn.cursor()

            # Calculate current Cocoa timestamp (seconds since 2001-01-01)
            cocoa_now = int(time.time() - 978307200)

            sql = """
                SELECT
                    a.Z_PK AS id,
                    a.ZNAME AS name,
                    a.ZTYPENAME AS type,
                    a.ZACTIVE AS active,
                    COALESCE(
                        a.ZONLINEBANKINGLEDGERBALANCEAMOUNT,
                        SUM(CASE 
                            WHEN t.ZPOSTEDDATE IS NOT NULL AND t.ZPOSTEDDATE <= ?
                            THEN t.ZAMOUNT 
                            ELSE 0 
                        END),
                        0
                    ) AS balance
                FROM ZACCOUNT a
                LEFT JOIN ZTRANSACTION t ON t.ZACCOUNT = a.Z_PK
                WHERE a.ZACTIVE = 1
                  AND a.ZTYPENAME IN ('CREDITCARD','CHECKING','SAVINGS')
                GROUP BY a.Z_PK, a.ZNAME, a.ZTYPENAME, a.ZACTIVE, a.ZONLINEBANKINGLEDGERBALANCEAMOUNT
            """
            cur.execute(sql, (cocoa_now,))
            rows = cur.fetchall()

            accounts_by_name = {}
            credit_card_names = []
            checking_names = []
            savings_names = []

            for r in rows:
                name = r["name"]
                acc_type = (r["type"] or "").upper().strip()
                bal = float(r["balance"] or 0.0)
                accounts_by_name[name] = {
                    "id": r["id"],
                    "type": acc_type,
                    "balance": bal,
                }
                if acc_type == "CREDITCARD":
                    credit_card_names.append(name)
                elif acc_type == "CHECKING":
                    checking_names.append(name)
                elif acc_type == "SAVINGS":
                    savings_names.append(name)

            return accounts_by_name, credit_card_names, checking_names, savings_names
        finally:
            conn.close()

    def _prompt_yes_no(self, question: str, default_no: bool = True) -> bool:
        """Prompt user for yes/no answer.

        Args:
            question: Question to ask
            default_no: Whether default is no

        Returns:
            True if user answered yes
        """
        default = "n" if default_no else "y"
        prompt = f"{question} [{'Y/n' if not default_no else 'y/N'}]: "
        try:
            ans = input(prompt).strip().lower()
        except EOFError:
            ans = ""
        if not ans:
            ans = default
        return ans in ("y", "yes")

    def update_debts_sheet(
        self, ws, accounts_by_name: Dict[str, Dict], credit_card_names: List[str]
    ) -> List[Dict]:
        """Update debt balances in Debts sheet.

        Args:
            ws: openpyxl worksheet
            accounts_by_name: Dictionary of account data by name
            credit_card_names: List of credit card account names

        Returns:
            List of update records
        """
        updates = []

        if ws.max_row < 2:
            return updates

        if not credit_card_names:
            return updates

        # Iterate rows starting from row 2
        for row in range(2, ws.max_row + 1):
            excel_name_cell = ws.cell(row=row, column=1)  # Column A
            balance_cell = ws.cell(row=row, column=2)  # Column B

            excel_name = (excel_name_cell.value or "").strip()
            if not excel_name:
                continue

            old_balance = balance_cell.value

            # If exact match, update without prompt
            if excel_name in credit_card_names:
                qname = excel_name
                qb = accounts_by_name[qname]["balance"]
                new_balance = abs(qb)

                # Only update if balance changed
                if old_balance != new_balance:
                    balance_cell.value = float(new_balance)
                    excel_name_cell.value = qname
                    updates.append(
                        {
                            "row": row,
                            "excel_name_old": excel_name,
                            "excel_name_new": qname,
                            "old_balance": old_balance,
                            "new_balance": new_balance,
                            "score": 100,
                            "auto": True,
                        }
                    )
                continue

            # Fuzzy match to credit card names
            match = process.extractOne(
                excel_name, credit_card_names, scorer=fuzz.WRatio
            )
            if not match:
                continue

            # Handle both rapidfuzz (2-tuple) and thefuzz (3-tuple) return formats
            if len(match) == 3:
                candidate, score, _ = match
            else:
                candidate, score = match

            if score < self.fuzzy_threshold:
                continue

            print("\n[debts] Potential match found:")
            print(f"  Excel name   : {excel_name}")
            print(f"  Quicken name : {candidate}")
            print(f"  Score        : {score}")

            if self._prompt_yes_no("Approve this match?", default_no=True):
                qb = accounts_by_name[candidate]["balance"]
                new_balance = abs(qb)

                # Only update if balance changed or name changed
                if old_balance != new_balance or excel_name != candidate:
                    balance_cell.value = float(new_balance)
                    excel_name_cell.value = candidate
                    updates.append(
                        {
                            "row": row,
                            "excel_name_old": excel_name,
                            "excel_name_new": candidate,
                            "old_balance": old_balance,
                            "new_balance": new_balance,
                            "score": score,
                            "auto": False,
                        }
                    )

        return updates

    def update_settings_sheet(
        self, ws, accounts_by_name: Dict[str, Dict], checking_names: List[str]
    ) -> Optional[Dict]:
        """Update current bank balance in Settings sheet.

        Args:
            ws: openpyxl worksheet
            accounts_by_name: Dictionary of account data by name
            checking_names: List of checking account names

        Returns:
            Update record or None if not updated
        """
        target_name = self.bank_account_name

        # Try exact match
        if (
            target_name in accounts_by_name
            and accounts_by_name[target_name]["type"] == "CHECKING"
        ):
            bal = float(accounts_by_name[target_name]["balance"])
            old_balance = ws.cell(row=3, column=2).value

            # Only update if balance changed
            if old_balance != bal:
                ws.cell(row=3, column=2).value = bal
                return {"name": target_name, "balance": bal, "matched": "exact"}
            return None

        # Try fuzzy match
        if checking_names:
            match = process.extractOne(target_name, checking_names, scorer=fuzz.WRatio)
            if match:
                # Handle both rapidfuzz (2-tuple) and thefuzz (3-tuple) return formats
                if len(match) == 3:
                    candidate, score, _ = match
                else:
                    candidate, score = match

                print(f"\n[settings] Could not find exact '{target_name}'.")
                print(f"  Closest checking account: {candidate} (score {score})")

                if self._prompt_yes_no(
                    "Use this account for Current Bank Balance?", default_no=True
                ):
                    bal = float(accounts_by_name[candidate]["balance"])
                    old_balance = ws.cell(row=3, column=2).value

                    # Only update if balance changed
                    if old_balance != bal:
                        ws.cell(row=3, column=2).value = bal
                        return {
                            "name": candidate,
                            "balance": bal,
                            "matched": f"fuzzy:{score}",
                        }
                    return None

        return None

    def update_workbook(
        self, xlsx_path: Path, interactive: bool = True
    ) -> Dict[str, any]:
        """Update Excel workbook with balances from Quicken database.

        Args:
            xlsx_path: Path to Excel workbook
            interactive: Whether to prompt for fuzzy matches

        Returns:
            Dictionary with update summary

        Raises:
            FileNotFoundError: If workbook doesn't exist
            BalanceUpdaterError: If update fails
        """
        if not xlsx_path.exists():
            raise FileNotFoundError(f"Excel workbook not found: {xlsx_path}")

        # Load Quicken data
        accounts_by_name, cc_names, checking_names, _ = self.load_quicken_balances()

        # Load workbook
        wb = openpyxl.load_workbook(str(xlsx_path))

        # Create backup
        backup_path = None
        if self.auto_backup:
            backup_path = self.backup_excel(xlsx_path)

        # Get sheets
        if "Debts" not in wb.sheetnames:
            raise BalanceUpdaterError("'Debts' sheet not found in workbook")

        ws_debts = wb["Debts"]
        ws_settings = wb["Settings"] if "Settings" in wb.sheetnames else None

        # Update Debts
        debt_updates = self.update_debts_sheet(ws_debts, accounts_by_name, cc_names)

        # Update Settings
        settings_update = None
        if ws_settings is not None:
            settings_update = self.update_settings_sheet(
                ws_settings, accounts_by_name, checking_names
            )

        # Save workbook
        wb.save(str(xlsx_path))

        return {
            "backup_path": backup_path,
            "debt_updates": debt_updates,
            "settings_update": settings_update,
            "workbook_path": xlsx_path,
        }
