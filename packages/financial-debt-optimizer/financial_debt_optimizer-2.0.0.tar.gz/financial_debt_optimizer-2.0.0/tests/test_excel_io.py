"""
Comprehensive tests for excel_io modules.

Tests ExcelReader, ExcelWriter, ExcelTemplateGenerator, and related functionality
for reading, writing, and generating Excel files.
"""

import shutil

# Import the classes to test
import sys
import tempfile
from datetime import date
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

src_path = Path(__file__).parent.parent / "debt-optimizer"
sys.path.insert(0, str(src_path))

from core.debt_optimizer import DebtOptimizer, OptimizationGoal, OptimizationResult
from core.financial_calc import (
    Debt,
    FutureExpense,
    FutureIncome,
    Income,
    RecurringExpense,
)
from excel_io.excel_reader import ExcelReader, ExcelTemplateGenerator
from excel_io.excel_writer import ExcelReportWriter, generate_simple_summary_report


class TestExcelTemplateGenerator:
    """Test cases for the ExcelTemplateGenerator class."""

    def setup_method(self):
        """Set up test fixtures."""
        self.temp_dir = Path(tempfile.mkdtemp())

    def teardown_method(self):
        """Clean up test fixtures."""
        shutil.rmtree(self.temp_dir)

    @pytest.mark.excel
    def test_generate_template_with_sample_data(self):
        """Test template generation with sample data."""
        template_path = self.temp_dir / "test_template_with_data.xlsx"

        ExcelTemplateGenerator.generate_template(
            str(template_path), include_sample_data=True
        )

        # Verify file was created
        assert template_path.exists()

        # Load workbook and verify sheets
        workbook = load_workbook(template_path)
        expected_sheets = [
            "Debts",
            "Income",
            "Recurring Expenses",
            "Future Income",
            "Future Expenses",
            "Settings",
        ]

        assert len(workbook.sheetnames) == 6
        for sheet_name in expected_sheets:
            assert sheet_name in workbook.sheetnames

    @pytest.mark.excel
    def test_generate_template_without_sample_data(self):
        """Test template generation without sample data."""
        template_path = self.temp_dir / "test_template_empty.xlsx"

        ExcelTemplateGenerator.generate_template(
            str(template_path), include_sample_data=False
        )

        # Verify file was created
        assert template_path.exists()

        # Load workbook and verify structure
        workbook = load_workbook(template_path)
        assert len(workbook.sheetnames) == 6

    @pytest.mark.excel
    def test_debts_sheet_structure(self):
        """Test that the Debts sheet has correct structure."""
        template_path = self.temp_dir / "test_debts_structure.xlsx"

        ExcelTemplateGenerator.generate_template(
            str(template_path), include_sample_data=True
        )

        workbook = load_workbook(template_path)
        debts_sheet = workbook["Debts"]

        # Check headers
        expected_headers = [
            "Name",
            "Balance",
            "Min Payment",
            "Interest Rate",
            "Due Date",
        ]
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = debts_sheet.cell(row=1, column=col).value
            assert actual_header == expected_header

        # Check sample data exists
        assert debts_sheet.cell(row=2, column=1).value is not None  # First debt name
        assert isinstance(
            debts_sheet.cell(row=2, column=2).value, (int, float)
        )  # Balance

    @pytest.mark.excel
    def test_income_sheet_structure(self):
        """Test that the Income sheet has correct structure."""
        template_path = self.temp_dir / "test_income_structure.xlsx"

        ExcelTemplateGenerator.generate_template(
            str(template_path), include_sample_data=True
        )

        workbook = load_workbook(template_path)
        income_sheet = workbook["Income"]

        # Check headers
        expected_headers = ["Source", "Amount", "Frequency", "Start Date"]
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = income_sheet.cell(row=1, column=col).value
            assert actual_header == expected_header

    @pytest.mark.excel
    def test_recurring_expenses_sheet_structure(self):
        """Test that the Recurring Expenses sheet has correct structure."""
        template_path = self.temp_dir / "test_expenses_structure.xlsx"

        ExcelTemplateGenerator.generate_template(
            str(template_path), include_sample_data=True
        )

        workbook = load_workbook(template_path)
        expenses_sheet = workbook["Recurring Expenses"]

        # Check headers
        expected_headers = [
            "Description",
            "Amount",
            "Frequency",
            "Due Date",
            "Start Date",
        ]
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = expenses_sheet.cell(row=1, column=col).value
            assert actual_header == expected_header

    @pytest.mark.excel
    def test_future_income_sheet_structure(self):
        """Test that the Future Income sheet has correct structure."""
        template_path = self.temp_dir / "test_future_income_structure.xlsx"

        ExcelTemplateGenerator.generate_template(
            str(template_path), include_sample_data=True
        )

        workbook = load_workbook(template_path)
        future_income_sheet = workbook["Future Income"]

        # Check headers
        expected_headers = [
            "Description",
            "Amount",
            "Start Date",
            "Frequency",
            "End Date",
        ]
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = future_income_sheet.cell(row=1, column=col).value
            assert actual_header == expected_header

    @pytest.mark.excel
    def test_future_expenses_sheet_structure(self):
        """Test that the Future Expenses sheet has correct structure."""
        template_path = self.temp_dir / "test_future_expenses_structure.xlsx"

        ExcelTemplateGenerator.generate_template(
            str(template_path), include_sample_data=True
        )

        workbook = load_workbook(template_path)
        future_expenses_sheet = workbook["Future Expenses"]

        # Check headers
        expected_headers = [
            "Description",
            "Amount",
            "Start Date",
            "Frequency",
            "End Date",
        ]
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = future_expenses_sheet.cell(row=1, column=col).value
            assert actual_header == expected_header

    @pytest.mark.excel
    def test_settings_sheet_structure(self):
        """Test that the Settings sheet has correct structure."""
        template_path = self.temp_dir / "test_settings_structure.xlsx"

        ExcelTemplateGenerator.generate_template(
            str(template_path), include_sample_data=True
        )

        workbook = load_workbook(template_path)
        settings_sheet = workbook["Settings"]

        # Check headers
        expected_headers = ["Setting", "Value"]
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = settings_sheet.cell(row=1, column=col).value
            assert actual_header == expected_header

        # Check default settings
        expected_settings = [
            "Emergency Fund",
            "Current Bank Balance",
            "Optimization Goal",
        ]
        for row in range(2, 5):
            setting_name = settings_sheet.cell(row=row, column=1).value
            assert setting_name in expected_settings


class TestExcelReader:
    """Test cases for the ExcelReader class."""

    def setup_method(self):
        """Set up test fixtures."""
        self.temp_dir = Path(tempfile.mkdtemp())
        self.template_path = self.temp_dir / "test_template.xlsx"

        # Create a test template with sample data
        ExcelTemplateGenerator.generate_template(
            str(self.template_path), include_sample_data=True
        )
        self.reader = ExcelReader(str(self.template_path))

    def teardown_method(self):
        """Clean up test fixtures."""
        shutil.rmtree(self.temp_dir)

    @pytest.mark.excel
    def test_excel_reader_initialization(self):
        """Test ExcelReader initialization."""
        assert self.reader.file_path == str(self.template_path)
        assert self.template_path.exists()

    @pytest.mark.excel
    def test_read_debts_sheet(self):
        """Test reading debts from Excel file."""
        debts = self.reader.read_debts()

        assert isinstance(debts, list)
        assert len(debts) > 0

        # Check that debts are Debt objects
        for debt in debts:
            assert isinstance(debt, Debt)
            assert debt.name is not None
            assert debt.balance > 0
            assert debt.minimum_payment > 0
            assert debt.interest_rate >= 0
            assert 1 <= debt.due_date <= 31

    @pytest.mark.excel
    def test_read_income_sheet(self):
        """Test reading income data from Excel file."""
        income_sources = self.reader.read_income()

        assert isinstance(income_sources, list)
        assert len(income_sources) > 0

        # Check that income sources are Income objects
        for income in income_sources:
            assert isinstance(income, Income)
            assert income.source is not None
            assert income.amount > 0
            assert income.frequency in [
                "weekly",
                "bi-weekly",
                "monthly",
                "quarterly",
                "annually",
            ]
            assert isinstance(income.start_date, date)

    @pytest.mark.excel
    def test_read_recurring_expenses_sheet(self):
        """Test reading recurring expenses from Excel file."""
        expenses = self.reader.read_recurring_expenses()

        assert isinstance(expenses, list)
        assert len(expenses) > 0

        # Check that expenses are RecurringExpense objects
        for expense in expenses:
            assert isinstance(expense, RecurringExpense)
            assert expense.description is not None
            assert expense.amount > 0
            assert expense.frequency in [
                "bi-weekly",
                "monthly",
                "quarterly",
                "annually",
            ]
            assert isinstance(expense.start_date, date)

    @pytest.mark.excel
    def test_read_future_income_sheet(self):
        """Test reading future income from Excel file."""
        future_income = self.reader.read_future_income()

        assert isinstance(future_income, list)
        assert len(future_income) > 0

        # Check that future income items are FutureIncome objects
        for income in future_income:
            assert isinstance(income, FutureIncome)
            assert income.description is not None
            assert income.amount > 0
            assert isinstance(income.start_date, date)

    @pytest.mark.excel
    def test_read_future_expenses_sheet(self):
        """Test reading future expenses from Excel file."""
        future_expenses = self.reader.read_future_expenses()

        assert isinstance(future_expenses, list)
        assert len(future_expenses) > 0

        # Check that future expense items are FutureExpense objects
        for expense in future_expenses:
            assert isinstance(expense, FutureExpense)
            assert expense.description is not None
            assert expense.amount > 0
            assert isinstance(expense.start_date, date)

    @pytest.mark.excel
    def test_read_settings_sheet(self):
        """Test reading settings from Excel file."""
        settings = self.reader.read_settings()

        assert isinstance(settings, dict)
        assert "emergency_fund" in settings
        assert "current_bank_balance" in settings
        assert "optimization_goal" in settings

        # Check data types
        assert isinstance(settings["emergency_fund"], (int, float))
        assert isinstance(settings["current_bank_balance"], (int, float))
        assert isinstance(settings["optimization_goal"], str)

    @pytest.mark.excel
    def test_read_all_data(self):
        """Test reading all data from Excel file at once."""
        debts, income, expenses, future_income, future_expenses, settings = (
            self.reader.read_all_data()
        )

        # Verify each component
        assert isinstance(debts, list) and len(debts) > 0
        assert isinstance(income, list) and len(income) > 0
        assert isinstance(expenses, list) and len(expenses) > 0
        assert isinstance(future_income, list) and len(future_income) > 0
        assert isinstance(future_expenses, list) and len(future_expenses) > 0
        assert isinstance(settings, dict)

    @pytest.mark.excel
    def test_read_empty_template(self):
        """Test reading from template without sample data."""
        empty_template_path = self.temp_dir / "empty_template.xlsx"
        ExcelTemplateGenerator.generate_template(
            str(empty_template_path), include_sample_data=False
        )

        empty_reader = ExcelReader(str(empty_template_path))
        debts = empty_reader.read_debts()

        # Should return empty list, not error
        assert isinstance(debts, list)
        assert len(debts) == 0

    @pytest.mark.excel
    def test_read_nonexistent_file_error(self):
        """Test that reading nonexistent file raises appropriate error."""
        nonexistent_path = self.temp_dir / "nonexistent.xlsx"

        with pytest.raises(FileNotFoundError):
            ExcelReader(str(nonexistent_path))


class TestExcelWriter:
    """Test cases for the ExcelReportWriter class."""

    def setup_method(self):
        """Set up test fixtures."""
        self.temp_dir = Path(tempfile.mkdtemp())
        self.output_path = self.temp_dir / "test_report.xlsx"

        # Create sample data for testing
        self.sample_debts = [
            Debt("Credit Card", 5000.0, 150.0, 18.99, 15),
            Debt("Auto Loan", 12000.0, 325.0, 4.5, 10),
        ]

        self.sample_income = [Income("Salary", 3500.0, "bi-weekly", date(2024, 1, 5))]

        # Create a sample optimization result
        optimizer = DebtOptimizer(self.sample_debts, self.sample_income)
        self.optimization_result = optimizer.optimize_debt_strategy(
            OptimizationGoal.MINIMIZE_INTEREST, 200.0
        )

        self.debt_summary = optimizer.generate_debt_summary()

    def teardown_method(self):
        """Clean up test fixtures."""
        shutil.rmtree(self.temp_dir)

    @pytest.mark.excel
    def test_excel_writer_initialization(self):
        """Test ExcelReportWriter initialization."""
        writer = ExcelReportWriter(str(self.output_path))
        assert writer.output_path == str(self.output_path)

    @pytest.mark.excel
    def test_create_comprehensive_report(self):
        """Test creating a comprehensive Excel report."""
        writer = ExcelReportWriter(str(self.output_path))
        writer.create_comprehensive_report(self.optimization_result, self.debt_summary)

        # Verify file was created
        assert self.output_path.exists()

        # Load and verify structure
        workbook = load_workbook(self.output_path)
        expected_sheets = ["Executive Summary", "Payment Schedule", "Monthly Summary"]

        # Check that key sheets exist
        for sheet_name in expected_sheets:
            assert sheet_name in workbook.sheetnames

    @pytest.mark.excel
    def test_create_comprehensive_report_with_comparison(self):
        """Test creating comprehensive report with strategy comparison."""
        optimizer = DebtOptimizer(self.sample_debts, self.sample_income)
        strategy_comparison = optimizer.compare_strategies(200.0)

        writer = ExcelReportWriter(str(self.output_path))
        writer.create_comprehensive_report(
            self.optimization_result, self.debt_summary, strategy_comparison
        )

        # Verify file was created
        assert self.output_path.exists()

        # Check that comparison sheet exists
        workbook = load_workbook(self.output_path)
        assert "Strategy Comparison" in workbook.sheetnames

    @pytest.mark.excel
    def test_generate_simple_summary_report(self):
        """Test generating a simple summary report."""
        simple_report_path = self.temp_dir / "simple_report.xlsx"

        generate_simple_summary_report(
            str(simple_report_path), self.optimization_result, self.debt_summary
        )

        # Verify file was created
        assert simple_report_path.exists()

        # Load and verify it's a simple report (fewer sheets)
        workbook = load_workbook(simple_report_path)
        # Simple report should have fewer sheets than comprehensive
        assert len(workbook.sheetnames) <= 3

    @pytest.mark.excel
    def test_payment_schedule_sheet_content(self):
        """Test that payment schedule sheet contains correct data."""
        writer = ExcelReportWriter(str(self.output_path))
        writer.create_comprehensive_report(self.optimization_result, self.debt_summary)

        workbook = load_workbook(self.output_path)
        schedule_sheet = workbook["Payment Schedule"]

        # Check headers (headers are in row 3, as row 1 is the merged title)
        expected_headers = [
            "Date",
            "Type",
            "Description",
            "Amount",
            "Interest",
            "Principal",
            "Total Debt Balance",
            "Debt Name",
            "Debt Balance",
            "Bank Balance",
        ]
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = schedule_sheet.cell(row=3, column=col).value
            assert actual_header == expected_header

        # Check that data exists (data starts at row 4)
        assert schedule_sheet.cell(row=4, column=1).value is not None  # First date
        assert isinstance(
            schedule_sheet.cell(row=4, column=4).value, (int, float)
        )  # Amount

    @pytest.mark.excel
    def test_summary_sheet_content(self):
        """Test that summary sheet contains correct information."""
        writer = ExcelReportWriter(str(self.output_path))
        writer.create_comprehensive_report(self.optimization_result, self.debt_summary)

        workbook = load_workbook(self.output_path)
        summary_sheet = workbook["Executive Summary"]

        # Check that summary information is present
        # (Exact content depends on implementation, but should have key metrics)
        assert summary_sheet.cell(row=1, column=1).value is not None

        # Should contain financial summary data
        found_total_debt = False
        found_strategy = False
        for row in range(1, 20):  # Check first 20 rows
            cell_value = summary_sheet.cell(row=row, column=1).value
            if cell_value and "debt" in str(cell_value).lower():
                found_total_debt = True
            if cell_value and "strategy" in str(cell_value).lower():
                found_strategy = True

        assert found_total_debt or found_strategy  # Should find at least one key metric

    @pytest.mark.excel
    def test_write_to_existing_file_overwrites(self):
        """Test that writing to existing file overwrites it."""
        # Create initial file
        writer1 = ExcelReportWriter(str(self.output_path))
        writer1.create_comprehensive_report(self.optimization_result, self.debt_summary)

        initial_size = self.output_path.stat().st_size

        # Write to same path again
        writer2 = ExcelReportWriter(str(self.output_path))
        writer2.create_comprehensive_report(self.optimization_result, self.debt_summary)

        # File should still exist and be valid
        assert self.output_path.exists()
        final_size = self.output_path.stat().st_size

        # Sizes should be similar (allowing for minor differences)
        assert abs(final_size - initial_size) < initial_size * 0.1


class TestExcelIOIntegration:
    """Integration tests for Excel I/O operations."""

    def setup_method(self):
        """Set up test fixtures."""
        self.temp_dir = Path(tempfile.mkdtemp())

    def teardown_method(self):
        """Clean up test fixtures."""
        shutil.rmtree(self.temp_dir)

    @pytest.mark.integration
    @pytest.mark.excel
    def test_template_to_analysis_workflow(self):
        """Test complete workflow from template generation to analysis."""
        # Step 1: Generate template
        template_path = self.temp_dir / "workflow_template.xlsx"
        ExcelTemplateGenerator.generate_template(
            str(template_path), include_sample_data=True
        )

        # Step 2: Read data from template
        reader = ExcelReader(str(template_path))
        debts, income, expenses, future_income, future_expenses, settings = (
            reader.read_all_data()
        )

        # Step 3: Run optimization
        optimizer = DebtOptimizer(
            debts, income, expenses, future_income, future_expenses, settings
        )
        result = optimizer.optimize_debt_strategy(
            OptimizationGoal.MINIMIZE_INTEREST, 200.0
        )
        debt_summary = optimizer.generate_debt_summary()

        # Step 4: Generate report
        report_path = self.temp_dir / "workflow_report.xlsx"
        writer = ExcelReportWriter(str(report_path))
        writer.create_comprehensive_report(result, debt_summary)

        # Verify entire workflow completed successfully
        assert template_path.exists()
        assert report_path.exists()
        assert isinstance(result, OptimizationResult)
        assert isinstance(debt_summary, dict)

    @pytest.mark.integration
    @pytest.mark.excel
    def test_round_trip_data_integrity(self):
        """Test that data maintains integrity through read/write operations."""
        # Create template with sample data
        template_path = self.temp_dir / "integrity_template.xlsx"
        ExcelTemplateGenerator.generate_template(
            str(template_path), include_sample_data=True
        )

        # Read the data
        reader = ExcelReader(str(template_path))
        original_debts = reader.read_debts()
        original_income = reader.read_income()

        # Verify we have the expected sample data
        assert len(original_debts) > 0
        assert len(original_income) > 0

        # Check specific values to ensure data integrity
        assert all(debt.balance > 0 for debt in original_debts)
        assert all(income.amount > 0 for income in original_income)
        assert all(debt.interest_rate >= 0 for debt in original_debts)

    @pytest.mark.integration
    @pytest.mark.excel
    def test_error_handling_corrupted_file(self):
        """Test error handling with corrupted Excel file."""
        # Create a file that looks like Excel but isn't
        fake_excel_path = self.temp_dir / "fake.xlsx"
        with open(fake_excel_path, "w") as f:
            f.write("This is not an Excel file")

        # Should raise appropriate error when trying to read
        with pytest.raises(Exception):  # Could be various Excel-related exceptions
            reader = ExcelReader(str(fake_excel_path))
            reader.read_debts()  # This should trigger the exception
