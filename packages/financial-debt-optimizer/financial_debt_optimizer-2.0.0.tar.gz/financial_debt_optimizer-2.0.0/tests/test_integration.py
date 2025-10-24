"""
Comprehensive integration tests for the Financial Debt Optimizer.

Tests end-to-end workflows from Excel template generation through analysis
and report creation, ensuring all components work together correctly.
"""

import shutil

# Import the classes to test
import sys
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
import pytest
from click.testing import CliRunner

src_path = Path(__file__).parent.parent / "debt-optimizer"
sys.path.insert(0, str(src_path))

from cli.commands import analyze, generate_template, main, validate
from core.debt_optimizer import DebtOptimizer, OptimizationGoal
from core.financial_calc import (
    Debt,
    FutureExpense,
    FutureIncome,
    Income,
    RecurringExpense,
)
from core.validation import validate_financial_scenario
from excel_io.excel_reader import ExcelReader, ExcelTemplateGenerator
from excel_io.excel_writer import ExcelReportWriter, generate_simple_summary_report
from visualization.charts import DebtVisualization


class TestCompleteWorkflow:
    """Test complete end-to-end workflows."""

    def setup_method(self):
        """Set up test fixtures."""
        self.runner = CliRunner()
        self.temp_dir = Path(tempfile.mkdtemp())

    def teardown_method(self):
        """Clean up test fixtures."""
        shutil.rmtree(self.temp_dir)

    @pytest.mark.integration
    def test_complete_cli_workflow(self):
        """Test complete CLI workflow from template to analysis."""
        template_path = self.temp_dir / "integration_template.xlsx"
        analysis_path = self.temp_dir / "integration_analysis.xlsx"

        with self.runner.isolated_filesystem():
            # Step 1: Generate template with sample data
            result1 = self.runner.invoke(
                generate_template, ["--output", str(template_path), "--sample-data"]
            )
            assert result1.exit_code == 0
            assert template_path.exists()

            # Step 2: Validate the template
            result2 = self.runner.invoke(validate, [str(template_path)])
            assert result2.exit_code == 0
            assert "Debts:" in result2.output
            assert "Income Sources:" in result2.output

            # Step 3: Run analysis with strategy comparison
            result3 = self.runner.invoke(
                main,
                [
                    "analyze",
                    "--input",
                    str(template_path),
                    "--output",
                    str(analysis_path),
                    "--goal",
                    "minimize_interest",
                    "--extra-payment",
                    "200",
                    "--compare-strategies",
                ],
            )
            assert result3.exit_code == 0
            assert analysis_path.exists()
            assert "Report generated" in result3.output
            assert "Optimization Results:" in result3.output

    @pytest.mark.integration
    def test_programmatic_workflow(self):
        """Test complete programmatic workflow using classes directly."""
        # Step 1: Generate template
        template_path = self.temp_dir / "programmatic_template.xlsx"
        ExcelTemplateGenerator.generate_template(
            str(template_path), include_sample_data=True
        )
        assert template_path.exists()

        # Step 2: Read data from template
        reader = ExcelReader(str(template_path))
        debts, income, expenses, future_income, future_expenses, settings = (
            reader.read_all_data()
        )

        # Verify data was read correctly
        assert len(debts) > 0
        assert len(income) > 0
        assert all(isinstance(debt, Debt) for debt in debts)
        assert all(isinstance(inc, Income) for inc in income)

        # Step 3: Validate the financial scenario
        is_valid, messages = validate_financial_scenario(
            debts, income, expenses, settings
        )
        assert (
            is_valid is True
            or len([m for m in messages if not m.startswith("Warning:")]) == 0
        )

        # Step 4: Run optimization
        optimizer = DebtOptimizer(
            debts, income, expenses, future_income, future_expenses, settings
        )
        result = optimizer.optimize_debt_strategy(
            OptimizationGoal.MINIMIZE_INTEREST, 300.0
        )

        # Verify optimization results
        assert result.total_interest_paid >= 0
        assert result.total_months_to_freedom > 0
        assert isinstance(result.payment_schedule, pd.DataFrame)
        assert len(result.payment_schedule) > 0

        # Step 5: Generate comprehensive report
        report_path = self.temp_dir / "programmatic_report.xlsx"
        writer = ExcelReportWriter(str(report_path))
        debt_summary = optimizer.generate_debt_summary()
        comparison_data = optimizer.compare_strategies(300.0)

        writer.create_comprehensive_report(result, debt_summary, comparison_data)
        assert report_path.exists()

        # Step 6: Generate visualizations
        viz = DebtVisualization()

        # Create debt progression data from payment schedule
        if len(result.payment_schedule) > 0:
            debt_progression = pd.DataFrame(
                {
                    "month": range(1, min(13, len(result.payment_schedule) + 1)),
                    "Total Debt": [
                        1000 * i
                        for i in range(1, min(13, len(result.payment_schedule) + 1))
                    ],
                }
            )
            fig1 = viz.plot_debt_progression(debt_progression)
            assert fig1 is not None

        fig2 = viz.plot_strategy_comparison(comparison_data)
        assert fig2 is not None

    @pytest.mark.integration
    def test_workflow_with_custom_data(self):
        """Test workflow with manually created custom financial data."""
        # Create custom financial scenario
        custom_debts = [
            Debt("High Interest Card", 8000.0, 200.0, 24.99, 15),
            Debt("Personal Loan", 15000.0, 450.0, 12.5, 5),
            Debt("Auto Loan", 25000.0, 520.0, 6.2, 10),
        ]

        custom_income = [
            Income("Primary Job", 4500.0, "bi-weekly", date(2024, 1, 1)),
            Income("Side Business", 1500.0, "monthly", date(2024, 1, 1)),
        ]

        custom_expenses = [
            RecurringExpense("Rent", 2000.0, "monthly", 1, date(2024, 1, 1)),
            RecurringExpense("Utilities", 300.0, "monthly", 15, date(2024, 1, 1)),
            RecurringExpense("Insurance", 250.0, "monthly", 25, date(2024, 1, 1)),
        ]

        custom_future_income = [
            FutureIncome(
                "Annual Bonus", 8000.0, date.today() + timedelta(days=90), "once"
            ),
            FutureIncome(
                "Promotion Raise", 800.0, date.today() + timedelta(days=180), "monthly"
            ),
        ]

        custom_future_expenses = [
            FutureExpense(
                "Home Repair", 5000.0, date.today() + timedelta(days=120), "once"
            ),
            FutureExpense(
                "Insurance Increase", 50.0, date.today() + timedelta(days=30), "monthly"
            ),
        ]

        custom_settings = {
            "emergency_fund": 10000.0,
            "current_bank_balance": 5000.0,
            "optimization_goal": "minimize_time",
        }

        # Validate the custom scenario
        is_valid, messages = validate_financial_scenario(
            custom_debts, custom_income, custom_expenses, custom_settings
        )
        assert isinstance(is_valid, bool)
        assert isinstance(messages, list)

        # Run optimization on custom data
        optimizer = DebtOptimizer(
            custom_debts,
            custom_income,
            custom_expenses,
            custom_future_income,
            custom_future_expenses,
            custom_settings,
        )

        # Test all three optimization goals
        goals = [
            OptimizationGoal.MINIMIZE_INTEREST,
            OptimizationGoal.MINIMIZE_TIME,
            OptimizationGoal.MAXIMIZE_CASHFLOW,
        ]

        results = {}
        for goal in goals:
            result = optimizer.optimize_debt_strategy(goal, 500.0)
            results[goal.value] = result

            # Verify each result is valid
            assert result.total_months_to_freedom > 0
            assert result.total_interest_paid >= 0
            assert len(result.payment_schedule) > 0

        # Compare results - different goals should produce different outcomes
        interest_result = results["minimize_interest"]
        time_result = results["minimize_time"]
        cashflow_result = results["maximize_cashflow"]

        # Interest minimization should generally pay less interest
        # (though not always, depending on the scenario)
        assert isinstance(interest_result.total_interest_paid, (int, float))
        assert isinstance(time_result.total_months_to_freedom, int)

        # Generate reports for each strategy
        for goal_name, result in results.items():
            report_path = self.temp_dir / f"custom_{goal_name}_report.xlsx"
            debt_summary = optimizer.generate_debt_summary()

            writer = ExcelReportWriter(str(report_path))
            writer.create_comprehensive_report(result, debt_summary)
            assert report_path.exists()


class TestDataFlowIntegration:
    """Test data flow integration between components."""

    def setup_method(self):
        """Set up test fixtures."""
        self.temp_dir = Path(tempfile.mkdtemp())

    def teardown_method(self):
        """Clean up test fixtures."""
        shutil.rmtree(self.temp_dir)

    @pytest.mark.integration
    def test_excel_round_trip_data_integrity(self):
        """Test data integrity through Excel write/read cycle."""
        # Create template and read original sample data
        template_path = self.temp_dir / "round_trip_template.xlsx"
        ExcelTemplateGenerator.generate_template(
            str(template_path), include_sample_data=True
        )

        reader = ExcelReader(str(template_path))
        original_debts, original_income, original_expenses, _, _, original_settings = (
            reader.read_all_data()
        )

        # Run optimization and generate report
        optimizer = DebtOptimizer(original_debts, original_income, original_expenses)
        result = optimizer.optimize_debt_strategy(
            OptimizationGoal.MINIMIZE_INTEREST, 100.0
        )
        debt_summary = optimizer.generate_debt_summary()

        # Write comprehensive report
        report_path = self.temp_dir / "round_trip_report.xlsx"
        writer = ExcelReportWriter(str(report_path))
        writer.create_comprehensive_report(result, debt_summary)

        # Verify report exists and has expected structure
        assert report_path.exists()

        # Load the report and verify it contains expected data
        from openpyxl import load_workbook

        workbook = load_workbook(report_path)

        # Should have multiple sheets
        assert len(workbook.sheetnames) >= 2

        # Should contain summary and payment schedule sheets
        expected_sheets = ["Executive Summary", "Payment Schedule"]
        for sheet_name in expected_sheets:
            assert sheet_name in workbook.sheetnames
            sheet = workbook[sheet_name]
            # Sheet should have content (not empty)
            assert sheet.max_row > 1

    @pytest.mark.integration
    def test_validation_integration_with_optimization(self):
        """Test integration between validation and optimization components."""
        # Create scenarios with different validation outcomes
        scenarios = [
            # Valid scenario
            {
                "debts": [Debt("Valid Debt", 5000.0, 150.0, 18.0, 15)],
                "income": [Income("Valid Income", 4000.0, "monthly", date(2024, 1, 1))],
                "expenses": [
                    RecurringExpense("Rent", 1500.0, "monthly", 1, date(2024, 1, 1))
                ],
                "settings": {
                    "emergency_fund": 1000.0,
                    "current_bank_balance": 2000.0,
                    "optimization_goal": "minimize_interest",
                },
            },
            # Scenario with warnings but still valid
            {
                "debts": [Debt("High Rate Debt", 3000.0, 100.0, 29.99, 15)],
                "income": [Income("Low Income", 2500.0, "monthly", date(2024, 1, 1))],
                "expenses": [
                    RecurringExpense(
                        "Basic Expenses", 800.0, "monthly", 1, date(2024, 1, 1)
                    )
                ],
                "settings": {
                    "emergency_fund": 500.0,
                    "current_bank_balance": 100.0,
                    "optimization_goal": "minimize_time",
                },
            },
        ]

        for i, scenario in enumerate(scenarios):
            # Validate scenario
            is_valid, messages = validate_financial_scenario(
                scenario["debts"],
                scenario["income"],
                scenario["expenses"],
                scenario["settings"],
            )

            # If valid or only warnings, should be able to optimize
            if is_valid or all(msg.startswith("Warning:") for msg in messages):
                optimizer = DebtOptimizer(
                    scenario["debts"],
                    scenario["income"],
                    scenario["expenses"],
                    [],
                    [],
                    scenario["settings"],
                )

                result = optimizer.optimize_debt_strategy(
                    OptimizationGoal(scenario["settings"]["optimization_goal"]), 100.0
                )

                # Should produce valid results
                assert result.total_months_to_freedom > 0
                assert result.total_interest_paid >= 0
                assert len(result.payment_schedule) > 0

    @pytest.mark.integration
    def test_multi_optimization_strategy_consistency(self):
        """Test consistency across multiple optimization strategies."""
        # Use realistic financial scenario
        debts = [
            Debt("Credit Card 1", 3000.0, 100.0, 22.99, 15),
            Debt("Credit Card 2", 2000.0, 75.0, 19.99, 10),
            Debt("Car Loan", 18000.0, 380.0, 5.5, 5),
        ]

        income = [
            Income("Salary", 3200.0, "bi-weekly", date(2024, 1, 1)),
            Income("Freelance", 600.0, "monthly", date(2024, 1, 1)),
        ]

        expenses = [
            RecurringExpense("Rent", 1800.0, "monthly", 1, date(2024, 1, 1)),
            RecurringExpense("Utilities", 200.0, "monthly", 15, date(2024, 1, 1)),
        ]

        settings = {
            "emergency_fund": 2000.0,
            "current_bank_balance": 1500.0,
            "optimization_goal": "minimize_interest",
        }

        optimizer = DebtOptimizer(debts, income, expenses, [], [], settings)

        # Test different extra payment amounts
        extra_payments = [0.0, 200.0, 500.0, 1000.0]
        results = {}

        for extra_payment in extra_payments:
            result = optimizer.optimize_debt_strategy(
                OptimizationGoal.MINIMIZE_INTEREST, extra_payment
            )
            results[extra_payment] = result

            # Higher extra payments should generally result in:
            # - Less total interest paid
            # - Fewer months to freedom
            assert result.total_months_to_freedom > 0
            assert result.total_interest_paid >= 0

        # Verify that more extra payment generally leads to less interest and time
        for i in range(1, len(extra_payments)):
            current = results[extra_payments[i]]
            previous = results[extra_payments[i - 1]]

            # More extra payment should reduce time to freedom
            assert current.total_months_to_freedom <= previous.total_months_to_freedom
            # And usually reduce total interest (though not always guaranteed)
            # This is just a sanity check that the optimization is working


class TestErrorHandlingIntegration:
    """Test error handling integration across components."""

    def setup_method(self):
        """Set up test fixtures."""
        self.temp_dir = Path(tempfile.mkdtemp())

    def teardown_method(self):
        """Clean up test fixtures."""
        shutil.rmtree(self.temp_dir)

    @pytest.mark.integration
    def test_invalid_excel_file_handling(self):
        """Test handling of invalid Excel files throughout the workflow."""
        # Create a fake Excel file
        fake_excel_path = self.temp_dir / "fake.xlsx"
        with open(fake_excel_path, "w") as f:
            f.write("This is not a real Excel file")

        # ExcelReader should handle this gracefully
        with pytest.raises(Exception):  # Should raise some form of exception
            reader = ExcelReader(str(fake_excel_path))
            reader.read_all_data()

        # CLI should also handle this gracefully
        runner = CliRunner()
        with runner.isolated_filesystem():
            result = runner.invoke(
                analyze, ["--input", str(fake_excel_path), "--output", "output.xlsx"]
            )
            assert result.exit_code != 0

    @pytest.mark.integration
    def test_insufficient_data_scenario(self):
        """Test handling of insufficient data scenarios."""
        # Create minimal invalid scenario
        minimal_debts = []  # No debts
        minimal_income = [Income("Low Income", 100.0, "monthly", date(2024, 1, 1))]
        minimal_expenses = []
        minimal_settings = {
            "emergency_fund": 0.0,
            "current_bank_balance": 0.0,
            "optimization_goal": "minimize_interest",
        }

        # Validation should catch this
        is_valid, messages = validate_financial_scenario(
            minimal_debts, minimal_income, minimal_expenses, minimal_settings
        )
        assert is_valid is False
        assert len(messages) > 0
        assert any("no debts" in msg.lower() for msg in messages)

    @pytest.mark.integration
    def test_extreme_values_handling(self):
        """Test handling of extreme financial values."""
        # Scenario with extreme values
        extreme_debts = [
            Debt("Huge Debt", 1000000.0, 50000.0, 50.0, 15),  # Very high values
            Debt("Tiny Debt", 0.01, 0.01, 0.01, 1),  # Very small values
        ]

        extreme_income = [
            Income("High Income", 100000.0, "monthly", date(2024, 1, 1)),
        ]

        extreme_expenses = [
            RecurringExpense("High Expense", 50000.0, "monthly", 1, date(2024, 1, 1)),
        ]

        extreme_settings = {
            "emergency_fund": 100000.0,
            "current_bank_balance": 50000.0,
            "optimization_goal": "minimize_interest",
        }

        # Should handle extreme values without crashing
        is_valid, messages = validate_financial_scenario(
            extreme_debts, extreme_income, extreme_expenses, extreme_settings
        )

        # May not be valid, but shouldn't crash
        assert isinstance(is_valid, bool)
        assert isinstance(messages, list)

        # If valid, optimization should also not crash
        if is_valid or all(msg.startswith("Warning:") for msg in messages):
            optimizer = DebtOptimizer(extreme_debts, extreme_income, extreme_expenses)
            try:
                result = optimizer.optimize_debt_strategy(
                    OptimizationGoal.MINIMIZE_INTEREST, 1000.0
                )
                assert isinstance(result.total_months_to_freedom, int)
                assert result.total_months_to_freedom > 0
            except Exception as e:
                # If it does fail, it should be a reasonable exception
                assert isinstance(e, (ValueError, OverflowError, ZeroDivisionError))


class TestPerformanceIntegration:
    """Test performance integration with realistic data sizes."""

    @pytest.mark.integration
    @pytest.mark.slow
    def test_large_dataset_performance(self):
        """Test performance with large datasets."""
        # Create a scenario with many debts and long payoff periods
        many_debts = []
        for i in range(20):  # 20 debts
            many_debts.append(
                Debt(
                    f"Debt {i+1}",
                    5000.0 + (i * 1000),
                    100.0 + (i * 10),
                    5.0 + (i * 0.5),
                    (i % 28) + 1,
                )
            )

        high_income = [
            Income("High Salary", 15000.0, "monthly", date(2024, 1, 1)),
        ]

        many_expenses = []
        for i in range(10):  # 10 recurring expenses
            many_expenses.append(
                RecurringExpense(
                    f"Expense {i+1}",
                    200.0 + (i * 50),
                    "monthly",
                    (i % 28) + 1,
                    date(2024, 1, 1),
                )
            )

        settings = {
            "emergency_fund": 10000.0,
            "current_bank_balance": 25000.0,
            "optimization_goal": "minimize_interest",
        }

        # Should handle large dataset efficiently
        optimizer = DebtOptimizer(
            many_debts, high_income, many_expenses, [], [], settings
        )

        import time

        start_time = time.time()

        result = optimizer.optimize_debt_strategy(
            OptimizationGoal.MINIMIZE_INTEREST, 2000.0
        )

        end_time = time.time()
        processing_time = end_time - start_time

        # Should complete in reasonable time (less than 30 seconds for this size)
        assert processing_time < 30.0

        # Should produce valid results
        assert result.total_months_to_freedom > 0
        assert len(result.payment_schedule) > 0
        # Payment schedule includes debt payments, income events, expenses, etc.
        # So it can be much larger than just debt payments
        # Reasonable upper bound is monthly events * months * (debts + income + expenses)
        max_expected_entries = (
            result.total_months_to_freedom
            * 30
            * (len(many_debts) + len(high_income) + len(many_expenses))
        )
        assert len(result.payment_schedule) <= max_expected_entries

    @pytest.mark.integration
    @pytest.mark.slow
    def test_memory_usage_stability(self):
        """Test memory usage stability across multiple operations."""
        import gc

        # Run multiple optimization cycles
        for cycle in range(5):
            # Create fresh data for each cycle
            debts = [
                Debt(f"Debt {cycle}_1", 5000.0, 150.0, 18.0, 15),
                Debt(f"Debt {cycle}_2", 8000.0, 200.0, 22.0, 10),
            ]

            income = [
                Income(f"Income {cycle}", 5000.0, "monthly", date(2024, 1, 1)),
            ]

            expenses = [
                RecurringExpense(
                    f"Expense {cycle}", 2000.0, "monthly", 1, date(2024, 1, 1)
                ),
            ]

            # Run complete workflow
            optimizer = DebtOptimizer(debts, income, expenses)
            result = optimizer.optimize_debt_strategy(
                OptimizationGoal.MINIMIZE_INTEREST, 300.0
            )
            comparison = optimizer.compare_strategies(300.0)

            # Create charts
            viz = DebtVisualization()
            fig1 = viz.plot_debt_progression(
                pd.DataFrame({"month": [1, 2, 3], "Total Debt": [1000, 800, 600]})
            )
            fig2 = viz.plot_strategy_comparison(comparison)

            # Clean up explicitly
            import matplotlib.pyplot as plt

            plt.close(fig1)
            plt.close(fig2)
            del optimizer, result, comparison

            # Force garbage collection
            gc.collect()

        # If we get here without memory errors, test passes
        assert True


@pytest.mark.integration
def test_end_to_end_realistic_scenario():
    """Test a completely realistic end-to-end scenario."""
    temp_dir = Path(tempfile.mkdtemp())

    try:
        # Realistic financial scenario for a typical household
        realistic_debts = [
            Debt("Chase Freedom Card", 4500.0, 135.0, 21.99, 15),
            Debt("Discover Card", 2800.0, 84.0, 18.99, 22),
            Debt("Car Loan", 28000.0, 485.0, 4.9, 8),
            Debt("Student Loan", 45000.0, 350.0, 5.5, 25),
        ]

        realistic_income = [
            Income("Primary Salary", 3800.0, "bi-weekly", date(2024, 1, 1)),
            Income("Spouse Salary", 2200.0, "bi-weekly", date(2024, 1, 1)),
            Income("Freelance Work", 650.0, "monthly", date(2024, 1, 1)),
        ]

        realistic_expenses = [
            RecurringExpense("Mortgage", 2200.0, "monthly", 1, date(2024, 1, 1)),
            RecurringExpense("Utilities", 280.0, "monthly", 15, date(2024, 1, 1)),
            RecurringExpense("Groceries", 180.0, "weekly", 1, date(2024, 1, 1)),
            RecurringExpense("Childcare", 1200.0, "monthly", 5, date(2024, 1, 1)),
            RecurringExpense("Insurance", 320.0, "monthly", 10, date(2024, 1, 1)),
        ]

        realistic_future_income = [
            FutureIncome(
                "Annual Bonus", 6000.0, date.today() + timedelta(days=60), "once"
            ),
            FutureIncome(
                "Tax Refund", 2800.0, date.today() + timedelta(days=90), "once"
            ),
            FutureIncome(
                "Promotion Raise", 400.0, date.today() + timedelta(days=210), "monthly"
            ),
        ]

        realistic_future_expenses = [
            FutureExpense(
                "Vacation", 4000.0, date.today() + timedelta(days=180), "once"
            ),
            FutureExpense(
                "Home Repair", 8000.0, date.today() + timedelta(days=120), "once"
            ),
            FutureExpense(
                "Car Maintenance", 150.0, date.today() + timedelta(days=30), "quarterly"
            ),
        ]

        realistic_settings = {
            "emergency_fund": 15000.0,
            "current_bank_balance": 8500.0,
            "optimization_goal": "minimize_interest",
        }

        # Complete workflow
        # 1. Validate scenario
        is_valid, messages = validate_financial_scenario(
            realistic_debts, realistic_income, realistic_expenses, realistic_settings
        )
        assert (
            is_valid is True
            or len([m for m in messages if not m.startswith("Warning:")]) == 0
        )

        # 2. Run optimization
        optimizer = DebtOptimizer(
            realistic_debts,
            realistic_income,
            realistic_expenses,
            realistic_future_income,
            realistic_future_expenses,
            realistic_settings,
        )

        # 3. Test all optimization strategies
        strategies = [
            OptimizationGoal.MINIMIZE_INTEREST,
            OptimizationGoal.MINIMIZE_TIME,
            OptimizationGoal.MAXIMIZE_CASHFLOW,
        ]

        results = {}
        for strategy in strategies:
            result = optimizer.optimize_debt_strategy(strategy, 400.0)
            results[strategy.value] = result

            # Each result should be reasonable for this scenario
            # With high income, debt can be paid off quickly, so allow shorter times
            assert (
                1 <= result.total_months_to_freedom <= 240
            )  # Allow 1 month to 20 years
            assert result.total_interest_paid >= 0
            assert len(result.payment_schedule) > 0

        # 4. Generate comprehensive report
        debt_summary = optimizer.generate_debt_summary()
        comparison_data = optimizer.compare_strategies(400.0)

        assert debt_summary["total_debt"] > 75000  # Should be around 80k total debt
        assert debt_summary["monthly_income"] > 12000  # Should be substantial income
        assert len(comparison_data) >= 3  # Should compare multiple strategies

        # 5. Create and save report
        report_path = temp_dir / "realistic_scenario_report.xlsx"
        writer = ExcelReportWriter(str(report_path))
        writer.create_comprehensive_report(
            results["minimize_interest"], debt_summary, comparison_data
        )
        assert report_path.exists()

        # 6. Generate visualizations
        viz = DebtVisualization()
        fig1 = viz.plot_debt_progression(
            pd.DataFrame(
                {
                    "month": [1, 2, 3, 4, 5],
                    "Total Debt": [80000, 75000, 70000, 65000, 60000],
                }
            )
        )
        fig2 = viz.plot_strategy_comparison(comparison_data)

        assert fig1 is not None
        assert fig2 is not None

        import matplotlib.pyplot as plt

        plt.close(fig1)
        plt.close(fig2)

        # 7. Test simple report generation
        simple_report_path = temp_dir / "realistic_simple_report.xlsx"
        generate_simple_summary_report(
            str(simple_report_path), results["minimize_time"], debt_summary
        )
        assert simple_report_path.exists()

    finally:
        shutil.rmtree(temp_dir)
