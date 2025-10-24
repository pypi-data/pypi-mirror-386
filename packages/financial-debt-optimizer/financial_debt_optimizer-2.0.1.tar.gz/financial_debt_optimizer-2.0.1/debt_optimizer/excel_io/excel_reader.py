from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.financial_calc import (
    Debt,
    FutureExpense,
    FutureIncome,
    Income,
    RecurringExpense,
)


class ExcelReader:
    """Read financial data from Excel files."""

    def __init__(self, file_path: str):
        """Initialize with path to Excel file."""
        self.file_path = file_path
        path_obj = Path(file_path)
        if not path_obj.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

    def read_all_data(
        self,
    ) -> Tuple[
        List[Debt],
        List[Income],
        List[RecurringExpense],
        List[FutureIncome],
        List[FutureExpense],
        Dict[str, Any],
    ]:
        """Read all data from Excel file and return debts, income, recurring expenses,
        future income, future expenses, and settings.
        """
        debts = self.read_debts()
        income = self.read_income()
        recurring_expenses = self.read_recurring_expenses()
        future_income = self.read_future_income()
        future_expenses = self.read_future_expenses()
        settings = self.read_settings()
        return (
            debts,
            income,
            recurring_expenses,
            future_income,
            future_expenses,
            settings,
        )

    def read_debts(self, sheet_name: str = "Debts") -> List[Debt]:
        """Read debt information from Excel file."""
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        except ValueError as e:
            raise ValueError(f"Could not read '{sheet_name}' sheet: {e}")

        # Normalize column names first
        df.columns = df.columns.str.lower().str.replace(" ", "_")

        # Validate required columns
        required_columns = [
            "name",
            "balance",
            "min_payment",
            "interest_rate",
            "due_date",
        ]
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            raise ValueError(
                f"Missing required columns in {sheet_name} sheet: {missing_columns}"
            )

        # Filter out rows with empty names or instruction text
        df = df.dropna(subset=["name"])
        df = df[~df["name"].astype(str).str.contains("Instructions:", na=False)]
        df = df[~df["name"].astype(str).str.startswith("•", na=False)]

        debts = []
        for index, row in df.iterrows():
            try:
                # Handle interest rate format detection
                raw_interest_rate = float(row["interest_rate"])
                # If the rate is between 0 and 1, it's already in decimal format (from Excel %)  # noqa: E501
                # Convert back to percentage for internal storage
                if 0 <= raw_interest_rate <= 1:
                    interest_rate = raw_interest_rate * 100
                else:
                    # Otherwise assume it's in percentage format already
                    interest_rate = raw_interest_rate

                debt = Debt(
                    name=str(row["name"]).strip(),
                    balance=float(row["balance"]),
                    minimum_payment=float(row["min_payment"]),
                    interest_rate=interest_rate,
                    due_date=int(row["due_date"]),
                )
                debts.append(debt)
            except (ValueError, TypeError) as e:
                raise ValueError(f"Error reading debt data at row {index + 2}: {e}")

        # Return empty list for empty templates instead of raising error
        # This allows the calling code to handle validation appropriately

        return debts

    def read_income(self, sheet_name: str = "Income") -> List[Income]:
        """Read income information from Excel file."""
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        except ValueError as e:
            raise ValueError(f"Could not read '{sheet_name}' sheet: {e}")

        # Normalize column names first
        df.columns = df.columns.str.lower().str.replace(" ", "_")

        # Validate required columns
        required_columns = ["source", "amount", "frequency", "start_date"]
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            raise ValueError(
                f"Missing required columns in {sheet_name} sheet: {missing_columns}"
            )

        # Filter out rows with empty sources or instruction text
        df = df.dropna(subset=["source"])
        df = df[~df["source"].astype(str).str.contains("Instructions:", na=False)]
        df = df[~df["source"].astype(str).str.startswith("•", na=False)]

        income_sources = []
        for index, row in df.iterrows():
            try:
                # Parse start_date with better error handling
                start_date_val = row["start_date"]
                if pd.isna(start_date_val):
                    start_date_val = date.today()
                elif isinstance(start_date_val, str):
                    try:
                        start_date_val = datetime.strptime(
                            start_date_val, "%Y-%m-%d"
                        ).date()
                    except ValueError:
                        # Try alternative formats if the first fails
                        try:
                            start_date_val = pd.to_datetime(start_date_val).date()
                        except (ValueError, TypeError, pd.errors.ParserError):
                            start_date_val = date.today()
                elif isinstance(start_date_val, datetime):
                    # Validate datetime is reasonable
                    if start_date_val.year > 9999 or start_date_val.year < 1900:
                        start_date_val = date.today()
                    else:
                        start_date_val = start_date_val.date()
                elif isinstance(start_date_val, pd.Timestamp):
                    # Validate timestamp is reasonable
                    if start_date_val.year > 9999 or start_date_val.year < 1900:
                        start_date_val = date.today()
                    else:
                        start_date_val = start_date_val.date()

                income = Income(
                    source=str(row["source"]).strip(),
                    amount=float(row["amount"]),
                    frequency=str(row["frequency"]).strip().lower(),
                    start_date=start_date_val,
                )
                income_sources.append(income)
            except (ValueError, TypeError) as e:
                raise ValueError(f"Error reading income data at row {index + 2}: {e}")

        if not income_sources:
            raise ValueError("No valid income records found")

        return income_sources

    def read_recurring_expenses(
        self, sheet_name: str = "Recurring Expenses"
    ) -> List[RecurringExpense]:
        """Read recurring expenses from Excel file."""
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        except ValueError:
            # Recurring Expenses sheet is optional, return empty list
            return []

        # Normalize column names first
        df.columns = df.columns.str.lower().str.replace(" ", "_")

        # Validate required columns
        required_columns = [
            "description",
            "amount",
            "frequency",
            "due_date",
            "start_date",
        ]
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            raise ValueError(
                f"Missing required columns in {sheet_name} sheet: {missing_columns}"
            )

        # Filter out rows with empty descriptions or instruction text
        df = df.dropna(subset=["description"])
        df = df[~df["description"].astype(str).str.contains("Instructions:", na=False)]
        df = df[~df["description"].astype(str).str.startswith("•", na=False)]

        recurring_expenses = []
        for index, row in df.iterrows():
            try:
                # Parse start_date with better error handling
                start_date_val = row["start_date"]
                if pd.isna(start_date_val):
                    start_date_val = date.today()
                elif isinstance(start_date_val, str):
                    try:
                        start_date_val = datetime.strptime(
                            start_date_val, "%Y-%m-%d"
                        ).date()
                    except ValueError:
                        try:
                            start_date_val = pd.to_datetime(start_date_val).date()
                        except (ValueError, TypeError, pd.errors.ParserError):
                            start_date_val = date.today()
                elif isinstance(start_date_val, datetime):
                    if start_date_val.year > 9999 or start_date_val.year < 1900:
                        start_date_val = date.today()
                    else:
                        start_date_val = start_date_val.date()
                elif isinstance(start_date_val, pd.Timestamp):
                    if start_date_val.year > 9999 or start_date_val.year < 1900:
                        start_date_val = date.today()
                    else:
                        start_date_val = start_date_val.date()

                expense = RecurringExpense(
                    description=str(row["description"]).strip(),
                    amount=float(row["amount"]),
                    frequency=str(row["frequency"]).strip().lower(),
                    due_date=int(row["due_date"]),
                    start_date=start_date_val,
                )
                recurring_expenses.append(expense)
            except (ValueError, TypeError) as e:
                raise ValueError(
                    f"Error reading recurring expense data at row {index + 2}: {e}"
                )

        return recurring_expenses

    def read_future_income(
        self, sheet_name: str = "Future Income"
    ) -> List[FutureIncome]:
        """Read future income events from Excel file - supports both one-time and recurring income."""  # noqa: E501
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        except ValueError:
            # Future Income sheet is optional, return empty list
            return []

        # Normalize column names first
        df.columns = df.columns.str.lower().str.replace(" ", "_")

        # Check for new format (with frequency support) vs legacy format
        has_new_format = "start_date" in df.columns and "frequency" in df.columns
        has_legacy_format = "date" in df.columns

        if has_new_format:
            # New format with recurrence support
            required_columns = ["description", "amount", "start_date"]
        elif has_legacy_format:
            # Legacy format - one-time events only
            required_columns = ["description", "amount", "date"]
        else:
            raise ValueError(
                "Future Income sheet must have either (description, amount, date) for "
                "one-time events or (description, amount, start_date) for recurring events"  # noqa: E501
            )

        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(
                f"Missing required columns in {sheet_name} sheet: {missing_columns}"
            )

        # Filter out rows with empty descriptions or instruction text
        df = df.dropna(subset=["description"])
        df = df[~df["description"].astype(str).str.contains("Instructions:", na=False)]
        df = df[~df["description"].astype(str).str.startswith("•", na=False)]

        future_income = []
        for index, row in df.iterrows():
            try:
                if has_new_format:
                    # Parse start_date
                    start_date = self._parse_date(
                        row["start_date"], f"start_date at row {index + 2}"
                    )
                    if not start_date or start_date <= date.today():
                        continue  # Skip past or invalid dates

                    # Parse optional frequency
                    frequency = None
                    if "frequency" in row and pd.notna(row["frequency"]):
                        frequency = str(row["frequency"]).strip().lower()
                        if frequency == "" or frequency == "none":
                            frequency = None
                        elif frequency == "once":
                            frequency = "once"  # Keep 'once' as explicit frequency

                    # Parse optional end_date
                    end_date = None
                    if "end_date" in row and pd.notna(row["end_date"]):
                        end_date = self._parse_date(
                            row["end_date"], f"end_date at row {index + 2}"
                        )

                    income = FutureIncome(
                        description=str(row["description"]).strip(),
                        amount=float(row["amount"]),
                        start_date=start_date,
                        frequency=frequency,
                        end_date=end_date,
                    )
                else:
                    # Legacy format - one-time event
                    income_date = self._parse_date(
                        row["date"], f"date at row {index + 2}"
                    )
                    if not income_date or income_date <= date.today():
                        continue  # Skip past or invalid dates

                    income = FutureIncome(
                        description=str(row["description"]).strip(),
                        amount=float(row["amount"]),
                        date=income_date,  # Uses legacy date field
                    )

                future_income.append(income)

            except (ValueError, TypeError) as e:
                raise ValueError(
                    f"Error reading future income data at row {index + 2}: {e}"
                )

        return future_income

    def _parse_date(self, date_value, context: str) -> Optional[date]:
        """Parse a date value from Excel with comprehensive error handling."""
        if pd.isna(date_value):
            return None

        try:
            if isinstance(date_value, str):
                # Try multiple date formats
                for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"]:
                    try:
                        return datetime.strptime(date_value, fmt).date()
                    except ValueError:
                        continue
                # If specific formats fail, try pandas
                try:
                    return pd.to_datetime(date_value).date()
                except (ValueError, TypeError, pd.errors.ParserError):
                    return None
            elif isinstance(date_value, datetime):
                if 1900 <= date_value.year <= 9999:
                    return date_value.date()
                else:
                    return None
            elif isinstance(date_value, pd.Timestamp):
                if 1900 <= date_value.year <= 9999:
                    return date_value.date()
                else:
                    return None
            elif hasattr(date_value, "date"):
                return date_value.date()
        except (ValueError, TypeError, AttributeError, OverflowError):
            # Log the specific error for debugging purposes
            pass

        return None

    def read_future_expenses(
        self, sheet_name: str = "Future Expenses"
    ) -> List[FutureExpense]:
        """Read future expense events from Excel file - supports both one-time and recurring expenses."""  # noqa: E501
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        except ValueError:
            # Future Expenses sheet is optional, return empty list
            return []

        # Normalize column names first
        df.columns = df.columns.str.lower().str.replace(" ", "_")

        # Check for new format (with frequency support) vs legacy format
        has_new_format = "start_date" in df.columns and "frequency" in df.columns
        has_legacy_format = "date" in df.columns

        if has_new_format:
            # New format with recurrence support
            required_columns = ["description", "amount", "start_date"]
        elif has_legacy_format:
            # Legacy format - one-time events only
            required_columns = ["description", "amount", "date"]
        else:
            raise ValueError(
                "Future Expenses sheet must have either (description, amount, date) for "  # noqa: E501
                "one-time events or (description, amount, start_date) for recurring events"  # noqa: E501
            )

        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(
                f"Missing required columns in {sheet_name} sheet: {missing_columns}"
            )

        # Filter out rows with empty descriptions or instruction text
        df = df.dropna(subset=["description"])
        df = df[~df["description"].astype(str).str.contains("Instructions:", na=False)]
        df = df[~df["description"].astype(str).str.startswith("•", na=False)]

        future_expenses = []
        for index, row in df.iterrows():
            try:
                if has_new_format:
                    # Parse start_date
                    start_date = self._parse_date(
                        row["start_date"], f"start_date at row {index + 2}"
                    )
                    if not start_date or start_date <= date.today():
                        continue  # Skip past or invalid dates

                    # Parse optional frequency
                    frequency = None
                    if "frequency" in row and pd.notna(row["frequency"]):
                        frequency = str(row["frequency"]).strip().lower()
                        if frequency == "" or frequency == "none":
                            frequency = None
                        elif frequency == "once":
                            frequency = "once"  # Keep 'once' as explicit frequency

                    # Parse optional end_date
                    end_date = None
                    if "end_date" in row and pd.notna(row["end_date"]):
                        end_date = self._parse_date(
                            row["end_date"], f"end_date at row {index + 2}"
                        )

                    expense = FutureExpense(
                        description=str(row["description"]).strip(),
                        amount=float(row["amount"]),
                        start_date=start_date,
                        frequency=frequency,
                        end_date=end_date,
                    )
                else:
                    # Legacy format - one-time event
                    expense_date = self._parse_date(
                        row["date"], f"date at row {index + 2}"
                    )
                    if not expense_date or expense_date <= date.today():
                        continue  # Skip past or invalid dates

                    expense = FutureExpense(
                        description=str(row["description"]).strip(),
                        amount=float(row["amount"]),
                        date=expense_date,  # Uses legacy date field
                    )

                future_expenses.append(expense)

            except (ValueError, TypeError) as e:
                raise ValueError(
                    f"Error reading future expense data at row {index + 2}: {e}"
                )

        return future_expenses

    def read_settings(self, sheet_name: str = "Settings") -> Dict[str, Any]:
        """Read settings from Excel file."""
        default_settings = {
            "current_bank_balance": 2000.0,
            "optimization_goal": "minimize_interest",
        }

        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        except ValueError:
            # Settings sheet is optional, return defaults
            return default_settings

        settings = default_settings.copy()

        # Convert settings from key-value pairs
        if "setting" in df.columns.str.lower() and "value" in df.columns.str.lower():
            df.columns = df.columns.str.lower()
            for _, row in df.iterrows():
                setting_key = str(row["setting"]).lower().replace(" ", "_")
                setting_value = row["value"]

                # Type conversion for known settings
                if setting_key in ["current_bank_balance"]:
                    try:
                        settings[setting_key] = float(setting_value)
                    except (ValueError, TypeError):
                        pass
                elif setting_key == "optimization_goal":
                    valid_goals = [
                        "minimize_interest",
                        "minimize_time",
                        "maximize_cashflow",
                    ]
                    if str(setting_value).lower() in valid_goals:
                        settings[setting_key] = str(setting_value).lower()
                else:
                    settings[setting_key] = setting_value

        return settings


class ExcelTemplateGenerator:
    """Generate Excel templates for data input."""

    @staticmethod
    def generate_template(output_path: str, include_sample_data: bool = True):
        """Generate a template Excel file with all required sheets."""
        workbook = Workbook()

        # Remove default sheet
        workbook.remove(workbook.active)

        # Create sheets
        ExcelTemplateGenerator._create_debts_sheet(workbook, include_sample_data)
        ExcelTemplateGenerator._create_income_sheet(workbook, include_sample_data)
        ExcelTemplateGenerator._create_recurring_expenses_sheet(
            workbook, include_sample_data
        )
        ExcelTemplateGenerator._create_future_income_sheet(
            workbook, include_sample_data
        )
        ExcelTemplateGenerator._create_future_expenses_sheet(
            workbook, include_sample_data
        )
        ExcelTemplateGenerator._create_settings_sheet(workbook, include_sample_data)

        # Save workbook
        workbook.save(output_path)

    @staticmethod
    def _create_debts_sheet(workbook: Workbook, include_sample: bool):
        """Create the debts sheet with headers and formatting."""
        sheet = workbook.create_sheet("Debts", 0)

        # Headers
        headers = ["Name", "Balance", "Min Payment", "Interest Rate", "Due Date"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        if include_sample:
            # Sample data
            sample_data = [
                ["Credit Card 1", 5000.00, 150.00, 18.99, 15],
                ["Auto Loan", 12000.00, 325.00, 4.50, 10],
                ["Personal Loan", 3000.00, 120.00, 12.50, 25],
            ]

            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    if col_idx in [2, 3]:  # Balance and Min Payment columns
                        cell.number_format = "$#,##0.00"
                    elif col_idx == 4:  # Interest Rate column
                        cell.number_format = "0.00%"

        # Column widths
        column_widths = [20, 15, 15, 15, 12]
        for col, width in enumerate(column_widths, 1):
            sheet.column_dimensions[
                sheet.cell(row=1, column=col).column_letter
            ].width = width

        # Add instructions
        instructions = [
            "",
            "Instructions:",
            "• Name: Descriptive name for the debt",
            "• Balance: Current outstanding balance",
            "• Min Payment: Required minimum monthly payment",
            "• Interest Rate: Annual percentage rate (e.g., 18.99 for 18.99%)",
            "• Due Date: Day of month payment is due (1-31)",
        ]

        for idx, instruction in enumerate(instructions):
            cell = sheet.cell(row=6 + idx, column=1)
            cell.value = instruction
            if idx == 1:  # "Instructions:" header
                cell.font = Font(bold=True)

    @staticmethod
    def _create_income_sheet(workbook: Workbook, include_sample: bool):
        """Create the income sheet with headers and formatting."""
        sheet = workbook.create_sheet("Income", 1)

        # Headers
        headers = ["Source", "Amount", "Frequency", "Start Date"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        if include_sample:
            # Sample data
            sample_data = [
                ["Salary", 3500.00, "bi-weekly", "2024-01-05"],
                ["Freelance", 800.00, "monthly", "2024-01-01"],
            ]

            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    if col_idx == 2:  # Amount column
                        cell.number_format = "$#,##0.00"
                    elif col_idx == 4:  # Date column
                        cell.number_format = "yyyy-mm-dd"

        # Column widths
        column_widths = [20, 15, 15, 15]
        for col, width in enumerate(column_widths, 1):
            sheet.column_dimensions[
                sheet.cell(row=1, column=col).column_letter
            ].width = width

        # Add instructions
        instructions = [
            "",
            "Instructions:",
            "• Source: Description of income source",
            "• Amount: Income amount per frequency period",
            "• Frequency: weekly, bi-weekly, monthly, etc.",
            "• Start Date: When this income starts (YYYY-MM-DD format)",
        ]

        for idx, instruction in enumerate(instructions):
            cell = sheet.cell(row=5 + idx, column=1)
            cell.value = instruction
            if idx == 1:  # "Instructions:" header
                cell.font = Font(bold=True)

    @staticmethod
    def _create_recurring_expenses_sheet(workbook: Workbook, include_sample: bool):
        """Create the recurring expenses sheet with headers and formatting."""
        sheet = workbook.create_sheet("Recurring Expenses", 2)

        # Headers
        headers = ["Description", "Amount", "Frequency", "Due Date", "Start Date"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        if include_sample:
            # Sample data
            sample_data = [
                ["Bank Fee", 2.00, "monthly", 1, "2024-01-01"],
                ["Netflix", 15.99, "monthly", 15, "2024-01-01"],
                ["Car Insurance", 85.00, "monthly", 25, "2024-01-01"],
                ["Additional Cost", 2.00, "bi-weekly", 15, "2024-01-01"],
            ]

            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    if col_idx == 2:  # Amount column
                        cell.number_format = "$#,##0.00"
                    elif col_idx == 5:  # Date column
                        cell.number_format = "yyyy-mm-dd"

        # Column widths
        column_widths = [25, 15, 15, 12, 15]
        for col, width in enumerate(column_widths, 1):
            sheet.column_dimensions[
                sheet.cell(row=1, column=col).column_letter
            ].width = width

        # Add instructions
        instructions = [
            "",
            "Instructions:",
            "• Description: What the expense is for",
            "• Amount: Cost per frequency period",
            "• Frequency: bi-weekly, monthly, quarterly, annually",
            "• Due Date: Day of month/period when payment is due (1-31) - ignored for bi-weekly",  # noqa: E501
            "• Start Date: When this expense starts (YYYY-MM-DD format)",
        ]

        for idx, instruction in enumerate(instructions):
            cell = sheet.cell(row=6 + idx, column=1)
            cell.value = instruction
            if idx == 1:  # "Instructions:" header
                cell.font = Font(bold=True)

    @staticmethod
    def _create_future_income_sheet(workbook: Workbook, include_sample: bool):
        """Create the future income sheet with headers and formatting - supports recurring income."""  # noqa: E501
        sheet = workbook.create_sheet("Future Income", 3)

        # Headers for new format with recurrence support
        headers = ["Description", "Amount", "Start Date", "Frequency", "End Date"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        if include_sample:
            # Sample data showing both one-time and recurring income
            sample_data = [
                [
                    "Annual Bonus",
                    5000.00,
                    "2026-03-15",
                    "",
                    "",
                ],  # One-time (no frequency)
                ["Tax Refund", 1200.00, "2025-04-01", "", ""],  # One-time
                [
                    "Salary Increase",
                    500.00,
                    "2026-01-19",
                    "monthly",
                    "",
                ],  # Your requested increase
                [
                    "Side Income",
                    250.00,
                    "2025-12-01",
                    "weekly",
                    "2026-06-30",
                ],  # Recurring with end
                [
                    "Quarterly Bonus",
                    1000.00,
                    "2025-12-15",
                    "quarterly",
                    "",
                ],  # Recurring indefinitely
            ]

            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    if col_idx == 2:  # Amount column
                        cell.number_format = "$#,##0.00"
                    elif col_idx in [3, 5]:  # Date columns
                        if value:  # Only format if there's a value
                            cell.number_format = "yyyy-mm-dd"

        # Column widths
        column_widths = [25, 15, 15, 15, 15]
        for col, width in enumerate(column_widths, 1):
            sheet.column_dimensions[
                sheet.cell(row=1, column=col).column_letter
            ].width = width

        # Add instructions
        instructions = [
            "",
            "Instructions:",
            "• Description: What the income is for",
            "• Amount: Income amount per occurrence",
            "• Start Date: When income begins (YYYY-MM-DD format)",
            "• Frequency: Use 'once' for one-time, or: daily, weekly, bi-weekly, monthly, quarterly, semi-annually, annually",  # noqa: E501
            "• End Date: When recurring income stops (leave blank for indefinite, ignored for one-time)",  # noqa: E501
            "",
            "Examples:",
            "• One-time: Description='Tax Refund', Amount=1200, Start Date=2025-04-01, Frequency=once",  # noqa: E501
            "• Recurring: Description='Raise', Amount=500, Start Date=2026-01-19, Frequency=monthly",  # noqa: E501
            "• Limited recurring: Description='Contract', Amount=2000, Start Date=2025-12-01, "  # noqa: E501
            "Frequency=monthly, End Date=2026-06-30",
        ]

        for idx, instruction in enumerate(instructions):
            cell = sheet.cell(row=7 + idx, column=1)
            cell.value = instruction
            if idx == 1:  # "Instructions:" header
                cell.font = Font(bold=True)
            elif idx in [
                8,
            ]:  # "Examples:" header
                cell.font = Font(bold=True)

    @staticmethod
    def _create_future_expenses_sheet(workbook: Workbook, include_sample: bool):
        """Create the future expenses sheet with headers and formatting - supports recurring expenses."""  # noqa: E501
        sheet = workbook.create_sheet("Future Expenses", 4)

        # Headers for new format with recurrence support
        headers = ["Description", "Amount", "Start Date", "Frequency", "End Date"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        if include_sample:
            # Sample data showing both one-time and recurring expenses
            sample_data = [
                ["Car Repair", 800.00, "2025-12-15", "", ""],  # One-time expense
                ["Home Improvement", 2500.00, "2026-03-01", "", ""],  # One-time expense
                [
                    "New Subscription",
                    9.99,
                    "2025-11-01",
                    "monthly",
                    "",
                ],  # Monthly subscription
                [
                    "Insurance Increase",
                    25.00,
                    "2026-01-01",
                    "monthly",
                    "2026-12-31",
                ],  # Limited recurring
                ["Annual Fee", 99.00, "2025-12-01", "annually", ""],  # Annual recurring
            ]

            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    if col_idx == 2:  # Amount column
                        cell.number_format = "$#,##0.00"
                    elif col_idx in [3, 5]:  # Date columns
                        if value:  # Only format if there's a value
                            cell.number_format = "yyyy-mm-dd"

        # Column widths
        column_widths = [25, 15, 15, 15, 15]
        for col, width in enumerate(column_widths, 1):
            sheet.column_dimensions[
                sheet.cell(row=1, column=col).column_letter
            ].width = width

        # Add instructions
        instructions = [
            "",
            "Instructions:",
            "• Description: What the expense is for",
            "• Amount: Expense amount per occurrence",
            "• Start Date: When expense begins (YYYY-MM-DD format)",
            "• Frequency: Use 'once' for one-time, or: daily, weekly, bi-weekly, monthly, quarterly, semi-annually, annually",  # noqa: E501
            "• End Date: When recurring expense stops (leave blank for indefinite, ignored for one-time)",  # noqa: E501
            "",
            "Examples:",
            "• One-time: Description='Car Repair', Amount=800, Start Date=2025-12-15, Frequency=once",  # noqa: E501
            "• Recurring: Description='New Subscription', Amount=9.99, Start Date=2025-11-01, Frequency=monthly",  # noqa: E501
            "• Limited recurring: Description='Temp Insurance', Amount=25, Start Date=2026-01-01, "  # noqa: E501
            "Frequency=monthly, End Date=2026-12-31",
        ]

        for idx, instruction in enumerate(instructions):
            cell = sheet.cell(row=7 + idx, column=1)
            cell.value = instruction
            if idx == 1:  # "Instructions:" header
                cell.font = Font(bold=True)
            elif idx in [
                8,
            ]:  # "Examples:" header
                cell.font = Font(bold=True)

    @staticmethod
    def _create_settings_sheet(workbook: Workbook, include_sample: bool):
        """Create the settings sheet with default values."""
        sheet = workbook.create_sheet("Settings", 5)

        # Headers
        headers = ["Setting", "Value"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Default settings
        default_settings = [
            ["Emergency Fund", 1000.00],
            ["Current Bank Balance", 2000.00],
            ["Optimization Goal", "minimize_interest"],
        ]

        from typing import List, Tuple, Union, cast

        setting_list = cast(List[Tuple[str, Union[str, float]]], default_settings)
        for row_idx, setting_data in enumerate(setting_list, 2):
            setting: str
            value: Union[str, float]
            setting, value = setting_data
            sheet.cell(row=row_idx, column=1).value = setting
            cell = sheet.cell(row=row_idx, column=2)
            cell.value = value
            if isinstance(value, (int, float)):
                cell.number_format = "$#,##0.00"

        # Column widths
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 20

        # Add instructions
        instructions = [
            "",
            "Settings Help:",
            "• Emergency Fund: Amount to keep as emergency buffer",
            "• Current Bank Balance: Starting cash balance in your account",
            "• Optimization Goal: minimize_interest, minimize_time, or maximize_cashflow",  # noqa: E501
        ]

        for idx, instruction in enumerate(instructions):
            cell = sheet.cell(row=7 + idx, column=1)
            cell.value = instruction
            if idx == 1:  # "Settings Help:" header
                cell.font = Font(bold=True)
