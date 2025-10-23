import os
import sqlite3
from abc import ABC, abstractmethod
from typing import override

import pandas as pd
import requests
from openpyxl import load_workbook  # type: ignore
from pandas import DataFrame


class Datasource(ABC):
    """Base class for all data sources"""

    @abstractmethod
    def extract_data(self, **kwargs):
        """Method to extract data from datasources"""
        raise NotImplementedError("Subclasses must implement this method")


def download_csv(url: str):
    response = requests.get(url)

    if response.status_code == 200:
        local_file_path = os.path.join(os.getcwd(), "downloaded_data.csv")
        with open(local_file_path, "wb") as file:
            file.write(response.content)
            print("Downloaded file Successfully📎")
            return local_file_path
    else:
        raise Exception(
            f"Failed to download file. Status code: {response.status_code}"
        )


class CSVSource(Datasource):
    """CSV file data source"""

    def __init__(self, file_path: str | None = None, url: str | None = None):
        super().__init__()
        assert file_path is not None or url is not None, ("File path or URL must be provided")
        self.file_path = file_path
        self.url = url

    @override
    def extract_data(self) -> DataFrame:
        if self.url:
            return pd.read_csv(self.url)
        elif self.file_path:
            return pd.read_csv(self.file_path)
        else:
            raise ValueError("Either file_path or url must be provided")


class SQLiteSource(Datasource):
    """sqlite file data source"""

    def __init__(self, connection_string: str, query: str):
        super().__init__()
        self.connection_string = connection_string
        self.query = query
    
    @override
    def extract_data(self) -> DataFrame:
        with sqlite3.connect(self.connection_string) as conn:
            return pd.read_sql_query(self.query, conn)


class XLSsource(Datasource):
    """Excel worksheet datasource"""

    def __init__(self, file_path: str = "", sheet_name: str = "", *args):
        super().__init__()
        self.file_path = file_path
        self.sheet_name = sheet_name

    @override
    def extract_data(self) -> DataFrame:
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"File {self.file_path} does not exist")
        workbook = load_workbook(self.file_path)
        sheet = workbook[self.sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        workbook.close()
        return pd.DataFrame(data)

        
class ParquetSource(Datasource):
    """
    Parquet datasource
    """
    def __init__(self):
        super().__init__()
        pass
    
    @override
    def extract_data(self):
        pass
