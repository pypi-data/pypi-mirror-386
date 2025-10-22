from __future__ import annotations

import os
import tempfile


def ensure_parent_dir(path: str) -> None:
    """确保目标路径的父目录存在，若不存在则创建。

    参数：
    - `path`: 文件路径。

    返回：
    - 无。
    """
    d = os.path.dirname(os.path.abspath(path))
    if d and not os.path.exists(d):
        os.makedirs(d, exist_ok=True)


def get_column_letter_safe(col_idx: int) -> str:
    """安全地将列号（1 基）转换为列字母（如 1 -> A）。

    参数：
    - `col_idx`: 列号，1 基。

    返回：
    - 列字母字符串。
    """
    try:
        from openpyxl.utils import get_column_letter
    except Exception:
        # 本地回退：仅支持 1..26 的列号
        letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        if 1 <= col_idx <= 26:
            return letters[col_idx - 1]
        raise
    return get_column_letter(col_idx)


def convert_image_to_bmp_temp(path: str) -> str:
    """使用 Pillow 将图片转换为临时 BMP 文件，并返回其路径。"""
    # 中文说明：将 PNG/JPG 等图片转换为临时 BMP 文件，以便在 `.xls` 中插入。
    try:
        from PIL import Image
    except Exception as exc:
        raise RuntimeError(
            "生成 BMP 需要 Pillow 依赖。请安装：pip install 'sheetcraft[images]'"
        ) from exc

    img = Image.open(path)
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".bmp")
    os.close(tmp_fd)
    img.convert("RGB").save(tmp_path, format="BMP")
    return tmp_path
