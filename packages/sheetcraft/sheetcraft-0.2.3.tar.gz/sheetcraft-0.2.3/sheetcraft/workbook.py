from __future__ import annotations

import os
from dataclasses import dataclass
import io
from contextlib import contextmanager
from typing import Any, Dict, Iterable, Optional, Sequence, Tuple, Union

from .styles import (
    apply_openpyxl_style,
    xlsxwriter_format_from_dict,
    xlwt_style_from_dict,
)
from .utils import ensure_parent_dir, get_column_letter_safe
from .validation import (
    add_data_validation_openpyxl,
    add_data_validation_xlsxwriter,
)
from .images import (
    insert_image_openpyxl,
    insert_image_xlsxwriter,
    insert_image_xlwt,
)


CellValue = Union[str, int, float, bool, None]
StyleDict = Dict[str, Any]


@dataclass
class DataValidationSpec:
    """数据有效性（Data Validation）规则规格。

    字段与 openpyxl/xlsxwriter 选项基本一致：
    - `type` 示例：'list'、'whole'、'decimal'、'date'、'time'、'textLength'、'custom'。
    - `operator` 示例：'between'、'notBetween'、'equal'、'notEqual'、'lessThan' 等。
    """

    type: str
    formula1: Optional[str] = None
    formula2: Optional[str] = None
    operator: Optional[str] = None
    allow_blank: bool = True
    show_input_message: bool = False
    prompt_title: Optional[str] = None
    prompt: Optional[str] = None
    show_error_message: bool = True
    error_title: Optional[str] = None
    error: Optional[str] = None


class ExcelWorkbook:
    """统一的 Excel 工作簿操作接口，支持 `.xlsx` 与 `.xls`，并可选启用快速写入。

    功能说明：
    - `.xlsx` 默认使用 openpyxl；设置 `fast=True` 时使用 xlsxwriter 以获得更快写入性能。
    - `.xls` 使用 xlwt（兼容性支持），受格式限制部分功能有所收缩。

    常用方法（语义化 API）：
    - `add_sheet(name)`: 新增工作表。
    - `sheet(name=None)`: 获取或创建工作表（别名，更便捷）。
    - `write_cell(sheet, row, col, value, style=None)`: 写入单个单元格。
    - `write_row(sheet, row_idx, values, styles=None)`: 写入一整行。
    - `write_rows(sheet, start_row, rows, styles=None)`: 批量写入多行。
    - `write_table(sheet, start_row, headers, rows, header_style=None, row_style=None)`: 写入表格（带表头与数据行）。
    - `export_dicts(sheet, start_row, data, header_map=None, order=None)`: 导出字典列表为表格。
    - 行列尺寸与合并：`set_row_height`、`set_column_width`、`merge_cells`。
    - 图片：`insert_image`（支持 PNG/JPG；`.xls` 自动转 BMP）与 `insert_image_in_cell`（按单元格尺寸适配）。
    - 数据有效性：`add_data_validation`。
    - 公式：`set_formula` 或直接传入以 `=` 开头的字符串。
    - 保存：`save(path)`。

    参数说明：
    - `output_path`: 输出文件路径；若使用 xlsxwriter 必填。
    - `file_format`: 文件格式，`xlsx` 或 `xls`；默认根据路径推断。
    - `fast`: 是否启用 xlsxwriter 快速写入（仅 `.xlsx`）。
    - `write_only`: openpyxl 流式写入模式（仅 `.xlsx`）。
    - `apply_format_fix_on_save`: 保存后是否应用格式修复（仅 `.xlsx`）。
    - `format_fix_config`: 格式修复配置（默认启用前缀修复）。
    """

    def __init__(
        self,
        output_path: Optional[str] = None,
        file_format: Optional[str] = None,
        fast: bool = False,
        write_only: bool = False,
        apply_format_fix_on_save: bool = False,
        format_fix_config: Optional[Any] = None,
    ) -> None:
        self.output_path = output_path
        self.file_format = (file_format or self._infer_format(output_path)) or "xlsx"
        self.fast = fast
        self.write_only = write_only
        # 新增：格式修复参数
        self._apply_format_fix_on_save = apply_format_fix_on_save
        self._format_fix_config = format_fix_config

        self._engine: str = self._choose_engine()
        self._wb = None
        self._sheets: Dict[str, Any] = {}
        # 记录尺寸（供 xlsxwriter 计算插入图片缩放）
        self._col_widths: Dict[str, Dict[int, float]] = {}
        self._row_heights: Dict[str, Dict[int, float]] = {}

        self._init_workbook()

    # region engine selection
    def _infer_format(self, path: Optional[str]) -> Optional[str]:
        if not path:
            return None
        ext = os.path.splitext(path)[1].lower()
        if ext == ".xlsx":
            return "xlsx"
        if ext == ".xls":
            return "xls"
        return None

    def _choose_engine(self) -> str:
        if self.file_format == "xlsx" and self.fast:
            return "xlsxwriter"
        if self.file_format == "xls":
            return "xlwt"
        return "openpyxl"

    # endregion

    # region init
    def _init_workbook(self) -> None:
        if self._engine == "openpyxl":
            from openpyxl import Workbook

            self._wb = Workbook(write_only=self.write_only)
            # 在 write_only 模式下，openpyxl 不提供 active 工作表；显式创建一个默认工作表。
            if self.write_only:
                ws = self._wb.create_sheet(title="Sheet1")
                self._sheets[ws.title] = ws
            else:
                # 常规模式下保留默认活动表
                default = self._wb.active
                self._sheets[default.title] = default
            # 标记是否处于内存模式（未提供输出路径）
            self._buffer = None
            self._in_memory = self.output_path is None
        elif self._engine == "xlsxwriter":
            import xlsxwriter

            if not self.output_path:
                # 内存模式：绑定到 BytesIO 以便预览获取字节
                self._buffer = io.BytesIO()
                self._wb = xlsxwriter.Workbook(self._buffer, {"in_memory": True})
                self._in_memory = True
            else:
                ensure_parent_dir(self.output_path)
                self._wb = xlsxwriter.Workbook(self.output_path)
                self._buffer = None
                self._in_memory = False
        elif self._engine == "xlwt":
            try:
                import xlwt  # noqa: F401
            except Exception as exc:
                raise RuntimeError(
                    "xlwt is required for .xls support. Install with: pip install 'sheetcraft[xls]'"
                ) from exc
            import xlwt

            self._wb = xlwt.Workbook(encoding="utf-8")
            self._buffer = None
            self._in_memory = self.output_path is None
        else:
            raise RuntimeError(f"Unknown engine: {self._engine}")

    # endregion

    # region sheet operations
    def add_sheet(self, name: str) -> Any:
        """新增工作表。

        参数：
        - `name`: 工作表名称。

        返回：
        - 新增的工作表对象。
        """
        if self._engine == "openpyxl":
            ws = self._wb.create_sheet(title=name)
            self._sheets[name] = ws
            return ws
        if self._engine == "xlsxwriter":
            ws = self._wb.add_worksheet(name)
            self._sheets[name] = ws
            return ws
        if self._engine == "xlwt":
            ws = self._wb.add_sheet(name)
            self._sheets[name] = ws
            return ws
        raise RuntimeError("Unsupported engine")

    def get_sheet(self, name: Optional[str] = None) -> Any:
        """获取工作表；若 `name` 为空，则返回当前活动或首个工作表。

        参数：
        - `name`: 工作表名称，可为空。

        返回：
        - 工作表对象。
        """
        if name is None:
            # return active or first
            if self._engine == "openpyxl":
                return self._wb.active
            # xlsxwriter/xlwt have no 'active'; return first created
            if self._sheets:
                return next(iter(self._sheets.values()))
            # create default
            return self.add_sheet("Sheet1")
        # 若指定名称不存在，则创建
        if name not in self._sheets:
            return self.add_sheet(name)
        return self._sheets[name]

    # 语义化别名，等价于 `get_sheet`
    def sheet(self, name: Optional[str] = None) -> Any:
        """语义化别名：获取或创建工作表。

        参数：
        - `name`: 工作表名称，可为空。

        返回：
        - 工作表对象。
        """
        return self.get_sheet(name)

    # endregion

    # region write operations
    def write_cell(
        self,
        sheet: Union[str, Any],
        row: int,
        col: int,
        value: CellValue,
        style: Optional[StyleDict] = None,
    ) -> None:
        """写入单个单元格。

        参数：
        - `sheet`: 工作表对象或名称字符串。
        - `row`: 行号（1 基）。
        - `col`: 列号（1 基）。
        - `value`: 单元格值，可为字符串、数值、布尔或 `None`。
        - `style`: 样式字典，参考 `styles.py`。

        返回：
        - 无。
        """
        ws = self._resolve_sheet(sheet)
        if self._engine == "openpyxl":
            c = ws.cell(row=row, column=col, value=value)
            if style:
                apply_openpyxl_style(c, style)
            return
        if self._engine == "xlsxwriter":
            fmt = xlsxwriter_format_from_dict(self._wb, style) if style else None
            if isinstance(value, str) and value.startswith("="):
                ws.write_formula(row - 1, col - 1, value, fmt)
            else:
                ws.write(row - 1, col - 1, value, fmt)
            return
        if self._engine == "xlwt":
            fmt = xlwt_style_from_dict(style) if style else None
            ws.write(row - 1, col - 1, value, fmt)
            return
        raise RuntimeError("Unsupported engine")

    def write_row(
        self,
        sheet: Union[str, Any],
        row_idx: int,
        values: Sequence[CellValue],
        styles: Optional[Sequence[Optional[StyleDict]]] = None,
    ) -> None:
        """写入一整行数据。

        参数：
        - `sheet`: 工作表对象或名称。
        - `row_idx`: 行号（1 基）。
        - `values`: 一行的值序列。
        - `styles`: 与 `values` 对应的样式序列，可为空。

        返回：
        - 无。
        """
        ws = self._resolve_sheet(sheet)
        if self._engine == "openpyxl" and self.write_only:
            # Use WriteOnlyCell for styled streaming
            from openpyxl.cell import WriteOnlyCell

            row_cells = []
            for col_idx, val in enumerate(values, start=1):
                cell = WriteOnlyCell(ws, value=val)
                if styles and styles[col_idx - 1]:
                    apply_openpyxl_style(cell, styles[col_idx - 1])
                row_cells.append(cell)
            ws.append(row_cells)
            return

        for col_idx, val in enumerate(values, start=1):
            style = (
                styles[col_idx - 1] if styles and col_idx - 1 < len(styles) else None
            )
            self.write_cell(ws, row_idx, col_idx, val, style)

    def write_rows(
        self,
        sheet: Union[str, Any],
        start_row: int,
        rows: Iterable[Sequence[CellValue]],
        styles: Optional[Sequence[Optional[StyleDict]]] = None,
    ) -> None:
        """批量写入多行数据。

        参数：
        - `sheet`: 工作表对象或名称。
        - `start_row`: 起始行（1 基）。
        - `rows`: 行数据的可迭代对象，每个元素为一行的值序列。
        - `styles`: 样式序列（应用到每一行的每一列），可为空。

        返回：
        - 无。
        """
        for offset, row_values in enumerate(rows):
            self.write_row(sheet, start_row + offset, row_values, styles)

    # endregion

    # region layout
    def set_row_height(
        self, sheet: Union[str, Any], row_idx: int, height: float
    ) -> None:
        """设置行高。

        参数：
        - `sheet`: 工作表对象或名称。
        - `row_idx`: 行号（1 基）。
        - `height`: 行高，单位为点（`.xls` 下内部转换为 twips）。

        返回：
        - 无。
        """
        ws = self._resolve_sheet(sheet)
        if self._engine == "openpyxl":
            ws.row_dimensions[row_idx].height = height
            # 记录行高
            name = getattr(ws, "title", "Sheet1")
            self._row_heights.setdefault(name, {})[row_idx] = height
            return
        if self._engine == "xlsxwriter":
            ws.set_row(row_idx - 1, height)
            # 记录行高
            name = getattr(ws, "get_name", lambda: next((n for n, w in self._sheets.items() if w is ws), "Sheet1"))()
            self._row_heights.setdefault(name, {})[row_idx] = height
            return
        if self._engine == "xlwt":
            # xlwt height is in twips (1/20 of a point)
            ws.row(row_idx - 1).height = int(height * 20)
            # 记录行高
            name = getattr(ws, "name", next((n for n, w in self._sheets.items() if w is ws), "Sheet1"))
            self._row_heights.setdefault(name, {})[row_idx] = height
            return

    def set_column_width(
        self, sheet: Union[str, Any], col_idx: int, width: float
    ) -> None:
        """设置列宽。

        参数：
        - `sheet`: 工作表对象或名称。
        - `col_idx`: 列号（1 基）。
        - `width`: 列宽（`.xls` 下内部转换为 1/256 字符宽单位）。

        返回：
        - 无。
        """
        ws = self._resolve_sheet(sheet)
        if self._engine == "openpyxl":
            letter = get_column_letter_safe(col_idx)
            ws.column_dimensions[letter].width = width
            # 记录列宽
            name = getattr(ws, "title", "Sheet1")
            self._col_widths.setdefault(name, {})[col_idx] = width
            return
        if self._engine == "xlsxwriter":
            ws.set_column(col_idx - 1, col_idx - 1, width)
            # 记录列宽
            name = getattr(ws, "get_name", lambda: next((n for n, w in self._sheets.items() if w is ws), "Sheet1"))()
            self._col_widths.setdefault(name, {})[col_idx] = width
            return
        if self._engine == "xlwt":
            # approx conversion; xlwt column width is 1/256 of the character width
            ws.col(col_idx - 1).width = int(width * 256)
            # 记录列宽
            name = getattr(ws, "name", next((n for n, w in self._sheets.items() if w is ws), "Sheet1"))
            self._col_widths.setdefault(name, {})[col_idx] = width
            return

    def merge_cells(
        self,
        sheet: Union[str, Any],
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
        style: Optional[StyleDict] = None,
        value: Optional[CellValue] = None,
    ) -> None:
        """合并单元格范围并可选写入值/应用样式。

        参数：
        - `sheet`: 工作表对象或名称。
        - `start_row`, `start_col`, `end_row`, `end_col`: 合并范围（1 基）。
        - `style`: 样式字典，可为空。
        - `value`: 合并单元格的显示值，可为空。

        返回：
        - 无。
        """
        ws = self._resolve_sheet(sheet)
        if self._engine == "openpyxl":
            ws.merge_cells(
                start_row=start_row,
                start_column=start_col,
                end_row=end_row,
                end_column=end_col,
            )
            if value is not None:
                c = ws.cell(row=start_row, column=start_col, value=value)
                if style:
                    apply_openpyxl_style(c, style)
            return
        if self._engine == "xlsxwriter":
            fmt = xlsxwriter_format_from_dict(self._wb, style) if style else None
            # xlsxwriter merge_range requires a value
            ws.merge_range(
                start_row - 1, start_col - 1, end_row - 1, end_col - 1, value or "", fmt
            )
            return
        if self._engine == "xlwt":
            ws.merge(start_row - 1, end_row - 1, start_col - 1, end_col - 1)
            if value is not None:
                fmt = xlwt_style_from_dict(style) if style else None
                ws.write(start_row - 1, start_col - 1, value, fmt)
            return

    # endregion

    # region images
    def insert_image(
        self,
        sheet: Union[str, Any],
        row: int,
        col: int,
        image_path: str,
        width: Optional[int] = None,
        height: Optional[int] = None,
        scale_x: Optional[float] = None,
        scale_y: Optional[float] = None,
    ) -> None:
        """插入图片到指定单元格锚点。

        参数：
        - `sheet`: 工作表对象或名称。
        - `row`, `col`: 锚点（1 基）。
        - `image_path`: 图片路径（支持 PNG/JPG；`.xls` 自动转 BMP）。
        - `width`, `height`: openpyxl 下可设置图片尺寸。
        - `scale_x`, `scale_y`: xlsxwriter 下可设置缩放。

        返回：
        - 无。
        """
        ws = self._resolve_sheet(sheet)
        if self._engine == "openpyxl":
            insert_image_openpyxl(ws, row, col, image_path, width=width, height=height)
            return
        if self._engine == "xlsxwriter":
            insert_image_xlsxwriter(
                ws, row, col, image_path, scale_x=scale_x, scale_y=scale_y
            )
            return
        if self._engine == "xlwt":
            insert_image_xlwt(ws, row, col, image_path)
            return

    # endregion

    # region validation
    def add_data_validation(
        self,
        sheet: Union[str, Any],
        cell_range: str,
        spec: DataValidationSpec,
    ) -> None:
        """为指定范围添加数据有效性校验（`.xlsx` 全量支持；`.xls` 不支持）。

        参数：
        - `sheet`: 工作表对象或名称。
        - `cell_range`: 单元格范围字符串，如 `A1:D10`。
        - `spec`: 数据有效性规则规格，详见 `DataValidationSpec`。

        返回：
        - 无。
        """
        ws = self._resolve_sheet(sheet)
        if self._engine == "openpyxl":
            add_data_validation_openpyxl(ws, cell_range, spec)
            return
        if self._engine == "xlsxwriter":
            first_row, first_col, last_row, last_col = self._parse_range(cell_range)
            add_data_validation_xlsxwriter(
                ws, first_row, first_col, last_row, last_col, spec
            )
            return
        if self._engine == "xlwt":
            # Not supported by xlwt; no-op for legacy format.
            return

    # endregion

    # region formulas
    def set_formula(
        self,
        sheet: Union[str, Any],
        row: int,
        col: int,
        formula: str,
        style: Optional[StyleDict] = None,
    ) -> None:
        """设置单元格公式（自动补全前导 `=`）。

        参数：
        - `sheet`: 工作表对象或名称。
        - `row`, `col`: 目标单元格位置（1 基）。
        - `formula`: 公式字符串，可带或不带 `=`。
        - `style`: 样式字典，可为空。

        返回：
        - 无。
        """
        ws = self._resolve_sheet(sheet)
        if self._engine == "openpyxl":
            c = ws.cell(row=row, column=col, value=f"={formula.lstrip('=')}")
            if style:
                apply_openpyxl_style(c, style)
            return
        if self._engine == "xlsxwriter":
            fmt = xlsxwriter_format_from_dict(self._wb, style) if style else None
            ws.write_formula(row - 1, col - 1, f"={formula.lstrip('=')}", fmt)
            return
        if self._engine == "xlwt":
            import xlwt

            fmt = xlwt_style_from_dict(style) if style else None
            ws.write(row - 1, col - 1, xlwt.Formula(formula.lstrip("=")), fmt)
            return

    # endregion

    # region save
    def save(self, path: Optional[str] = None) -> None:
        """保存工作簿到目标路径。

        参数：
        - `path`: 输出路径；为空时使用初始化时设置的 `output_path`。

        返回：
        - 无。
        """
        target = path or self.output_path
        if not target:
            raise ValueError("Output path must be provided to save the workbook")
        ensure_parent_dir(target)

        if self._engine == "openpyxl":
            self._wb.save(target)
        elif self._engine == "xlsxwriter":
            # xlsxwriter saves on close
            self._wb.close()
        elif self._engine == "xlwt":
            self._wb.save(target)
        else:
            raise RuntimeError("Unsupported engine")

        # 可选：保存后应用格式修复（仅 .xlsx）
        if self.file_format == "xlsx" and self._apply_format_fix_on_save:
            try:
                from .format_fix import fix_xlsx, FormatFixConfig
                cfg = self._format_fix_config or FormatFixConfig()
                fix_xlsx(target, None, cfg)
            except Exception:
                # 修复失败不影响主流程
                pass

    # endregion

    # region preview / bytes
    def get_bytes(self) -> bytes:
        """以字节流形式返回当前工作簿内容，无需写入磁盘。

        注意：
        - 对 `xlsxwriter` 引擎，调用本方法会关闭工作簿（`close()`），后续不可再写。
        - 对 `openpyxl` 与 `xlwt`，会将内容保存到内存缓冲区并返回字节。
        """
        if self._engine == "openpyxl":
            buf = io.BytesIO()
            self._wb.save(buf)
            return buf.getvalue()
        if self._engine == "xlsxwriter":
            # 若为内存模式，数据写入到 BytesIO；否则需要关闭并读取已写入的目标文件
            if getattr(self, "_buffer", None) is None:
                # 非内存模式：关闭后从目标文件读取字节（会写入磁盘）
                self._wb.close()
                with open(self.output_path, "rb") as f:
                    return f.read()
            else:
                self._wb.close()
                return self._buffer.getvalue()
        if self._engine == "xlwt":
            buf = io.BytesIO()
            self._wb.save(buf)
            return buf.getvalue()
        raise RuntimeError("Unsupported engine")

    @contextmanager
    def preview_temp(self):
        """生成临时文件进行预览，退出上下文后自动删除。

        用法：
        >>> wb = ExcelWorkbook()
        >>> with wb.preview_temp() as path:
        ...     # 预览 path 文件
        ...     pass
        """
        import tempfile

        suffix = ".xlsx" if self.file_format == "xlsx" else ".xls"
        fd, path = tempfile.mkstemp(suffix=suffix)
        os.close(fd)
        try:
            # 直接使用内存字节写入临时文件，避免 xlsxwriter 重绑定目标路径
            data = self.get_bytes()
            with open(path, "wb") as f:
                f.write(data)
            # 在预览文件上应用可选格式修复
            if self.file_format == "xlsx" and self._apply_format_fix_on_save:
                try:
                    from .format_fix import fix_xlsx, FormatFixConfig
                    cfg = self._format_fix_config or FormatFixConfig()
                    fix_xlsx(path, None, cfg)
                except Exception:
                    pass
            yield path
        finally:
            try:
                os.remove(path)
            except OSError:
                pass

    # endregion

    # region helpers
    def _resolve_sheet(self, sheet: Union[str, Any]) -> Any:
        """内部辅助：根据名称或对象解析工作表。"""
        if isinstance(sheet, str):
            return self._sheets.get(sheet) or self.add_sheet(sheet)
        return sheet

    def _parse_range(self, cell_range: str) -> Tuple[int, int, int, int]:
        """解析范围字符串（如 `A1:D10`）为 1 基行列边界。"""
        from openpyxl.utils import coordinate_to_tuple

        parts = cell_range.split(":")
        if len(parts) == 1:
            r, c = coordinate_to_tuple(parts[0])
            return r, c, r, c
        r1, c1 = coordinate_to_tuple(parts[0])
        r2, c2 = coordinate_to_tuple(parts[1])
        return r1, c1, r2, c2

    # endregion

    # region semantic helpers
    def write_table(
        self,
        sheet: Union[str, Any],
        start_row: int,
        headers: Sequence[str],
        rows: Iterable[Sequence[CellValue]],
        header_style: Optional[StyleDict] = None,
        row_style: Optional[StyleDict] = None,
    ) -> None:
        """语义化写表：写入表头与数据行。

        参数：
        - `sheet`: 工作表对象或名称。
        - `start_row`: 起始行（1 基）。
        - `headers`: 表头文本序列。
        - `rows`: 数据行的可迭代对象。
        - `header_style`: 表头样式，可为空。
        - `row_style`: 数据行统一样式，可为空。

        返回：
        - 无。
        """
        ws = self._resolve_sheet(sheet)
        self.write_row(
            ws,
            start_row,
            list(headers),
            styles=[header_style] * len(headers) if header_style else None,
        )
        for i, r in enumerate(rows):
            self.write_row(
                ws,
                start_row + 1 + i,
                list(r),
                styles=[row_style] * len(r) if row_style else None,
            )

    def export_dicts(
        self,
        sheet: Union[str, Any],
        start_row: int,
        data: Sequence[Dict[str, Any]],
        header_map: Optional[Dict[str, str]] = None,
        order: Optional[Sequence[str]] = None,
        header_style: Optional[StyleDict] = None,
        row_style: Optional[StyleDict] = None,
    ) -> None:
        """导出字典列表为表格。

        参数：
        - `sheet`: 工作表对象或名称。
        - `start_row`: 起始行（1 基）。
        - `data`: 字典列表，每个字典代表一行数据。
        - `header_map`: 字段到表头文本的映射，若为空则使用字段名。
        - `order`: 字段顺序，若为空则按首个字典的键顺序。
        - `header_style`: 表头样式，可为空。
        - `row_style`: 数据行样式，可为空。

        返回：
        - 无。
        """
        if not data:
            return
        if order is None:
            order = list(data[0].keys())
        headers = [header_map.get(k, k) if header_map else k for k in order]
        self.write_table(
            sheet,
            start_row,
            headers,
            [[row.get(k) for k in order] for row in data],
            header_style,
            row_style,
        )

    # endregion

    def insert_image_in_cell(
        self,
        sheet: Union[str, Any],
        row: int,
        col: int,
        image_path: str,
        keep_ratio: bool = True,
    ) -> None:
        """将图片按单元格尺寸嵌入（覆盖单元格区域），默认保持比例以适配单元格。

        - openpyxl：根据列宽/行高换算像素，计算目标宽高，调用图片插入。
        - xlsxwriter：记录的列宽/行高用于计算缩放比例（若未设置则使用默认值）。
        - xlwt：插入 BMP（无法缩放），作为最小兼容。
        """
        ws = self._resolve_sheet(sheet)
        if self._engine == "openpyxl":
            # 计算单元格像素尺寸
            letter = get_column_letter_safe(col)
            col_dim = ws.column_dimensions.get(letter)
            char_w = getattr(col_dim, "width", None) or 8.43
            cell_w_px = int(round(char_w * 7 + 5))
            row_dim = ws.row_dimensions.get(row)
            pt_h = getattr(row_dim, "height", None) or 15.0
            cell_h_px = int(round(pt_h * 96 / 72))
            w, h = cell_w_px, cell_h_px
            if keep_ratio:
                try:
                    from PIL import Image  # type: ignore

                    with Image.open(image_path) as im:
                        iw, ih = im.size
                    ratio = min(cell_w_px / max(iw, 1), cell_h_px / max(ih, 1))
                    w = int(iw * ratio)
                    h = int(ih * ratio)
                except Exception:
                    pass
            insert_image_openpyxl(ws, row, col, image_path, width=w, height=h)
            return
        if self._engine == "xlsxwriter":
            # 使用已记录的列宽/行高（若未设置则使用默认值）
            name = getattr(ws, "get_name", lambda: next((n for n, w in self._sheets.items() if w is ws), "Sheet1"))()
            char_w = (self._col_widths.get(name, {}).get(col) or 8.43)
            cell_w_px = int(round(char_w * 7 + 5))
            pt_h = (self._row_heights.get(name, {}).get(row) or 15.0)
            cell_h_px = int(round(pt_h * 96 / 72))
            scale_x = None
            scale_y = None
            try:
                from PIL import Image  # type: ignore

                with Image.open(image_path) as im:
                    iw, ih = im.size
                sx = cell_w_px / max(iw, 1)
                sy = cell_h_px / max(ih, 1)
                if keep_ratio:
                    r = min(sx, sy)
                    scale_x = r
                    scale_y = r
                else:
                    scale_x = sx
                    scale_y = sy
            except Exception:
                pass
            insert_image_xlsxwriter(
                ws, row, col, image_path, scale_x=scale_x, scale_y=scale_y
            )
            return
        if self._engine == "xlwt":
            insert_image_xlwt(ws, row, col, image_path)
            return

    # endregion

    # endregion
