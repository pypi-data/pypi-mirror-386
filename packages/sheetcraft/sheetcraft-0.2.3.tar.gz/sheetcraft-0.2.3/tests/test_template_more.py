from pathlib import Path

from sheetcraft.template import ExcelTemplate


def test_jinja_for_dash_variant_and_endfor_dash(tmp_path):
    """覆盖 `{%- for ... %}` 与 `{%- endfor %}` 的解析与清理分支。"""
    from openpyxl import Workbook, load_workbook

    tpl = tmp_path / "tpl_dash.xlsx"
    out = tmp_path / "out_dash.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "DASH"
    ws.cell(row=1, column=1, value="{%- for item in items %}")
    ws.cell(row=2, column=1, value="{{ item.name }}")
    ws.cell(row=3, column=1, value="{%- endfor %}")
    wb.save(str(tpl))

    ExcelTemplate().render(str(tpl), {"items": [{"name": "A"}, {"name": "B"}]}, str(out))

    rwb = load_workbook(str(out))
    aws = rwb.active
    # 仅容量 1，渲染首项，边界标签已清理
    assert aws.cell(row=2, column=1).value == "A"
    assert aws.cell(row=1, column=1).value in (None, "")
    assert aws.cell(row=3, column=1).value in (None, "")


def test_end_row_has_additional_text_is_kept_in_template(tmp_path):
    """多行块 end 行包含除 endfor 外的文本时，被视为模板行参与渲染。"""
    from openpyxl import Workbook, load_workbook

    tpl = tmp_path / "tpl_end_text.xlsx"
    out = tmp_path / "out_end_text.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "ENDTEXT"
    ws.cell(row=1, column=1, value="{% for item in items %}")
    ws.cell(row=2, column=1, value="{{ item.name }}")
    # end 行：A1 是 endfor，B1 放置额外文本
    ws.cell(row=3, column=1, value="{% endfor %}")
    ws.cell(row=3, column=2, value="ENDTEXT")
    wb.save(str(tpl))

    # 仅一条数据，第二模板行应写入纯文本 ENDTEXT
    ExcelTemplate().render(str(tpl), {"items": [{"name": "X"}]}, str(out))

    rwb = load_workbook(str(out))
    aws = rwb.active
    assert aws.cell(row=2, column=1).value == "X"
    assert aws.cell(row=3, column=2).value == "ENDTEXT"
    # 边界标签被清空
    assert aws.cell(row=1, column=1).value in (None, "")
    assert aws.cell(row=3, column=1).value in (None, "")


def test_template_img_explicit_size_clears_cell(tmp_path):
    """占位符包含显式 width/height 时走默认插入分支，并清空单元格。"""
    from openpyxl import Workbook, load_workbook

    repo_root = Path(__file__).resolve().parents[1]
    img_path = repo_root / "examples" / "templates" / "img.png"
    assert img_path.exists()

    tpl = tmp_path / "tpl_img_size.xlsx"
    out = tmp_path / "out_img_size.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "IMGSIZE"
    ws.cell(row=1, column=1, value=f'__SHEETCRAFT_IMG__{{"path":"{img_path}","width":50,"height":30}}')
    wb.save(str(tpl))

    ExcelTemplate().render(str(tpl), {}, str(out))

    rwb = load_workbook(str(out))
    aws = rwb.active
    assert aws.cell(row=1, column=1).value is None