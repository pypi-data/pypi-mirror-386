from openpyxl import Workbook, load_workbook
from sheetcraft.template import ExcelTemplate


def test_template_render_with_format_fix_enabled(tmp_path):
    tpl = tmp_path / "tpl.xlsx"
    out = tmp_path / "out.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "T"
    ws.cell(row=1, column=1, value="{{ v }}")
    wb.save(str(tpl))

    # 启用格式修复：即使模板不包含 drawings，分支也应执行且不报错
    tpl_engine = ExcelTemplate(apply_format_fix=True)
    tpl_engine.render(str(tpl), {"v": "ok"}, str(out))

    rwb = load_workbook(str(out))
    aws = rwb.active
    assert aws.cell(row=1, column=1).value == "ok"