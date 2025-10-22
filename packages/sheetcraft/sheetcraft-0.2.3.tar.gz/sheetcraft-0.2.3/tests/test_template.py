import json
from pathlib import Path
from sheetcraft import ExcelTemplate, ExcelWorkbook


def build_block_template(path: str):
    """构建包含 Jinja2 for 块语法的模板工作簿。"""
    wb = ExcelWorkbook(output_path=path)
    ws = wb.sheet("T")
    # Jinja2 for 块开始（容量 2 行）
    wb.write_cell(ws, 1, 1, "{% for item in items %}")
    wb.write_row(ws, 2, ["{{ item.name }}", "{{ item.qty }}", "{{ item.price }}"])
    wb.write_row(ws, 3, ["Subtotal", "=B3*C3", ""])
    wb.write_cell(ws, 4, 1, "{% endfor %}")
    # 其他变量
    wb.write_cell(ws, 6, 1, "{{ title }}")
    wb.save()


def test_template_block_repeat(tmp_path):
    """验证 Jinja2 for 块语法渲染。"""
    tpl = tmp_path / "tpl.xlsx"
    out = tmp_path / "out.xlsx"
    build_block_template(str(tpl))

    data = {
        "title": "Sales Report",
        "items": [
            {"name": "A", "qty": 1, "price": 1.5},
            {"name": "B", "qty": 2, "price": 2.0},
        ],
    }
    ExcelTemplate().render(str(tpl), data, str(out))

    assert out.exists()


def test_template_simple_repeat(tmp_path):
    """验证 Jinja2 单行 for 语法渲染首条数据。"""
    tpl = tmp_path / "tpl2.xlsx"
    out = tmp_path / "out2.xlsx"
    wb = ExcelWorkbook(output_path=str(tpl))
    ws = wb.sheet("T")
    wb.write_cell(ws, 1, 1, "{% for item in items %}")
    wb.write_row(ws, 2, ["{{ item.name }}", "{{ item.qty }}"])
    wb.write_cell(ws, 3, 1, "{% endfor %}")
    wb.save()

    data = {"items": [{"name": "A", "qty": 1}, {"name": "B", "qty": 2}]}
    ExcelTemplate().render(str(tpl), data, str(out))
    assert out.exists()


def build_json_based_template(path: str):
    """使用示例数据的字段构建一个可渲染的 Jinja2 for 模板。"""
    wb = ExcelWorkbook(output_path=path)
    ws = wb.sheet("T")
    # 顶部基本信息
    wb.write_row(ws, 1, ["{{ id }}", "{{ project_name }}", "{{ receiver }}"])

    # 到货明细块（容量 3 行）
    wb.write_cell(ws, 3, 1, "{% for item in arrival_details %}")
    wb.write_row(
        ws,
        4,
        [
            "{{ item.equipment }}",
            "{{ item.quantity }}",
            "{{ item.unit }}",
            "{{ item.supplier }}",
            "{{ item.equipment_model }}",
        ],
    )
    wb.write_row(
        ws,
        5,
        [
            "{{ item.equipment }}",
            "{{ item.quantity }}",
            "{{ item.unit }}",
            "{{ item.supplier }}",
            "{{ item.equipment_model }}",
        ],
    )
    wb.write_row(
        ws,
        6,
        [
            "{{ item.equipment }}",
            "{{ item.quantity }}",
            "{{ item.unit }}",
            "{{ item.supplier }}",
            "{{ item.equipment_model }}",
        ],
    )
    wb.write_cell(ws, 7, 1, "{% endfor %}")

    # 验收明细块（容量 4 行）
    wb.write_cell(ws, 9, 1, "{% for row in acceptance_details %}")
    wb.write_row(
        ws, 10, ["{{ row.item }}", "{{ row.content }}", "{{ row.conclusion }}"]
    )
    wb.write_row(
        ws, 11, ["{{ row.item }}", "{{ row.content }}", "{{ row.conclusion }}"]
    )
    wb.write_row(
        ws, 12, ["{{ row.item }}", "{{ row.content }}", "{{ row.conclusion }}"]
    )
    wb.write_row(
        ws, 13, ["{{ row.item }}", "{{ row.content }}", "{{ row.conclusion }}"]
    )
    wb.write_cell(ws, 14, 1, "{% endfor %}")
    wb.save()


def test_examples_data_json_render(tmp_path):
    """加载 examples/templates/data.json，渲染到自建模板并验证循环与变量替换。"""
    base = Path(__file__).resolve().parents[1] / "examples" / "templates"
    data_path = base / "data.json"
    assert data_path.exists()

    with open(data_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    tpl = tmp_path / "tpl_json.xlsx"
    out = tmp_path / "out_json.xlsx"
    build_json_based_template(str(tpl))
    ExcelTemplate().render(str(tpl), data, str(out))
    assert out.exists()

    # 校验输出包含项目名称、至少一条到货明细与一条验收明细内容
    from openpyxl import load_workbook

    wb = load_workbook(str(out))

    def contains_text(s: str) -> bool:
        for ws in wb.worksheets:
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(row=r, column=c).value
                    if isinstance(v, str) and s in v:
                        return True
        return False

    assert contains_text(data["project_name"])
    assert contains_text(data["arrival_details"][0]["equipment"])
    assert contains_text(data["acceptance_details"][0]["content"])


def test_jinja_single_line_for_preserves_first_item_and_clears_end(tmp_path):
    from openpyxl import Workbook

    tpl = tmp_path / "tpl_single_for.xlsx"
    out = tmp_path / "out_single_for.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.cell(row=1, column=1, value="{% for item in items %}")
    ws.cell(row=1, column=2, value="{{ item.name }}")
    ws.cell(row=1, column=3, value="{{ loop.index }}")
    ws.cell(row=2, column=1, value="{% endfor %}")
    wb.save(str(tpl))

    data = {"items": [{"name": "X"}, {"name": "Y"}]}
    ExcelTemplate().render(str(tpl), data, str(out))

    from openpyxl import load_workbook

    rwb = load_workbook(str(out))
    ws2 = rwb.active
    assert ws2.cell(row=1, column=2).value == "X"
    assert ws2.cell(row=1, column=3).value == "1"
    assert ws2.cell(row=2, column=1).value is None


def test_jinja_multi_line_for_with_merges_and_loop_ctx(tmp_path):
    from openpyxl import Workbook

    tpl = tmp_path / "tpl_multi_for.xlsx"
    out = tmp_path / "out_multi_for.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "M"
    ws.cell(row=1, column=1, value="{% for item in items %}")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3)
    ws.cell(row=2, column=2, value="{{ item.name }}-{{ loop.index }}/{{ loop.length }}")
    ws.cell(row=3, column=1, value='{{ item.code|default("N/A") }}')
    ws.cell(row=3, column=2, value="{{ loop.first }}")
    ws.cell(row=3, column=3, value="{{ loop.last }}")
    ws.cell(row=4, column=1, value="{% endfor %}")
    wb.save(str(tpl))

    data = {"items": [{"name": "Foo", "code": "A1"}, {"name": "Bar"}]}
    ExcelTemplate().render(str(tpl), data, str(out))

    from openpyxl import load_workbook

    rwb = load_workbook(str(out))
    ws2 = rwb.active
    assert ws2.cell(row=2, column=2).value == "Foo-1/2"
    # 第二项（行3）code 使用默认值
    assert ws2.cell(row=3, column=1).value == "N/A"
    assert ws2.cell(row=3, column=2).value == "False"
    assert ws2.cell(row=3, column=3).value == "True"
    assert ws2.cell(row=1, column=1).value is None
    assert ws2.cell(row=4, column=1).value is None


def test_jinja_capacity_clears_remaining_rows(tmp_path):
    from openpyxl import Workbook

    tpl = tmp_path / "tpl_capacity.xlsx"
    out = tmp_path / "out_capacity.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "C"
    ws.cell(row=1, column=1, value="{% for item in items %}")
    ws.cell(row=2, column=1, value="{{ item.name }}")
    ws.cell(row=3, column=1, value="{{ item.name }}")
    ws.cell(row=4, column=1, value="{{ item.name }}")
    ws.cell(row=5, column=1, value="{% endfor %}")
    wb.save(str(tpl))

    data = {"items": [{"name": "A"}, {"name": "B"}]}
    ExcelTemplate().render(str(tpl), data, str(out))

    from openpyxl import load_workbook

    rwb = load_workbook(str(out))
    ws2 = rwb.active
    assert ws2.cell(row=2, column=1).value == "A"
    assert ws2.cell(row=3, column=1).value == "B"
    assert ws2.cell(row=4, column=1).value in (None, "")


def test_render_string_fallback_for_invalid_jinja(tmp_path):
    from openpyxl import Workbook

    tpl = tmp_path / "tpl_invalid_jinja.xlsx"
    out = tmp_path / "out_invalid_jinja.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "I"
    ws.cell(row=1, column=1, value="{% for item in items %")
    ws.cell(row=1, column=2, value="{{ project.name }}")
    wb.save(str(tpl))

    data = {"items": [], "project": {"name": "P"}}
    ExcelTemplate().render(str(tpl), data, str(out))

    from openpyxl import load_workbook

    rwb = load_workbook(str(out))
    ws2 = rwb.active
    assert ws2.cell(row=1, column=1).value == "{% for item in items %"
    assert ws2.cell(row=1, column=2).value == "P"


def test_template_image_tag_inserts_and_clears_cell(tmp_path):
    from openpyxl import Workbook, load_workbook
    from sheetcraft.template import ExcelTemplate
    from pathlib import Path

    tpl_path = tmp_path / "tpl_img_tag.xlsx"
    out_path = tmp_path / "out_img_tag.xlsx"

    # 计算仓库内示例图片的绝对路径
    repo_root = Path(__file__).resolve().parents[1]
    img_path = repo_root / "examples/templates/img.png"

    wb = Workbook()
    ws = wb.active
    ws.title = "IMGEXT"
    ws.cell(row=1, column=1, value=f'__SHEETCRAFT_IMG__{{"path":"{img_path}","fit":true}}')
    wb.save(str(tpl_path))

    ExcelTemplate().render(str(tpl_path), {}, str(out_path))

    rwb = load_workbook(str(out_path))
    aws = rwb.active
    assert aws.cell(row=1, column=1).value is None


def test_image_placeholder_disabled_clears_cell(tmp_path):
    from openpyxl import Workbook

    tpl = tmp_path / "tpl_img.xlsx"
    out = tmp_path / "out_img.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "IMG"
    ws.cell(row=1, column=1, value='__SHEETCRAFT_IMG__{"path":"fake.png"}')
    wb.save(str(tpl))

    ExcelTemplate(enable_images=False).render(str(tpl), {}, str(out))

    from openpyxl import load_workbook

    rwb = load_workbook(str(out))
    ws2 = rwb.active
    assert ws2.cell(row=1, column=1).value is None


def test_template_image_payload_invalid_json_falls_back(tmp_path):
    from openpyxl import Workbook, load_workbook
    from sheetcraft.template import ExcelTemplate

    tpl_path = tmp_path / "tpl_img_payload.xlsx"
    out_path = tmp_path / "out_img_payload.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "IMGPAYLOAD"
    # 构造无效 JSON 的占位符，触发异常分支并回退为原字符串
    payload = '__SHEETCRAFT_IMG__{"path":}'
    ws.cell(row=1, column=1, value=payload)
    wb.save(str(tpl_path))

    ExcelTemplate().render(str(tpl_path), {}, str(out_path))

    rwb = load_workbook(str(out_path))
    aws = rwb.active
    assert aws.cell(row=1, column=1).value == payload


def test_jinja_invalid_for_tag_skips_block(tmp_path):
    from openpyxl import Workbook, load_workbook
    from sheetcraft.template import ExcelTemplate

    tpl_path = tmp_path / "tpl_invalid_for.xlsx"
    out_path = tmp_path / "out_invalid_for.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "INVFOR"
    # 无效但满足扫描条件的 for 标签（有 {%, %}, 且包含 for 与 in）
    invalid_tag = "{% for item in %}"
    ws.cell(row=1, column=1, value=invalid_tag)
    # 同行添加一个正常变量，验证第二阶段渲染仍正常
    ws.cell(row=1, column=2, value="{{ project.name }}")
    wb.save(str(tpl_path))

    ExcelTemplate().render(str(tpl_path), {"project": {"name": "P"}}, str(out_path))

    rwb = load_workbook(str(out_path))
    aws = rwb.active
    assert aws.cell(row=1, column=1).value == invalid_tag
    assert aws.cell(row=1, column=2).value == "P"


def test_jinja_for_without_endfor_is_ignored(tmp_path):
    from openpyxl import Workbook, load_workbook
    from sheetcraft.template import ExcelTemplate

    tpl_path = tmp_path / "tpl_for_no_end.xlsx"
    out_path = tmp_path / "out_for_no_end.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "FORNOEND"
    # 正常语法的 for 开始标签，但没有 endfor 行
    tag = "{% for item in items %}"
    ws.cell(row=1, column=1, value=tag)
    wb.save(str(tpl_path))

    ExcelTemplate().render(str(tpl_path), {"items": [1, 2]}, str(out_path))

    rwb = load_workbook(str(out_path))
    aws = rwb.active
    # 未找到 endfor 时，应忽略该块并保留原字符串（由字符串渲染回退确保）
    assert aws.cell(row=1, column=1).value == tag


def test_template_init_filters_applied(tmp_path):
    from openpyxl import Workbook, load_workbook
    from sheetcraft.template import ExcelTemplate

    tpl_path = tmp_path / "tpl_filters.xlsx"
    out_path = tmp_path / "out_filters.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "FILTERS"
    ws.cell(row=1, column=1, value="{{ name|shout }}")
    wb.save(str(tpl_path))

    tpl = ExcelTemplate(filters={"shout": lambda s: str(s).upper()})
    tpl.render(str(tpl_path), {"name": "hello"}, str(out_path))

    rwb = load_workbook(str(out_path))
    aws = rwb.active
    assert aws.cell(row=1, column=1).value == "HELLO"
