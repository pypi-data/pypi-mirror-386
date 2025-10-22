from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple


class ExcelTemplate:
    """基于 Jinja2 的 `.xlsx` 模板渲染器，支持灵活变量与块重复语法。

    功能：
    - 变量替换：单元格中含有 `{{ var }}` 的内容将按传入数据渲染。

    限制：
    - 仅支持 `.xlsx` 模板渲染；`.xls` 请使用导出 API。
    - 暂不支持嵌套重复块；若需多个表格可使用多个独立块。
    """

    def __init__(
        self, filters: Optional[Dict[str, Any]] = None, enable_images: bool = True,
        apply_format_fix: bool = False,
        format_fix_config: Optional[Any] = None,
    ) -> None:
        """初始化渲染器。

        参数：
        - `filters`: 自定义 Jinja2 过滤器字典（名称到函数）。
        - `enable_images`: 是否处理 `{% img %}` 图片占位符（默认 True）。若为 False，仅清空占位单元格，不插入图片。
        - `apply_format_fix`: 渲染保存后是否应用格式修复（仅 `.xlsx`）。
        - `format_fix_config`: 格式修复配置（默认启用前缀修复）。

        返回：
        - 无。
        """
        try:
            import jinja2  # noqa: F401
        except Exception as exc:
            raise RuntimeError(
                "Jinja2 is required for template rendering. Install with: pip install 'sheetcraft[template]'"
            ) from exc
        import jinja2

        # 在变量最终输出阶段将 None 归一为空字符串，避免显示 "None"
        self._env = jinja2.Environment(autoescape=False, finalize=self._finalize_value)
        # 图片处理开关
        self._enable_images = enable_images
        # 新增：格式修复参数
        self._apply_format_fix = apply_format_fix
        self._format_fix_config = format_fix_config
        # 注册图片标签扩展
        try:
            from .jinja_ext import ImageTagExtension

            self._env.add_extension(ImageTagExtension)
        except Exception:
            # 环境缺少 jinja2 或扩展失败时，继续基础渲染能力
            pass
        if filters:
            self._env.filters.update(filters)

    def add_filters(self, filters: Dict[str, Any]) -> None:
        """添加或更新 Jinja2 过滤器。

        参数：
        - `filters`: 过滤器字典（名称到函数）。

        返回：
        - 无。
        """
        self._env.filters.update(filters)

    def render(
        self, template_path: str, data: Dict[str, Any], output_path: str
    ) -> None:
        """渲染模板工作簿并保存至目标路径。"""
        from openpyxl import load_workbook
        import os
        import json

        wb = load_workbook(template_path)

        for ws in wb.worksheets:
            # 先处理 Jinja `{% for %}` 块
            self._process_jinja_for_blocks(ws, data)

            # 第二阶段：渲染剩余单元格中的模板变量 + 识别图片占位符
            prefix = "__SHEETCRAFT_IMG__"
            base_dir = os.path.dirname(os.path.abspath(template_path))
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(row=r, column=c).value
                    if isinstance(v, str):
                        rendered = self._render_string(v, data)
                        if isinstance(rendered, str) and rendered.startswith(prefix):
                            try:
                                payload = json.loads(rendered[len(prefix) :])
                                p = str(payload.get("path", ""))
                                # 若禁用图片处理，或未提供路径，直接清空单元格
                                if not self._enable_images or not p:
                                    ws.cell(row=r, column=c).value = None
                                    continue
                                width = payload.get("width")
                                height = payload.get("height")
                                fit = payload.get("fit")
                                in_cell = payload.get("in_cell")
                                keep_ratio = payload.get("keep_ratio", True)
                                # 解析相对路径：相对于模板所在目录
                                if p and not os.path.isabs(p):
                                    p = os.path.join(base_dir, p)
                                # 按单元格适配尺寸（宽充满，可能跨行）：旧逻辑保留
                                if fit or (
                                    width is None
                                    and height is None
                                    and payload.get("scale_x") is None
                                    and payload.get("scale_y") is None
                                    and not in_cell
                                ):
                                    from .images import calc_fit_size_openpyxl

                                    width, height = calc_fit_size_openpyxl(ws, r, c, p)
                                    from .images import insert_image_openpyxl
                                    insert_image_openpyxl(
                                        ws, r, c, p, width=width, height=height
                                    )
                                    ws.cell(row=r, column=c).value = None
                                    continue
                                # 新增：按单元格内嵌（宽/高双向适配），默认保持比例
                                if in_cell:
                                    # 计算单元格像素尺寸
                                    from .utils import get_column_letter_safe
                                    col_letter = get_column_letter_safe(c)
                                    col_dim = ws.column_dimensions.get(col_letter)
                                    char_w = getattr(col_dim, "width", None) or 8.43
                                    cell_w_px = int(round(char_w * 7 + 5))
                                    row_dim = ws.row_dimensions.get(r)
                                    pt_h = getattr(row_dim, "height", None) or 15.0
                                    cell_h_px = int(round(pt_h * 96 / 72))
                                    w, h = cell_w_px, cell_h_px
                                    if keep_ratio:
                                        try:
                                            from PIL import Image  # type: ignore

                                            with Image.open(p) as im:
                                                iw, ih = im.size
                                            ratio = min(cell_w_px / max(iw, 1), cell_h_px / max(ih, 1))
                                            w = int(iw * ratio)
                                            h = int(ih * ratio)
                                        except Exception:
                                            # 无 Pillow 或读取失败时，回退为单元格尺寸
                                            pass
                                    from .images import insert_image_openpyxl
                                    insert_image_openpyxl(ws, r, c, p, width=w, height=h)
                                    ws.cell(row=r, column=c).value = None
                                    continue
                                # 默认：依据显式尺寸/缩放插入
                                from .images import insert_image_openpyxl
                                insert_image_openpyxl(
                                    ws, r, c, p, width=width, height=height
                                )
                                ws.cell(row=r, column=c).value = None
                            except Exception:
                                # 未能正确解析占位符时（例如模板语法异常），降级为普通字符串渲染结果
                                ws.cell(row=r, column=c).value = (
                                    None if (isinstance(rendered, str) and rendered == "") else rendered
                                )
                        else:
                            ws.cell(row=r, column=c).value = (
                                None if (isinstance(rendered, str) and rendered == "") else rendered
                            )
            wb.save(output_path)
            # 可选：保存后应用格式修复（仅 .xlsx）
            if self._apply_format_fix:
                try:
                    from .format_fix import fix_xlsx, FormatFixConfig
                    cfg = self._format_fix_config or FormatFixConfig()
                    fix_xlsx(output_path, None, cfg)
                except Exception:
                    pass

    def _render_string(self, tpl: str, ctx: Dict[str, Any]) -> str:
        """渲染字符串模板。"""
        try:
            tmpl = self._env.from_string(tpl)
            return tmpl.render(**ctx)
        except Exception:
            # 当单元格中包含未闭合的 Jinja 控制语句（如 {% for %}）时，优雅降级：原样返回，避免整体渲染失败。
            return tpl

    # 新增：在变量最终输出阶段将 None -> ""
    def _finalize_value(self, v: Any) -> Any:
        return "" if v is None else v

    # 新增：内部辅助，提取单元格文本中的 `{% for ... %}` 子串
    def _extract_for_inner(self, s: str) -> Optional[str]:
        s2 = str(s)
        i = s2.find("{")
        if i == -1:
            return None
        # 定位 Jinja 起止标记
        i = s2.find("{%", i)
        if i == -1:
            return None
        j = s2.find("%}", i + 2)
        if j == -1:
            return None
        inner = s2[i + 2 : j].strip()
        # 支持 `{%- ... %}` 变体：去除前导 `-`
        inner = inner.lstrip("-").strip()
        # 更宽松的匹配：以 "for " 开头且包含 " in "
        if inner.startswith("for ") and " in " in inner:
            return inner
        return None

    def _has_endfor(self, s: str) -> bool:
        s2 = str(s)
        i = s2.find("{%")
        if i == -1:
            return False
        j = s2.find("%}", i + 2)
        if j == -1:
            return False
        inner = s2[i + 2 : j].strip()
        inner = inner.lstrip("-").strip()
        return inner.startswith("endfor")

    def _find_jinja_for_blocks(self, ws) -> List[JinjaForBlock]:
        """扫描并解析 Jinja `{% for x in y %}` 块，返回块边界与键信息。"""
        blocks: List[JinjaForBlock] = []
        r = 1
        max_r = ws.max_row
        max_c = ws.max_column
        while r <= max_r:
            tag_inner = None
            for c in range(1, max_c + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    inner = self._extract_for_inner(v)
                    if inner is not None:
                        tag_inner = inner
                        break
            if tag_inner is None:
                r += 1
                continue
            # 解析 `{% for alias in key %}`
            try:
                parts = tag_inner.split()
                alias = parts[1]
                key = parts[3]
            except Exception:
                r += 1
                continue
            # 向下查找 `{% endfor %}` 行（允许同一行任意位置出现）
            end_row = None
            rr = r + 1
            while rr <= max_r:
                hit = False
                for cc in range(1, max_c + 1):
                    vv = ws.cell(row=rr, column=cc).value
                    if isinstance(vv, str) and self._has_endfor(vv):
                        end_row = rr
                        hit = True
                        break
                if hit:
                    break
                rr += 1
            if end_row:
                blocks.append(
                    JinjaForBlock(start_row=r, end_row=end_row, key=key, alias=alias)
                )
                r = end_row + 1
            else:
                r += 1
        return blocks

    def _process_jinja_for_blocks(self, ws, data: Dict[str, Any]) -> None:
        """在不增加行/列的前提下展开 Jinja `{% for %}` 块。"""
        blocks = self._find_jinja_for_blocks(ws)
        for blk in blocks:
            items = data.get(blk.key)
            if not isinstance(items, list):
                continue

            # 计算合并单元格锚点（返回锚点坐标，未合并时返回自身坐标）
            def merged_anchor(row: int, col: int):
                for rng in ws.merged_cells.ranges:
                    if (
                        rng.min_row <= row <= rng.max_row
                        and rng.min_col <= col <= rng.max_col
                    ):
                        return rng.min_row, rng.min_col
                return (row, col)

            # 捕获模板行（按锚点归一，避免非锚单元格被跳过）
            has_inner = blk.end_row > blk.start_row + 1
            template_rows: List[Dict[Tuple[int, int], Any]] = []
            if has_inner:
                for r in range(blk.start_row + 1, blk.end_row):
                    row_map: Dict[Tuple[int, int], Any] = {}
                    for c in range(1, ws.max_column + 1):
                        val = ws.cell(row=r, column=c).value
                        ar, ac = merged_anchor(r, c)
                        # 优先保留非空字符串模板（若锚点已有占位，则以字符串模板覆盖）
                        if (ar, ac) not in row_map or (
                            isinstance(val, str) and val and val.strip()
                        ):
                            row_map[(ar, ac)] = val
                    template_rows.append(row_map)
                # 若 end_row 存在内容（除 endfor 标签外），也作为模板行加入
                end_row_has_content = False
                end_row_map: Dict[Tuple[int, int], Any] = {}
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=blk.end_row, column=c).value
                    if isinstance(val, str) and self._has_endfor(val):
                        continue
                    ar, ac = merged_anchor(blk.end_row, c)
                    if (ar, ac) not in end_row_map or (
                        isinstance(val, str) and val and val.strip()
                    ):
                        end_row_map[(ar, ac)] = val
                for v in end_row_map.values():
                    if v not in (None, ""):
                        end_row_has_content = True
                        break
                if end_row_has_content:
                    template_rows.append(end_row_map)
            else:
                # 邻接边界：优先选择包含内容的那一行作为模板（start_row 或 end_row）
                start_map: Dict[Tuple[int, int], Any] = {}
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=blk.start_row, column=c).value
                    if isinstance(val, str):
                        s_full = str(val)
                        inner = self._extract_for_inner(s_full)
                        if inner is not None:
                            # 去除单元格中的 for 标签子串，保留其余文本
                            i = s_full.find("{%")
                            j = s_full.find("%}", i + 2)
                            before = s_full[:i]
                            after = s_full[j + 2 :]
                            cleaned = (before + after).strip()
                            val = cleaned if cleaned else None
                    ar, ac = merged_anchor(blk.start_row, c)
                    if (ar, ac) not in start_map or (
                        isinstance(val, str) and val and val.strip()
                    ):
                        start_map[(ar, ac)] = val
                # 构造 end 行的模板映射（跳过 endfor 标签单元格）
                end_map: Dict[Tuple[int, int], Any] = {}
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=blk.end_row, column=c).value
                    if isinstance(val, str) and self._has_endfor(val):
                        continue
                    ar, ac = merged_anchor(blk.end_row, c)
                    if (ar, ac) not in end_map or (
                        isinstance(val, str) and val and val.strip()
                    ):
                        end_map[(ar, ac)] = val
                # 判定选择哪一行作为容量 1 的模板行
                def _has_non_empty(m: Dict[Tuple[int, int], Any]) -> bool:
                    for v in m.values():
                        if v not in (None, ""):
                            return True
                    return False
                template_rows.append(end_map if _has_non_empty(end_map) else start_map)
            capacity = len(template_rows)
            length = len(items)

            def loop_ctx(i: int):
                return {
                    "index": i + 1,
                    "index0": i,
                    "length": length,
                    "first": i == 0,
                    "last": i == length - 1,
                }

            # 写入数据或置空（直接写入锚点单元格，避免合并区域被忽略）
            for i in range(capacity):
                item = items[i] if i < length else {}
                row_map = template_rows[i]
                for (ar, ac), tpl in row_map.items():
                    if isinstance(tpl, str):
                        ctx = {blk.alias: item, **data, "loop": loop_ctx(i)}
                        _val = self._render_string(tpl, ctx)
                        ws.cell(row=ar, column=ac).value = (
                            None if (isinstance(_val, str) and _val == "") else _val
                        )
                    else:
                        ws.cell(row=ar, column=ac).value = tpl if i < length else None

            # 清空边界所在行
            if has_inner:
                # 多行块：仅删除边界行中的标签单元格，保留其他文本
                # 处理 start_row：去除 for 标签子串
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=blk.start_row, column=c).value
                    if isinstance(val, str):
                        s_full = str(val)
                        inner = self._extract_for_inner(s_full)
                        if inner is not None:
                            i = s_full.find("{%")
                            j = s_full.find("%}", i + 2)
                            before = s_full[:i]
                            after = s_full[j + 2 :]
                            cleaned = (before + after).strip()
                            ar, ac = merged_anchor(blk.start_row, c)
                            ws.cell(row=ar, column=ac).value = cleaned if cleaned else None
                # 处理 end_row：清空 endfor 标签单元格
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=blk.end_row, column=c).value
                    if isinstance(val, str) and self._has_endfor(val):
                        ar, ac = merged_anchor(blk.end_row, c)
                        ws.cell(row=ar, column=ac).value = None
            else:
                # 单行块：保留模板行渲染结果，仅清空 endfor 行
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=blk.end_row, column=c).value
                    if isinstance(val, str) and self._has_endfor(val):
                        a2 = merged_anchor(blk.end_row, c)
                        ws.cell(row=a2[0], column=a2[1]).value = None
                # 同时处理 start_row：去除 for 标签子串，保留其它文本（若有）
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=blk.start_row, column=c).value
                    if isinstance(val, str):
                        s_full = str(val)
                        inner = self._extract_for_inner(s_full)
                        if inner is not None:
                            i = s_full.find("{%")
                            j = s_full.find("%}", i + 2)
                            before = s_full[:i]
                            after = s_full[j + 2 :]
                            cleaned = (before + after).strip()
                            a1 = merged_anchor(blk.start_row, c)
                            ws.cell(row=a1[0], column=a1[1]).value = cleaned if cleaned else None

    # endregion


@dataclass
class JinjaForBlock:
    start_row: int
    end_row: int
    key: str
    alias: str
