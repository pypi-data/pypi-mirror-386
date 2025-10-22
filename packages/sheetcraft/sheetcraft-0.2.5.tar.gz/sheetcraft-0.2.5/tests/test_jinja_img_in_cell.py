from pathlib import Path

from sheetcraft.template import ExcelTemplate


def test_jinja_img_in_cell_renders_and_embeds(tmp_path):
    """使用 Jinja img 标签生成 in_cell 占位并成功渲染嵌入。"""
    from openpyxl import Workbook, load_workbook
    from jinja2 import Environment
    from sheetcraft.jinja_ext import ImageTagExtension

    tpl = tmp_path / "tpl_jinja_in_cell.xlsx"
    out = tmp_path / "out_jinja_in_cell.xlsx"

    repo_root = Path(__file__).resolve().parents[1]
    img_path = repo_root / "examples/templates/img.png"

    # 构造模板：在 A1 放置通过扩展生成的占位
    wb = Workbook()
    ws = wb.active
    ws.title = "JINJA"

    env = Environment(extensions=[ImageTagExtension])
    tpl_cell = env.from_string("{% img '" + str(img_path) + "' in_cell=true %}").render({})
    ws.cell(row=1, column=1, value=tpl_cell)
    wb.save(str(tpl))

    ExcelTemplate().render(str(tpl), {}, str(out))

    rwb = load_workbook(str(out))
    aws = rwb.active
    assert aws.cell(row=1, column=1).value is None