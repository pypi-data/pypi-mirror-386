from pathlib import Path

from sheetcraft.template import ExcelTemplate


def test_template_image_in_cell_clears_and_saves(tmp_path):
    """模板占位符包含 in_cell=true 时，按单元格内嵌插入，并清空占位单元格。"""
    from openpyxl import Workbook, load_workbook

    tpl = tmp_path / "tpl_in_cell.xlsx"
    out = tmp_path / "out_in_cell.xlsx"

    repo_root = Path(__file__).resolve().parents[1]
    img_path = repo_root / "examples/templates/img.png"

    wb = Workbook()
    ws = wb.active
    ws.title = "IN_CELL"
    # 直接写入占位符字符串，跳过 Jinja 解析，覆盖渲染路径
    ws.cell(row=2, column=2, value=f'__SHEETCRAFT_IMG__{{"path":"{img_path}","in_cell":true}}')
    wb.save(str(tpl))

    ExcelTemplate().render(str(tpl), {}, str(out))

    rwb = load_workbook(str(out))
    aws = rwb.active
    # 占位符应被清空
    assert aws.cell(row=2, column=2).value is None