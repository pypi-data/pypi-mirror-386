import json
import pandas as pd
import io
import os
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from backlin.config.env import CachePathConfig


def bytes2human(n, format_str="%(value).1f%(symbol)s"):
    """Used by various scripts. See:
    http://goo.gl/zeJZl

    >>> bytes2human(10000)
    '9.8K'
    >>> bytes2human(100001221)
    '95.4M'
    """
    symbols = ('B', 'KB', 'MB', 'GB', 'TB', 'PB', 'EB', 'ZB', 'YB')
    prefix = {}
    for i, s in enumerate(symbols[1:]):
        prefix[s] = 1 << (i + 1) * 10
    for symbol in reversed(symbols[1:]):
        if n >= prefix[symbol]:
            value = float(n) / prefix[symbol]
            return format_str % locals()
    return format_str % dict(symbol=symbols[0], value=n)


def bytes2file_response(bytes_info):
    yield bytes_info


def export_list2excel(list_data: List):
    """
    工具方法：将需要导出的list数据转化为对应excel的二进制数据
    :param list_data: 数据列表
    :return: 字典信息对应excel的二进制数据
    """
    df = pd.DataFrame(list_data)
    binary_data = io.BytesIO()
    df.to_excel(binary_data, index=False, engine='openpyxl')
    binary_data = binary_data.getvalue()

    return binary_data


def get_jsonl_templage(json_object):

    # 保存Excel文件为字节类型的数据
    file = io.BytesIO()
    for i in range(3):
        json.dump(json_object, file, ensure_ascii=False, indent=2)
        file.write("\n")
    file.seek(0)

    # 读取字节数据
    excel_data = file.getvalue()

    return excel_data

def get_json_templage(json_object):

    # 保存Excel文件为字节类型的数据
    file = io.BytesIO()
    json.dump(json_object, file, ensure_ascii=False, indent=2)
    file.seek(0)

    # 读取字节数据
    excel_data = file.getvalue()

    return excel_data

def get_excel_template(header_list: List, selector_header_list: List, option_list: List[dict]):
    """
    工具方法：将需要导出的list数据转化为对应excel的二进制数据
    :param header_list: 表头数据列表
    :param selector_header_list: 需要设置为选择器格式的表头数据列表
    :param option_list: 选择器格式的表头预设的选项列表
    :return: 模板excel的二进制数据
    """
    # 创建Excel工作簿
    wb = Workbook()
    # 选择默认的活动工作表
    ws = wb.active

    # 设置表头文字
    headers = header_list

    # 设置表头背景样式为灰色，前景色为白色
    header_fill = PatternFill(start_color="ababab", end_color="ababab", fill_type="solid")

    # 将表头写入第一行
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        # 设置列宽度为16
        ws.column_dimensions[chr(64 + col_num)].width = 12
        # 设置水平居中对齐
        cell.alignment = Alignment(horizontal='center')

    # 设置选择器的预设选项
    options = option_list

    # 获取selector_header的字母索引
    for selector_header in selector_header_list:
        column_selector_header_index = headers.index(selector_header) + 1

        # 创建数据有效性规则
        header_option = []
        for option in options:
            if option.get(selector_header):
                header_option = option.get(selector_header)
        dv = DataValidation(type="list", formula1=f'"{",".join(header_option)}"')
        # 设置数据有效性规则的起始单元格和结束单元格
        dv.add(
            f'{get_column_letter(column_selector_header_index)}2:{get_column_letter(column_selector_header_index)}1048576')
        # 添加数据有效性规则到工作表
        ws.add_data_validation(dv)

    # 保存Excel文件为字节类型的数据
    file = io.BytesIO()
    wb.save(file)
    file.seek(0)

    # 读取字节数据
    excel_data = file.getvalue()

    return excel_data


def get_filepath_from_url(url: str):
    """
    工具方法:根据请求参数获取文件路径
    :param url: 请求参数中的url参数
    :return: 文件路径
    """
    file_info = url.split("?")[1].split("&")
    task_id = file_info[0].split("=")[1]
    file_name = file_info[1].split("=")[1]
    task_path = file_info[2].split("=")[1]
    cache_path = CachePathConfig.get_path()
    filepath = os.path.join(cache_path, task_path, task_id, file_name)

    return filepath
