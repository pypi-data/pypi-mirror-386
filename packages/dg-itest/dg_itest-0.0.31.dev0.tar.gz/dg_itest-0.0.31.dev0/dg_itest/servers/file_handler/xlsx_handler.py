#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/4/19 19:01
# @Author  : yaitza
# @Email   : yaitza@foxmail.com
import json
import traceback
from openpyxl import *
from dg_itest.utils.logger import logger

class XlsxHandler():
	def __init__(self, file_path):
		self.file_path = file_path

	def load(self):
		raw = []
		workbook = load_workbook(self.file_path)
		table = workbook.active
		# 抛出测试用例编写时各种的语法不规范错误
		error_row, error_col, error_type, error_value = 0, 0, None, None
		try:
			for row in range(2, table.max_row + 1):
				tester = {"name": None, "request": {}, "validate": [], "expect": None}
				for col in range(table.max_column):
					col_title = table.cell(1, col + 1).value
					error_row = row
					error_col = col
					error_type = col_title
					error_value = table.cell(row, col + 1).value
					if col_title in ['url', 'method', 'params']:
						request = tester.get("request")
						if col_title == "params":
							request_params = eval(table.cell(row, col + 1).value)
							request.update({table.cell(1, col + 1).value: request_params})
						else:
							request.update({table.cell(1, col + 1).value: table.cell(row, col + 1).value})
						tester.update({"request": request})
					elif col_title == "validate":
						if table.cell(row, col + 1).value is not None and table.cell(row, col + 1).value != '':
							tester.get("validate").append(json.loads(table.cell(row, col + 1).value))
					elif col_title == "expect":
						tester.update({table.cell(1, col + 1).value: json.loads(table.cell(row, col + 1).value)})
					else:
						tester.update({table.cell(1, col + 1).value: table.cell(row, col + 1).value})
				raw.append({"test": tester})
		except Exception:
			logger.error(f"Excel表格中第{error_row}行，第{error_col}列对应参数{error_type}的值有错，请检查:\n{error_value}\n")
			logger.error(traceback.format_exc())
			raise Exception("用例收集错误!")
		return raw
